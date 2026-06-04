#!/usr/bin/env python3
"""
Engineering Pipeline - IFR Sync + Version Management + Sharepoint Sync + Deliverable Cross-Check
工程文档自动化管线工具 v7.1

v7.1 New Features:
    - Sharepoint Sync stage: auto-sync IFR(Client) → Client Sharepoint for non-approved files
      Skips files whose doc-ID is already in Approved to IFC/, or already present in target

v7.0 New Features (merged pipeline):
    - Merged ifr_automation_v6 + version_manager_v4 into single script
    - Deliverable cross-check: auto-detect Excel layout, compare files vs Excel,
      auto-insert new items, auto-update revisions, highlight changes
    - Pipeline orchestrator: sequential IFR Sync → Version Mgmt → Sharepoint Sync → Deliverable per project
    - Unified interactive menu with 10 options
    - CLI: --pipeline, --deliverable-only, --deliverable-check-only, --stages

v6.0 Features:
    - IFC/IFR file classification: prevents IFC files from IFR(Client) directories
v5.0 Features:
    - Folder-based file classification fallback (any naming convention)
v4.0 Features:
    - IFC(Client) folder copy to external target
    - Version manager: PDF old version cleanup, Native/Reports/Schedule management
    - Folder relocator: detect/fix misplaced structural folders
v3.0 Features:
    - Smart project validation, hierarchical scanning, multi-source directories
    - Whitelist/Blacklist, safety checker, comprehensive validation reports

Usage:
    # Interactive mode (recommended)
    python ifr_automation_v7.py

    # Full pipeline for all projects
    python ifr_automation_v7.py --pipeline

    # Deliverable cross-check only
    python ifr_automation_v7.py --deliverable-only

    # Specific stages
    python ifr_automation_v7.py --stages ifr_sync version_mgmt sharepoint_sync deliverable

    # Validate only (no changes)
    python ifr_automation_v7.py --validate-only

Author: Generated with Claude Code
Date: 2026-03-11
Version: 7.0
"""

import os
import sys
import argparse
import logging
import json
import re
import shutil
import fnmatch
import time
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional, Set
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from itertools import groupby

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

try:
    from colorama import init, Fore, Style, Back
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import requests as _requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

try:
    import win32com.client
    import pythoncom
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False


# =============================================================================
# Telegram Notification Helper
# =============================================================================

def send_telegram_notification(message: str, parse_mode: str = "HTML") -> bool:
    """Send a notification message to the configured Telegram chat.

    Reads TELEGRAM_TOKEN and TELEGRAM_CHAT_ID from the .env file at:
      D:/1. SOP/SOP_Project Status&Tasks√/V3 Manul&Auto CSV to excel/Automatic Export√/.env

    Returns True if sent successfully, False otherwise (silently fails).
    """
    if not REQUESTS_AVAILABLE:
        return False

    env_path = Path(r"D:\1. SOP\SOP_Project Status&Tasks√\V3 Manul&Auto CSV to excel\Automatic Export√\.env")
    if not env_path.exists():
        return False

    token = chat_id = None
    try:
        for line in env_path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if line.startswith('TELEGRAM_TOKEN='):
                token = line.split('=', 1)[1].strip()
            elif line.startswith('TELEGRAM_CHAT_ID='):
                chat_id = line.split('=', 1)[1].strip()
    except Exception:
        return False

    if not token or not chat_id:
        return False

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = _requests.post(url, json={
            "chat_id": chat_id,
            "text": message,
            "parse_mode": parse_mode,
        }, timeout=10)
        return resp.status_code == 200
    except Exception:
        return False


def format_pipeline_result(results: Dict) -> str:
    """Format a pipeline result dict into a Telegram-friendly message."""
    project = results.get('project_name', 'Unknown')
    success = results.get('success', False)
    icon = "✅" if success else "❌"

    lines = [f"{icon} <b>Pipeline: {project}</b>"]

    ifr = results.get('ifr_sync')
    if ifr:
        if ifr.get('success'):
            lines.append(f"  📄 IFR Sync: 文件夹={ifr.get('folders',0)}, "
                         f"图纸={ifr.get('drawings',0)}, 报告={ifr.get('reports',0)}")
        elif 'error' in ifr:
            lines.append(f"  ❌ IFR Sync: {ifr['error']}")

    vm = results.get('version_mgmt')
    if vm and not vm.get('error'):
        lines.append(f"  🗂 Version Mgmt: 扫描={vm.get('scanned',0)}, 移动={vm.get('moved',0)}")
    elif vm and vm.get('error'):
        lines.append(f"  ❌ Version Mgmt: {vm['error']}")

    sp = results.get('sharepoint_sync')
    if sp and not sp.get('error'):
        sp_parts = [f"复制={sp.get('copied',0)}"]
        if sp.get('archived', 0) > 0:
            sp_parts.append(f"归档={sp['archived']}")
        sp_parts.append(f"已审批跳过={sp.get('skipped_approved',0)}")
        lines.append(f"  📤 Sharepoint Sync: {', '.join(sp_parts)}")
    elif sp and sp.get('error'):
        lines.append(f"  ❌ Sharepoint Sync: {sp['error']}")

    dlv = results.get('deliverable')
    if dlv and not dlv.get('error') and not dlv.get('skipped'):
        lines.append(f"  📋 Deliverable: 新增={dlv.get('new_items',0)}, "
                     f"更新={dlv.get('rev_mismatches',0)}, "
                     f"插入={dlv.get('inserted',0)}, 更新={dlv.get('updated',0)}")
    elif dlv and dlv.get('skipped'):
        lines.append(f"  ⚠️ Deliverable: {dlv.get('reason','skipped')}")
    elif dlv and dlv.get('error'):
        lines.append(f"  ❌ Deliverable: {dlv['error']}")

    return "\n".join(lines)


# =============================================================================
# Utility Functions
# =============================================================================

def to_long_path(path: Path) -> Path:
    """Convert path to long path format on Windows to support paths > 260 chars."""
    if os.name == 'nt':
        path_str = str(path.resolve())
        if not path_str.startswith('\\\\?\\'):
            path_str = '\\\\?\\' + path_str
        return Path(path_str)
    return path


# IFC/IFR classification patterns
_RE_IFC_FILE = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?:[_\s-]?IFC)?(?=[_.\s]|$)', re.IGNORECASE)
_RE_IFR_FILE = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*([A-Z])(?=[_.\s]|$)', re.IGNORECASE)


def _is_ifc_file(filename: str) -> bool:
    """判断文件是否为 IFC 文件（数字版本号）"""
    stem = Path(filename).stem
    if _RE_IFC_FILE.search(stem):
        return True
    if 'IFC' in stem.upper():
        return True
    return False


def _is_ifr_file(filename: str) -> bool:
    """判断文件是否为 IFR 文件（字母版本号）"""
    stem = Path(filename).stem
    return bool(_RE_IFR_FILE.search(stem))


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class SourceDirectory:
    """Information about a source directory."""
    name: str
    path: str
    exists: bool = False
    file_count: int = 0
    files: List[str] = field(default_factory=list)
    priority: int = 1


@dataclass
class ProjectValidation:
    """Validation result for a project."""
    valid: bool
    project_name: str
    project_path: str
    region: str = ""
    missing_dirs: List[str] = field(default_factory=list)
    source_dirs_found: List[SourceDirectory] = field(default_factory=list)
    has_legacy_structure: bool = False
    has_existing_ifr_client: bool = False
    has_existing_ifc_client: bool = False
    ifc_file_count: int = 0
    last_modified: str = ""
    days_since_modified: int = 0
    recommended_action: str = "skip"  # process/confirm/skip
    warning_message: str = ""
    drawings_count: int = 0
    reports_count: int = 0
    folders_to_create: int = 0

    def to_dict(self) -> Dict:
        result = asdict(self)
        result['source_dirs_found'] = [asdict(s) for s in self.source_dirs_found]
        return result


@dataclass
class ProcessResult:
    """Result of processing a single project."""
    project_name: str
    project_path: str
    success: bool
    folders_created: int = 0
    drawings_copied: int = 0
    drawings_skipped: int = 0
    reports_copied: int = 0
    reports_skipped: int = 0
    ifc_files_copied: int = 0
    ifc_files_skipped: int = 0
    ifc_target_path: str = ""
    errors: List[str] = field(default_factory=list)
    skipped: bool = False
    skip_reason: str = ""


# =============================================================================
# Configuration Management
# =============================================================================

class ConfigManager:
    """Manages configuration file for IFR automation."""

    DEFAULT_CONFIG = {
        "default_root_path": "",
        "recent_projects": [],
        "auto_backup": True,
        "skip_existing_files": True,
        "show_progress_bar": True,
        "log_level": "INFO",
        "color_output": True,
        "max_recent_projects": 10,
        "last_operation": None,
        "last_run_time": None,

        # V3: Project filters
        "project_filters": {
            "whitelist": [],
            "blacklist": [],
            "auto_process_only_validated": True,
            "require_confirmation_for_legacy": True
        },

        # V3: Drawing sources configuration
        "drawing_sources": [
            {
                "path": "Design/Engineering/1. Drawings/2. IFR_internal",
                "enabled": True,
                "priority": 1,
                "description": "Internal drawings directory (primary source)"
            },
            {
                "path": "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
                "enabled": True,
                "priority": 2,
                "description": "Civil reports directory (some drawings here)"
            }
        ],

        # V3: Validation rules
        "validation_rules": {
            "require_ifr_internal": False,
            "require_at_least_one_source": True,
            "min_source_dirs": 1,
            "max_days_since_modified": 180  # 6 months
        },

        # V4: IFC(Client) source path (relative to project root)
        "ifc_client_path": "Design/Engineering/1. Drawings/4. IFC(Client)"
    }

    def __init__(self, config_path: Optional[Path] = None):
        if config_path is None:
            self.config_path = Path(__file__).parent / "config.json"
        else:
            self.config_path = Path(config_path)
        self.config = self.load()

    def load(self) -> Dict:
        if self.config_path.exists():
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                config = self._deep_merge(self.DEFAULT_CONFIG.copy(), loaded)
                return config
            except Exception as e:
                print(f"Warning: Could not load config: {e}")
                return self.DEFAULT_CONFIG.copy()
        else:
            return self.DEFAULT_CONFIG.copy()

    def _deep_merge(self, base: Dict, override: Dict) -> Dict:
        """Deep merge two dictionaries."""
        result = base.copy()
        for key, value in override.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = self._deep_merge(result[key], value)
            else:
                result[key] = value
        return result

    def save(self):
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Warning: Could not save config: {e}")

    def get(self, key: str, default=None):
        keys = key.split('.')
        value = self.config
        for k in keys:
            if isinstance(value, dict) and k in value:
                value = value[k]
            else:
                return default
        return value

    def set(self, key: str, value):
        keys = key.split('.')
        config = self.config
        for k in keys[:-1]:
            if k not in config:
                config[k] = {}
            config = config[k]
        config[keys[-1]] = value
        self.save()

    def add_recent_project(self, project_path: str):
        recent = self.config.get("recent_projects", [])
        if project_path in recent:
            recent.remove(project_path)
        recent.insert(0, project_path)
        max_recent = self.config.get("max_recent_projects", 10)
        self.config["recent_projects"] = recent[:max_recent]
        self.save()


# =============================================================================
# UI Utilities
# =============================================================================

class UIHelper:
    """Helper class for console UI elements."""

    @staticmethod
    def clear_screen():
        os.system('cls' if os.name == 'nt' else 'clear')

    @staticmethod
    def colorize(text: str, color: str, bold: bool = False) -> str:
        if not COLORAMA_AVAILABLE:
            return text
        color_map = {
            "green": Fore.GREEN,
            "red": Fore.RED,
            "yellow": Fore.YELLOW,
            "blue": Fore.BLUE,
            "cyan": Fore.CYAN,
            "magenta": Fore.MAGENTA,
            "white": Fore.WHITE,
        }
        prefix = color_map.get(color, "")
        if bold:
            prefix = Style.BRIGHT + prefix
        return f"{prefix}{text}{Style.RESET_ALL}"

    @staticmethod
    def print_header(title: str, subtitle: str = ""):
        width = 72
        print()
        if COLORAMA_AVAILABLE:
            print(Fore.CYAN + "+" + "=" * width + "+")
            print(Fore.CYAN + "|" + Style.BRIGHT + Fore.WHITE + title.center(width) + Style.RESET_ALL + Fore.CYAN + "|")
            if subtitle:
                print(Fore.CYAN + "|" + Fore.WHITE + subtitle.center(width) + Fore.CYAN + "|")
            print(Fore.CYAN + "+" + "=" * width + "+" + Style.RESET_ALL)
        else:
            print("+" + "=" * width + "+")
            print("|" + title.center(width) + "|")
            if subtitle:
                print("|" + subtitle.center(width) + "|")
            print("+" + "=" * width + "+")
        print()

    @staticmethod
    def print_separator(char: str = "-", width: int = 72):
        if COLORAMA_AVAILABLE:
            print(Fore.CYAN + char * width + Style.RESET_ALL)
        else:
            print(char * width)

    @staticmethod
    def print_menu(options: List[Tuple[str, str]], prompt: str = "请输入选项") -> str:
        print()
        for key, label in options:
            if COLORAMA_AVAILABLE:
                print(f"  {Fore.YELLOW}[{key}]{Style.RESET_ALL} {label}")
            else:
                print(f"  [{key}] {label}")
        print()
        UIHelper.print_separator()
        return input(f"\n{prompt}: ").strip()

    @staticmethod
    def print_success(message: str):
        symbol = "[OK]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.GREEN}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def print_warning(message: str):
        symbol = "[!]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.YELLOW}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def print_error(message: str):
        symbol = "[X]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.RED}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def print_info(message: str):
        symbol = "[i]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.CYAN}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def confirm(message: str, default: bool = True) -> bool:
        suffix = " [Y/n]: " if default else " [y/N]: "
        response = input(message + suffix).strip().lower()
        if not response:
            return default
        return response in ('y', 'yes')

    @staticmethod
    def status_icon(status: str) -> str:
        """Get status icon with color."""
        icons = {
            "ready": (Fore.GREEN + "[v]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[v]", "完整"),
            "confirm": (Fore.YELLOW + "[!]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[!]", "需确认"),
            "legacy": (Fore.RED + "[x]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[x]", "旧结构"),
            "skip": (Fore.RED + "[-]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[-]", "跳过"),
        }
        return icons.get(status, ("[?]", "未知"))

    @staticmethod
    def action_icon(action: str) -> str:
        """Get action recommendation icon."""
        if COLORAMA_AVAILABLE:
            icons = {
                "process": Fore.GREEN + "[SAFE]" + Style.RESET_ALL + " 可以安全处理",
                "confirm": Fore.YELLOW + "[WARN]" + Style.RESET_ALL + " 需要人工确认",
                "skip": Fore.RED + "[SKIP]" + Style.RESET_ALL + " 建议跳过",
            }
        else:
            icons = {
                "process": "[SAFE] 可以安全处理",
                "confirm": "[WARN] 需要人工确认",
                "skip": "[SKIP] 建议跳过",
            }
        return icons.get(action, "[?] 未知")


# =============================================================================
# Project Validator
# =============================================================================

class ProjectValidator:
    """Project validator - ensures only valid project structures are processed."""

    # Required directory structure
    REQUIRED_STRUCTURE = [
        "Design/Engineering/1. Drawings",
        "Design/Engineering/2. Calcs & Reports"
    ]

    # Source directories to check
    SOURCE_DIRECTORIES = [
        {"path": "Design/Engineering/1. Drawings/2. IFR_internal", "name": "IFR_internal", "priority": 1},
        {"path": "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure", "name": "Civil Reports", "priority": 2},
        {"path": "Design/Engineering/2. Calcs & Reports/Reports/Electrical", "name": "Electrical Reports", "priority": 2},
        {"path": "Design/Engineering/2. Calcs & Reports/Schedule", "name": "Schedule", "priority": 3},
    ]

    # Drawing patterns
    DRAWING_PATTERNS = [
        r'-C-PLN-\d+', r'-C-GEN-\d+', r'-C-SEC-\d+',
        r'-E-PLN-\d+', r'-E-SLD-\d+', r'-E-BLD-\d+',
        r'-E-CFG-\d+', r'-E-GAD-\d+',
    ]

    # Report patterns
    REPORT_PATTERNS = [
        r'-C-RPT-', r'-E-RPT-', r'-E-SCH-',
    ]

    # V4: IFC(Client) path
    IFC_CLIENT_REL_PATH = "Design/Engineering/1. Drawings/4. IFC(Client)"

    def __init__(self, config: ConfigManager):
        self.config = config
        self.validation_rules = config.get("validation_rules", {})
        self.ifc_client_rel_path = config.get("ifc_client_path", self.IFC_CLIENT_REL_PATH)

    def validate_project(self, project_path: Path, region: str = "") -> ProjectValidation:
        """Validate a project's structure and return detailed validation result."""
        validation = ProjectValidation(
            valid=False,
            project_name=project_path.name,
            project_path=str(project_path),
            region=region
        )

        # Check basic structure
        missing_dirs = []
        for required in self.REQUIRED_STRUCTURE:
            if not (project_path / required).exists():
                missing_dirs.append(required)

        validation.missing_dirs = missing_dirs

        if missing_dirs:
            validation.has_legacy_structure = True
            validation.warning_message = f"缺少必要目录: {', '.join(missing_dirs)}"
            validation.recommended_action = "skip"
            return validation

        # Check source directories
        source_dirs_found = []
        total_drawings = 0
        total_reports = 0

        for source in self.SOURCE_DIRECTORIES:
            source_path = project_path / source["path"]
            source_dir = SourceDirectory(
                name=source["name"],
                path=source["path"],
                exists=source_path.exists(),
                priority=source["priority"]
            )

            if source_path.exists():
                # Count files (use rglob to search subdirectories)
                pdf_files = list(source_path.rglob("*.pdf"))
                source_dir.file_count = len(pdf_files)
                source_dir.files = [f.name for f in pdf_files]

                # Categorize files using folder-based fallback
                is_drawing_source = "1. Drawings" in source["path"]
                is_report_source = "2. Calcs & Reports" in source["path"]
                for pdf_file in pdf_files:
                    name = pdf_file.name
                    if self._is_drawing(name):
                        total_drawings += 1
                    elif self._is_report(name):
                        total_reports += 1
                    elif is_drawing_source:
                        total_drawings += 1
                    elif is_report_source:
                        # Skip files in non-deliverable subdirs
                        try:
                            rel_parts = pdf_file.relative_to(source_path).parts[:-1]
                            in_excluded = any(
                                part.lower() in ('ss', 'superseded', '_export', 'data')
                                or part.lower().startswith(('appendix', 'stk'))
                                for part in rel_parts
                            )
                        except Exception:
                            rel_parts = ()
                            in_excluded = False
                        if in_excluded:
                            pass
                        elif rel_parts:
                            # Parent prefix check for subdirectory files
                            parent_name = rel_parts[0]
                            parent_prefix = parent_name.split('_')[0]
                            if '_' in parent_name and name.startswith(parent_prefix):
                                total_reports += 1
                        else:
                            total_reports += 1

            source_dirs_found.append(source_dir)

        validation.source_dirs_found = source_dirs_found
        validation.drawings_count = total_drawings
        validation.reports_count = total_reports

        # Check for existing IFR(Client)
        ifr_client = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)"
        validation.has_existing_ifr_client = ifr_client.exists()

        # V4: Check for existing IFC(Client)
        ifc_client = project_path / self.ifc_client_rel_path
        validation.has_existing_ifc_client = ifc_client.exists()
        if ifc_client.exists():
            try:
                validation.ifc_file_count = sum(1 for _ in ifc_client.rglob("*") if _.is_file())
            except Exception:
                validation.ifc_file_count = 0

        # Calculate folders to create
        if not validation.has_existing_ifr_client:
            validation.folders_to_create = 7  # Main + 3 subs + 3 SS folders
        else:
            # Count missing subfolders
            subfolders = ["1.Drawing/SS", "2.Reports/SS", "3.Deliverables/SS"]
            validation.folders_to_create = sum(
                1 for sf in subfolders if not (ifr_client / sf).exists()
            )

        # Get last modified time
        try:
            mtime = datetime.fromtimestamp(project_path.stat().st_mtime)
            validation.last_modified = mtime.strftime("%Y-%m-%d")
            validation.days_since_modified = (datetime.now() - mtime).days
        except:
            validation.last_modified = "未知"
            validation.days_since_modified = 999

        # Determine recommendation
        active_sources = sum(1 for s in source_dirs_found if s.exists and s.file_count > 0)
        min_sources = self.validation_rules.get("min_source_dirs", 1)
        max_days = self.validation_rules.get("max_days_since_modified", 180)

        if active_sources >= min_sources:
            if validation.days_since_modified > max_days:
                validation.valid = True
                validation.recommended_action = "confirm"
                validation.warning_message = f"项目超过 {max_days} 天未修改，可能是旧项目"
            elif validation.has_existing_ifr_client:
                validation.valid = True
                validation.recommended_action = "confirm"
                validation.warning_message = "目标目录已存在，可能会覆盖现有内容"
            else:
                validation.valid = True
                validation.recommended_action = "process"
        else:
            validation.has_legacy_structure = True
            validation.recommended_action = "skip"
            validation.warning_message = f"有效源目录不足 (找到 {active_sources}, 需要 {min_sources})"

        return validation

    def _is_drawing(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.DRAWING_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                # Exclude reports
                for exclude in self.REPORT_PATTERNS:
                    if re.search(exclude, filename, re.IGNORECASE):
                        return False
                return True
        return False

    def _is_report(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.REPORT_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return True
        return False


# =============================================================================
# Safety Checker
# =============================================================================

class SafetyChecker:
    """Safety checker - prevents accidental operations."""

    def __init__(self, config: ConfigManager):
        self.config = config

    def check_before_process(self, validation: ProjectValidation) -> Dict:
        """Perform safety checks before processing."""
        warnings = []
        checks = {
            "has_recent_activity": validation.days_since_modified < 180,
            "has_existing_ifr_client": validation.has_existing_ifr_client,
            "source_dirs_not_empty": validation.drawings_count > 0 or validation.reports_count > 0,
            "structure_complete": not validation.has_legacy_structure
        }

        if validation.has_existing_ifr_client:
            warnings.append("目标目录已存在文件，可能会覆盖现有内容")

        if not checks["has_recent_activity"]:
            warnings.append(f"项目超过 {validation.days_since_modified} 天未修改，可能是旧项目")

        if not checks["source_dirs_not_empty"]:
            warnings.append("源目录中没有找到任何文件")

        if not checks["structure_complete"]:
            warnings.append("项目结构不完整")

        return {
            "safe": len(warnings) == 0,
            "warnings": warnings,
            "checks": checks
        }


# =============================================================================
# Project Scanner
# =============================================================================

class ProjectScanner:
    """Hierarchical project scanner."""

    # Known region prefixes
    REGION_PATTERNS = [
        r'^(\d+)\.',      # 1.NSW, 2.SA
        r'^([A-Z]{2,3})', # NSW, SA, VIC
    ]

    def __init__(self, config: ConfigManager, validator: ProjectValidator):
        self.config = config
        self.validator = validator

    def scan_hierarchical(self, root_path: Path) -> Dict[str, List[ProjectValidation]]:
        """Scan projects hierarchically by region."""
        projects_by_region = defaultdict(list)

        if not root_path.exists():
            return projects_by_region

        # Get whitelist and blacklist
        filters = self.config.get("project_filters", {})
        whitelist = filters.get("whitelist", [])
        blacklist = filters.get("blacklist", [])

        # First level: look for regions or direct projects
        for item in root_path.iterdir():
            if not item.is_dir():
                continue

            # Check if this is a region folder
            if self._is_region_folder(item):
                region_name = item.name
                # Scan projects within region
                for project_dir in item.iterdir():
                    if not project_dir.is_dir():
                        continue

                    # Apply whitelist/blacklist
                    if not self._passes_filter(project_dir.name, whitelist, blacklist):
                        continue

                    # Check if it's a valid project
                    if self._is_project_folder(project_dir):
                        validation = self.validator.validate_project(project_dir, region_name)
                        projects_by_region[region_name].append(validation)

            # Check if this is a direct project folder
            elif self._is_project_folder(item):
                if not self._passes_filter(item.name, whitelist, blacklist):
                    continue

                validation = self.validator.validate_project(item, "Root")
                projects_by_region["Root"].append(validation)

        # Include additional (non-standard structure) projects from config
        additional = self.config.get("additional_projects", [])
        already_paths = {v.project_path for vals in projects_by_region.values() for v in vals}
        for entry in additional:
            proj_path = Path(entry["path"])
            if str(proj_path) in already_paths or not proj_path.exists():
                continue
            region = entry.get("region", "Other")
            validation = ProjectValidation(
                valid=True,
                project_name=proj_path.name,
                project_path=str(proj_path),
                region=region,
            )
            validation.source_dirs_found = []
            # Auto-detect drawings root for non-standard structure
            drawings_root = entry.get("drawings_root", "1. Drawings")
            drw_path = proj_path / drawings_root
            if drw_path.exists():
                sd = SourceDirectory(
                    name=drawings_root,
                    path=drawings_root,
                    exists=True,
                    priority=1,
                )
                pdf_files = list(drw_path.rglob("*.pdf"))
                sd.file_count = len(pdf_files)
                sd.files = [f.name for f in pdf_files[:50]]
                validation.source_dirs_found.append(sd)
            projects_by_region[region].append(validation)

        # Sort projects within each region
        for region in projects_by_region:
            projects_by_region[region].sort(key=lambda x: x.project_name)

        return dict(projects_by_region)

    def _is_region_folder(self, path: Path) -> bool:
        """Check if folder is a region container (like 1.NSW, 2.SA)."""
        name = path.name

        # Check if matches region patterns
        for pattern in self.REGION_PATTERNS:
            if re.match(pattern, name):
                # Check if it contains project-like subfolders
                for subdir in path.iterdir():
                    if subdir.is_dir() and self._is_project_folder(subdir):
                        return True

        return False

    def _is_project_folder(self, path: Path) -> bool:
        """Check if folder is a project (has Design/Engineering structure)."""
        design_path = path / "Design" / "Engineering"
        return design_path.exists()

    def _passes_filter(self, name: str, whitelist: List[str], blacklist: List[str]) -> bool:
        """Check if project name passes whitelist/blacklist filters."""
        # If whitelist is set, only include matching projects
        if whitelist:
            if not any(fnmatch.fnmatch(name, pattern) for pattern in whitelist):
                return False

        # Exclude blacklisted projects
        if blacklist:
            if any(fnmatch.fnmatch(name, pattern) for pattern in blacklist):
                return False

        return True


# =============================================================================
# Drawing Collector
# =============================================================================

class DrawingCollector:
    """Collects drawings from multiple source directories."""

    DRAWING_PATTERNS = [
        r'-C-PLN-\d+', r'-C-GEN-\d+', r'-C-SEC-\d+',
        r'-E-PLN-\d+', r'-E-SLD-\d+', r'-E-BLD-\d+',
        r'-E-CFG-\d+', r'-E-GAD-\d+',
    ]

    EXCLUDE_PATTERNS = [
        r'-C-RPT-', r'-E-RPT-', r'-E-SCH-',
    ]

    def __init__(self, config: ConfigManager):
        self.config = config
        self.sources = config.get("drawing_sources", [])

    def collect_drawings(self, project_path: Path) -> Dict:
        """Collect drawings from all configured source directories."""
        result = {
            "files": [],
            "total_count": 0,
            "by_source": {},
            "duplicates_removed": 0
        }

        seen_files = set()

        for source in self.sources:
            if not source.get("enabled", True):
                continue

            source_path = project_path / source["path"]
            if not source_path.exists():
                continue

            source_name = Path(source["path"]).name
            source_count = 0
            is_drawing_source = "1. Drawings" in source["path"]

            try:
                pdf_files = list(source_path.glob("*.pdf"))
            except Exception as e:
                continue

            for pdf_file in pdf_files:
                if not self._should_collect_as_drawing(pdf_file.name, is_drawing_source):
                    continue

                # Check for duplicates
                if pdf_file.name in seen_files:
                    result["duplicates_removed"] += 1
                    continue

                seen_files.add(pdf_file.name)

                try:
                    # Use long path for stat operations
                    pdf_file_long = to_long_path(pdf_file)
                    file_info = {
                        "filename": pdf_file.name,
                        "source_path": str(pdf_file_long),
                        "source_dir": source_name,
                        "size": pdf_file_long.stat().st_size,
                        "modified": datetime.fromtimestamp(pdf_file_long.stat().st_mtime).strftime("%Y-%m-%d")
                    }
                    result["files"].append(file_info)
                    source_count += 1
                except Exception:
                    # Fallback without long path
                    try:
                        file_info = {
                            "filename": pdf_file.name,
                            "source_path": str(pdf_file),
                            "source_dir": source_name,
                            "size": 0,
                            "modified": "unknown"
                        }
                        result["files"].append(file_info)
                        source_count += 1
                    except Exception:
                        pass

            result["by_source"][source_name] = source_count

        result["total_count"] = len(result["files"])
        return result

    def _should_collect_as_drawing(self, filename: str, is_drawing_source: bool) -> bool:
        """Check if file should be collected as a drawing.

        Uses pattern matching first, falls back to folder-based classification
        for projects with unrecognized naming conventions (e.g. 50023-xx, TFS-xx).
        """
        if not filename.lower().endswith('.pdf'):
            return False
        # Pattern match: explicitly a drawing
        if self._is_drawing(filename):
            return True
        # Pattern match: explicitly a report → exclude
        for pattern in self.EXCLUDE_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return False
        # No pattern match: trust folder structure
        return is_drawing_source

    def _is_drawing(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.DRAWING_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                for exclude in self.EXCLUDE_PATTERNS:
                    if re.search(exclude, filename, re.IGNORECASE):
                        return False
                return True
        return False


# =============================================================================
# Report Generator
# =============================================================================

class ValidationReportGenerator:
    """Generates validation reports."""

    def generate_report(self, projects_by_region: Dict[str, List[ProjectValidation]],
                       root_path: str) -> str:
        """Generate a detailed validation report."""
        lines = []
        lines.append("=" * 70)
        lines.append("项目验证报告")
        lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"扫描根目录: {root_path}")
        lines.append("=" * 70)
        lines.append("")

        # Summary
        total = sum(len(projects) for projects in projects_by_region.values())
        ready = sum(1 for projects in projects_by_region.values()
                   for p in projects if p.recommended_action == "process")
        confirm = sum(1 for projects in projects_by_region.values()
                     for p in projects if p.recommended_action == "confirm")
        skip = sum(1 for projects in projects_by_region.values()
                  for p in projects if p.recommended_action == "skip")

        lines.append(f"总计: {total} 个项目")
        lines.append(f"  [v] 可处理: {ready} 个")
        lines.append(f"  [!] 需确认: {confirm} 个")
        lines.append(f"  [x] 建议跳过: {skip} 个")
        lines.append("")

        # Ready projects
        if ready > 0:
            lines.append("-" * 70)
            lines.append(f"[v] 可以安全处理的项目 ({ready})")
            lines.append("-" * 70)
            lines.append("")

            idx = 1
            for region, projects in projects_by_region.items():
                for p in projects:
                    if p.recommended_action == "process":
                        lines.extend(self._format_project(idx, p))
                        idx += 1

        # Confirm projects
        if confirm > 0:
            lines.append("-" * 70)
            lines.append(f"[!] 需要人工确认的项目 ({confirm})")
            lines.append("-" * 70)
            lines.append("")

            idx = 1
            for region, projects in projects_by_region.items():
                for p in projects:
                    if p.recommended_action == "confirm":
                        lines.extend(self._format_project(idx, p))
                        idx += 1

        # Skip projects
        if skip > 0:
            lines.append("-" * 70)
            lines.append(f"[x] 建议跳过的项目 ({skip})")
            lines.append("-" * 70)
            lines.append("")

            idx = 1
            for region, projects in projects_by_region.items():
                for p in projects:
                    if p.recommended_action == "skip":
                        lines.extend(self._format_project(idx, p))
                        idx += 1

        return "\n".join(lines)

    def _format_project(self, idx: int, p: ProjectValidation) -> List[str]:
        lines = []
        lines.append(f"{idx}. {p.project_name}")
        lines.append(f"   路径: {p.project_path}")
        lines.append(f"   区域: {p.region}")
        lines.append(f"   状态: {'完整结构' if p.valid else '不完整结构'}")

        if p.warning_message:
            lines.append(f"   警告: {p.warning_message}")

        lines.append("   源目录:")
        for source in p.source_dirs_found:
            status = "[v]" if source.exists and source.file_count > 0 else "[x]"
            count = f"({source.file_count} 个文件)" if source.exists else "(不存在)"
            lines.append(f"     {status} {source.name} {count}")

        # V4: Show IFC(Client) status
        if p.has_existing_ifc_client:
            lines.append(f"   IFC(Client): [v] 存在 ({p.ifc_file_count} 个文件)")
        else:
            lines.append(f"   IFC(Client): [x] 不存在")

        if p.recommended_action == "process":
            lines.append("   预计操作:")
            lines.append(f"     - 创建 {p.folders_to_create} 个文件夹")
            lines.append(f"     - 复制 {p.drawings_count} 个图纸")
            lines.append(f"     - 复制 {p.reports_count} 个报告")

        lines.append("")
        return lines

    def save_report(self, content: str, path: Path):
        """Save report to file."""
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)


# =============================================================================
# Main IFR Automation Class
# =============================================================================

class IFRAutomation:
    """Main IFR automation class with v4 features."""

    IFR_STRUCTURE = {
        "1.Drawing": ["SS"],
        "2.Reports": ["SS"],
        "3.Deliverables": ["SS"]
    }

    DRAWING_PATTERNS = [
        r'-C-PLN-\d+', r'-C-GEN-\d+', r'-C-SEC-\d+',
        r'-E-PLN-\d+', r'-E-SLD-\d+', r'-E-BLD-\d+',
        r'-E-CFG-\d+', r'-E-GAD-\d+',
    ]

    REPORT_PATTERNS = [
        r'-C-RPT-', r'-E-RPT-', r'-E-SCH-',
    ]

    # V4: IFC(Client) source path (relative to project root)
    IFC_CLIENT_REL_PATH = "Design/Engineering/1. Drawings/4. IFC(Client)"

    def __init__(self, root_path: str, config: ConfigManager,
                 dry_run: bool = False, create_folders_only: bool = False,
                 mirror_files_only: bool = False, validate_only: bool = False,
                 yes_to_all: bool = False, auto_safe_only: bool = False,
                 log_level: str = "INFO", output_json: bool = False,
                 interactive: bool = False,
                 excel_doc_ids: Optional[Set[str]] = None):

        self.root_path = Path(root_path).resolve()
        self.config = config
        self.dry_run = dry_run
        self.create_folders_only = create_folders_only
        self.mirror_files_only = mirror_files_only
        self.validate_only = validate_only
        self.yes_to_all = yes_to_all
        self.auto_safe_only = auto_safe_only
        self.log_level = log_level
        self.output_json = output_json
        self.interactive = interactive
        self.excel_doc_ids = excel_doc_ids or set()

        self.validator = ProjectValidator(config)
        self.scanner = ProjectScanner(config, self.validator)
        self.safety_checker = SafetyChecker(config)
        self.drawing_collector = DrawingCollector(config)
        self.report_generator = ValidationReportGenerator()

        self.ifc_client_rel_path = config.get("ifc_client_path", self.IFC_CLIENT_REL_PATH)

        self.results: List[ProcessResult] = []
        self._setup_logging()

    def _setup_logging(self):
        log_dir = Path(__file__).parent / "logs"
        log_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f"ifr_automation_{timestamp}.log"

        log_format = '%(asctime)s - %(levelname)s - %(message)s'

        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter(log_format))

        console_handler = logging.StreamHandler(sys.stdout)
        if self.interactive:
            console_handler.setLevel(logging.WARNING)
        else:
            console_handler.setLevel(getattr(logging, self.log_level.upper()))
        console_handler.setFormatter(logging.Formatter('%(levelname)s - %(message)s'))

        self.logger = logging.getLogger(f'IFRAutomation_{timestamp}')
        self.logger.setLevel(logging.DEBUG)
        self.logger.handlers.clear()
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)

        self.log_file = log_file

    def scan_projects(self) -> Dict[str, List[ProjectValidation]]:
        """Scan all projects hierarchically."""
        return self.scanner.scan_hierarchical(self.root_path)

    def process_project(self, validation: ProjectValidation) -> ProcessResult:
        """Process a single validated project."""
        project_path = Path(validation.project_path)
        result = ProcessResult(
            project_name=validation.project_name,
            project_path=validation.project_path,
            success=False
        )

        try:
            # Create folder structure
            if not self.mirror_files_only:
                folders_created = self._create_folder_structure(project_path)
                result.folders_created = folders_created

            # Mirror files
            if not self.create_folders_only:
                drawings = self._mirror_drawings(project_path)
                result.drawings_copied = drawings["copied"]
                result.drawings_skipped = drawings["skipped"]

                reports = self._mirror_reports(project_path)
                result.reports_copied = reports["copied"]
                result.reports_skipped = reports["skipped"]

            result.success = True

        except Exception as e:
            result.errors.append(str(e))
            self.logger.error(f"Error processing {validation.project_name}: {e}")

        return result

    def copy_ifc_to_target(self, project_path: Path, target_path: Path) -> Dict[str, int]:
        """Copy the entire 4. IFC(Client) folder to the target path.

        Copies all files and subdirectories preserving folder structure.
        Uses copy (not move). Skips files that already exist with same size/date.

        Returns dict with copied/skipped/error counts.
        """
        result = {"copied": 0, "skipped": 0, "errors": []}

        ifc_source = project_path / self.ifc_client_rel_path
        ifc_source_long = self._to_long_path(ifc_source)

        if not ifc_source_long.exists():
            result["errors"].append(f"IFC(Client) 源目录不存在: {ifc_source}")
            return result

        target_long = self._to_long_path(target_path)

        # Create target directory if it doesn't exist
        if not self.dry_run:
            target_long.mkdir(parents=True, exist_ok=True)

        # Walk through all files in source and copy preserving structure
        try:
            source_files = list(ifc_source_long.rglob("*"))
        except Exception as e:
            result["errors"].append(f"无法扫描 IFC(Client) 目录: {e}")
            return result

        for source_item in source_files:
            # Calculate relative path from ifc_source
            try:
                rel_path = source_item.relative_to(ifc_source_long)
            except ValueError:
                # Fallback: try without long path prefix
                try:
                    rel_path = source_item.relative_to(ifc_source)
                except ValueError:
                    continue

            dest_item = target_long / rel_path

            if source_item.is_dir():
                # Create subdirectory
                if not self.dry_run:
                    dest_item.mkdir(parents=True, exist_ok=True)
                continue

            # It's a file - copy it
            try:
                # Ensure parent directory exists
                if not self.dry_run:
                    dest_item.parent.mkdir(parents=True, exist_ok=True)

                if self._should_copy(source_item, dest_item):
                    if not self.dry_run:
                        shutil.copy2(str(source_item), str(dest_item))
                    result["copied"] += 1
                    self.logger.info(f"IFC 复制: {rel_path}")
                else:
                    result["skipped"] += 1
            except OSError as e:
                self.logger.warning(f"无法复制 IFC 文件 (路径过长?): {rel_path} - {e}")
                result["errors"].append(f"{rel_path}: {e}")
            except Exception as e:
                self.logger.warning(f"复制 IFC 文件出错: {rel_path} - {e}")
                result["errors"].append(f"{rel_path}: {e}")

        return result

    def _create_folder_structure(self, project_path: Path) -> int:
        """Create IFR folder structure."""
        created_count = 0
        ifr_client = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)"

        # Use long path format on Windows
        ifr_client_long = self._to_long_path(ifr_client)

        if not ifr_client_long.exists():
            if not self.dry_run:
                ifr_client_long.mkdir(parents=True, exist_ok=True)
            created_count += 1

        for main_folder, subfolders in self.IFR_STRUCTURE.items():
            main_path = self._to_long_path(ifr_client / main_folder)
            if not main_path.exists():
                if not self.dry_run:
                    main_path.mkdir(parents=True, exist_ok=True)
                created_count += 1

            for subfolder in subfolders:
                sub_path = self._to_long_path(ifr_client / main_folder / subfolder)
                if not sub_path.exists():
                    if not self.dry_run:
                        sub_path.mkdir(parents=True, exist_ok=True)
                    created_count += 1

        return created_count

    def _mirror_drawings(self, project_path: Path) -> Dict[str, int]:
        """Mirror drawings from source directories."""
        result = {"copied": 0, "skipped": 0, "errors": []}
        dest_dir = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)/1.Drawing"

        if not dest_dir.exists():
            return result

        drawings = self.drawing_collector.collect_drawings(project_path)

        for file_info in drawings["files"]:
            # IFR(Client) 目录只接受 IFR 文件，排除 IFC 文件
            if _is_ifc_file(file_info["filename"]) and not _is_ifr_file(file_info["filename"]):
                continue

            source = Path(file_info["source_path"])
            dest = dest_dir / file_info["filename"]

            try:
                # Use long path format on Windows
                source_long = self._to_long_path(source)
                dest_long = self._to_long_path(dest)

                if self._should_copy(source_long, dest_long):
                    if not self.dry_run:
                        shutil.copy2(str(source_long), str(dest_long))
                    result["copied"] += 1
                else:
                    result["skipped"] += 1
            except OSError as e:
                self.logger.warning(f"无法复制图纸 (路径过长?): {file_info['filename']} - {e}")
                result["errors"].append(f"{file_info['filename']}: {e}")
            except Exception as e:
                self.logger.warning(f"复制图纸出错: {file_info['filename']} - {e}")
                result["errors"].append(f"{file_info['filename']}: {e}")

        return result

    def _to_long_path(self, path: Path) -> Path:
        """Convert path to long path format on Windows to support paths > 260 chars."""
        return to_long_path(path)

    def _mirror_reports(self, project_path: Path) -> Dict[str, int]:
        """Mirror reports from source directories.

        Only syncs the LATEST version per doc-ID — older revisions in the same
        source folder are skipped to avoid polluting IFR(Client)/2.Reports with
        outdated files.
        """
        result = {"copied": 0, "skipped": 0, "errors": []}
        dest_dir = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)/2.Reports"

        if not dest_dir.exists():
            return result

        source_dirs = [
            project_path / "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
            project_path / "Design/Engineering/2. Calcs & Reports/Reports/Electrical",
            project_path / "Design/Engineering/2. Calcs & Reports/Schedule",
            project_path / "Design/Engineering/1. Drawings/2. IFR_internal",
        ]

        # Phase 1: collect all candidate files with metadata
        candidates = []  # list of (pdf_path, mtime)

        for source_dir in source_dirs:
            if not source_dir.exists():
                continue

            is_report_source = "2. Calcs & Reports" in str(source_dir)

            try:
                pdf_files = list(source_dir.rglob("*.pdf"))
            except Exception as e:
                self.logger.warning(f"无法扫描目录 {source_dir}: {e}")
                continue

            for pdf_file in pdf_files:
                # Check if file is in a non-deliverable subdirectory
                try:
                    rel_parts = pdf_file.relative_to(source_dir).parts[:-1]
                    # Always exclude SS/superseded
                    in_ss = any(
                        part.lower() in ('ss', 'superseded', '_export', 'data')
                        for part in rel_parts
                    )
                    if in_ss:
                        in_excluded_subdir = True
                    else:
                        # Appendix: exclude unless file has a valid doc-ID
                        in_appendix = any(
                            part.lower().startswith(('appendix', 'stk'))
                            for part in rel_parts
                        )
                        if in_appendix:
                            _doc_id_pat = re.match(r'^(GG\d{2}-[A-Z]-[A-Z]{3}-\d{3})', pdf_file.name, re.IGNORECASE)
                            in_excluded_subdir = not bool(_doc_id_pat)
                        else:
                            in_excluded_subdir = False
                except Exception:
                    in_excluded_subdir = False

                if not self._should_collect_as_report(pdf_file.name, is_report_source,
                                                       in_excluded_subdir, rel_parts,
                                                       self.excel_doc_ids):
                    continue

                # IFR(Client) 目录只接受 IFR 文件，排除 IFC 文件
                if _is_ifc_file(pdf_file.name) and not _is_ifr_file(pdf_file.name):
                    continue

                try:
                    source_long = self._to_long_path(pdf_file)
                    mtime = source_long.stat().st_mtime
                except Exception:
                    mtime = 0
                candidates.append((pdf_file, mtime))

        # Phase 2: group by doc-ID, keep only the latest version per doc-ID
        # This prevents syncing outdated revisions (e.g. RevA when RevB exists)
        doc_id_groups: Dict[str, List[Tuple[Path, float]]] = defaultdict(list)
        no_doc_id = []  # files without extractable doc-ID → keep all
        for pdf_file, mtime in candidates:
            doc_id = _extract_doc_id_standalone(pdf_file.name)
            if doc_id:
                doc_id_groups[doc_id].append((pdf_file, mtime))
            else:
                no_doc_id.append((pdf_file, mtime))

        # For each doc-ID group, keep only the file with the newest mtime
        files_to_sync = []
        for doc_id, group in doc_id_groups.items():
            group.sort(key=lambda x: x[1], reverse=True)
            files_to_sync.append(group[0][0])  # newest by mtime
        for pdf_file, _ in no_doc_id:
            files_to_sync.append(pdf_file)

        # Phase 3: deduplicate by filename and copy
        seen_files = set()
        for pdf_file in files_to_sync:
            if pdf_file.name in seen_files:
                continue
            seen_files.add(pdf_file.name)

            dest = dest_dir / pdf_file.name

            try:
                source_long = self._to_long_path(pdf_file)
                dest_long = self._to_long_path(dest)

                if self._should_copy(source_long, dest_long):
                    if not self.dry_run:
                        shutil.copy2(str(source_long), str(dest_long))
                    result["copied"] += 1
                else:
                    result["skipped"] += 1
            except OSError as e:
                self.logger.warning(f"无法复制文件 (路径过长?): {pdf_file.name} - {e}")
                result["errors"].append(f"{pdf_file.name}: {e}")
            except Exception as e:
                self.logger.warning(f"复制文件出错: {pdf_file.name} - {e}")
                result["errors"].append(f"{pdf_file.name}: {e}")

        return result

    def _should_collect_as_report(self, filename: str, is_report_source: bool,
                                    in_excluded_subdir: bool = False,
                                    rel_parts: tuple = (),
                                    excel_doc_ids: Optional[Set[str]] = None) -> bool:
        """Check if file should be collected as a report.

        Uses pattern matching first, falls back to folder-based classification
        for projects with unrecognized naming conventions.
        Files in excluded subdirs (Appendix, SS, STK, etc.) only included
        if they explicitly match a report pattern.
        In fallback mode, files in subdirectories must match parent folder's
        doc-ID prefix to avoid picking up third-party reference documents.
        Phase 4b: if prefix mismatch but file's doc-ID exists in the
        Deliverable Excel, it IS a project deliverable (LMS-style naming).
        """
        if not filename.lower().endswith('.pdf'):
            return False
        # Pattern match: explicitly a report → always include
        if self._is_report(filename):
            return True
        # Pattern match: explicitly a drawing → exclude
        for pattern in self.DRAWING_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return False
        # No pattern match: trust folder structure, but skip excluded subdirs
        if not is_report_source or in_excluded_subdir:
            return False
        # For files in subdirectories: check parent folder prefix match
        if rel_parts:
            parent_name = rel_parts[0]
            parent_prefix = parent_name.split('_')[0]
            # Only include if parent has a doc-ID style name and filename matches
            if '_' in parent_name:
                if not filename.startswith(parent_prefix):
                    # Phase 4b: prefix mismatch — check Deliverable Excel as
                    # fallback.  LMS folders group related deliverables under
                    # a single folder whose doc-ID differs from the files
                    # inside (e.g. KE-300 folder containing RE-305 file).
                    # Only files whose doc-ID is in the Deliverable Excel are
                    # project deliverables; others are third-party references.
                    if excel_doc_ids:
                        file_doc_id = _extract_doc_id_standalone(filename)
                        if file_doc_id and file_doc_id.upper() in excel_doc_ids:
                            return True
                    return False
            else:
                # Non-standard folder (e.g. "Geotech Report") → skip in fallback
                return False
        return True

    def _is_report(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.REPORT_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return True
        return False

    def _should_copy(self, source: Path, dest: Path) -> bool:
        if not dest.exists():
            return True
        try:
            src_stat = source.stat()
            dst_stat = dest.stat()
            if src_stat.st_size != dst_stat.st_size:
                return True
            if src_stat.st_mtime > dst_stat.st_mtime:
                return True
            return False
        except:
            return True

    def archive_approved_in_ifr_client(self, project_path: Path) -> Dict[str, int]:
        """Archive approved files in IFR(Client) to Approved to IFC/ subfolders.

        Two archive triggers:
          1. File has '-Approved' suffix in IFR(Client)  (local trigger)
          2. File's doc-ID already approved in Client Sharepoint  (reverse feedback)
             — checks both '-Approved' files and 'Approved to IFC/' in Sharepoint

        This gives employees visibility: opening IFR(Client)/1.Drawing shows
        only pending files; approved ones are in Approved to IFC/.

        Returns {"archived": N, "skipped": N, "errors": [...]}.
        """
        result = {"archived": 0, "skipped": 0, "errors": []}
        ifr_client_base = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)"
        if not ifr_client_base.exists():
            return result

        sharepoint_base = project_path / "Design/Engineering/13. Client Sharepoint/1.IFR"

        # Mapping: IFR(Client) subfolder → Client Sharepoint subfolder
        pairs = [
            ("1.Drawing", "2.Drawing"),
            ("2.Reports", "1.Report"),
        ]

        for ifr_sub, sp_sub in pairs:
            sub_dir = self._to_long_path(ifr_client_base / ifr_sub)
            if not sub_dir.exists():
                continue

            approved_dir = sub_dir / "Approved to IFC"

            # ── Reverse feedback: collect approved doc-IDs with revision from Client Sharepoint ──
            # Maps doc_id -> highest approved revision letter (e.g. 'A', 'B')
            sp_approved_doc_revs: Dict[str, str] = {}
            sp_dir = sharepoint_base / sp_sub if sharepoint_base.exists() else None
            if sp_dir and sp_dir.exists():
                # Check Approved to IFC/ subfolder in Sharepoint
                sp_approved_dir = sp_dir / "Approved to IFC"
                if sp_approved_dir.exists():
                    try:
                        for f in sp_approved_dir.iterdir():
                            if f.is_file() and f.suffix.lower() == '.pdf':
                                doc_id = _extract_doc_id_standalone(f.name)
                                if doc_id:
                                    did = doc_id.upper()
                                    # Strip -Approved suffix before extracting revision
                                    clean_stem = _RE_APPROVED_SUFFIX.sub('', f.stem)
                                    rev_m = _RE_IFR_FILE.search(clean_stem)
                                    rev = rev_m.group(1).upper() if rev_m else ''
                                    if did not in sp_approved_doc_revs or rev > sp_approved_doc_revs[did]:
                                        sp_approved_doc_revs[did] = rev
                    except (OSError, PermissionError):
                        pass
                # Check -Approved suffix files still in Sharepoint directory
                try:
                    for f in sp_dir.iterdir():
                        if (f.is_file() and f.suffix.lower() == '.pdf'
                                and _RE_APPROVED_SUFFIX.search(f.stem)):
                            doc_id = _extract_doc_id_standalone(f.name)
                            if doc_id:
                                did = doc_id.upper()
                                clean_stem = _RE_APPROVED_SUFFIX.sub('', f.stem)
                                rev_m = _RE_IFR_FILE.search(clean_stem)
                                rev = rev_m.group(1).upper() if rev_m else ''
                                if did not in sp_approved_doc_revs or rev > sp_approved_doc_revs[did]:
                                    sp_approved_doc_revs[did] = rev
                except (OSError, PermissionError):
                    pass

            # ── Scan IFR(Client) files ──
            try:
                pdf_files = [f for f in sub_dir.iterdir()
                             if f.is_file() and f.suffix.lower() == '.pdf'
                             and not f.name.startswith('~$')]
            except (OSError, PermissionError) as e:
                self.logger.warning(f"扫描 {ifr_sub} 失败: {e}")
                result["errors"].append(f"{ifr_sub}: {e}")
                continue

            for f in pdf_files:
                # Trigger 1: file itself has -Approved suffix
                has_approved_suffix = _RE_APPROVED_SUFFIX.search(f.stem) is not None

                # Trigger 2: doc-ID approved in Client Sharepoint (reverse feedback)
                # Only archive if file's revision <= highest approved revision
                # (newer revisions are new IFR submissions, not yet approved)
                doc_id = _extract_doc_id_standalone(f.name)
                is_sp_approved = False
                if doc_id and doc_id.upper() in sp_approved_doc_revs:
                    did = doc_id.upper()
                    file_rev_m = _RE_IFR_FILE.search(f.stem)
                    file_rev = file_rev_m.group(1).upper() if file_rev_m else ''
                    if file_rev <= sp_approved_doc_revs[did]:
                        is_sp_approved = True

                if not has_approved_suffix and not is_sp_approved:
                    continue

                reason = "Approved后缀" if has_approved_suffix else "Sharepoint已审批"
                dest = approved_dir / f.name

                if self.dry_run:
                    self.logger.info(
                        f"[DRY-RUN] 归档 IFR(Client): {f.name} → {ifr_sub}/Approved to IFC/ ({reason})")
                    print(f"    [预览] 归档({reason}): {f.name} → {ifr_sub}/Approved to IFC/")
                    result["archived"] += 1
                    continue

                try:
                    self._to_long_path(approved_dir).mkdir(parents=True, exist_ok=True)
                    if dest.exists():
                        self.logger.info(f"已存在于归档目录，跳过: {f.name}")
                        result["skipped"] += 1
                        continue
                    shutil.move(str(self._to_long_path(f)), str(self._to_long_path(dest)))
                    self.logger.info(
                        f"归档 IFR(Client): {f.name} → {ifr_sub}/Approved to IFC/ ({reason})")
                    result["archived"] += 1
                except Exception as e:
                    self.logger.error(f"归档失败: {f.name} - {e}")
                    result["errors"].append(f"{f.name}: {e}")

        return result

    def sync_to_sharepoint(self, project_path: Path) -> Dict[str, int]:
        """Sync files to Client Sharepoint for files not yet approved or present.

        Source mapping:
          2. IFR_internal           → 13. Client Sharepoint/1.IFR/2.Drawing  (version-managed, unique per doc-ID)
          IFR(Client)/2.Reports     → 13. Client Sharepoint/1.IFR/1.Report

        Before sync, archives -Approved files in IFR(Client)/1.Drawing to Approved to IFC/.

        Skip rules:
          1. File has '-Approved' suffix → managed by ApprovedIFCManager
          2. File's doc-ID is already in target's 'Approved to IFC/' subfolder
          3. File's doc-ID is already in IFR(Client) source's 'Approved to IFC/' subfolder
          4. Exact filename already exists in target with same size/mtime
        """
        result = {"copied": 0, "skipped": 0, "skipped_approved": 0,
                  "archived": 0, "errors": []}

        # Step 0: archive -Approved files in IFR(Client) first
        archive_result = self.archive_approved_in_ifr_client(project_path)
        result["archived"] = archive_result["archived"]
        result["errors"].extend(archive_result.get("errors", []))

        # Mapping: (source_path_relative, sharepoint_subfolder, ifr_client_subfolder_for_approved_check)
        # Drawing: source = IFR_internal (version-managed); also check IFR(Client)/1.Drawing/Approved to IFC
        # Reports: source = IFR(Client)/2.Reports; also check IFR(Client)/2.Reports/Approved to IFC
        ifr_client_base = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)"
        ifr_internal = project_path / "Design/Engineering/1. Drawings/2. IFR_internal"
        sharepoint_base = project_path / "Design/Engineering/13. Client Sharepoint/1.IFR"

        if not sharepoint_base.exists():
            result["errors"].append(f"Client Sharepoint IFR 目录不存在: {sharepoint_base}")
            return result

        # Define sync pairs: (source_dir, sp_subfolder, ifr_client_subfolder_for_approved)
        sync_pairs = []
        # Drawing: use IFR_internal as source (already version-managed, unique per doc-ID)
        if ifr_internal.exists():
            sync_pairs.append((ifr_internal, "2.Drawing", "1.Drawing"))
        elif (ifr_client_base / "1.Drawing").exists():
            # Fallback to IFR(Client)/1.Drawing if IFR_internal doesn't exist
            sync_pairs.append((ifr_client_base / "1.Drawing", "2.Drawing", "1.Drawing"))
        # Reports: use IFR(Client)/2.Reports as source
        if (ifr_client_base / "2.Reports").exists():
            sync_pairs.append((ifr_client_base / "2.Reports", "1.Report", "2.Reports"))

        for source_dir, sp_sub, ifr_sub in sync_pairs:
            source_dir = self._to_long_path(source_dir)
            target_dir = self._to_long_path(sharepoint_base / sp_sub)

            if not source_dir.exists():
                continue
            if not target_dir.exists():
                continue

            # Collect approved doc-IDs with their revision from BOTH target's and IFR(Client)'s Approved to IFC/
            # Maps doc_id -> highest approved revision letter (e.g. 'A', 'B')
            approved_doc_revs: Dict[str, str] = {}
            approved_dirs = [target_dir / "Approved to IFC"]
            if ifr_client_base.exists():
                approved_dirs.append(ifr_client_base / ifr_sub / "Approved to IFC")

            for approved_dir in approved_dirs:
                if approved_dir.exists():
                    try:
                        for f in approved_dir.iterdir():
                            if f.is_file() and f.suffix.lower() == '.pdf':
                                doc_id = _extract_doc_id_standalone(f.name)
                                if doc_id:
                                    did = doc_id.upper()
                                    # Extract letter revision from approved filename
                                    rev_m = _RE_IFR_FILE.search(f.stem)
                                    rev = rev_m.group(1).upper() if rev_m else ''
                                    # Keep highest revision per doc-ID
                                    if did not in approved_doc_revs or rev > approved_doc_revs[did]:
                                        approved_doc_revs[did] = rev
                    except (OSError, PermissionError) as e:
                        self.logger.warning(f"扫描 Approved to IFC 失败: {e}")

            # Collect existing filenames in target directory (excluding subdirs)
            existing_filenames = set()
            try:
                for f in target_dir.iterdir():
                    if f.is_file():
                        existing_filenames.add(f.name)
            except (OSError, PermissionError) as e:
                self.logger.warning(f"扫描目标目录失败: {e}")

            # Scan source and sync eligible files
            try:
                source_files = [f for f in source_dir.iterdir()
                                if f.is_file() and f.suffix.lower() == '.pdf'
                                and not f.name.startswith('~$')]
            except (OSError, PermissionError) as e:
                self.logger.warning(f"扫描源目录失败: {source_dir} - {e}")
                result["errors"].append(f"{sp_sub}: {e}")
                continue

            for src_file in source_files:
                filename = src_file.name

                # Skip -Approved files (managed by ApprovedIFCManager)
                if _RE_APPROVED_SUFFIX.search(Path(filename).stem):
                    result["skipped"] += 1
                    continue

                # Skip if doc-ID is already approved AND source revision <= approved revision
                # (newer revisions after approval should still be synced for client review)
                doc_id = _extract_doc_id_standalone(filename)
                if doc_id:
                    did = doc_id.upper()
                    if did in approved_doc_revs:
                        src_rev_m = _RE_IFR_FILE.search(Path(filename).stem)
                        src_rev = src_rev_m.group(1).upper() if src_rev_m else ''
                        if src_rev <= approved_doc_revs[did]:
                            result["skipped_approved"] += 1
                            continue

                # Copy if newer/different or not present
                dest = target_dir / filename
                try:
                    source_long = self._to_long_path(src_file)
                    dest_long = self._to_long_path(dest)

                    if self._should_copy(source_long, dest_long):
                        if not self.dry_run:
                            shutil.copy2(str(source_long), str(dest_long))
                        result["copied"] += 1
                        self.logger.info(f"Sharepoint同步: {filename} → {sp_sub}/")
                    else:
                        result["skipped"] += 1
                except OSError as e:
                    self.logger.warning(f"无法复制到Sharepoint (路径过长?): {filename} - {e}")
                    result["errors"].append(f"{filename}: {e}")
                except Exception as e:
                    self.logger.warning(f"复制到Sharepoint出错: {filename} - {e}")
                    result["errors"].append(f"{filename}: {e}")

        return result


# =============================================================================
# Version Manager — PDF Mode (merged from version_manager_v4.py)
# =============================================================================

class VersionManager:
    """Manage file versions by moving older versions to SS folder."""

    VERSION_PATTERNS = [
        r'[_\s]Rev\.?\s*([A-Z0-9]+)',
        r'[_\s]R([A-Z])(?=[_\.\s]|$)',
        r'[_\s]R(\d+)(?=[_\.\s]|$)',
        r'[_\s]V(\d+)',
        r'[_\s]v(\d+)',
        r'-([A-Z])(?=[_\.\s]|$)',
    ]

    TARGET_SUBDIRS = [
        r"Design\Engineering\1. Drawings\2. IFR_internal",
        r"Design\Engineering\1. Drawings\3. IFR(Client)\1.Drawing",
        r"Design\Engineering\1. Drawings\3. IFR(Client)\2.Reports",
        r"Design\Engineering\1. Drawings\3. IFR(Client)\3.Deliverables",
        r"Design\Engineering\1. Drawings\4. IFC(Client)",
        r"Design\Engineering\13. Client Sharepoint\1.IFR\1.Report",
        r"Design\Engineering\13. Client Sharepoint\1.IFR\2.Drawing",
        r"Design\Engineering\13. Client Sharepoint\2.IFC",
    ]

    PROJECT_PATTERNS = [
        r'^GG-?\d+',
        r'^NSW\d+',
        r'^[A-Z]{2,4}-\d+',
    ]

    def __init__(self, root_path: str, dry_run: bool = False):
        self.root_path = Path(root_path)
        self.dry_run = dry_run
        self.total_stats = {"scanned": 0, "groups": 0, "moved": 0}

    def extract_base_name_and_version(self, filename: str) -> Tuple[str, Optional[str]]:
        name_without_ext = Path(filename).stem
        extension = Path(filename).suffix
        # Strip browser download-duplicate suffix like " (1)", " (2)" etc.
        name_without_ext = re.sub(r'\s*\(\d+\)\s*$', '', name_without_ext)
        for pattern in self.VERSION_PATTERNS:
            match = re.search(pattern, name_without_ext, re.IGNORECASE)
            if match:
                version = match.group(1).upper()
                base_name = re.sub(pattern, '', name_without_ext, flags=re.IGNORECASE)
                base_name = re.sub(r'[_\s]+$', '', base_name)
                return (base_name + extension, version)
        return (Path(name_without_ext).stem + extension if name_without_ext != Path(filename).stem else filename, None)

    def scan_for_projects(self) -> List[Path]:
        projects = []
        if not self.root_path.exists():
            return projects

        def scan_level(path: Path, depth: int = 0, max_depth: int = 3):
            if depth > max_depth:
                return
            try:
                for item in path.iterdir():
                    if not item.is_dir():
                        continue
                    try:
                        ifr_path = item / "Design" / "Engineering" / "1. Drawings"
                        if ifr_path.exists():
                            for pattern in self.PROJECT_PATTERNS:
                                if re.match(pattern, item.name, re.IGNORECASE):
                                    projects.append(item)
                                    break
                            else:
                                if (ifr_path / "2. IFR_internal").exists() or (ifr_path / "3. IFR(Client)").exists():
                                    projects.append(item)
                        else:
                            scan_level(item, depth + 1, max_depth)
                    except (OSError, PermissionError):
                        continue
            except (OSError, PermissionError):
                pass

        scan_level(self.root_path)
        projects = sorted(set(projects), key=lambda p: p.name)
        return projects

    @staticmethod
    def _find_ss_folder(target_dir: Path) -> Path:
        """Find existing SS/Superseded/Superceded folder, or return default SS/ path."""
        try:
            for item in target_dir.iterdir():
                if item.is_dir() and item.name.lower() in ('ss', 'superseded', 'superceded'):
                    return item
        except (OSError, PermissionError):
            pass
        return target_dir / 'SS'

    def scan_files_pdf(self, target_dir: Path) -> Dict[str, List[Tuple[Path, str, datetime]]]:
        if not target_dir.exists():
            return {}
        file_groups: Dict[str, List[Tuple[Path, str, datetime]]] = defaultdict(list)
        try:
            pdf_files = list(target_dir.glob("*.pdf"))
        except (OSError, PermissionError):
            return file_groups
        for file_path in pdf_files:
            try:
                if file_path.is_file():
                    base_name, version = self.extract_base_name_and_version(file_path.name)
                    pdf_long = to_long_path(file_path)
                    modified_time = datetime.fromtimestamp(pdf_long.stat().st_mtime)
                    version_str = version if version else "NO_VERSION"
                    file_groups[base_name].append((file_path, version_str, modified_time))
            except (OSError, PermissionError):
                continue
        return file_groups

    def identify_old_versions(self, file_groups: Dict, ss_folder: Path) -> List[Tuple[Path, Path, str]]:
        files_to_move = []
        for base_name, files in file_groups.items():
            if len(files) <= 1:
                continue
            files_sorted = sorted(files, key=lambda x: x[2], reverse=True)
            newest_file = files_sorted[0]
            for file_path, version, modified_time in files_sorted[1:]:
                dest_path = ss_folder / file_path.name
                reason = f"较旧版本 (最新: {newest_file[0].name})"
                files_to_move.append((file_path, dest_path, reason))
        return files_to_move

    def identify_ifc_in_ifr_client(self, target_dir: Path, ss_folder: Path) -> List[Tuple[Path, Path, str]]:
        """识别 IFR(Client) 目录中的 IFC 文件，标记移动到 SS。"""
        files_to_move = []
        if "IFR(Client)" not in str(target_dir):
            return files_to_move
        try:
            pdf_files = list(target_dir.glob("*.pdf"))
        except (OSError, PermissionError):
            return files_to_move
        for pdf_file in pdf_files:
            stem = pdf_file.stem
            is_ifc = bool(_RE_IFC_FILE.search(stem))
            is_ifr = bool(_RE_IFR_FILE.search(stem))
            if is_ifc and not is_ifr:
                dest_path = ss_folder / pdf_file.name
                reason = "IFC文件不应在IFR(Client)目录中"
                files_to_move.append((pdf_file, dest_path, reason))
        return files_to_move

    def move_files(self, files_to_move: List[Tuple[Path, Path, str]], ss_folder: Path) -> int:
        if not files_to_move:
            return 0
        if not self.dry_run:
            try:
                ss_folder.mkdir(exist_ok=True)
            except (OSError, PermissionError):
                try:
                    to_long_path(ss_folder).mkdir(exist_ok=True)
                except Exception as e:
                    print(f"      [!] 无法创建 SS 文件夹: {e}")
                    return 0
        ss_name = ss_folder.name
        moved_count = 0
        for source, dest, reason in files_to_move:
            if self.dry_run:
                print(f"      [预览] {source.name} -> {ss_name}/")
            else:
                try:
                    if dest.exists():
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        new_name = f"{dest.stem}_{timestamp}{dest.suffix}"
                        dest = ss_folder / new_name
                    shutil.move(str(to_long_path(source)), str(to_long_path(dest)))
                    print(f"      [v] 已移动: {source.name} -> {ss_name}/")
                    moved_count += 1
                except Exception as e:
                    print(f"      [!] 移动失败: {source.name} - {e}")

        # Cleanup old timestamped versions in SS (keep max 3 per base file)
        if not self.dry_run and ss_folder.exists():
            self._cleanup_ss_folder(ss_folder, max_versions=3)

        return moved_count

    @staticmethod
    def _cleanup_ss_folder(ss_path: Path, max_versions: int = 3):
        """Remove old timestamped copies in SS, keeping only the most recent N per base file."""
        ts_pattern = re.compile(r'^(.+?)_(\d{8}_\d{6})(\.\w+)$')
        groups = defaultdict(list)
        for f in ss_path.iterdir():
            if not f.is_file():
                continue
            m = ts_pattern.match(f.name)
            if m:
                base = m.group(1) + m.group(3)
                groups[base].append(f)
        for base, files in groups.items():
            if len(files) <= max_versions:
                continue
            files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            for old_file in files[max_versions:]:
                try:
                    old_file.unlink()
                except Exception:
                    pass

    def analyze_directory(self, target_dir: Path) -> Dict:
        result = {
            "exists": target_dir.exists(), "total_files": 0,
            "multi_version_groups": {}, "files_to_move": 0,
            "ifc_files": [], "ifc_files_count": 0
        }
        if not target_dir.exists():
            return result
        file_groups = self.scan_files_pdf(target_dir)
        result["total_files"] = sum(len(files) for files in file_groups.values())
        multi_version_groups = {k: v for k, v in file_groups.items() if len(v) > 1}
        result["multi_version_groups"] = multi_version_groups
        result["files_to_move"] = sum(len(v) - 1 for v in multi_version_groups.values())
        ss_folder = self._find_ss_folder(target_dir)
        ifc_files = self.identify_ifc_in_ifr_client(target_dir, ss_folder)
        result["ifc_files"] = ifc_files
        result["ifc_files_count"] = len(ifc_files)
        result["files_to_move"] += len(ifc_files)
        return result

    def process_directory(self, target_dir: Path, show_details: bool = True) -> Dict[str, int]:
        stats = {"scanned": 0, "groups": 0, "moved": 0}
        if not target_dir.exists():
            if show_details:
                print(f"    [!] 目录不存在，跳过")
            return stats
        ss_folder = self._find_ss_folder(target_dir)
        file_groups = self.scan_files_pdf(target_dir)
        total_files = sum(len(files) for files in file_groups.values())
        stats["scanned"] = total_files
        if show_details:
            print(f"    找到 {total_files} 个 PDF 文件")
        multi_version_groups = {k: v for k, v in file_groups.items() if len(v) > 1}
        stats["groups"] = len(multi_version_groups)
        ifc_files = self.identify_ifc_in_ifr_client(target_dir, ss_folder)
        if not multi_version_groups and not ifc_files:
            if show_details:
                print(f"    [v] 没有重复版本文件")
            return stats
        if multi_version_groups and show_details:
            print(f"    发现 {len(multi_version_groups)} 组多版本文件:")
            for base_name, files in multi_version_groups.items():
                files_sorted = sorted(files, key=lambda x: x[2], reverse=True)
                print(f"\n      基础文件: {base_name}")
                for i, (file_path, version, modified_time) in enumerate(files_sorted):
                    status = "[保留]" if i == 0 else "[->SS]"
                    time_str = modified_time.strftime('%Y-%m-%d %H:%M')
                    print(f"        {status} {file_path.name} (版本:{version}, {time_str})")
        files_to_move = self.identify_old_versions(file_groups, ss_folder)
        if files_to_move:
            if show_details:
                print(f"\n    移动 {len(files_to_move)} 个旧版本到 SS 文件夹:")
            moved_count = self.move_files(files_to_move, ss_folder)
            stats["moved"] = moved_count
        if ifc_files:
            if show_details:
                print(f"\n    发现 {len(ifc_files)} 个 IFC 文件在 IFR(Client) 目录中:")
                for src, dst, reason in ifc_files:
                    print(f"        [->SS] {src.name} ({reason})")
            moved_ifc = self.move_files(ifc_files, ss_folder)
            stats["moved"] += moved_ifc
        return stats

    def process_project(self, project_path: Path, show_details: bool = True) -> Dict[str, int]:
        project_stats = {"scanned": 0, "groups": 0, "moved": 0}
        for subdir in self.TARGET_SUBDIRS:
            target_dir = project_path / subdir
            if show_details:
                short_path = subdir.split("\\")[-1] if "\\" in subdir else subdir
                print(f"\n  [{short_path}]")
            stats = self.process_directory(target_dir, show_details)
            project_stats["scanned"] += stats["scanned"]
            project_stats["groups"] += stats["groups"]
            project_stats["moved"] += stats["moved"]
        return project_stats


# =============================================================================
# Native Version Manager (merged from version_manager_v4.py)
# =============================================================================

# Scan root configs
SCAN_ROOTS = [
    ("Design/Engineering/1. Drawings/1. Native", 1, "1. Native"),
    ("Design/Engineering/2. Calcs & Reports/Reports", 2, "Reports"),
    ("Design/Engineering/2. Calcs & Reports/Schedule", 1, "Schedule"),
]

_SKIP_FOLDER_NAMES = {
    'ss', 'superseded', 'superceded',
    'bom', 'appendix', 'approved',
}
_SKIP_FOLDER_PREFIXES = ('sy supply',)
_IGNORE_EXT = {'.dwl', '.dwl2', '.err', '.log', '.tmp', '.lnk', '.db', '.ini'}


@dataclass
class DwgFile:
    """A versioned file with parsed metadata."""
    path: Path
    filename: str
    rev_type: str        # 'IFR', 'IFC', or 'OTHER'
    revision: str
    mtime: datetime
    bak_path: Optional[Path] = None

    @property
    def rev_display(self) -> str:
        if self.rev_type == 'IFR':
            return f"Rev{self.revision}"
        elif self.rev_type == 'IFC':
            return f"Rev{self.revision}_IFC"
        return "no-rev"


@dataclass
class NativeFileAction:
    source: Path
    dest: Path
    action: str   # 'rename', 'move_to_ss'
    reason: str


@dataclass
class FolderRelocation:
    source: Path
    dest: Path
    reason: str
    action: str   # 'move', 'merge', 'warn'


@dataclass
class NativeFolderResult:
    folder_name: str
    folder_path: Path
    doc_id: str
    description: str
    dwg_files: List[DwgFile] = field(default_factory=list)
    actions: List[NativeFileAction] = field(default_factory=list)
    kept_ifr: Optional[DwgFile] = None
    kept_ifc: Optional[DwgFile] = None
    kept_ifr_all: List[DwgFile] = field(default_factory=list)  # v10: 每种扩展名的最新 IFR
    kept_ifc_all: List[DwgFile] = field(default_factory=list)  # v10: 每种扩展名的最新 IFC


def _classify_dwg(filename: str) -> Tuple[str, str]:
    """Classify a filename as IFR, IFC, or OTHER."""
    stem = Path(filename).stem
    m = _RE_IFC_FILE.search(stem)
    if m:
        return ('IFC', m.group(1))
    m = _RE_IFR_FILE.search(stem)
    if m:
        return ('IFR', m.group(1).upper())
    return ('OTHER', '')


def _parse_folder_name(folder_name: str) -> Tuple[str, str]:
    cleaned = re.sub(r'^[^\w]+', '', folder_name)
    m = re.match(r'^([A-Z0-9][\w-]+-\d{2,3})[_\s](.+?)(?:[_\s]Rev.*)?$', cleaned, re.IGNORECASE)
    if m:
        return (m.group(1), m.group(2).strip())
    if '_' in cleaned:
        parts = cleaned.split('_', 1)
        return (parts[0], parts[1].rstrip().rstrip('_Rev').strip())
    return (cleaned if cleaned else folder_name, '')


def _make_standard_dwg_name(doc_id: str, description: str, rev_type: str,
                            revision: str, ext: str = '.dwg') -> str:
    if rev_type == 'IFC':
        return f"{doc_id}_{description}_Rev{revision}_IFC{ext}"
    return f"{doc_id}_{description}_Rev{revision.upper()}{ext}"


class NativeVersionManager:
    """Manage file versions across Native, Reports, and Schedule folders."""

    def __init__(self, project_path: str, dry_run: bool = True):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.results: List[Tuple[str, NativeFolderResult]] = []

    @property
    def native_root(self) -> Path:
        return self.project_path / "Design" / "Engineering" / "1. Drawings" / "1. Native"

    def find_doc_folders(self, folder_filter: Optional[str] = None,
                         scope: str = 'all') -> List[Tuple[str, Path]]:
        results = []
        for rel_path, depth, display_name in SCAN_ROOTS:
            if scope != 'all':
                if scope == 'native' and 'Native' not in rel_path:
                    continue
                if scope == 'reports' and 'Reports' not in rel_path:
                    continue
                if scope == 'schedule' and 'Schedule' not in rel_path:
                    continue
            root = self.project_path / rel_path
            if not root.exists():
                continue
            if depth == 1:
                candidates = self._list_doc_dirs(root)
            elif depth == 2:
                candidates = []
                for category_dir in sorted(root.iterdir()):
                    if category_dir.is_dir() and not self._should_skip(category_dir.name):
                        candidates.extend(self._list_doc_dirs(category_dir))
            else:
                candidates = []
            for folder in candidates:
                if folder_filter and folder_filter.lower() not in folder.name.lower():
                    continue
                results.append((display_name, folder))
        return results

    def _list_doc_dirs(self, parent: Path) -> List[Path]:
        dirs = []
        try:
            for item in sorted(parent.iterdir(), key=lambda p: p.name):
                if item.is_dir() and not self._should_skip(item.name):
                    dirs.append(item)
        except (OSError, PermissionError):
            pass
        return dirs

    @staticmethod
    def _should_skip(name: str) -> bool:
        if name.startswith('.'):
            return True
        low = name.lower()
        if low in _SKIP_FOLDER_NAMES:
            return True
        for prefix in _SKIP_FOLDER_PREFIXES:
            if low.startswith(prefix):
                return True
        return False

    @staticmethod
    def _find_or_create_ss_folder(folder: Path, create: bool = False) -> Path:
        try:
            for item in folder.iterdir():
                if item.is_dir() and item.name.lower() in ('ss', 'superseded', 'superceded'):
                    return item
        except (OSError, PermissionError):
            pass
        ss_path = folder / 'SS'
        if create:
            to_long_path(ss_path).mkdir(exist_ok=True)
        return ss_path

    def scan_files(self, folder: Path) -> List[DwgFile]:
        files = []
        bak_map: Dict[str, Path] = {}
        ss_names = {'ss', 'superseded', 'superceded'}
        try:
            all_items = list(folder.iterdir())
        except (OSError, PermissionError):
            return []
        for f in all_items:
            if f.is_file() and f.suffix.lower() == '.bak':
                bak_map[f.stem.lower()] = f
        for f in all_items:
            if not f.is_file():
                continue
            if f.parent.name.lower() in ss_names:
                continue
            ext = f.suffix.lower()
            if ext in _IGNORE_EXT or ext == '.bak':
                continue
            rev_type, revision = _classify_dwg(f.name)
            try:
                mtime = datetime.fromtimestamp(to_long_path(f).stat().st_mtime)
            except (OSError, PermissionError):
                mtime = datetime.min
            files.append(DwgFile(
                path=f, filename=f.name,
                rev_type=rev_type, revision=revision,
                mtime=mtime, bak_path=bak_map.get(f.stem.lower()),
            ))
        return files

    def process_folder(self, folder: Path) -> NativeFolderResult:
        """v10 改进: 每种扩展名各自保留最新版本，保护唯一的辅助文件。"""
        doc_id, description = _parse_folder_name(folder.name)
        result = NativeFolderResult(
            folder_name=folder.name, folder_path=folder,
            doc_id=doc_id or folder.name, description=description or '',
        )
        files = self.scan_files(folder)
        result.dwg_files = files
        if not files:
            return result
        ifr_files = [f for f in files if f.rev_type == 'IFR']
        ifc_files = [f for f in files if f.rev_type == 'IFC']
        other_files = [f for f in files if f.rev_type == 'OTHER']
        ss_folder = self._find_or_create_ss_folder(folder)

        # v10: IFR — 按扩展名分组，每组保留 mtime 最新的文件
        if ifr_files:
            ifr_by_ext: dict[str, list] = {}
            for f in ifr_files:
                ifr_by_ext.setdefault(f.path.suffix.lower(), []).append(f)
            for ext, group in ifr_by_ext.items():
                group_sorted = sorted(group, key=lambda f: f.mtime, reverse=True)
                result.kept_ifr_all.append(group_sorted[0])
                for old in group_sorted[1:]:
                    self._plan_move(result, old, ss_folder, "older IFR revision")
            result.kept_ifr = max(result.kept_ifr_all, key=lambda f: f.mtime)

        # v10: IFC — 按扩展名分组，每组保留 mtime 最新的文件
        if ifc_files:
            ifc_by_ext: dict[str, list] = {}
            for f in ifc_files:
                ifc_by_ext.setdefault(f.path.suffix.lower(), []).append(f)
            for ext, group in ifc_by_ext.items():
                group_sorted = sorted(group, key=lambda f: f.mtime, reverse=True)
                result.kept_ifc_all.append(group_sorted[0])
                for old in group_sorted[1:]:
                    self._plan_move(result, old, ss_folder, "older IFC revision")
            result.kept_ifc = max(result.kept_ifc_all, key=lambda f: f.mtime)

        # v10: OTHER — 仅当存在同扩展名的版本化文件时才移到 SS
        for other in other_files:
            ext = other.path.suffix.lower()
            has_versioned_same_ext = any(
                f.path.suffix.lower() == ext and f.rev_type != 'OTHER'
                for f in files
            )
            if has_versioned_same_ext:
                self._plan_move(result, other, ss_folder, "unversioned/legacy (versioned copy exists)")

        # v10: 保护最新 DWG — 按 doc-ID 分组，有 IFC 则只保留 IFC，否则保留最新 IFR
        dwg_files = [f for f in files if f.path.suffix.lower() == '.dwg']
        if dwg_files:
            _did_pat = re.compile(r'^([A-Z0-9][\w-]+-\d{3})', re.IGNORECASE)
            dwg_by_docid: dict[str, list] = {}
            for f in dwg_files:
                m = _did_pat.match(f.filename)
                did = m.group(1) if m else '__no_docid__'
                dwg_by_docid.setdefault(did, []).append(f)
            protected_paths: set[str] = set()
            movable_paths: set[str] = set()
            for did, group in dwg_by_docid.items():
                ifr_dwgs = [f for f in group if f.rev_type == 'IFR']
                ifc_dwgs = [f for f in group if f.rev_type == 'IFC']
                if ifc_dwgs:
                    # 有 IFC DWG → 只保留最新 IFC，IFR 字母版本可移到 SS
                    protected_paths.add(str(max(ifc_dwgs, key=lambda f: f.mtime).path))
                    for f in ifr_dwgs:
                        movable_paths.add(str(f.path))
                elif ifr_dwgs:
                    # 无 IFC → 保留最新 IFR
                    protected_paths.add(str(max(ifr_dwgs, key=lambda f: f.mtime).path))
                else:
                    # 无 IFR/IFC 分类 → 保留最新
                    protected_paths.add(str(max(group, key=lambda f: f.mtime).path))
            # 从移动列表中移除受保护的 DWG
            if protected_paths:
                result.actions = [
                    a for a in result.actions
                    if not (a.action == 'move_to_ss' and str(a.source) in protected_paths)
                ]
            # 将有 IFC 时多余的 IFR DWG 加入移动列表
            for f in dwg_files:
                fp = str(f.path)
                if fp in movable_paths and fp not in protected_paths:
                    already_planned = any(
                        a.action == 'move_to_ss' and str(a.source) == fp
                        for a in result.actions
                    )
                    if not already_planned:
                        self._plan_move(result, f, ss_folder,
                                        "IFR DWG superseded by IFC DWG")

        # v10: 对所有保留的文件执行重命名
        if doc_id and description:
            for kept in result.kept_ifr_all:
                self._plan_rename(result, kept, doc_id, description)
            for kept in result.kept_ifc_all:
                self._plan_rename(result, kept, doc_id, description)
        return result

    def _plan_move(self, result, dwg, ss_folder, reason):
        result.actions.append(NativeFileAction(
            source=dwg.path, dest=ss_folder / dwg.filename,
            action='move_to_ss', reason=reason
        ))
        if dwg.bak_path and dwg.bak_path.exists():
            result.actions.append(NativeFileAction(
                source=dwg.bak_path, dest=ss_folder / dwg.bak_path.name,
                action='move_to_ss', reason=f".bak follows {dwg.filename}"
            ))

    def _plan_rename(self, result, dwg, doc_id, description):
        ext = dwg.path.suffix
        standard = _make_standard_dwg_name(doc_id, description, dwg.rev_type,
                                           dwg.revision, ext)
        if dwg.filename != standard:
            result.actions.append(NativeFileAction(
                source=dwg.path, dest=dwg.path.parent / standard,
                action='rename', reason=f"标准化 -> {standard}"
            ))
            if dwg.bak_path and dwg.bak_path.exists():
                bak_std = _make_standard_dwg_name(doc_id, description,
                                                  dwg.rev_type, dwg.revision, '.bak')
                if dwg.bak_path.name != bak_std:
                    result.actions.append(NativeFileAction(
                        source=dwg.bak_path, dest=dwg.bak_path.parent / bak_std,
                        action='rename', reason=f".bak 标准化 -> {bak_std}"
                    ))

    def process_all(self, folder_filter=None, scope='all'):
        doc_folders = self.find_doc_folders(folder_filter, scope)
        results = []
        for group_name, folder in doc_folders:
            result = self.process_folder(folder)
            results.append((group_name, result))
        self.results = results
        return results

    def execute_actions(self, results):
        stats = {'renamed': 0, 'moved': 0, 'errors': 0, 'ss_created': 0}
        for _group_name, result in results:
            if not result.actions:
                continue
            moves = [a for a in result.actions if a.action == 'move_to_ss']
            if moves:
                ss_folder = self._find_or_create_ss_folder(result.folder_path, create=True)
                if not ss_folder.exists():
                    try:
                        to_long_path(ss_folder).mkdir(exist_ok=True)
                        stats['ss_created'] += 1
                    except Exception as e:
                        print(f"      [!] 无法创建 SS: {e}")
                        stats['errors'] += 1
                        continue
                for action in moves:
                    action.dest = ss_folder / action.source.name
            for action in result.actions:
                try:
                    src = str(to_long_path(action.source))
                    dst_path = action.dest
                    if to_long_path(dst_path).exists():
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        dst_path = dst_path.parent / f"{dst_path.stem}_{ts}{dst_path.suffix}"
                    dst = str(to_long_path(dst_path))
                    shutil.move(src, dst)
                    if action.action == 'rename':
                        stats['renamed'] += 1
                    else:
                        stats['moved'] += 1
                except Exception as e:
                    print(f"      [!] 失败 {action.source.name}: {e}")
                    stats['errors'] += 1
        return stats


# =============================================================================
# Folder Relocator (merged from version_manager_v4.py)
# =============================================================================

_L1_PATTERNS = [
    (re.compile(r'^1[\.\s]*Native', re.IGNORECASE),        '1. Native'),
    (re.compile(r'^2[\.\s]*IFR[_\s]*internal', re.IGNORECASE), '2. IFR_internal'),
    (re.compile(r'^3[\.\s]*IFR\s*\(Client\)', re.IGNORECASE), '3. IFR(Client)'),
    (re.compile(r'^4[\.\s]*IFC\s*\(Client\)', re.IGNORECASE), '4. IFC(Client)'),
]

_L2_PATTERNS = [
    (re.compile(r'^1[\.\s]*Drawing', re.IGNORECASE),       '1.Drawing'),
    (re.compile(r'^2[\.\s]*Reports?', re.IGNORECASE),      '2.Reports'),
    (re.compile(r'^3[\.\s]*Deliverables?', re.IGNORECASE), '3.Deliverables'),
]

_FLAT_L1 = {'1. Native', '2. IFR_internal', '4. IFC(Client)'}
_IFR_CLIENT_CHILDREN = {'1.Drawing', '2.Reports', '3.Deliverables'}


class FolderRelocator:
    """Detect and relocate misplaced structural folders under 1. Drawings/."""

    def __init__(self, project_path: Path, dry_run: bool = True):
        self.project_path = project_path
        self.drawings_root = project_path / "Design" / "Engineering" / "1. Drawings"
        self.dry_run = dry_run

    @staticmethod
    def _match_l1(name: str) -> Optional[str]:
        for pat, canonical in _L1_PATTERNS:
            if pat.match(name):
                return canonical
        return None

    @staticmethod
    def _match_l2(name: str) -> Optional[str]:
        for pat, canonical in _L2_PATTERNS:
            if pat.match(name):
                return canonical
        return None

    def _ifr_client_path(self) -> Path:
        try:
            for item in self.drawings_root.iterdir():
                if item.is_dir() and self._match_l1(item.name) == '3. IFR(Client)':
                    return item
        except (OSError, PermissionError):
            pass
        return self.drawings_root / '3. IFR(Client)'

    def scan(self) -> List[FolderRelocation]:
        relocations: List[FolderRelocation] = []
        if not self.drawings_root.exists():
            return relocations
        ifr_client = self._ifr_client_path()
        try:
            depth0_items = sorted(self.drawings_root.iterdir(), key=lambda p: p.name)
        except (OSError, PermissionError):
            return relocations
        for item in depth0_items:
            if not item.is_dir():
                continue
            l1_match = self._match_l1(item.name)
            l2_match = self._match_l2(item.name)
            if l2_match:
                dest = ifr_client / l2_match
                relocations.append(FolderRelocation(
                    source=item, dest=dest,
                    reason=f"Level-2 文件夹 '{item.name}' 不应直接在 1. Drawings/ 下",
                    action='merge' if dest.exists() else 'move',
                ))
                continue
            if not l1_match:
                continue
            try:
                depth1_items = sorted(item.iterdir(), key=lambda p: p.name)
            except (OSError, PermissionError):
                continue
            for child in depth1_items:
                if not child.is_dir():
                    continue
                child_l1 = self._match_l1(child.name)
                child_l2 = self._match_l2(child.name)
                if child_l1:
                    dest = self.drawings_root / child_l1
                    relocations.append(FolderRelocation(
                        source=child, dest=dest,
                        reason=f"Level-1 文件夹 '{child.name}' 被误放在 '{item.name}/' 内",
                        action='merge' if dest.exists() else 'move',
                    ))
                elif child_l2 and l1_match in _FLAT_L1:
                    dest = ifr_client / child_l2
                    relocations.append(FolderRelocation(
                        source=child, dest=dest,
                        reason=f"'{child.name}' 不应在 flat 文件夹 '{item.name}/' 内",
                        action='merge' if dest.exists() else 'move',
                    ))
                elif child_l2 and l1_match == '3. IFR(Client)':
                    self._scan_deep(child, item, ifr_client, relocations)
        if ifr_client.exists():
            existing_children = set()
            try:
                for c in ifr_client.iterdir():
                    if c.is_dir():
                        m = self._match_l2(c.name)
                        if m:
                            existing_children.add(m)
            except (OSError, PermissionError):
                pass
            for expected in _IFR_CLIENT_CHILDREN:
                if expected not in existing_children:
                    relocations.append(FolderRelocation(
                        source=ifr_client / expected,
                        dest=ifr_client / expected,
                        reason=f"3. IFR(Client)/ 缺少子文件夹 '{expected}'",
                        action='warn',
                    ))
        return relocations

    def _scan_deep(self, folder, parent_l1, ifr_client, relocations):
        try:
            children = sorted(folder.iterdir(), key=lambda p: p.name)
        except (OSError, PermissionError):
            return
        for child in children:
            if not child.is_dir():
                continue
            child_l1 = self._match_l1(child.name)
            child_l2 = self._match_l2(child.name)
            if child_l1:
                dest = self.drawings_root / child_l1
                relocations.append(FolderRelocation(
                    source=child, dest=dest,
                    reason=f"Level-1 '{child.name}' 嵌套在 '{folder.name}/' 内 (深层错位)",
                    action='merge' if dest.exists() else 'move',
                ))
            elif child_l2:
                dest = ifr_client / child_l2
                relocations.append(FolderRelocation(
                    source=child, dest=dest,
                    reason=f"Level-2 '{child.name}' 嵌套在 '{folder.name}/' 内 (深层错位)",
                    action='merge' if dest.exists() else 'move',
                ))

    def execute(self, relocations):
        stats = {'moved': 0, 'merged': 0, 'warnings': 0, 'errors': 0}
        for rel in relocations:
            if rel.action == 'warn':
                stats['warnings'] += 1
                continue
            try:
                src = to_long_path(rel.source)
                dst = to_long_path(rel.dest)
                if rel.action == 'move':
                    dst.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(src), str(dst))
                    print(f"  [v] 已移动: {rel.source.name} -> {rel.dest}")
                    stats['moved'] += 1
                elif rel.action == 'merge':
                    dst.mkdir(parents=True, exist_ok=True)
                    moved_children = 0
                    try:
                        for child in src.iterdir():
                            child_dst = dst / child.name
                            if child_dst.exists():
                                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                                if child.is_file():
                                    child_dst = dst / f"{child.stem}_{ts}{child.suffix}"
                                else:
                                    child_dst = dst / f"{child.name}_{ts}"
                            shutil.move(str(child), str(child_dst))
                            moved_children += 1
                    except Exception as e:
                        print(f"  [!] 合并部分失败 {rel.source.name}: {e}")
                        stats['errors'] += 1
                    try:
                        if src.exists() and not any(src.iterdir()):
                            src.rmdir()
                    except (OSError, PermissionError):
                        pass
                    print(f"  [v] 已合并: {rel.source.name} -> {rel.dest} ({moved_children} 项)")
                    stats['merged'] += 1
            except Exception as e:
                print(f"  [!] 失败: {rel.source.name} - {e}")
                stats['errors'] += 1
        return stats


# =============================================================================
# Deliverable Manager (NEW in v7)
# =============================================================================

class FormatChangeWarning(Exception):
    """Raised when deliverable Excel layout cannot be auto-detected."""
    pass


@dataclass
class DeliverableLayout:
    """Detected Excel layout for a deliverable file."""
    header_row: int           # Row number where headers are found
    rev_col: int              # Column index for Revision (1-based)
    date_col: int             # Column index for Submission Date
    status_col: int           # Column index for Status
    doc_id_col: int           # Column index for Doc ID (usually B=2)
    desc_col: int             # Column index for Description (usually C=3)
    file_rev_cell: str        # Cell address for file revision (e.g. 'K1')
    last_updated_cell: str    # Cell address for last updated date (e.g. 'C79')


@dataclass
class DeliverableCrossCheckResult:
    """Result of deliverable cross-check."""
    project_name: str
    excel_path: str
    items_in_folders_not_excel: List[Dict] = field(default_factory=list)
    items_in_excel_not_folders: List[str] = field(default_factory=list)
    revision_mismatches: List[Dict] = field(default_factory=list)
    doc_id_corrections: List[Dict] = field(default_factory=list)  # rows where FILE NO was corrected
    status_updates: List[Dict] = field(default_factory=list)  # IFC status updates
    naming_warnings: List[Dict] = field(default_factory=list)  # filename normalization warnings
    rows_inserted: int = 0
    rows_updated: int = 0
    new_file_rev: str = ""
    errors: List[str] = field(default_factory=list)


class DeliverableManager:
    """Cross-check and update deliverable Excel files against project folders."""

    # Doc ID extraction patterns
    _RE_DOC_ID_GG = re.compile(r'^(GG\d{2}-[A-Z]-[A-Z]{3}-\d{3})', re.IGNORECASE)
    _RE_DOC_ID_LMS = re.compile(r'^(\d{5}-[A-Z]{2}-\d{3})', re.IGNORECASE)
    _RE_DOC_ID_GENERIC = re.compile(r'^([A-Z0-9][\w]+-[A-Z]+-[A-Z]*-?\d{3})', re.IGNORECASE)

    # Revision extraction from filename
    _RE_REV_LETTER = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*([A-Z])(?=[_.\s]|$)', re.IGNORECASE)
    _RE_REV_NUMBER = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?=[_.\s]|$)', re.IGNORECASE)

    # Header pattern to detect column layout
    HEADER_KEYWORDS = ['revision', 'submission date', 'status']

    HIGHLIGHT_FILL = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99',
                                 fill_type='solid') if OPENPYXL_AVAILABLE else None
    CHANGE_MARKER_FILL = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99',
                                     fill_type='solid') if OPENPYXL_AVAILABLE else None
    STATUS_APPROVED_FILL = PatternFill(start_color='FF00B050', end_color='FF00B050',
                                       fill_type='solid') if OPENPYXL_AVAILABLE else None
    STATUS_SUBMITTED_FILL = PatternFill(start_color='FFFFC000', end_color='FFFFC000',
                                        fill_type='solid') if OPENPYXL_AVAILABLE else None

    # IFC folder path (relative to project)
    IFC_FOLDER = "Design/Engineering/1. Drawings/4. IFC(Client)"

    # Source folders to scan for deliverable items
    SOURCE_FOLDERS = [
        "Design/Engineering/1. Drawings/2. IFR_internal",
        "Design/Engineering/2. Calcs & Reports/Reports/Electrical",
        "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
        "Design/Engineering/2. Calcs & Reports/Schedule",
        "Design/Engineering/1. Drawings/3. IFR(Client)/2.Reports",  # fallback for reports
    ]

    # Deliverable folder path (relative to project)
    DELIVERABLE_REL_PATH = "Design/Engineering/1. Drawings/3. IFR(Client)/3.Deliverables"

    # Sync target paths
    PRIMARY_SYNC = "Design/Engineering/1. Drawings/3. IFR(Client)/3.Deliverables"
    SECONDARY_SYNC = "Design/Engineering/13. Client Sharepoint/1.IFR/3.Deliverables"

    def __init__(self, project_path: Path, dry_run: bool = False):
        self.project_path = project_path
        self.dry_run = dry_run

    # All possible locations for the deliverable Excel (searched in order)
    DELIVERABLE_SEARCH_PATHS = [
        "Design/Engineering/8. Deliverables",
        "Design/Engineering/1. Drawings/3. IFR(Client)/3.Deliverables",
        "Design/Engineering/13. Client Sharepoint/1.IFR/3.Deliverables",
    ]

    def find_deliverable_excel(self) -> Optional[Path]:
        """Find the deliverable Excel file across known locations.

        Searches: 8. Deliverables (master) → 3.Deliverables → 13. Client Sharepoint.
        """
        for rel_path in self.DELIVERABLE_SEARCH_PATHS:
            dlv_path = self.project_path / rel_path
            found = self._find_excel_in_folder(dlv_path)
            if found:
                return found
        return None

    def _find_excel_in_folder(self, folder: Path) -> Optional[Path]:
        """Find a deliverable Excel in a specific folder."""
        if not folder.exists():
            return None
        # Priority 1: filename contains 'dlv' or 'deliverable'
        for f in folder.iterdir():
            if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm') and not f.name.startswith('~$'):
                if 'dlv' in f.name.lower() or 'deliverable' in f.name.lower():
                    return f
        # Priority 2: any xlsx that isn't a temp file
        for f in folder.iterdir():
            if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm') and not f.name.startswith('~$'):
                return f
        return None

    def detect_layout(self, ws) -> DeliverableLayout:
        """Auto-detect Excel column layout by scanning rows 1-20 for header pattern."""
        for row_idx in range(1, 21):
            for col_idx in range(1, 20):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val and isinstance(cell_val, str) and 'revision' in cell_val.lower():
                    # Check if next two columns match pattern
                    next1 = ws.cell(row=row_idx, column=col_idx + 1).value
                    next2 = ws.cell(row=row_idx, column=col_idx + 2).value
                    if (next1 and isinstance(next1, str) and 'date' in next1.lower() and
                        next2 and isinstance(next2, str) and 'status' in next2.lower()):
                        # Found the header pattern
                        layout = DeliverableLayout(
                            header_row=row_idx,
                            rev_col=col_idx,
                            date_col=col_idx + 1,
                            status_col=col_idx + 2,
                            doc_id_col=2,  # Column B
                            desc_col=3,    # Column C
                            file_rev_cell=f"{openpyxl.utils.get_column_letter(col_idx)}1",
                            last_updated_cell=self._detect_last_updated_cell(ws, col_idx),
                        )
                        return layout
        raise FormatChangeWarning(
            f"无法在前20行检测到 'Revision | Submission Date | Status' 列头模式"
        )

    def _detect_last_updated_cell(self, ws, rev_col: int) -> str:
        """Try to find the 'last updated' date cell by scanning for the label."""
        # Search rows 1-10 for "Last Updated" label, return the cell to its right
        for row in range(1, 11):
            for col in range(1, 20):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and 'last updated' in val.lower():
                    date_col = col + 1
                    return f'{openpyxl.utils.get_column_letter(date_col)}{row}'
        # Fallback: search near top for date-like values
        for row in range(1, 10):
            for col in range(1, 20):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, datetime):
                    return f'{openpyxl.utils.get_column_letter(col)}{row}'
        return 'L2'  # Reasonable default

    def extract_doc_id(self, filename: str) -> Optional[str]:
        """Extract doc ID from filename."""
        stem = Path(filename).stem
        for pattern in [self._RE_DOC_ID_GG, self._RE_DOC_ID_LMS, self._RE_DOC_ID_GENERIC]:
            m = pattern.match(stem)
            if m:
                return m.group(1)
        # Fallback: take everything before first _Rev or _rev
        m = re.match(r'^(.+?)(?:[_\s]-?[Rr]ev)', stem)
        if m:
            return m.group(1).rstrip('_- ')
        return None

    def extract_revision(self, filename: str) -> Optional[str]:
        """Extract revision from filename."""
        stem = Path(filename).stem
        m = self._RE_REV_LETTER.search(stem)
        if m:
            return m.group(1).upper()
        m = self._RE_REV_NUMBER.search(stem)
        if m:
            return m.group(1)
        return None

    def extract_description(self, filename: str) -> str:
        """Extract description from filename (between doc_id and revision)."""
        stem = Path(filename).stem
        doc_id = self.extract_doc_id(filename)
        if not doc_id:
            return stem
        remainder = stem[len(doc_id):].lstrip('_- ')
        # Remove revision suffix
        remainder = re.sub(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*[A-Z0-9]+.*$', '', remainder, flags=re.IGNORECASE)
        return remainder.strip('_- ') or stem

    def scan_source_folders(self) -> Dict[str, Dict]:
        """Scan source folders and extract doc IDs with their revisions."""
        items = {}  # doc_id -> {revision, filename, source, description}
        for rel_path in self.SOURCE_FOLDERS:
            source_dir = self.project_path / rel_path
            if not source_dir.exists():
                continue
            try:
                for f in source_dir.rglob("*.pdf"):
                    if f.name.startswith('~$'):
                        continue
                    # Skip files in excluded subfolders
                    try:
                        rel_parts = f.relative_to(source_dir).parts
                        skip_exact = ('ss', 'superseded', 'superceded', 'approved to ifc')
                        skip_prefix = ('appendix', 'reference', 'app ')
                        # ALWAYS skip SS/superseded — even inside appendix folders
                        if any(p.lower() in skip_exact for p in rel_parts[:-1]):
                            continue
                        # In appendix subfolders, only include files with a valid doc-ID
                        in_appendix = any(
                            any(p.lower().startswith(pfx) for pfx in skip_prefix)
                            for p in rel_parts[:-1]
                        )
                        if in_appendix and not self.extract_doc_id(f.name):
                            continue
                    except ValueError:
                        pass
                    # Skip IFC files (only IFR goes into deliverables)
                    if _is_ifc_file(f.name) and not _is_ifr_file(f.name):
                        continue
                    doc_id = self.extract_doc_id(f.name)
                    if not doc_id:
                        continue
                    rev = self.extract_revision(f.name)
                    # Keep the highest revision per doc_id
                    if doc_id in items:
                        existing_rev = items[doc_id]['revision']
                        if rev and existing_rev and self._compare_revisions(rev, existing_rev) > 0:
                            items[doc_id] = {
                                'revision': rev, 'filename': f.name,
                                'source': rel_path, 'description': self.extract_description(f.name)
                            }
                    else:
                        items[doc_id] = {
                            'revision': rev, 'filename': f.name,
                            'source': rel_path, 'description': self.extract_description(f.name)
                        }
            except (OSError, PermissionError):
                continue
        return items

    @staticmethod
    def _normalize_desc(desc: str) -> str:
        """Normalize description for fuzzy matching: lowercase, strip punctuation/spaces."""
        return re.sub(r'[^a-z0-9]', '', desc.lower())

    @staticmethod
    def _desc_similarity(a: str, b: str) -> float:
        """Sequence-based similarity ratio using difflib-style matching."""
        if not a or not b:
            return 0.0
        if a == b:
            return 1.0
        # Use difflib SequenceMatcher for accurate similarity
        from difflib import SequenceMatcher
        return SequenceMatcher(None, a, b).ratio()

    def _find_match_by_description(self, folder_doc_id: str, folder_desc: str,
                                    all_excel_rows: List[Dict],
                                    matched_rows: Set[int]) -> Optional[Dict]:
        """Try to find an Excel row matching by description when doc-ID doesn't match.
        Also requires that the doc-ID category prefix matches (e.g. both are E-BLD).
        Returns the row_info dict if a match is found, else None."""
        norm_folder = self._normalize_desc(folder_desc)
        if not norm_folder or len(norm_folder) < 5:
            return None
        # Extract category prefix (e.g. "E-BLD" from "GG31-E-BLD-003")
        folder_cat = re.sub(r'^[A-Z0-9]+-', '', folder_doc_id, count=1)  # strip project prefix
        folder_cat = re.sub(r'-\d+$', '', folder_cat)  # strip trailing number
        best_match = None
        best_ratio = 0.0
        for row_info in all_excel_rows:
            if row_info['row'] in matched_rows:
                continue
            excel_id = row_info['doc_id']
            # Check category prefix matches
            excel_cat = re.sub(r'^[A-Z0-9]+-', '', excel_id, count=1)
            excel_cat = re.sub(r'-\d+$', '', excel_cat)
            if folder_cat.upper() != excel_cat.upper():
                continue
            norm_excel = self._normalize_desc(row_info.get('description', ''))
            if not norm_excel:
                continue
            # Exact match
            if norm_folder == norm_excel:
                return row_info
            # Similarity check (handles minor differences like "communication" vs "communications")
            ratio = self._desc_similarity(norm_folder, norm_excel)
            if ratio > best_ratio and ratio >= 0.85:
                best_ratio = ratio
                best_match = row_info
        return best_match

    def _compare_revisions(self, rev_a: str, rev_b: str) -> int:
        """Compare two revision strings. Returns >0 if a>b, <0 if a<b, 0 if equal."""
        if rev_a == rev_b:
            return 0
        # Try numeric comparison
        try:
            return int(rev_a) - int(rev_b)
        except ValueError:
            pass
        # Letter comparison
        if rev_a.isalpha() and rev_b.isalpha():
            return ord(rev_a.upper()) - ord(rev_b.upper())
        # Mixed: letters come after numbers in IFR convention
        if rev_a.isalpha() and rev_b.isdigit():
            return 1
        if rev_a.isdigit() and rev_b.isalpha():
            return -1
        return 0

    def read_excel_items(self, ws, layout: DeliverableLayout) -> Tuple[Dict[str, Dict], List[Dict]]:
        """Read doc IDs and revisions from Excel.
        Returns (items_dict, all_rows_list).
        items_dict: doc_id -> info (last row wins for duplicates, used for revision matching)
        all_rows_list: every row's data (used for description matching with duplicates)
        """
        items = {}
        all_rows = []
        for row in range(layout.header_row + 1, ws.max_row + 1):
            doc_id_val = ws.cell(row=row, column=layout.doc_id_col).value
            if not doc_id_val or not isinstance(doc_id_val, str):
                continue
            doc_id = doc_id_val.strip()
            if not doc_id:
                continue
            rev_val = ws.cell(row=row, column=layout.rev_col).value
            status_val = ws.cell(row=row, column=layout.status_col).value
            desc_val = ws.cell(row=row, column=layout.desc_col).value
            row_info = {
                'doc_id': doc_id,
                'row': row,
                'revision': str(rev_val).strip() if rev_val else '',
                'status': str(status_val).strip() if status_val else '',
                'description': str(desc_val).strip() if desc_val else '',
            }
            items[doc_id] = row_info
            all_rows.append(row_info)
        return items, all_rows

    def preload_doc_ids(self) -> Set[str]:
        """Lightweight: read only the doc-ID column from Deliverable Excel.

        Returns a set of all doc-IDs (uppercased) found in Excel.
        Used by _mirror_reports() to validate files whose parent folder
        prefix doesn't match (e.g. LMS client naming convention).
        """
        excel_path = self.find_deliverable_excel()
        if not excel_path:
            return set()
        try:
            wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
            ws = wb.active
            layout = self.detect_layout(ws)
            doc_ids = set()
            for row in range(layout.header_row + 1, ws.max_row + 1):
                val = ws.cell(row=row, column=layout.doc_id_col).value
                if val and isinstance(val, str) and val.strip():
                    doc_ids.add(val.strip().upper())
            wb.close()
            return doc_ids
        except Exception:
            return set()

    def _is_yellow_fill(self, cell) -> bool:
        """Check if a cell has yellow highlight fill."""
        if not cell.fill or not cell.fill.fgColor:
            return False
        color = cell.fill.fgColor
        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb)
            # Check for yellow-ish colors (FFFF99, FFFFFF00, etc)
            return rgb in ('FFFFFF99', '00FFFF99', 'FFFFFF00', '00FFFF00')
        return False

    def clear_previous_highlights(self, ws, layout: DeliverableLayout):
        """Clear change marker highlights (J=10 and N=14) and status fills (K-M) from previous runs."""
        no_fill = PatternFill(fill_type=None)
        clear_cols = [10, 14, layout.rev_col, layout.date_col, layout.status_col]
        for row in range(layout.header_row + 1, ws.max_row + 1):
            for col in clear_cols:
                cell = ws.cell(row=row, column=col)
                if self._is_yellow_fill(cell) or self._is_status_fill(cell):
                    cell.fill = no_fill

    def _is_status_fill(self, cell) -> bool:
        """Check if a cell has green or amber status fill."""
        if not cell.fill or not cell.fill.fgColor:
            return False
        color = cell.fill.fgColor
        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb)
            return rgb in ('FF00B050', '00FF00B050', 'FFFFC000', '00FFC000')
        return False

    # Approved to IFC folders (scanned for IFC status)
    APPROVED_IFC_FOLDERS = [
        "Design/Engineering/13. Client Sharepoint/1.IFR/1.Report/Approved to IFC",
        "Design/Engineering/13. Client Sharepoint/1.IFR/2.Drawing/Approved to IFC",
        "Design/Engineering/1. Drawings/3. IFR(Client)/1.Drawing/Approved to IFC",
        "Design/Engineering/1. Drawings/3. IFR(Client)/2.Reports/Approved to IFC",
    ]

    def scan_ifc_folder(self) -> Dict[str, Dict]:
        """Scan IFC(Client) folder AND all Approved to IFC folders.

        Returns {doc_id: {'revision': int_or_str, 'filename': str}}.
        For each doc_id, keeps only the highest revision found.
        Sources:
          1. 4. IFC(Client)/ — Drawing IFC PDFs (numeric revision)
          2. 13. Client Sharepoint/1.IFR/{1.Report,2.Drawing}/Approved to IFC/
          3. 3. IFR(Client)/{1.Drawing,2.Reports}/Approved to IFC/
        """
        ifc_info: Dict[str, Dict] = {}

        # Source 1: IFC(Client) folder (numeric revisions)
        ifc_dir = self.project_path / self.IFC_FOLDER
        if ifc_dir.exists():
            try:
                for f in ifc_dir.rglob("*.pdf"):
                    if f.name.startswith('~$'):
                        continue
                    doc_id = self.extract_doc_id(f.name)
                    if not doc_id:
                        continue
                    m = self._RE_REV_NUMBER.search(f.name)
                    rev_num = int(m.group(1)) if m else 0
                    if doc_id not in ifc_info or rev_num > ifc_info[doc_id]['revision']:
                        ifc_info[doc_id] = {'revision': rev_num, 'filename': f.name}
            except (OSError, PermissionError):
                pass

        # Source 2: Client Sharepoint Approved to IFC folders
        for rel_path in self.APPROVED_IFC_FOLDERS:
            approved_dir = self.project_path / rel_path
            if not approved_dir.exists():
                continue
            try:
                for f in approved_dir.iterdir():
                    if not f.is_file() or f.suffix.lower() != '.pdf':
                        continue
                    if f.name.startswith('~$'):
                        continue
                    doc_id = self.extract_doc_id(f.name)
                    if not doc_id:
                        continue
                    # Extract revision: try numeric first, then letter
                    m = self._RE_REV_NUMBER.search(f.name)
                    if m:
                        rev = int(m.group(1))
                    else:
                        m = self._RE_REV_LETTER.search(f.name)
                        rev = m.group(1).upper() if m else 0
                    # Only add if not already found with a higher revision
                    if doc_id not in ifc_info:
                        ifc_info[doc_id] = {'revision': rev, 'filename': f.name}
            except (OSError, PermissionError):
                pass

        return ifc_info

    def normalize_filename(self, filename: str, deliverable_desc: Optional[str] = None) -> Tuple[str, List[str]]:
        """Normalize filename conventions. Returns (normalized_name, list_of_changes).
        Rules:
        1. After FILE NO, separator must be '_' (not '-' or ' ')
        2. Revision format: '_Rev' + letter/number, no space (e.g. _RevA, _RevC)
        3. Description should match deliverable Excel column C (optional)
        """
        stem = Path(filename).stem
        ext = Path(filename).suffix
        changes = []
        doc_id = self.extract_doc_id(filename)
        if not doc_id:
            return filename, changes

        remainder = stem[len(doc_id):]

        # Rule 1: separator after doc-ID must be '_'
        if remainder and remainder[0] in ('-', ' '):
            old_sep = remainder[0]
            remainder = '_' + remainder[1:]
            changes.append(f"separator '{old_sep}' → '_' after FILE NO")

        # Rule 2: normalize revision format
        # _rA → _RevA, _Rev A → _RevA, _rev0 → _Rev0, Rev C → _RevC
        def rev_normalizer(m):
            prefix = m.group(1)   # separator (_, -, space or empty)
            rev_part = m.group(2) # 'Rev', 'rev', 'r', 'R'
            space = m.group(3)    # optional space
            letter = m.group(4)   # version letter/number
            normalized = f'_Rev{letter.upper()}' if letter.isalpha() else f'_Rev{letter}'
            original = m.group(0)
            if original != normalized:
                changes.append(f"revision '{original.strip()}' → '{normalized.strip()}'")
            return normalized

        remainder = re.sub(
            r'([_\s-])([Rr]ev|[Rr])\.?(\s?)([A-Z0-9])(?=[_.\s]|$)',
            rev_normalizer,
            remainder,
            flags=re.IGNORECASE
        )

        # Rule 3: description matching (optional)
        # Replace abbreviated or wrong descriptions with deliverable Excel column C
        if deliverable_desc:
            current_desc = self.extract_description(doc_id + remainder + ext)
            norm_current = self._normalize_desc(current_desc)
            norm_expected = self._normalize_desc(deliverable_desc)
            if norm_current and norm_expected and norm_current != norm_expected:
                ratio = self._desc_similarity(norm_current, norm_expected)
                if ratio < 0.7:
                    # Description is significantly different — suggest replacement
                    # Replace the description portion in remainder
                    # remainder = _description_RevX
                    rev_match = re.search(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*[A-Z0-9](?=[_.\s]|$)', remainder)
                    if rev_match:
                        rev_suffix = remainder[rev_match.start():]
                        remainder = f'_{deliverable_desc}{rev_suffix}'
                    else:
                        remainder = f'_{deliverable_desc}'
                    changes.append(f"description '{current_desc}' -> '{deliverable_desc}'")

        new_name = doc_id + remainder + ext
        return new_name, changes

    def cleanup_ss_folder(self, ss_path: Path, max_versions: int = 3):
        """Remove old timestamped copies in SS folder, keeping only the most recent N per base file."""
        if not ss_path.exists():
            return
        # Group files by base name (without timestamp suffix)
        ts_pattern = re.compile(r'^(.+?)_(\d{8}_\d{6})(\.\w+)$')
        groups = defaultdict(list)
        for f in ss_path.iterdir():
            if not f.is_file():
                continue
            m = ts_pattern.match(f.name)
            if m:
                base = m.group(1) + m.group(3)
                groups[base].append(f)

        for base, files in groups.items():
            if len(files) <= max_versions:
                continue
            # Sort by modification time descending, keep newest max_versions
            files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            for old_file in files[max_versions:]:
                try:
                    old_file.unlink()
                except Exception:
                    pass

    def cross_check(self, excel_path: Path, external_ifc_map: Optional[Dict] = None) -> DeliverableCrossCheckResult:
        """Perform full cross-check between source folders and deliverable Excel.

        Args:
            external_ifc_map: If provided (from Pipeline Stage 3 IFC Transmittal),
                merges with internal scan_ifc_folder() results. Format:
                {doc_id: {'revision': int, 'filename': str}}
        """
        if not OPENPYXL_AVAILABLE:
            result = DeliverableCrossCheckResult(
                project_name=self.project_path.name,
                excel_path=str(excel_path)
            )
            result.errors.append("openpyxl 未安装。请运行: pip install openpyxl")
            return result

        result = DeliverableCrossCheckResult(
            project_name=self.project_path.name,
            excel_path=str(excel_path)
        )

        # Load workbook with retry for Dropbox locks
        wb = self._load_workbook_with_retry(excel_path)
        if wb is None:
            result.errors.append(f"无法打开 Excel 文件: {excel_path}")
            return result

        ws = wb.active
        try:
            layout = self.detect_layout(ws)
        except FormatChangeWarning as e:
            result.errors.append(str(e))
            wb.close()
            return result

        # Scan source folders
        folder_items = self.scan_source_folders()

        # Read Excel items
        excel_items, all_excel_rows = self.read_excel_items(ws, layout)

        # Dimension 1: Item presence (with smart matching by description)
        matched_folder_ids = set()  # folder doc_ids that matched an Excel row
        matched_excel_rows = set()  # excel row numbers that matched a folder item

        # First pass: exact doc-ID matches
        for doc_id in folder_items:
            if doc_id in excel_items:
                matched_folder_ids.add(doc_id)
                matched_excel_rows.add(excel_items[doc_id]['row'])

        # Second pass: try description matching for unmatched folder items
        # Uses all_excel_rows to handle duplicate doc-IDs in Excel
        for doc_id, info in folder_items.items():
            if doc_id in matched_folder_ids:
                continue
            match_row_info = self._find_match_by_description(
                doc_id, info['description'], all_excel_rows, matched_excel_rows
            )
            if match_row_info:
                # Found a match by description — the Excel FILE NO is wrong
                matched_folder_ids.add(doc_id)
                matched_excel_rows.add(match_row_info['row'])
                excel_info = match_row_info
                result.doc_id_corrections.append({
                    'doc_id': doc_id,              # correct doc-ID from file
                    'old_doc_id': excel_info['doc_id'],  # wrong doc-ID in Excel
                    'row': excel_info['row'],
                    'description': info['description'],
                    'filename': info['filename'],
                })
                # Also check revision for this matched row
                folder_rev = info['revision']
                excel_rev = excel_info['revision']
                if folder_rev and excel_rev and self._compare_revisions(folder_rev, excel_rev) > 0:
                    result.revision_mismatches.append({
                        'doc_id': doc_id,
                        'row': excel_info['row'],
                        'excel_rev': excel_rev,
                        'folder_rev': folder_rev,
                        'filename': info['filename'],
                    })
            else:
                result.items_in_folders_not_excel.append({
                    'doc_id': doc_id,
                    'revision': info['revision'],
                    'filename': info['filename'],
                    'description': info['description'],
                    'source': info['source'],
                })

        for doc_id, info in excel_items.items():
            if info['row'] in matched_excel_rows:
                continue
            if doc_id not in folder_items:
                if info['status'].lower() not in ('n/a', 'reserved', ''):
                    result.items_in_excel_not_folders.append(doc_id)

        # Dimension 2: Revision comparison (for exact doc-ID matches)
        for doc_id, folder_info in folder_items.items():
            if doc_id not in excel_items:
                continue
            excel_info = excel_items[doc_id]
            folder_rev = folder_info['revision']
            excel_rev = excel_info['revision']
            if excel_info['status'].lower() in ('n/a', 'reserved'):
                continue
            if not folder_rev:
                continue
            if not excel_rev:
                # Excel has no revision but folder does → treat as update
                result.revision_mismatches.append({
                    'doc_id': doc_id,
                    'row': excel_info['row'],
                    'excel_rev': excel_rev or '',
                    'folder_rev': folder_rev,
                    'filename': folder_info['filename'],
                })
            elif self._compare_revisions(folder_rev, excel_rev) > 0:
                result.revision_mismatches.append({
                    'doc_id': doc_id,
                    'row': excel_info['row'],
                    'excel_rev': excel_rev,
                    'folder_rev': folder_rev,
                    'filename': folder_info['filename'],
                })
            elif not excel_info.get('status', '').strip():
                # Revision matches but status is empty → set to Submitted
                result.revision_mismatches.append({
                    'doc_id': doc_id,
                    'row': excel_info['row'],
                    'excel_rev': excel_rev,
                    'folder_rev': folder_rev,
                    'filename': folder_info['filename'],
                })

        # IFC status tracking: scan IFC folder and flag items as "Approved IFC"
        ifc_info = self.scan_ifc_folder()
        if external_ifc_map:
            for doc_id, info in external_ifc_map.items():
                if doc_id not in ifc_info:
                    ifc_info[doc_id] = info
        for doc_id, info in ifc_info.items():
            if doc_id in excel_items:
                excel_info = excel_items[doc_id]
                current_status = excel_info.get('status', '').strip()
                if current_status.lower() != 'approved ifc':
                    result.status_updates.append({
                        'doc_id': doc_id,
                        'row': excel_info['row'],
                        'old_status': current_status,
                        'new_status': 'Approved IFC',
                        'ifc_rev': info['revision'],
                    })

        # Naming warnings: check filename conventions for matched items
        for doc_id, info in folder_items.items():
            excel_desc = None
            if doc_id in excel_items:
                excel_desc = excel_items[doc_id].get('description')
            normalized, changes = self.normalize_filename(info['filename'], excel_desc)
            if changes:
                result.naming_warnings.append({
                    'doc_id': doc_id,
                    'filename': info['filename'],
                    'suggested': normalized,
                    'changes': changes,
                    'source': info['source'],
                })

        wb.close()
        return result

    def apply_updates(self, excel_path: Path, check_result: DeliverableCrossCheckResult) -> DeliverableCrossCheckResult:
        """Apply cross-check results to the Excel file."""
        if not OPENPYXL_AVAILABLE:
            check_result.errors.append("openpyxl 未安装")
            return check_result

        if (not check_result.items_in_folders_not_excel and
                not check_result.revision_mismatches and
                not check_result.doc_id_corrections and
                not check_result.status_updates):
            return check_result

        wb = self._load_workbook_with_retry(excel_path)
        if wb is None:
            check_result.errors.append(f"无法打开 Excel 文件: {excel_path}")
            return check_result

        ws = wb.active
        try:
            layout = self.detect_layout(ws)
        except FormatChangeWarning as e:
            check_result.errors.append(str(e))
            wb.close()
            return check_result

        # Clear previous highlights
        self.clear_previous_highlights(ws, layout)

        # Apply doc-ID corrections (matched by description, FILE NO was wrong)
        for correction in check_result.doc_id_corrections:
            row = correction['row']
            ws.cell(row=row, column=layout.doc_id_col).value = correction['doc_id']
            # Marker highlights on J (col 10) and N (col 14)
            ws.cell(row=row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self.CHANGE_MARKER_FILL
            # Status-based fill on K-M
            ws.cell(row=row, column=layout.rev_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self.STATUS_SUBMITTED_FILL
            check_result.rows_updated += 1

        # Apply revision updates
        for mismatch in check_result.revision_mismatches:
            row = mismatch['row']
            ws.cell(row=row, column=layout.rev_col).value = mismatch['folder_rev']
            ws.cell(row=row, column=layout.status_col).value = 'Submitted'
            ws.cell(row=row, column=layout.date_col).value = datetime.now().strftime('%d/%m/%y')
            # Marker highlights on J and N
            ws.cell(row=row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self.CHANGE_MARKER_FILL
            # Status-based fill on K-M (amber for Submitted)
            ws.cell(row=row, column=layout.rev_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self.STATUS_SUBMITTED_FILL
            check_result.rows_updated += 1

        # Apply IFC status updates (revision, date, and status)
        for update in check_result.status_updates:
            row = update['row']
            ws.cell(row=row, column=layout.status_col).value = 'Approved IFC'
            # Write IFC numerical revision to K column (0, 1, 2, 3...)
            ifc_rev = update.get('ifc_rev')
            if ifc_rev is not None:
                ws.cell(row=row, column=layout.rev_col).value = str(ifc_rev)
            # Write submission date
            ws.cell(row=row, column=layout.date_col).value = datetime.now().strftime('%d/%m/%y')
            # Green fill on K-M for approved
            ws.cell(row=row, column=layout.rev_col).fill = self.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self.STATUS_APPROVED_FILL
            # Marker highlights on J and N
            ws.cell(row=row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self.CHANGE_MARKER_FILL
            check_result.rows_updated += 1

        # Insert new items
        for item in check_result.items_in_folders_not_excel:
            insert_row = ws.max_row + 1
            ws.cell(row=insert_row, column=layout.doc_id_col).value = item['doc_id']
            ws.cell(row=insert_row, column=layout.desc_col).value = item['description']
            if item['revision']:
                ws.cell(row=insert_row, column=layout.rev_col).value = item['revision']
            # Marker highlights on J and N
            ws.cell(row=insert_row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=insert_row, column=14).fill = self.CHANGE_MARKER_FILL
            # Status-based fill on K-M
            ws.cell(row=insert_row, column=layout.rev_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=insert_row, column=layout.date_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=insert_row, column=layout.status_col).fill = self.STATUS_SUBMITTED_FILL
            check_result.rows_inserted += 1

        # Update file revision and last updated date
        old_file_rev = ws[layout.file_rev_cell].value
        new_file_rev = self._increment_file_revision(old_file_rev)
        # Preserve prefix format (e.g. "Revision 1.9" → "Revision 2.0")
        old_str = str(old_file_rev).strip() if old_file_rev else ''
        prefix_match = re.match(r'^((?:Revision|Rev)\s*)', old_str, re.IGNORECASE)
        if prefix_match:
            ws[layout.file_rev_cell].value = f"{prefix_match.group(1)}{new_file_rev}"
        else:
            ws[layout.file_rev_cell].value = new_file_rev
        ws[layout.last_updated_cell].value = datetime.now().strftime('%d/%m/%y')
        check_result.new_file_rev = str(new_file_rev)

        if not self.dry_run:
            # Save to a new filename with updated revision
            new_path = self._get_new_filename(excel_path, new_file_rev)
            old_path = excel_path

            # Supersede old version
            self._supersede_file(old_path)

            # Save new version
            wb.save(str(to_long_path(new_path)))
            wb.close()

            # Sync to target locations
            self._sync_deliverable(new_path)

            # Cleanup old versions in SS folders (keep max 3)
            ss_folder = old_path.parent / "SS"
            if ss_folder.exists():
                self.cleanup_ss_folder(ss_folder, max_versions=3)

            check_result.excel_path = str(new_path)
        else:
            wb.close()

        return check_result

    def _load_workbook_with_retry(self, path: Path, max_retries: int = 3):
        """Load workbook with retry for Dropbox locks."""
        for attempt in range(max_retries):
            try:
                return openpyxl.load_workbook(str(to_long_path(path)))
            except PermissionError:
                if attempt < max_retries - 1:
                    wait = 2 ** attempt
                    print(f"    [!] 文件被锁定，{wait}秒后重试... (尝试 {attempt + 1}/{max_retries})")
                    time.sleep(wait)
                else:
                    return None
            except Exception:
                return None
        return None

    def _increment_file_revision(self, current_rev) -> str:
        """Increment file revision: 1.9→2.0, 14→15, 'Revision 1.9'→'2.0', etc."""
        if current_rev is None:
            return '1.0'
        rev_str = str(current_rev).strip()
        # Strip prefix like "Revision " or "Rev "
        rev_str = re.sub(r'^(?:Revision|Rev)\s*', '', rev_str, flags=re.IGNORECASE).strip()
        if '.' in rev_str:
            try:
                parts = rev_str.split('.')
                major = int(parts[0])
                minor = int(parts[1])
                if minor >= 9:
                    return f"{major + 1}.0"
                else:
                    return f"{major}.{minor + 1}"
            except (ValueError, IndexError):
                return rev_str
        else:
            try:
                return str(int(rev_str) + 1)
            except ValueError:
                return rev_str

    def _get_new_filename(self, old_path: Path, new_rev: str) -> Path:
        """Generate new filename with updated revision."""
        stem = old_path.stem
        ext = old_path.suffix
        # Match _rev or _Revision_ followed by version number (e.g., _rev1.9, _Revision_2.0)
        new_stem = re.sub(
            r'(_rev(?:ision_?)?)[\d.]+',
            rf'\g<1>{new_rev}',
            stem,
            flags=re.IGNORECASE,
            count=1
        )
        if new_stem == stem:
            # No revision pattern found, append it
            new_stem = f"{stem}_rev{new_rev}"
        return old_path.parent / (new_stem + ext)

    def _supersede_file(self, file_path: Path):
        """Move file to SUPERSEDED subfolder (copy-then-delete for Dropbox)."""
        ss_folder = file_path.parent / "SS"
        if not ss_folder.exists():
            # Check for existing superseded folder
            for item in file_path.parent.iterdir():
                if item.is_dir() and item.name.lower() in ('ss', 'superseded', 'superceded'):
                    ss_folder = item
                    break
            else:
                ss_folder.mkdir(exist_ok=True)

        dest = ss_folder / file_path.name
        if dest.exists():
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = ss_folder / f"{file_path.stem}_{ts}{file_path.suffix}"

        try:
            # Copy then delete (Dropbox-safe)
            shutil.copy2(str(to_long_path(file_path)), str(to_long_path(dest)))
            to_long_path(file_path).unlink()
        except Exception as e:
            print(f"    [!] 无法supersede文件: {file_path.name} - {e}")

    def _sync_deliverable(self, new_file: Path):
        """Sync new deliverable file to target locations."""
        # Primary sync target (already in the right place if file is in 3.Deliverables)
        primary_target = self.project_path / self.PRIMARY_SYNC
        if primary_target.exists() and new_file.parent != primary_target:
            dest = primary_target / new_file.name
            try:
                # Supersede old versions in target
                self._supersede_old_versions_in_folder(primary_target, new_file.name)
                shutil.copy2(str(to_long_path(new_file)), str(to_long_path(dest)))
            except Exception as e:
                print(f"    [!] 主同步失败: {e}")

        # Secondary sync target (only if exists)
        secondary_target = self.project_path / self.SECONDARY_SYNC
        if secondary_target.exists():
            dest = secondary_target / new_file.name
            try:
                self._supersede_old_versions_in_folder(secondary_target, new_file.name)
                shutil.copy2(str(to_long_path(new_file)), str(to_long_path(dest)))
                print(f"    [v] 已同步到 Client Sharepoint: {new_file.name}")
            except Exception as e:
                print(f"    [!] 次要同步失败: {e}")

    def _supersede_old_versions_in_folder(self, folder: Path, new_filename: str):
        """Move old versions of the deliverable in a folder to SS."""
        # Find the base name pattern (without revision)
        base_pattern = re.sub(r'(?:_rev|_Revision_?)[\d.]+', '', Path(new_filename).stem,
                             flags=re.IGNORECASE)
        ss_folder = folder / "SS"
        for f in folder.iterdir():
            if not f.is_file():
                continue
            if f.name == new_filename:
                continue
            f_base = re.sub(r'(?:_rev|_Revision_?)[\d.]+', '', f.stem, flags=re.IGNORECASE)
            if f_base.lower() == base_pattern.lower() and f.suffix.lower() == Path(new_filename).suffix.lower():
                if not ss_folder.exists():
                    ss_folder.mkdir(exist_ok=True)
                dest = ss_folder / f.name
                try:
                    shutil.copy2(str(to_long_path(f)), str(to_long_path(dest)))
                    to_long_path(f).unlink()
                except Exception:
                    pass


# =============================================================================
# IFC Transmittal Manager - Version Management & Transmittal Generation
# =============================================================================

@dataclass
class IFCResult:
    """Result of IFC management run."""
    total_ifc_files: int = 0
    duplicates_archived: int = 0
    new_files_since_last: int = 0
    deliverable_updates: int = 0
    transmittal_path: Optional[str] = None
    transmittal_number: int = 0
    errors: List[str] = field(default_factory=list)
    file_list: List[Dict] = field(default_factory=list)  # {name, action, doc_id, revision}


class IFCTransmittalManager:
    """Manage IFC file versions, update deliverable Excel, generate transmittals.

    Workflow:
      1. Scan 4. IFC(Client)/ for PDFs, group by doc-ID
      2. Identify duplicates → move old versions to SS/
      3. Collect new/updated files since last run
      4. Update deliverable Excel column L with IFC revision numbers
      5. Generate transmittal Excel with header info from config
      6. Save transmittal to 8. Deliverables/Transmittal/ + sync targets
      7. Save run state for incremental tracking
    """

    IFC_CLIENT_REL = "Design/Engineering/1. Drawings/4. IFC(Client)"
    TRANSMITTAL_REL = "Design/Engineering/8. Deliverables/Transmittal"

    # Reuse patterns from module level
    RE_DOC_ID = re.compile(r'^(\d{5}-[A-Z]{2}-\d{3})', re.IGNORECASE)
    RE_IFC_REV = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?:[_\s-]?IFC)?(?=[_.\s]|$)', re.IGNORECASE)

    # Transmittal styling
    HEADER_FONT = openpyxl.styles.Font(bold=True, size=14) if OPENPYXL_AVAILABLE else None
    SUBHEADER_FONT = openpyxl.styles.Font(bold=True, size=11) if OPENPYXL_AVAILABLE else None
    NORMAL_FONT = openpyxl.styles.Font(size=10) if OPENPYXL_AVAILABLE else None
    COL_HEADER_FONT = openpyxl.styles.Font(bold=True, size=10) if OPENPYXL_AVAILABLE else None
    COL_HEADER_FILL = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2',
                                   fill_type='solid') if OPENPYXL_AVAILABLE else None
    THIN_BORDER = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin'),
    ) if OPENPYXL_AVAILABLE else None
    IFC_APPROVED_FILL = PatternFill(start_color='FF00B050', end_color='FF00B050',
                                     fill_type='solid') if OPENPYXL_AVAILABLE else None

    def __init__(self, project_path, config=None, dry_run=False):
        self.project_path = Path(project_path)
        self.config = config or {}
        self.dry_run = dry_run
        self.logger = logging.getLogger(self.__class__.__name__)

        # Config-driven paths
        ifc_cfg = self.config.get("ifc_transmittal", {})
        self.transmittal_rel = ifc_cfg.get("transmittal_path", self.TRANSMITTAL_REL)
        self.sync_targets = ifc_cfg.get("sync_targets", [])
        self.state_file_name = ifc_cfg.get("state_file", "ifc_state.json")

        # Per-project transmittal header: check project_overrides first, then global
        project_folder_name = self.project_path.name  # e.g. "GG-31 Warnertown BESS"
        project_overrides = ifc_cfg.get("project_overrides", {})
        if project_folder_name in project_overrides:
            self.header_config = project_overrides[project_folder_name].get(
                "transmittal_header", ifc_cfg.get("transmittal_header", {}))
        else:
            self.header_config = ifc_cfg.get("transmittal_header", {})

        self.ifc_dir = self.project_path / self.IFC_CLIENT_REL
        self.transmittal_dir = self.project_path / self.transmittal_rel
        self.state_file = self.transmittal_dir / self.state_file_name

    # ── 2a. Version Management ──

    def scan_ifc_files(self) -> Dict[str, List[Path]]:
        """Scan 4. IFC(Client)/ for PDFs, group by doc-ID. Excludes SS/."""
        grouped = defaultdict(list)
        if not self.ifc_dir.exists():
            self.logger.warning(f"IFC directory not found: {self.ifc_dir}")
            return grouped
        for f in self.ifc_dir.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() != '.pdf':
                continue
            if f.name.startswith('~$'):
                continue
            doc_id = self._extract_doc_id(f.name)
            if doc_id:
                grouped[doc_id].append(f)
        return dict(grouped)

    def identify_duplicates(self, grouped: Dict[str, List[Path]]) -> List[Tuple[Path, Path]]:
        """For doc-IDs with >1 file, identify old versions to archive.

        Returns list of (source, destination) tuples for files to move to SS/.
        """
        ss_folder = self.ifc_dir / "SS"
        duplicates = []
        for doc_id, files in grouped.items():
            if len(files) <= 1:
                continue
            # Sort by revision number (desc), then by mtime (desc) for tie-breaking
            scored = []
            for f in files:
                rev = self._extract_revision_number(f.name)
                mtime = f.stat().st_mtime
                scored.append((f, rev, mtime))
            scored.sort(key=lambda x: (x[1], x[2]), reverse=True)
            # Keep first (highest rev, newest), archive the rest
            for f, rev, mtime in scored[1:]:
                dest = ss_folder / f.name
                duplicates.append((f, dest))
        return duplicates

    def move_old_to_ss(self, duplicates: List[Tuple[Path, Path]]) -> int:
        """Move old versions to SS/ subfolder."""
        if not duplicates:
            return 0
        ss_folder = self.ifc_dir / "SS"
        if not self.dry_run:
            try:
                ss_folder.mkdir(exist_ok=True)
            except (OSError, PermissionError):
                try:
                    to_long_path(ss_folder).mkdir(exist_ok=True)
                except Exception as e:
                    self.logger.error(f"Cannot create SS folder: {e}")
                    return 0
        moved = 0
        for source, dest in duplicates:
            if self.dry_run:
                self.logger.info(f"[DRY-RUN] Would move: {source.name} -> SS/")
                moved += 1
            else:
                try:
                    if dest.exists():
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        dest = ss_folder / f"{dest.stem}_{timestamp}{dest.suffix}"
                    shutil.move(str(to_long_path(source)), str(to_long_path(dest)))
                    self.logger.info(f"Archived: {source.name} -> SS/")
                    moved += 1
                except Exception as e:
                    self.logger.error(f"Move failed: {source.name} - {e}")
        return moved

    # ── 2b. Deliverable Update ──

    def update_deliverable_ifc_status(self, ifc_files: Dict[str, Dict]) -> int:
        """Update deliverable Excel column L with IFC revision numbers.

        Args:
            ifc_files: {doc_id: {revision: int, filename: str}}

        Returns:
            Number of rows updated.
        """
        if not ifc_files:
            return 0

        dm = DeliverableManager(self.project_path, dry_run=self.dry_run)
        excel_path = dm.find_deliverable_excel()
        if not excel_path:
            self.logger.warning("Deliverable Excel not found")
            return 0

        if self.dry_run:
            # In dry-run, just count matches
            try:
                wb = openpyxl.load_workbook(str(to_long_path(excel_path)))
                ws = wb.active
                layout = dm.detect_layout(ws)
                count = 0
                for row in range(layout.header_row + 1, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=layout.doc_id_col).value
                    if cell_val and str(cell_val).strip() in ifc_files:
                        count += 1
                wb.close()
                return count
            except Exception as e:
                self.logger.error(f"Dry-run deliverable scan failed: {e}")
                return 0

        # Actual update
        try:
            wb = openpyxl.load_workbook(str(to_long_path(excel_path)))
            ws = wb.active
            layout = dm.detect_layout(ws)
            updated = 0

            for row in range(layout.header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=layout.doc_id_col).value
                if not cell_val:
                    continue
                doc_id = str(cell_val).strip()
                if doc_id not in ifc_files:
                    continue

                rev_num = ifc_files[doc_id]["revision"]
                # Write IFC numerical revision to K column (0, 1, 2, 3...)
                if rev_num is not None:
                    ws.cell(row=row, column=layout.rev_col, value=str(rev_num))
                # Write submission date to date column (L)
                ws.cell(row=row, column=layout.date_col, value=datetime.now().strftime('%d/%m/%y'))
                # Set status to "Approved IFC" with green fill
                ws.cell(row=row, column=layout.status_col, value="Approved IFC")
                # Green fill on K-M
                ws.cell(row=row, column=layout.rev_col).fill = self.IFC_APPROVED_FILL
                ws.cell(row=row, column=layout.date_col).fill = self.IFC_APPROVED_FILL
                ws.cell(row=row, column=layout.status_col).fill = self.IFC_APPROVED_FILL
                updated += 1

            if updated > 0:
                # Version the deliverable: move old to SS, save new
                self._version_and_save_deliverable(wb, excel_path)
            else:
                wb.close()

            return updated
        except Exception as e:
            self.logger.error(f"Deliverable update failed: {e}")
            return 0

    def _version_and_save_deliverable(self, wb, excel_path: Path):
        """Save workbook, versioning old file to SS/, then sync to all deliverable locations."""
        ss_folder = excel_path.parent / "SS"
        ss_folder.mkdir(exist_ok=True)

        # Move current file to SS
        if excel_path.exists():
            ss_dest = ss_folder / excel_path.name
            if ss_dest.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                ss_dest = ss_folder / f"{excel_path.stem}_{timestamp}{excel_path.suffix}"
            try:
                shutil.copy2(str(to_long_path(excel_path)), str(to_long_path(ss_dest)))
            except Exception as e:
                self.logger.warning(f"Failed to backup deliverable: {e}")

        # Save updated workbook
        try:
            wb.save(str(to_long_path(excel_path)))
            self.logger.info(f"Deliverable updated: {excel_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to save deliverable: {e}")
        finally:
            wb.close()

        # Sync to all deliverable locations (same as DeliverableManager._sync_deliverable)
        dm = DeliverableManager(self.project_path, dry_run=self.dry_run)
        if not self.dry_run:
            dm._sync_deliverable(excel_path)

    # ── 2c. Transmittal Generation ──

    def get_last_run_state(self) -> Dict:
        """Read state from ifc_state.json."""
        if not self.state_file.exists():
            return {}
        try:
            with open(str(to_long_path(self.state_file)), 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def get_last_run_timestamp(self) -> Optional[datetime]:
        """Get timestamp of last run."""
        state = self.get_last_run_state()
        ts = state.get("last_run")
        if ts:
            try:
                return datetime.fromisoformat(ts)
            except (ValueError, TypeError):
                pass
        return None

    def collect_new_ifc_files(self, since: Optional[datetime],
                               transmitted: Set[str]) -> List[Dict]:
        """Collect IFC files new or updated since last run.

        Returns list of {path, doc_id, revision, action} dicts.
        action is 'New' or 'Update'.
        """
        results = []
        if not self.ifc_dir.exists():
            return results

        for f in self.ifc_dir.iterdir():
            if not f.is_file() or f.suffix.lower() != '.pdf':
                continue
            if f.name.startswith('~$'):
                continue

            doc_id = self._extract_doc_id(f.name)
            if not doc_id:
                continue

            rev = self._extract_revision_number(f.name)
            mtime = datetime.fromtimestamp(f.stat().st_mtime)

            # Determine if this file should be included
            include = False
            if since is None:
                include = True  # First run: include all
            elif mtime > since:
                include = True  # Modified since last run
            elif doc_id not in transmitted:
                include = True  # New doc-ID not previously transmitted

            if not include:
                continue

            action = "Update" if doc_id in transmitted else "New"
            results.append({
                "path": f,
                "doc_id": doc_id,
                "revision": rev,
                "action": action,
                "filename": f.name,
            })

        # Sort by doc_id for consistent ordering
        results.sort(key=lambda x: x["doc_id"])
        return results

    def determine_next_tsmt_number(self) -> int:
        """Find next transmittal number (shared sequence with IFR).

        Scans 8. Deliverables/Transmittal/ for existing TSMT files.
        """
        max_num = 0
        if self.transmittal_dir.exists():
            for f in self.transmittal_dir.iterdir():
                if not f.is_file():
                    continue
                m = re.search(r'TSMT[_-]?(\d+)', f.name, re.IGNORECASE)
                if m:
                    num = int(m.group(1))
                    if num > max_num:
                        max_num = num

        # Also check state file for last known number
        state = self.get_last_run_state()
        last_num = state.get("last_tsmt_number", 0)
        max_num = max(max_num, last_num)

        return max_num + 1

    def _get_project_number(self) -> str:
        """Extract project number from header config or project path.

        Supports formats: '50023' (5-digit), 'GG-31' (GG prefix), 'GG31' (no dash).
        """
        project_str = self.header_config.get("project", "")
        # Try 5-digit number first (e.g. "50023 ADAMS RD")
        m = re.match(r'(\d{5})', project_str)
        if m:
            return m.group(1)
        # Try GG-style (e.g. "GG-31 Warnertown BESS")
        m = re.match(r'(GG-?\d+)', project_str, re.IGNORECASE)
        if m:
            return m.group(1)
        # Fallback: try extracting from project path folder name
        folder_name = self.project_path.name  # e.g. "GG-31 Warnertown BESS"
        m2 = re.match(r'(GG-?\d+)', folder_name, re.IGNORECASE)
        if m2:
            return m2.group(1)
        for part in self.project_path.parts:
            m3 = re.match(r'(\d{5})', part)
            if m3:
                return m3.group(1)
        return "00000"

    def generate_transmittal(self, files: List[Dict], tsmt_num: int,
                              run_date: datetime) -> Optional[Path]:
        """Generate transmittal Excel workbook.

        Args:
            files: List of {path, doc_id, revision, action, filename}
            tsmt_num: Transmittal sequence number
            run_date: Date for the transmittal

        Returns:
            Path to generated transmittal file, or None on error.
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl not available, cannot generate transmittal")
            return None

        project_num = self._get_project_number()
        tsmt_filename = f"{project_num}-TSMT-{tsmt_num}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transmittal"

        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30

        # ── Header section (rows 1-11) ──
        ws.cell(row=1, column=2, value=self.header_config.get("company", "")).font = self.HEADER_FONT
        ws.merge_cells('B1:G1')

        # Company details
        details = self.header_config.get("company_details", "")
        ws.cell(row=2, column=2, value=details).font = self.NORMAL_FONT
        ws.merge_cells('B2:G2')
        ws.row_dimensions[2].height = 50

        # Transmittal number and date
        ws.cell(row=4, column=2, value="TRANSMITTAL").font = self.SUBHEADER_FONT
        ws.cell(row=4, column=5, value=f"No: {project_num}-TSMT-{tsmt_num}").font = self.SUBHEADER_FONT
        ws.cell(row=5, column=5, value=f"Date: {run_date.strftime('%d/%m/%Y')}").font = self.NORMAL_FONT

        # To / From
        ws.cell(row=6, column=2, value="To:").font = self.SUBHEADER_FONT
        ws.cell(row=6, column=3, value=self.header_config.get("to", "")).font = self.NORMAL_FONT
        ws.cell(row=7, column=3, value=self.header_config.get("to_address", "")).font = self.NORMAL_FONT
        ws.row_dimensions[7].height = 30

        ws.cell(row=8, column=2, value="Attention:").font = self.SUBHEADER_FONT
        ws.cell(row=8, column=3, value=self.header_config.get("attention", "")).font = self.NORMAL_FONT

        ws.cell(row=9, column=2, value="Project:").font = self.SUBHEADER_FONT
        ws.cell(row=9, column=3, value=self.header_config.get("project", "")).font = self.NORMAL_FONT

        ws.cell(row=10, column=2, value="Subject:").font = self.SUBHEADER_FONT
        ws.cell(row=10, column=3, value=self.header_config.get("subject", "")).font = self.NORMAL_FONT

        # ── Column headers (row 12) ──
        headers = [
            (2, "Item NO."),
            (3, "Doc ID"),
            (4, "Qty"),
            (5, "File Name"),
            (6, "Comment"),
            (7, "Description"),
        ]
        for col, text in headers:
            cell = ws.cell(row=12, column=col, value=text)
            cell.font = self.COL_HEADER_FONT
            cell.fill = self.COL_HEADER_FILL
            cell.border = self.THIN_BORDER

        # ── Data rows (14, 16, 18... alternating) ──
        data_row = 14
        for idx, file_info in enumerate(files, 1):
            ws.cell(row=data_row, column=2, value=idx).border = self.THIN_BORDER
            ws.cell(row=data_row, column=3, value=file_info["doc_id"]).border = self.THIN_BORDER
            ws.cell(row=data_row, column=4, value=1).border = self.THIN_BORDER
            ws.cell(row=data_row, column=5, value=file_info["filename"]).border = self.THIN_BORDER
            ws.cell(row=data_row, column=6, value=file_info["action"]).border = self.THIN_BORDER

            # Extract description from filename
            desc = self._extract_description(file_info["filename"], file_info["doc_id"])
            ws.cell(row=data_row, column=7, value=desc).border = self.THIN_BORDER

            data_row += 2  # Alternating rows

        # Save transmittal
        self.transmittal_dir.mkdir(parents=True, exist_ok=True)
        tsmt_path = self.transmittal_dir / tsmt_filename

        if self.dry_run:
            wb.close()
            return tsmt_path

        try:
            wb.save(str(to_long_path(tsmt_path)))
            self.logger.info(f"Transmittal saved: {tsmt_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to save transmittal: {e}")
            wb.close()
            return None
        wb.close()

        # Sync to targets
        self._sync_transmittal(tsmt_path)

        return tsmt_path

    def _sync_transmittal(self, tsmt_path: Path):
        """Copy transmittal to sync target directories."""
        for rel_target in self.sync_targets:
            target_dir = self.project_path / rel_target
            if not target_dir.exists():
                try:
                    target_dir.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    self.logger.warning(f"Cannot create sync target: {target_dir} - {e}")
                    continue
            dest = target_dir / tsmt_path.name
            try:
                shutil.copy2(str(to_long_path(tsmt_path)), str(to_long_path(dest)))
                self.logger.info(f"Synced transmittal to: {rel_target}")
            except Exception as e:
                self.logger.error(f"Transmittal sync failed to {rel_target}: {e}")

    def _sync_ifc_files_to_sharepoint(self):
        """Sync IFC PDFs from 4. IFC(Client)/ to 13. Client Sharepoint/2.IFC/."""
        if self.dry_run:
            return
        for rel_target in self.sync_targets:
            target_dir = self.project_path / rel_target
            if not target_dir.exists():
                try:
                    target_dir.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    self.logger.warning(f"Cannot create IFC sync target: {target_dir} - {e}")
                    continue
            if not self.ifc_dir.exists():
                continue
            synced = 0
            for pdf in self.ifc_dir.glob("*.pdf"):
                if pdf.name.startswith('~$'):
                    continue
                dest = target_dir / pdf.name
                # Only copy if newer or different size (idempotent)
                if dest.exists():
                    src_stat = pdf.stat()
                    dst_stat = dest.stat()
                    if abs(src_stat.st_size - dst_stat.st_size) < 2 and abs(src_stat.st_mtime - dst_stat.st_mtime) < 2:
                        continue
                try:
                    shutil.copy2(str(to_long_path(pdf)), str(to_long_path(dest)))
                    synced += 1
                except Exception as e:
                    self.logger.error(f"IFC sync failed: {pdf.name} - {e}")
            if synced:
                self.logger.info(f"Synced {synced} IFC file(s) to {rel_target}")

    def save_run_state(self, run_date: datetime, tsmt_num: int,
                        transmitted_files: List[str]):
        """Save run state to ifc_state.json."""
        if self.dry_run:
            return
        state = self.get_last_run_state()
        # Merge transmitted files with existing
        existing = set(state.get("transmitted_files", []))
        existing.update(transmitted_files)

        new_state = {
            "last_run": run_date.isoformat(),
            "last_tsmt_number": tsmt_num,
            "transmitted_files": sorted(existing),
        }
        try:
            self.transmittal_dir.mkdir(parents=True, exist_ok=True)
            with open(str(to_long_path(self.state_file)), 'w', encoding='utf-8') as f:
                json.dump(new_state, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"Failed to save state: {e}")

    # ── 2d. Orchestrator ──

    # Paths to version-manage (relative to project_path)
    VERSION_MANAGED_PATHS = [
        "Design/Engineering/1. Drawings/4. IFC(Client)",
        "Design/Engineering/13. Client Sharepoint/1.IFR/2.Drawing",
        "Design/Engineering/13. Client Sharepoint/2.IFC",
    ]

    def _scan_and_dedup(self) -> Tuple[Dict, Dict, int]:
        """Scan IFC files, group by doc-ID, dedup, build ifc_map.

        Returns: (grouped_files, all_ifc_map, duplicates_archived)
        - grouped_files: {doc_id: [Path, ...]}
        - all_ifc_map: {doc_id: {'revision': int, 'filename': str}}
        - duplicates_archived: count of old files moved to SS/
        """
        grouped = self.scan_ifc_files()
        dupes_archived = 0

        duplicates = self.identify_duplicates(grouped)
        if duplicates:
            extra_moved = self.move_old_to_ss(duplicates)
            dupes_archived += extra_moved
            grouped = self.scan_ifc_files()

        all_ifc_map = {}
        for doc_id, files in grouped.items():
            best_rev = -1
            best_file = None
            for f in files:
                rev = self._extract_revision_number(f.name)
                if rev > best_rev:
                    best_rev = rev
                    best_file = f
            if best_file:
                all_ifc_map[doc_id] = {
                    "revision": best_rev,
                    "filename": best_file.name,
                }

        return grouped, all_ifc_map, dupes_archived

    def _collect_and_transmit(self, run_date: datetime, all_ifc_map: Dict,
                               result: IFCResult) -> IFCResult:
        """Collect new files, generate transmittal, sync sharepoint, save state.

        Mutates and returns the given result object.
        """
        state = self.get_last_run_state()
        since = self.get_last_run_timestamp()
        transmitted = set(state.get("transmitted_files", []))
        new_files = self.collect_new_ifc_files(since, transmitted)
        result.new_files_since_last = len(new_files)

        if not new_files:
            return result

        tsmt_num = self.determine_next_tsmt_number()
        tsmt_path = self.generate_transmittal(new_files, tsmt_num, run_date)
        if tsmt_path:
            result.transmittal_path = str(tsmt_path)
            result.transmittal_number = tsmt_num

        self._sync_ifc_files_to_sharepoint()

        for fi in new_files:
            result.file_list.append({
                "name": fi["filename"],
                "action": fi["action"],
                "doc_id": fi["doc_id"],
                "revision": fi["revision"],
            })

        transmitted_filenames = [fi["filename"] for fi in new_files]
        self.save_run_state(run_date, tsmt_num, transmitted_filenames)

        return result

    def run(self, run_date: datetime = None) -> IFCResult:
        """Run the full IFC management workflow.

        1. Version-manage IFC(Client) + Client Sharepoint/2.Drawing
        2. Scan & deduplicate IFC files (by doc-ID)
        3. Update deliverable Excel
        4. Collect new/updated files since last run
        5. Generate transmittal + sync + save state
        """
        if run_date is None:
            run_date = datetime.now()

        result = IFCResult()

        # 1. Version Management — clean old revisions in IFC(Client) + Sharepoint Drawing
        vm = VersionManager(str(self.project_path.parent), dry_run=self.dry_run)
        vm_total_moved = 0
        for rel_path in self.VERSION_MANAGED_PATHS:
            target_dir = self.project_path / rel_path
            if not target_dir.exists():
                continue
            short_name = rel_path.rsplit("/", 1)[-1]
            print(f"\n  [版本管理] {short_name}")
            vm_stats = vm.process_directory(target_dir, show_details=True)
            vm_total_moved += vm_stats.get("moved", 0)
        result.duplicates_archived = vm_total_moved

        # 2. Scan + dedup + build ifc_map
        grouped, all_ifc_map, extra_dupes = self._scan_and_dedup()
        result.total_ifc_files = sum(len(files) for files in grouped.values())
        result.duplicates_archived += extra_dupes

        # 3. Update deliverable Excel with ALL current IFC revisions
        result.deliverable_updates = self.update_deliverable_ifc_status(all_ifc_map)

        # 4-5. Collect new files, generate transmittal, sync, save state
        result = self._collect_and_transmit(run_date, all_ifc_map, result)

        return result

    def run_for_pipeline(self) -> dict:
        """Pipeline context: scan + dedup only, return ifc_map for Stage 4.

        Does NOT write Excel, does NOT update ifc_state.json, does NOT generate
        transmittal. The ifc_map is passed to Stage 4 (DeliverableManager) for
        unified Excel write.
        """
        grouped, all_ifc_map, dupes = self._scan_and_dedup()
        return {
            'total_ifc_files': sum(len(files) for files in grouped.values()),
            'duplicates_archived': dupes,
            'ifc_map': all_ifc_map,
        }

    # ── Helper methods ──

    def _extract_doc_id(self, filename: str) -> Optional[str]:
        """Extract doc ID from filename using all known patterns (GG, LMS, TSF, generic)."""
        return _extract_doc_id_standalone(filename)

    def _extract_revision_number(self, filename: str) -> int:
        """Extract numeric revision from filename. Returns 0 if not found."""
        stem = Path(filename).stem
        m = self.RE_IFC_REV.search(stem)
        if m:
            return int(m.group(1))
        return 0

    def _extract_description(self, filename: str, doc_id: str) -> str:
        """Extract description from filename (part between doc_id and revision)."""
        stem = Path(filename).stem
        if doc_id and stem.startswith(doc_id):
            remainder = stem[len(doc_id):].lstrip('_- ')
        else:
            remainder = stem
        # Remove revision suffix
        remainder = re.sub(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*\d+.*$', '', remainder, flags=re.IGNORECASE)
        return remainder.strip('_- ') or stem


# =============================================================================
# IFC Stamp Mixin - Shared stamp logic for IFCManager & PanelIFCManager
# =============================================================================

class IFCStampMixin:
    """Mixin providing 'FOR CONSTRUCTION' stamp for IFC DWGs.

    Two stamp methods:
      - Block modify: Find existing 'IFR' block, change text to 'FOR CONSTRUCTION'
        (preferred — uses existing block geometry, most reliable)
      - COM draw: Draw new rectangle + MText on IFC_STAMP layer
        (fallback — for DWGs without an IFR block)

    Both IFCManager and PanelIFCManager inherit this.
    """

    # Reference stamp geometry (from Tatua_Standard_Frame.dwg 'IFR' block)
    # Title block (Frame_NEW) = 841 x 594, IFR block = 110.511 x 17.745
    # IFR block reference insertion in standard frame: (811.221, 73.259)
    # => stamp right edge at x=811.221, stamp bottom at y=73.259
    _REF_TB_WIDTH = 841.0
    _REF_TB_HEIGHT = 594.0
    _REF_RECT_W = 110.511
    _REF_RECT_H = 17.745
    _REF_TEXT_H = 7.0
    _REF_TEXT_W = 116.419
    _REF_TEXT_Y_OFFSET = 13.182
    # Position ratios within frame (from IFR block ref insertion in standard frame)
    _REF_X_RIGHT_OFFSET = 29.779   # 841.0 - 811.221 = offset from right edge
    _REF_Y_BOTTOM = 73.259         # offset from frame bottom

    # Per-title-block reference geometry overrides.
    # Maps title block block-definition name (str) to a dict of _REF_* key overrides.
    # All known projects (Warnertown, Tatua/Coleambally, LMS, Coleambally2) share the
    # default 841x594 reference, so this dict is empty for now.
    # To support a future project with different frame dimensions, add an entry:
    #   "MyTitleBlock": {"_REF_TB_WIDTH": 1189.0, "_REF_TB_HEIGHT": 841.0,
    #                    "_REF_X_RIGHT_OFFSET": 40.0, "_REF_Y_BOTTOM": 90.0, ...}
    # Subclasses may extend or replace this dict for project-specific overrides.
    _TB_REF_OVERRIDES: dict = {}

    # COLOUR stamp box (above FOR CONSTRUCTION): same width, taller for 2-line text
    _REF_COLOUR_RECT_H = 29.0     # calibrated to gold std PLN-005: COL/AB ratio=1.63 (was 26→72pt, now 29→81pt)
    _REF_COLOUR_GAP = 5.0         # vertical gap between AS BUILT and COLOUR boxes
                                  # (raised from 2.0 — boxes were touching; user
                                  #  wanted them separated a bit more)
    _REF_COLOUR_TEXT_H = 5.5      # smaller text height for 2-line COLOUR stamp

    # Polyline border width override.  None → use 0.5*scale. Subclasses set a
    # fixed value (e.g. AsBuiltManager uses thick red borders to match gold std).
    _STAMP_CW: Optional[float] = None

    # AutoCAD color index for stamp box borders + text. 7 = black/white (IFC).
    # AsBuiltManager overrides to 1 (red) to match the gold-standard AS BUILT look.
    _STAMP_COLOR: int = 7

    STAMP_LAYER = "IFC_STAMP"
    STAMP_TEXT = "{\\fArial Narrow|b1;FOR CONSTRUCTION}"
    COLOUR_TEXT = "{\\fArial Narrow|b1;DRAWINGS TO BE\\PPRINTED IN COLOUR}"

    # Known typos to fix during IFC conversion (case-insensitive search, exact replacement)
    _TYPO_FIXES = [
        ('Coulour', 'Colour'),
    ]

    def _fix_known_typos(self, doc):
        """Fix known typos in MText entities during IFC conversion.

        Uses SelectionSet with MTEXT filter (safe for large DWGs).
        """
        import pythoncom as _pythoncom
        ss_name = f"_TypoFix_{int(time.time() * 1000) % 1_000_000}"
        fixed_count = 0
        try:
            ss = doc.SelectionSets.Add(ss_name)
            filter_type = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_I2, [0])
            filter_data = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_VARIANT, ["MTEXT"])
            ss.Select(5, None, None, filter_type, filter_data)
            for i in range(ss.Count):
                try:
                    entity = ss.Item(i)
                    text = entity.TextString
                    new_text = text
                    for typo, correct in self._TYPO_FIXES:
                        # Case-insensitive replacement
                        pattern = re.compile(re.escape(typo), re.IGNORECASE)
                        if pattern.search(new_text):
                            # Preserve case pattern: if source is all-caps, replacement is all-caps
                            def _case_repl(m):
                                matched = m.group(0)
                                if matched.isupper():
                                    return correct.upper()
                                elif matched.islower():
                                    return correct.lower()
                                return correct
                            new_text = pattern.sub(_case_repl, new_text)
                    if new_text != text:
                        entity.TextString = new_text
                        fixed_count += 1
                except Exception:
                    continue
            ss.Delete()
        except Exception:
            try:
                doc.SelectionSets.Item(ss_name).Delete()
            except Exception:
                pass
        if fixed_count:
            print(f"    修正 {fixed_count} 处已知拼写错误")

    @staticmethod
    def _com_retry(func, max_retries=5, delay=2):
        """Retry a COM call that may fail with 'Call was rejected by callee'."""
        for attempt in range(max_retries):
            try:
                return func()
            except Exception as e:
                err_str = str(e)
                # -2147418111 = RPC_E_CALL_REJECTED (callee busy)
                # -2147417848 = RPC_E_SERVERFAULT
                if any(code in err_str for code in (
                    '-2147418111',  # RPC_E_CALL_REJECTED
                    '-2147417848',  # RPC_E_DISCONNECTED
                    '-2147417851',  # RPC_E_SERVERFAULT
                    '-2147352567',  # DISP_E_EXCEPTION
                )):
                    if attempt < max_retries - 1:
                        time.sleep(delay * (attempt + 1))
                        continue
                raise
        return None

    @staticmethod
    def _strip_mtext_formatting(text: str) -> str:
        """Strip MText formatting codes and return plain text content.

        Handles: {\\fFont|flags;content}, {\\Hheight;content}, \\P (paragraph),
        \\A alignment, \\L underline, \\O overline, etc.
        """
        import re
        if not text:
            return ''
        # Remove {\\f...;  and closing }
        s = re.sub(r'\{\\f[^;]*;', '', text)
        s = re.sub(r'\{\\H[^;]*;', '', s)
        s = re.sub(r'\{\\[A-Za-z][^;]*;', '', s)
        s = s.replace('}', '')
        # Remove inline codes: \\P, \\A0, \\A1, \\L, \\O, etc.
        s = re.sub(r'\\[PpLlOo]', '', s)
        s = re.sub(r'\\[Aa]\d', '', s)
        return s.strip()

    def _remove_ifc_stamp(self, doc):
        """Remove existing IFC/IFR stamps from all spaces.

        Pass 0: Delete IFR stamp block references (INSERT entities whose name
                contains 'IFR' or whose attributes contain 'FOR REVIEW').
        Pass 1: Delete all entities on IFC_STAMP layer via SelectionSet (fast).
        Pass 2: Delete legacy MText whose ENTIRE text (after stripping formatting)
                is exactly a stamp phrase. Does NOT delete MText that merely
                contains the phrase as a substring.
        """
        import pythoncom as _pythoncom

        # Pass 0: delete IFR stamp block references
        # Build set of known IFR stamp block names by inspecting block definitions
        _ifr_block_names: set = set()
        try:
            for bi in range(doc.Blocks.Count):
                try:
                    blk = doc.Blocks.Item(bi)
                    bname = blk.Name
                    if bname.startswith('*'):
                        continue  # skip anonymous/special blocks
                    bname_upper = bname.upper()
                    # Quick name check — catch both IFR and IFC stamp blocks
                    if 'IFR' in bname_upper or 'IFC' in bname_upper:
                        _ifr_block_names.add(bname)
                        continue
                    # Check block definition entities for review MText (only small blocks)
                    if blk.Count <= 10:
                        for ei in range(blk.Count):
                            try:
                                ent = blk.Item(ei)
                                if ent.EntityName in ('AcDbMText', 'AcDbText'):
                                    txt = ent.TextString.upper()
                                    plain = self._strip_mtext_formatting(txt).upper()
                                    if plain in ('ISSUED FOR REVIEW', 'FOR REVIEW',
                                                 'FOR CONSTRUCTION'):
                                        _ifr_block_names.add(bname)
                                        break
                            except Exception:
                                pass
                except Exception:
                    pass
        except Exception:
            pass

        ss_name0 = f"_StampCleanup0_{int(time.time() * 1000) % 1_000_000}"
        try:
            try:
                doc.SelectionSets.Item(ss_name0).Delete()
            except Exception:
                pass
            ss0 = doc.SelectionSets.Add(ss_name0)
            filter_type = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_I2, [0])  # entity type
            filter_data = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_VARIANT, ["INSERT"])
            ss0.Select(5, None, None, filter_type, filter_data)

            # Get title block name to avoid accidentally deleting it
            tb_name = getattr(self, 'title_block_name', None) or ''

            for i in range(ss0.Count - 1, -1, -1):
                try:
                    entity = ss0.Item(i)
                    block_name = entity.Name
                    # Skip the title block itself
                    if tb_name and block_name.upper() == tb_name.upper():
                        continue

                    # Check 1: block name in known IFR stamp names (from definition scan)
                    is_ifr_stamp = block_name in _ifr_block_names

                    # Check 2: block name contains 'IFR' or 'IFC' (case-insensitive)
                    if not is_ifr_stamp:
                        bname_u = block_name.upper()
                        is_ifr_stamp = 'IFR' in bname_u or 'IFC' in bname_u

                    # Check 3: block attributes contain stamp phrases
                    if not is_ifr_stamp:
                        try:
                            attrs = entity.GetAttributes()
                            if len(attrs) < 5:
                                for attr in attrs:
                                    try:
                                        val = attr.TextString.upper()
                                        if ('ISSUED FOR REVIEW' in val or
                                                'FOR REVIEW' in val or
                                                'FOR CONSTRUCTION' in val):
                                            is_ifr_stamp = True
                                            break
                                    except Exception:
                                        pass
                        except Exception:
                            pass

                    if is_ifr_stamp:
                        # Final safety: skip blocks with many attributes (likely title blocks)
                        try:
                            attr_count = len(entity.GetAttributes())
                        except Exception:
                            attr_count = 0
                        if attr_count >= 5:
                            continue
                        entity.Delete()
                except Exception:
                    pass
            ss0.Delete()
        except Exception:
            try:
                doc.SelectionSets.Item(ss_name0).Delete()
            except Exception:
                pass

        # Pass 0b: delete IFR/IFC stamp block references in PaperSpace layouts
        # (Pass 0 SelectionSet only searches active space — misses PaperSpace)
        try:
            _layouts_0b = list(doc.Layouts)
        except Exception:
            _layouts_0b = []
        for _layout_0b in _layouts_0b:
            try:
                if _layout_0b.Name.lower() == 'model':
                    continue
                _block_0b = _layout_0b.Block
                _to_delete_0b = []
                for _i_0b in range(_block_0b.Count):
                    try:
                        _ent_0b = _block_0b.Item(_i_0b)
                        if _ent_0b.EntityName != 'AcDbBlockReference':
                            continue
                        _bname_0b = _ent_0b.Name
                        if tb_name and _bname_0b.upper() == tb_name.upper():
                            continue
                        _bname_0b_u = _bname_0b.upper()
                        _is_stamp_0b = ('IFR' in _bname_0b_u or 'IFC' in _bname_0b_u or
                                        _bname_0b in _ifr_block_names)
                        if not _is_stamp_0b:
                            try:
                                _attrs_0b = _ent_0b.GetAttributes()
                                if len(_attrs_0b) < 5:
                                    for _a in _attrs_0b:
                                        _v = _a.TextString.upper()
                                        if ('FOR REVIEW' in _v or
                                                'FOR CONSTRUCTION' in _v):
                                            _is_stamp_0b = True
                                            break
                            except Exception:
                                pass
                        if _is_stamp_0b:
                            try:
                                if len(_ent_0b.GetAttributes()) >= 5:
                                    continue
                            except Exception:
                                pass
                            _to_delete_0b.append(_ent_0b)
                    except Exception:
                        continue
                for _e_0b in reversed(_to_delete_0b):
                    try:
                        _e_0b.Delete()
                    except Exception:
                        pass
            except Exception:
                continue

        # Pass 1: delete all entities on IFC_STAMP layer via SelectionSet
        ss_name = f"_StampCleanup_{int(time.time() * 1000) % 1_000_000}"
        try:
            try:
                doc.SelectionSets.Item(ss_name).Delete()
            except Exception:
                pass
            ss = doc.SelectionSets.Add(ss_name)
            # Filter by layer = IFC_STAMP
            filter_type = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_I2, [8])  # 8 = layer name
            filter_data = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_VARIANT, [self.STAMP_LAYER])
            ss.Select(5, None, None, filter_type, filter_data)
            for i in range(ss.Count - 1, -1, -1):
                try:
                    ss.Item(i).Delete()
                except Exception:
                    pass
            ss.Delete()
        except Exception:
            try:
                doc.SelectionSets.Item(ss_name).Delete()
            except Exception:
                pass

        # Pass 2: delete legacy MText with EXACT stamp content (filtered to MTEXT only)
        # Only delete if the entire text (after stripping formatting) matches a stamp phrase.
        _STAMP_PHRASES = {'FOR CONSTRUCTION', 'ISSUED FOR REVIEW', 'FOR REVIEW',
                          'DRAWINGS TO BE PRINTED IN COLOUR', 'AS BUILT', 'AS-BUILT'}
        ss_name2 = f"_StampCleanup2_{int(time.time() * 1000) % 1_000_000}"
        try:
            try:
                doc.SelectionSets.Item(ss_name2).Delete()
            except Exception:
                pass
            ss2 = doc.SelectionSets.Add(ss_name2)
            filter_type = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_I2, [0])  # entity type
            filter_data = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_VARIANT, ["MTEXT"])
            ss2.Select(5, None, None, filter_type, filter_data)
            for i in range(ss2.Count - 1, -1, -1):
                try:
                    entity = ss2.Item(i)
                    text = entity.TextString
                    if text:
                        plain = self._strip_mtext_formatting(text).upper().strip()
                        if plain in _STAMP_PHRASES or any(
                                p in plain for p in _STAMP_PHRASES):
                            entity.Delete()
                except Exception:
                    pass
            ss2.Delete()
        except Exception:
            try:
                doc.SelectionSets.Item(ss_name2).Delete()
            except Exception:
                pass

        # Pass 3: clean stamps from ALL PaperSpace layouts
        # SelectionSet only searches active space — PaperSpace layouts need direct iteration
        # Also clean stamp-like polylines (closed rectangles ~111x18 at x>500 y<200)
        _STAMP_PHRASES_UPPER = {'FOR CONSTRUCTION', 'ISSUED FOR REVIEW', 'FOR REVIEW',
                                'DRAWINGS TO BE PRINTED IN COLOUR', 'AS BUILT', 'AS-BUILT'}

        def _is_stamp_polyline(ent):
            """Check if a closed polyline looks like a stamp border."""
            try:
                if not ent.Closed:
                    return False
                coords = list(ent.Coordinates)
                if len(coords) != 8:  # rectangle = 4 points x 2 coords
                    return False
                xs = [coords[j] for j in range(0, 8, 2)]
                ys = [coords[j] for j in range(1, 8, 2)]
                w = max(xs) - min(xs)
                h = max(ys) - min(ys)
                if 50 < w < 200 and 8 < h < 40 and 3 < w / h < 15:
                    if min(xs) > 500 and min(ys) < 200:
                        return True
            except Exception:
                pass
            return False

        try:
            layouts = list(doc.Layouts)
        except Exception:
            layouts = []
        for layout in layouts:
            try:
                if layout.Name.lower() == 'model':
                    continue
                block = layout.Block
                to_delete = []
                for i in range(block.Count):
                    try:
                        entity = block.Item(i)
                        ename = entity.EntityName
                        # Delete entities on IFC_STAMP layer
                        try:
                            if entity.Layer == self.STAMP_LAYER:
                                to_delete.append(entity)
                                continue
                        except Exception:
                            pass
                        # Delete legacy stamp MText (exact match)
                        if ename in ('AcDbMText', 'AcDbText'):
                            try:
                                text = entity.TextString
                                if text:
                                    plain = self._strip_mtext_formatting(text).upper().strip()
                                    if plain in _STAMP_PHRASES_UPPER or any(
                                            p in plain for p in _STAMP_PHRASES_UPPER):
                                        to_delete.append(entity)
                            except Exception:
                                pass
                        # Delete stamp-like polylines (closed rect ~111x18 at bottom-right)
                        elif ename in ('AcDbPolyline', 'AcDbLwPolyline'):
                            if _is_stamp_polyline(entity):
                                to_delete.append(entity)
                    except Exception:
                        continue
                for entity in reversed(to_delete):
                    try:
                        entity.Delete()
                    except Exception:
                        pass
            except Exception:
                continue

    def _scan_has_colour(self, doc):
        """Scan DWG for existing 'DRAWINGS TO BE PRINTED IN COLOUR' MTEXT.

        Returns True if found. Also fixes typos (COULOUR→COLOUR) in-place.
        Searches BOTH ModelSpace (via SelectionSet) and all PaperSpace layouts
        (via direct iteration) — SelectionSet mode 5 only searches active space.
        """
        import re as _re
        has_colour = False

        def _check_and_fix(ent):
            nonlocal has_colour
            try:
                raw = self._com_retry(lambda e=ent: e.TextString) or ''
                txt = self._strip_mtext_formatting(raw).upper()
                if ('COLOUR' in txt or 'COULOUR' in txt or 'COLOR' in txt) and 'PRINT' in txt:
                    has_colour = True
                    fixed = _re.sub(r'[Cc][Oo][Uu]?[Ll][Oo][Uu]+[Rr]', 'COLOUR',
                                    raw, flags=_re.IGNORECASE)
                    if fixed != raw:
                        self._com_retry(lambda e=ent, v=fixed: setattr(e, 'TextString', v))
                        print(f"    COLOUR typo 修复: {raw[:40]}... → {fixed[:40]}...")
            except Exception:
                pass

        # Pass A: ModelSpace via SelectionSet (safe for 400K+ entity DWGs)
        try:
            ss_name = f"COLOUR_SCAN_{int(time.time()*1000)}"
            ss = doc.SelectionSets.Add(ss_name)
            try:
                ft = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_I2, [0])
                fv = win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_VARIANT, ["AcDbMText"])
                ss.Select(5, None, None, ft, fv)
                for i in range(ss.Count):
                    try:
                        _check_and_fix(ss.Item(i))
                    except Exception:
                        continue
            finally:
                try:
                    ss.Delete()
                except Exception:
                    pass
        except Exception as e:
            print(f"    COLOUR 扫描(MS)异常: {e}")

        # Pass B: Each PaperSpace layout via direct iteration (small, safe)
        try:
            for layout in doc.Layouts:
                if layout.Name.lower() == 'model':
                    continue
                block = layout.Block
                for i in range(block.Count):
                    try:
                        entity = block.Item(i)
                        if entity.EntityName == 'AcDbMText':
                            _check_and_fix(entity)
                    except Exception:
                        continue
        except Exception as e:
            print(f"    COLOUR 扫描(PS)异常: {e}")

        return has_colour

    def _check_colour_overlap(self, doc, stamp_left, stamp_right,
                              colour_bottom, colour_top, layout_name=None):
        """Check if the COLOUR stamp area overlaps with existing entities.

        Returns True if overlap found (should skip COLOUR drawing).
        Expands search area slightly to catch nearby entities.
        When layout_name is a PaperSpace name, iterates that layout directly
        (SelectionSet crossing window only searches the active space).
        """
        import pythoncom as _pycom
        skip_layers = {self.STAMP_LAYER, 'DEFPOINTS', 'ASHADE'}
        expand = 15.0

        def _entity_overlaps(ent):
            try:
                if ent.Layer.upper() in skip_layers:
                    return False
                if ent.EntityName == 'AcDbViewport':
                    return False
                mn, mx = ent.GetBoundingBox()
                el, eb = float(mn[0]), float(mn[1])
                er, et = float(mx[0]), float(mx[1])
                if (el < stamp_right + expand and er > stamp_left - expand and
                        eb < colour_top + expand and et > colour_bottom - expand):
                    return True
            except Exception:
                pass
            return False

        # PaperSpace: direct iteration (small layouts, safe)
        if layout_name and layout_name.lower() != 'model':
            try:
                for layout in doc.Layouts:
                    if layout.Name == layout_name:
                        block = layout.Block
                        for i in range(block.Count):
                            try:
                                if _entity_overlaps(block.Item(i)):
                                    return True
                            except Exception:
                                continue
                        break
            except Exception:
                pass
            return False

        # ModelSpace: SelectionSet crossing window (handles 400K+ entities)
        try:
            ss_name = f"_ColourOvlp_{int(time.time() * 1000) % 1_000_000}"
            ss = doc.SelectionSets.Add(ss_name)
            pt1 = win32com.client.VARIANT(_pycom.VT_ARRAY | _pycom.VT_R8,
                [stamp_left - expand, colour_bottom - expand, 0.0])
            pt2 = win32com.client.VARIANT(_pycom.VT_ARRAY | _pycom.VT_R8,
                [stamp_right + expand, colour_top + expand, 0.0])
            ss.Select(1, pt1, pt2)
            found = False
            for i in range(ss.Count):
                try:
                    ent = ss.Item(i)
                    if ent.Layer.upper() in skip_layers:
                        continue
                    if ent.EntityName == 'AcDbViewport':
                        continue
                    found = True
                    break
                except Exception:
                    continue
            ss.Delete()
            return found
        except Exception:
            pass
        return False

    def _ensure_colour_has_border(self, doc, draw_space, stamp_left, stamp_right,
                                   colour_bottom, colour_top, cw, layout_name=None,
                                   redraw=False):
        """Reconcile the existing COLOUR stamp with the bot's stamp position.

        redraw=False (IFC keep-existing): move the existing COLOUR MText to the
            stamp position and draw a border around it (Fix 1 + Fix 3).
        redraw=True (AsBuilt force/redraw): the caller will draw a FRESH COLOUR
            box + text, so DELETE the old COLOUR MText here (don't move/keep it)
            and skip drawing a border — otherwise the old + fresh duplicate.
        Either way, Fix 2 clears stale border entities in the zone.
        """
        colour_entity = None

        # Find the COLOUR MText entity
        if layout_name and layout_name.lower() != 'model':
            try:
                for layout in doc.Layouts:
                    if layout.Name == layout_name:
                        block = layout.Block
                        for i in range(block.Count):
                            try:
                                e = block.Item(i)
                                if e.EntityName == 'AcDbMText':
                                    raw = self._com_retry(lambda ent=e: ent.TextString) or ''
                                    txt = self._strip_mtext_formatting(raw).upper()
                                    if ('COLOUR' in txt or 'COLOR' in txt) and 'PRINT' in txt:
                                        colour_entity = e
                                        break
                            except Exception:
                                continue
                        break
            except Exception:
                pass
        else:
            import pythoncom as _pycom
            try:
                ss_name = f"_ColBorder_{int(time.time()*1000) % 1_000_000}"
                ss = doc.SelectionSets.Add(ss_name)
                ft = win32com.client.VARIANT(_pycom.VT_ARRAY | _pycom.VT_I2, [0])
                fv = win32com.client.VARIANT(_pycom.VT_ARRAY | _pycom.VT_VARIANT, ["AcDbMText"])
                ss.Select(5, None, None, ft, fv)
                for i in range(ss.Count):
                    try:
                        e = ss.Item(i)
                        raw = self._com_retry(lambda ent=e: ent.TextString) or ''
                        txt = self._strip_mtext_formatting(raw).upper()
                        if ('COLOUR' in txt or 'COLOR' in txt) and 'PRINT' in txt:
                            colour_entity = e
                            break
                    except Exception:
                        continue
                ss.Delete()
            except Exception:
                pass

        if colour_entity is None:
            # No existing COLOUR MText. In redraw mode still run Fix 2 (clear any
            # stale borders) below; in keep mode there's nothing to do.
            if not redraw:
                return

        # Fix 1: reconcile the existing COLOUR MText.
        colour_center_x = (stamp_left + stamp_right) / 2.0
        colour_center_y = (colour_bottom + colour_top) / 2.0
        if colour_entity is not None:
            if redraw:
                # Caller draws a fresh COLOUR box + text → delete the old MText so
                # it isn't duplicated.
                try:
                    colour_entity.Delete()
                except Exception:
                    pass
                colour_entity = None
            else:
                # Keep-existing: move it onto the stamp position.
                try:
                    new_ip = win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8,
                        [colour_center_x, colour_center_y, 0.0])
                    colour_entity.InsertionPoint = new_ip
                    colour_entity.Width = stamp_right - stamp_left
                    colour_entity.AttachmentPoint = 5  # MiddleCenter
                except Exception:
                    pass

        # Fix 2: Delete old border entities in the COLOUR+AS BUILT stamp zone.
        # Pre-bot stamps may use AcDb2dPolyline, AcDbSolid, etc. — handle all.
        # Use SelectionSet crossing window (safe for 400K+ entity DWGs — uses
        # spatial index, NOT full iteration). Filter entity type in Python.
        # Unlock layer before delete in case entity is on a locked layer.
        _COLOUR_ENTITY_TYPES = frozenset((
            'AcDbPolyline', 'AcDbLwPolyline', 'AcDb2dPolyline',
            'AcDbSolid', 'AcDbTrace',
        ))
        # ±300 DWG units covers any pre-bot border regardless of position offset
        _margin = 300.0
        _search_l = stamp_left   - _margin
        _search_r = stamp_right  + _margin
        _search_b = colour_bottom - _margin   # spans COLOUR zone and below
        _search_t = colour_top   + _margin
        _deleted_count = 0

        def _delete_entity_safe(e):
            """Unlock layer if needed, then delete."""
            try:
                lyr = doc.Layers.Item(e.Layer)
                was_locked = lyr.Lock
                if was_locked:
                    lyr.Lock = False
            except Exception:
                was_locked = False
                lyr = None
            try:
                e.Delete()
                return True
            except Exception:
                return False
            finally:
                if lyr and was_locked:
                    try:
                        lyr.Lock = True
                    except Exception:
                        pass

        if layout_name and layout_name.lower() != 'model':
            # PaperSpace: iterate layout block directly (small, safe)
            try:
                for layout in doc.Layouts:
                    if layout.Name == layout_name:
                        block = layout.Block
                        _to_del = []
                        for i in range(block.Count):
                            try:
                                e = block.Item(i)
                                if e.EntityName not in _COLOUR_ENTITY_TYPES:
                                    continue
                                # NEVER delete the bot's own freshly-drawn boxes
                                # (AS BUILT box is on IFC_STAMP, drawn just before
                                # this cleanup runs). Only old pre-bot borders on
                                # other layers ('0'/'QA'/etc) should go.
                                if e.Layer == self.STAMP_LAYER:
                                    continue
                                emn, emx = e.GetBoundingBox()
                                cx = (float(emn[0]) + float(emx[0])) / 2
                                cy = (float(emn[1]) + float(emx[1])) / 2
                                if (_search_l <= cx <= _search_r and
                                        _search_b <= cy <= _search_t):
                                    _to_del.append(e)
                            except Exception:
                                continue
                        for _e in reversed(_to_del):
                            if _delete_entity_safe(_e):
                                _deleted_count += 1
                        break
            except Exception:
                pass
        else:
            # ModelSpace: use SelectionSet crossing window (spatial index — safe for
            # large DWGs). No entity type filter on SS itself; filter type in Python.
            import pythoncom as _pycom
            try:
                ss2_name = f"_ColBChk_{int(time.time()*1000) % 1_000_000}"
                ss2 = doc.SelectionSets.Add(ss2_name)
                pt1 = win32com.client.VARIANT(
                    _pycom.VT_ARRAY | _pycom.VT_R8,
                    [_search_l, _search_b, 0.0])
                pt2 = win32com.client.VARIANT(
                    _pycom.VT_ARRAY | _pycom.VT_R8,
                    [_search_r, _search_t, 0.0])
                ss2.Select(1, pt1, pt2)   # crossing window, all entity types
                _to_del = []
                for i in range(ss2.Count):
                    try:
                        e = ss2.Item(i)
                        if e.EntityName in _COLOUR_ENTITY_TYPES \
                                and e.Layer != self.STAMP_LAYER:
                            _to_del.append(e)
                    except Exception:
                        continue
                ss2.Delete()
                for _e in reversed(_to_del):
                    if _delete_entity_safe(_e):
                        _deleted_count += 1
            except Exception:
                pass

        # Fix 3: draw a border around the EXISTING COLOUR text (keep mode only).
        # In redraw mode the caller draws a fresh box — drawing one here too would
        # duplicate it, so skip.
        if redraw:
            return
        try:
            colour_pts = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8,
                [stamp_left, colour_bottom,
                 stamp_right, colour_bottom,
                 stamp_right, colour_top,
                 stamp_left, colour_top])
            pline = draw_space.AddLightWeightPolyline(colour_pts)
            pline.Closed = True
            pline.Layer = self.STAMP_LAYER
            pline.color = getattr(self, '_STAMP_COLOR', 7)
            pline.ConstantWidth = cw
            print(f"    印章: COLOUR 边框已重绘 (cw={cw:.2f}, 清除旧框={_deleted_count})")
        except Exception as e:
            print(f"    印章: COLOUR 边框绘制失败 ({e})")

    def _stamp_via_com_draw(self, doc, block_ref, space, has_colour=True,
                            layout_name=None):
        """Draw rectangle + MText 'FOR CONSTRUCTION' stamp inside the frame.

        Position matches the original IFR stamp location (bottom-right area),
        scaled proportionally to the actual title block dimensions.
        Caller must call _remove_ifc_stamp(doc) ONCE before invoking this
        (avoids removing the first stamp when drawing the second in multi-sheet).

        Args:
            has_colour: If True, original DWG already has COLOUR stamp — skip drawing it.
                        If False, draw both FOR CONSTRUCTION and COLOUR boxes.
            layout_name: Layout name where stamp should be drawn. Used to get a
                        fresh space reference (avoids stale COM proxy issues).
        """
        # Get title block bounding box
        try:
            min_pt, max_pt = block_ref.GetBoundingBox()
            tb_left_x = float(min_pt[0])
            tb_bottom_y = float(min_pt[1])
            tb_right_x = float(max_pt[0])
            tb_top_y = float(max_pt[1])
            tb_width = tb_right_x - tb_left_x
            tb_height = tb_top_y - tb_bottom_y
        except Exception as e:
            print(f"    印章: 无法获取 title block 边界 ({e})，跳过")
            return

        if tb_width <= 0:
            print(f"    印章: title block 宽度异常 ({tb_width:.1f})，跳过")
            return

        # Resolve per-title-block reference geometry overrides.
        # Look up the block definition name and apply any project-specific _REF_* values.
        # To add a new project with different frame dimensions, populate _TB_REF_OVERRIDES
        # on the relevant subclass (IFCManager, AsBuiltManager, etc.) with the title block
        # name as key and a dict of _REF_* overrides as value.
        _tb_name = ''
        try:
            _tb_name = block_ref.Name or ''
        except Exception:
            pass
        _tb_overrides = (self._TB_REF_OVERRIDES or {}).get(_tb_name, {})
        _ref_tb_w = _tb_overrides.get('_REF_TB_WIDTH', self._REF_TB_WIDTH)
        _ref_tb_h = _tb_overrides.get('_REF_TB_HEIGHT', self._REF_TB_HEIGHT)
        _ref_rect_w = _tb_overrides.get('_REF_RECT_W', self._REF_RECT_W)
        _ref_rect_h = _tb_overrides.get('_REF_RECT_H', self._REF_RECT_H)
        _ref_text_h = _tb_overrides.get('_REF_TEXT_H', self._REF_TEXT_H)
        _ref_text_w = _tb_overrides.get('_REF_TEXT_W', self._REF_TEXT_W)
        _ref_text_y_offset = _tb_overrides.get('_REF_TEXT_Y_OFFSET', self._REF_TEXT_Y_OFFSET)
        _ref_x_right_offset = _tb_overrides.get('_REF_X_RIGHT_OFFSET', self._REF_X_RIGHT_OFFSET)
        _ref_y_bottom = _tb_overrides.get('_REF_Y_BOTTOM', self._REF_Y_BOTTOM)
        _ref_colour_rect_h = _tb_overrides.get('_REF_COLOUR_RECT_H', self._REF_COLOUR_RECT_H)
        _ref_colour_gap = _tb_overrides.get('_REF_COLOUR_GAP', self._REF_COLOUR_GAP)
        _ref_colour_text_h = _tb_overrides.get('_REF_COLOUR_TEXT_H', self._REF_COLOUR_TEXT_H)
        if _tb_overrides:
            print(f"    印章: 使用 title block '{_tb_name}' 专属参考尺寸 "
                  f"({_ref_tb_w:.0f}x{_ref_tb_h:.0f})")

        # Scale all dimensions proportionally to title block size
        scale = tb_width / _ref_tb_w
        rect_w = _ref_rect_w * scale
        rect_h = _ref_rect_h * scale
        text_h = _ref_text_h * scale
        text_w = _ref_text_w * scale
        text_y_offset = _ref_text_y_offset * scale

        # Position: INSIDE the frame, matching the original IFR stamp location
        # IFR stamp in standard frame: right edge at x=811.221, bottom at y=73.259
        # in a 841x594 frame. Scale proportionally to actual frame size.
        scale_y = tb_height / _ref_tb_h if tb_height > 0 else scale
        x_right_offset = _ref_x_right_offset * scale
        y_bottom_offset = _ref_y_bottom * scale_y

        # Stamp right edge = frame_right - offset; stamp insertion point is right edge
        stamp_right_x = tb_right_x - x_right_offset
        stamp_left_x = stamp_right_x - rect_w
        stamp_bottom_y = tb_bottom_y + y_bottom_offset
        stamp_top_y = stamp_bottom_y + rect_h

        # MText centered in rectangle
        mtext_x = (stamp_left_x + stamp_right_x) / 2.0
        mtext_y = (stamp_bottom_y + stamp_top_y) / 2.0

        # --- P2 (stamp-vs-content overlap): left at gold-standard position ---
        # On drawings whose bottom-right corner is full (E-BLD-003 battery table,
        # GAD-001 legend) the stamp lands on existing content. This resisted
        # reliable automation and is NOT auto-handled here:
        #   - auto-SHIFT is futile — the content fills the whole corner via a
        #     viewport, so there is no clear spot to move to;
        #   - reading ModelSpace-through-viewport at convert time needs an
        #     active-space toggle that was slow (649s) and broke PUBLISH (GAD-001);
        #   - a fitz text-in-box QA heuristic mis-judged (false-positive on the
        #     gold E-PLN-002, missed BLD-003) — worse than nothing.
        # So the stamp stays at the proportional gold-standard position (correct
        # for the ~21 drawings with an empty corner); the few dense-corner
        # drawings are handled MANUALLY. See [[asbuilt-ifr-stamp-standard]].
        print(f"    印章: COM 绘制 (scale={scale:.3f}, tb={tb_width:.0f}x{tb_height:.0f}, "
              f"pos=({stamp_left_x:.1f},{stamp_bottom_y:.1f})->({stamp_right_x:.1f},{stamp_top_y:.1f}))")

        # Ensure IFC_STAMP layer exists, set to stamp color, ON, thawed
        _stamp_color = getattr(self, '_STAMP_COLOR', 7)
        try:
            layer = doc.Layers.Add(self.STAMP_LAYER)
            layer.color = _stamp_color
            layer.LayerOn = True
            layer.Freeze = False
        except Exception:
            try:
                layer = doc.Layers.Item(self.STAMP_LAYER)
                layer.color = _stamp_color
                layer.LayerOn = True
                layer.Freeze = False
            except Exception:
                pass

        # Thaw IFC_STAMP layer in ALL PaperSpace viewports
        # Only needed when stamp is in ModelSpace — PaperSpace stamps are always
        # visible in PDF without viewport thaw (they exist in the layout directly).
        _stamp_in_paperspace = (layout_name is not None
                                and layout_name.lower() != 'model')
        _vp_thaw_failed = 0
        if not _stamp_in_paperspace:
            try:
                for layout in doc.Layouts:
                    if layout.Name.lower() == 'model':
                        continue
                    block = layout.Block
                    for i in range(block.Count):
                        try:
                            entity = block.Item(i)
                            if entity.EntityName == 'AcDbViewport':
                                try:
                                    frozen_layers = list(entity.GetFrozenLayers() or [])
                                    if self.STAMP_LAYER in frozen_layers:
                                        frozen_layers.remove(self.STAMP_LAYER)
                                        if frozen_layers:
                                            entity.PutFrozenLayers(frozen_layers)
                                        else:
                                            # IFC_STAMP was the only frozen layer — try empty array
                                            try:
                                                empty = win32com.client.VARIANT(
                                                    pythoncom.VT_ARRAY | pythoncom.VT_BSTR, [])
                                                entity.PutFrozenLayers(empty)
                                            except Exception:
                                                _vp_thaw_failed += 1
                                except Exception:
                                    _vp_thaw_failed += 1
                        except Exception:
                            continue
            except Exception:
                pass

            # Fallback: if COM viewport thaw failed, use VPLAYER SendCommand
            if _vp_thaw_failed > 0:
                try:
                    for layout in doc.Layouts:
                        if layout.Name.lower() == 'model':
                            continue
                        try:
                            doc.ActiveLayout = layout
                            time.sleep(0.5)
                            doc.SendCommand(
                                f'-VPLAYER\nThaw\n{self.STAMP_LAYER}\nAll\n\n')
                            time.sleep(1.5)
                        except Exception:
                            pass
                    # Switch back to Model
                    for layout in doc.Layouts:
                        if layout.Name.lower() == 'model':
                            doc.ActiveLayout = layout
                            break
                except Exception:
                    pass

        try:
            # --- Get FRESH space reference AFTER viewport thaw ---
            # The viewport thaw loop above does heavy COM work (iterates all
            # layouts/viewports, may change ActiveLayout via VPLAYER fallback).
            # This can invalidate previously-obtained space references.
            # Get a fresh reference right before creating entities.
            draw_space = space  # fallback
            if layout_name and layout_name.lower() != 'model':
                # Strategy 1: re-acquire layout.Block fresh
                try:
                    for _lay in doc.Layouts:
                        if _lay.Name == layout_name:
                            draw_space = win32com.client.Dispatch(_lay.Block)
                            break
                except Exception:
                    pass
                # Strategy 2: test it, if broken try doc.PaperSpace
                try:
                    _test_pts = win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 1, 0])
                    _test = draw_space.AddLightWeightPolyline(_test_pts)
                    _test.Delete()
                except Exception:
                    try:
                        doc.ActiveLayout = doc.Layouts.Item(layout_name)
                        time.sleep(0.5)
                        draw_space = doc.PaperSpace
                    except Exception:
                        pass
            else:
                # ModelSpace: always get fresh reference after viewport thaw
                try:
                    draw_space = doc.ModelSpace
                except Exception:
                    pass
                # Test it — if stale, try Dispatch wrapping
                try:
                    _test_pts = win32com.client.VARIANT(
                        pythoncom.VT_ARRAY | pythoncom.VT_R8, [0, 0, 1, 0])
                    _test = draw_space.AddLightWeightPolyline(_test_pts)
                    _test.Delete()
                except Exception:
                    try:
                        draw_space = win32com.client.Dispatch(doc.ModelSpace)
                    except Exception:
                        pass

            # --- FOR CONSTRUCTION stamp (lower box) ---
            rect_pts = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8,
                [stamp_left_x, stamp_bottom_y,
                 stamp_right_x, stamp_bottom_y,
                 stamp_right_x, stamp_top_y,
                 stamp_left_x, stamp_top_y])
            pline = draw_space.AddLightWeightPolyline(rect_pts)
            pline.Closed = True
            pline.Layer = self.STAMP_LAYER
            pline.color = _stamp_color  # explicit (ByLayer might inherit other)
            _cw = self._STAMP_CW if self._STAMP_CW is not None else (0.5 * scale)
            pline.ConstantWidth = _cw

            center_pt = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8,
                [mtext_x, mtext_y, 0.0])
            mtext = draw_space.AddMText(center_pt, rect_w, self.STAMP_TEXT)
            mtext.Height = text_h
            mtext.AttachmentPoint = 5  # MiddleCenter
            mtext.InsertionPoint = center_pt
            mtext.Layer = self.STAMP_LAYER
            mtext.color = _stamp_color

            # --- DRAWINGS TO BE PRINTED IN COLOUR stamp (upper box) ---
            # Only draw if: (1) original DWG has no COLOUR stamp, AND
            #               (2) no overlap with existing entities in the COLOUR area
            colour_rect_h = _ref_colour_rect_h * scale
            colour_gap = _ref_colour_gap * scale
            colour_bottom_y = stamp_top_y + colour_gap
            colour_top_y = colour_bottom_y + colour_rect_h

            # COLOUR stamp strategy:
            # 1) Decide whether we draw a FRESH COLOUR box.
            #    - has_colour / _FORCE_COLOUR (AsBuilt): always redraw aligned.
            #    - else: draw only if no genuine drawing content overlaps.
            # 2) Then reconcile existing COLOUR entities via _ensure_colour_has_border:
            #    - redraw=True  → DELETE old MText + clear borders (we draw fresh),
            #                     so the old + fresh never duplicate.
            #    - redraw=False → keep+border the existing COLOUR text.
            _draw_colour = False
            if has_colour or getattr(self, '_FORCE_COLOUR', False):
                _draw_colour = True
                print(f"    印章: COLOUR → 重绘对齐 (has_colour={has_colour}, "
                      f"force={getattr(self, '_FORCE_COLOUR', False)})")
            elif self._check_colour_overlap(doc, stamp_left_x, stamp_right_x,
                                             colour_bottom_y, colour_top_y,
                                             layout_name=layout_name):
                print(f"    印章: COLOUR 区域有真实重叠实体，跳过")
            else:
                _draw_colour = True

            # Reconcile old COLOUR entities (delete-if-redraw, keep-if-not).
            self._ensure_colour_has_border(
                doc, draw_space, stamp_left_x, stamp_right_x,
                colour_bottom_y, colour_top_y, _cw,
                layout_name=layout_name, redraw=_draw_colour)

            if _draw_colour:

                colour_pts = win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    [stamp_left_x, colour_bottom_y,
                     stamp_right_x, colour_bottom_y,
                     stamp_right_x, colour_top_y,
                     stamp_left_x, colour_top_y])
                pline2 = draw_space.AddLightWeightPolyline(colour_pts)
                pline2.Closed = True
                pline2.Layer = self.STAMP_LAYER
                pline2.color = _stamp_color  # same color as AS BUILT box
                pline2.ConstantWidth = _cw   # same width as AS BUILT box

                colour_center_x = (stamp_left_x + stamp_right_x) / 2.0
                colour_center_y = (colour_bottom_y + colour_top_y) / 2.0
                colour_center_pt = win32com.client.VARIANT(
                    pythoncom.VT_ARRAY | pythoncom.VT_R8,
                    [colour_center_x, colour_center_y, 0.0])
                colour_text_h = _ref_colour_text_h * scale
                mtext2 = draw_space.AddMText(colour_center_pt, rect_w, self.COLOUR_TEXT)
                mtext2.Height = colour_text_h
                mtext2.AttachmentPoint = 5  # MiddleCenter
                mtext2.InsertionPoint = colour_center_pt
                mtext2.Layer = self.STAMP_LAYER
                mtext2.color = _stamp_color

            doc.Regen(1)
            if has_colour:
                stamp_label = "FOR CONSTRUCTION (COLOUR 保留原DWG)"
            elif _draw_colour:
                stamp_label = "FOR CONSTRUCTION + COLOUR"
            else:
                stamp_label = "FOR CONSTRUCTION (COLOUR 跳过-重叠)"
            print(f"    印章: {stamp_label} 已添加 (COM, rect={rect_w:.1f}x{rect_h:.1f}, 黑色边框)")
        except Exception as e:
            print(f"    印章添加失败: {e}")

    def _add_ifc_stamp(self, doc, block_ref, space):
        """Add 'FOR CONSTRUCTION' stamp to the DWG (single title block).

        Removes existing stamps first, then draws new one.
        For multi-title-block DWGs, call _remove_ifc_stamp() once then
        _stamp_via_com_draw() per title block directly.
        """
        has_colour = self._scan_has_colour(doc)
        self._remove_ifc_stamp(doc)
        self._stamp_via_com_draw(doc, block_ref, space, has_colour=has_colour)


# =============================================================================
# IFC Manager - AutoCAD COM Automation (NEW in v7)
# =============================================================================

class IFCManager(IFCStampMixin):
    """Automate IFR→IFC conversion via AutoCAD COM API.

    Workflow per DWG:
      1. Open DWG in AutoCAD
      2. Find title block, read latest IFR revision row
      3. Update title block: set IFC revision, clear old rows, write IFC row
      4. SaveAs new IFC DWG
      5. Export PDF (all layouts) to IFC(Client) folder
      6. Close without saving original IFR DWG
    """

    NATIVE_ROOT = "Design/Engineering/1. Drawings/1. Native"
    IFC_OUTPUT = "Design/Engineering/1. Drawings/4. IFC(Client)"

    @staticmethod
    def _com_retry(func, max_retries=5, delay=2):
        """Retry a COM call that may fail with 'Call was rejected by callee'."""
        for attempt in range(max_retries):
            try:
                return func()
            except Exception as e:
                err_str = str(e)
                if any(code in err_str for code in (
                    '-2147418111',  # RPC_E_CALL_REJECTED
                    '-2147417848',  # RPC_E_DISCONNECTED
                    '-2147417851',  # RPC_E_SERVERFAULT
                    '-2147352567',  # DISP_E_EXCEPTION
                )):
                    if attempt < max_retries - 1:
                        time.sleep(delay * (attempt + 1))
                        continue
                raise
        return None

    # Title block attribute TAGs
    DEFAULT_TITLE_BLOCK = "ACE-Wanertown_Siyuan"
    REV_ROWS = 6  # max revision history rows in title block

    # Personnel TAG suffixes per row (prefixed by row number, e.g. 1DRAWN, 2DRAWN)
    PERSONNEL_TAGS = ['DRAWN', 'CHECK', 'ENGINEER', 'QA', 'PROJECT']

    def __init__(self, project_path, dry_run=False, title_block_name=None,
                 native_root=None, ifc_output=None, preserve_ifr=False):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.title_block_name = title_block_name or self.DEFAULT_TITLE_BLOCK
        self.preserve_ifr = preserve_ifr
        self._acad = None
        self._dm = DeliverableManager(self.project_path, dry_run=dry_run)
        self._native_root_override = Path(native_root) if native_root else None
        self._ifc_output_override = Path(ifc_output) if ifc_output else None

    @property
    def native_root(self) -> Path:
        if self._native_root_override:
            return self._native_root_override
        return self.project_path / self.NATIVE_ROOT

    @property
    def ifc_output(self) -> Path:
        if self._ifc_output_override:
            return self._ifc_output_override
        return self.project_path / self.IFC_OUTPUT

    def _get_acad(self):
        """Get or launch AutoCAD COM instance."""
        if not WIN32COM_AVAILABLE:
            raise RuntimeError("win32com 未安装，无法使用 AutoCAD COM API。请安装 pywin32。")
        if self._acad is not None:
            try:
                _ = self._acad.Visible
                return self._acad
            except Exception:
                self._acad = None
        # Try to connect to running instance first
        try:
            self._acad = win32com.client.GetActiveObject("AutoCAD.Application")
            # NOTE: Do NOT wrap with Dispatch() — early binding via type library
            # causes child objects (layout.Block, ModelSpace) to return <unknown>
            # for methods like AddLightWeightPolyline and AddMText.
            # Late binding (from GetActiveObject) works correctly for all methods.
        except Exception:
            try:
                print("  正在启动 AutoCAD（可能需要30秒）...")
                self._acad = win32com.client.Dispatch("AutoCAD.Application")
                # Cold-start race: a freshly-launched AutoCAD returns from Dispatch
                # before it's ready; setting .Visible too soon raises "Visible can
                # not be set". Wait until the app responds, then set Visible best-effort.
                for _start_wait in range(30):
                    try:
                        _ = self._acad.Documents
                        break
                    except Exception:
                        time.sleep(1)
                try:
                    self._acad.Visible = True
                except Exception:
                    pass
            except Exception as e:
                raise RuntimeError(f"无法连接或启动 AutoCAD: {e}")
        # Wait for Documents collection to be ready
        for _doc_wait in range(15):
            try:
                _ = self._acad.Documents.Count
                break
            except Exception:
                time.sleep(1)
        # Suppress all notification dialogs
        try:
            self._acad.Preferences.System.BeepOnError = False
        except Exception:
            pass
        try:
            doc = self._acad.ActiveDocument
            if doc:
                doc.SetVariable("LAYERNOTIFY", 0)
                doc.SetVariable("LAYEREVALCTL", 0)
        except Exception:
            pass
        return self._acad

    def scan_native_folders(self) -> List[Dict]:
        """Scan 1. Native/ subfolders and identify IFR DWGs ready for IFC conversion.

        Returns list of dicts:
          {doc_id, folder, latest_ifr_dwg, latest_ifr_rev,
           existing_ifc_rev, needs_ifc, description}
        """
        results = []
        native = self.native_root
        if not native.exists():
            return results

        # Also scan IFC output folder to know existing IFC revisions
        existing_ifc = self._scan_existing_ifc()

        for category_dir in sorted(native.iterdir()):
            if not category_dir.is_dir():
                continue
            if category_dir.name.lower() in ('ss', 'superseded', 'superceded'):
                continue
            for doc_folder in sorted(category_dir.iterdir()):
                if not doc_folder.is_dir():
                    continue
                if doc_folder.name.lower() in ('ss', 'superseded', 'superceded'):
                    continue
                doc_id, description = _parse_folder_name(doc_folder.name)
                if not doc_id:
                    continue

                # Find latest IFR DWG (highest letter revision)
                latest_ifr = None
                latest_rev = ''
                for f in doc_folder.iterdir():
                    if not f.is_file() or f.suffix.lower() != '.dwg':
                        continue
                    if f.name.startswith('~$'):
                        continue
                    rev_type, revision = _classify_dwg(f.name)
                    if rev_type == 'IFR' and revision > latest_rev:
                        latest_rev = revision
                        latest_ifr = f

                if not latest_ifr:
                    continue

                ifc_rev = existing_ifc.get(doc_id)
                results.append({
                    'doc_id': doc_id,
                    'folder': doc_folder,
                    'latest_ifr_dwg': latest_ifr,
                    'latest_ifr_rev': latest_rev,
                    'existing_ifc_rev': ifc_rev,
                    'needs_ifc': True,  # always allow re-conversion
                    'description': description,
                })

        return results

    def _scan_existing_ifc(self) -> Dict[str, int]:
        """Scan IFC output folder for existing IFC files, return {doc_id: max_rev_number}."""
        ifc_revs = {}
        ifc_dir = self.ifc_output
        if not ifc_dir.exists():
            return ifc_revs
        for f in ifc_dir.rglob("*"):
            if not f.is_file():
                continue
            if f.suffix.lower() not in ('.dwg', '.pdf'):
                continue
            doc_id = self._dm.extract_doc_id(f.name)
            if not doc_id:
                continue
            rev_type, revision = _classify_dwg(f.name)
            if rev_type == 'IFC':
                try:
                    rev_num = int(revision)
                    if doc_id not in ifc_revs or rev_num > ifc_revs[doc_id]:
                        ifc_revs[doc_id] = rev_num
                except ValueError:
                    pass
        return ifc_revs

    def _find_title_block(self, doc):
        """Find first title block. Returns (block_ref, attrs_dict, space).

        For multi-sheet DWGs, use _find_all_title_blocks() instead.
        """
        results = self._find_all_title_blocks(doc)
        if results:
            return results[0]
        return None, {}, None

    def _find_all_title_blocks(self, doc):
        """Find ALL title block references in AutoCAD document.

        Returns list of (block_ref, attrs_dict, space, layout_name) tuples.
        Multi-sheet DWGs may have title blocks in ModelSpace OR in separate
        PaperSpace layouts. SelectionSet only searches the ACTIVE space, so
        we must iterate ALL layouts individually.

        Strategy:
          1. Search ModelSpace via SelectionSet (INSERT filter) → _ms_tbs/_ms_fallback
          2. Search EACH PaperSpace layout directly → _ps_tbs/_ps_fallback
          3. Priority: _ps_tbs > _ps_fallback > _ms_tbs > _ms_fallback
             (PaperSpace TBs preferred — stamps drawn there are always visible
              in PDF without viewport thaw issues)
        """
        import pythoncom as _pythoncom

        _ms_tbs = []          # ModelSpace exact name matches
        _ms_fallback = []     # ModelSpace fallback (DRAWINGNUMBER+REVISION)
        _ps_tbs = []          # PaperSpace exact name matches
        _ps_fallback = []     # PaperSpace fallback (DRAWINGNUMBER+REVISION)

        # --- Pass 1: Search ModelSpace via SelectionSet (handles 400K+ entities) ---
        ss_name = f"_TBSearch_{int(time.time() * 1000) % 1_000_000}"
        for _cleanup_name in [ss_name, "_TitleBlockSearch"]:
            for _attempt in range(3):
                try:
                    doc.SelectionSets.Item(_cleanup_name).Delete()
                    break
                except Exception:
                    if _attempt < 2:
                        time.sleep(0.2)
                    break

        try:
            # Ensure ModelSpace is active for SelectionSet
            saved_layout = None
            try:
                saved_layout = doc.ActiveLayout.Name
                doc.ActiveSpace = 1  # acModelSpace = 1
            except Exception:
                pass

            ss = doc.SelectionSets.Add(ss_name)
            filter_type = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_I2, [0])
            filter_data = win32com.client.VARIANT(
                _pythoncom.VT_ARRAY | _pythoncom.VT_VARIANT, ["INSERT"])
            ss.Select(5, None, None, filter_type, filter_data)

            ms = doc.ModelSpace
            for i in range(ss.Count):
                # Dispatch-wrap entity for proper COM type resolution
                # (SelectionSet.Item may return late-bound objects without
                # type library info, causing <unknown>.TextString errors)
                entity = win32com.client.Dispatch(ss.Item(i))
                try:
                    block_name = entity.Name
                except Exception:
                    continue

                if block_name == self.title_block_name:
                    attrs = self._get_attrs_dict(entity)
                    _ms_tbs.append((entity, attrs, ms, 'Model'))
                else:
                    attrs = self._get_attrs_dict(entity)
                    if 'DRAWINGNUMBER' in attrs and 'REVISION' in attrs:
                        _ms_fallback.append((entity, attrs, ms, 'Model'))

            ss.Delete()
        except Exception as e:
            logging.warning(f"ModelSpace SelectionSet search failed: {e}")
            try:
                doc.SelectionSets.Item(ss_name).Delete()
            except Exception:
                pass

        # --- Pass 2: Search each PaperSpace layout directly ---
        # PaperSpace layouts typically have < 100 entities (viewports, title block, notes)
        # so direct iteration is safe and guaranteed to find all title blocks.
        # Always collect PaperSpace TBs (they take priority over ModelSpace).
        # Retry up to 3 times if COM fails (e.g. "Call was rejected by callee")
        for _ps_attempt in range(3):
            try:
                for layout in doc.Layouts:
                    if layout.Name.lower() == 'model':
                        continue
                    layout_name = layout.Name
                    block = win32com.client.Dispatch(layout.Block)
                    count = block.Count
                    for i in range(count):
                        try:
                            entity = win32com.client.Dispatch(block.Item(i))
                            if entity.EntityName != 'AcDbBlockReference':
                                continue
                            block_name = entity.Name
                        except Exception:
                            continue

                        if block_name == self.title_block_name:
                            attrs = self._get_attrs_dict(entity)
                            _ps_tbs.append((entity, attrs, block, layout_name))
                        else:
                            attrs = self._get_attrs_dict(entity)
                            if 'DRAWINGNUMBER' in attrs and 'REVISION' in attrs:
                                _ps_fallback.append((entity, attrs, block, layout_name))
                break  # success — exit retry loop
            except Exception as e:
                if _ps_attempt < 2:
                    logging.warning(f"PaperSpace search attempt {_ps_attempt+1}/3 failed: {e}")
                    time.sleep(2 * (_ps_attempt + 1))
                else:
                    logging.warning(f"PaperSpace layout search failed after 3 attempts: {e}")

        # Restore original active layout
        try:
            if saved_layout:
                doc.ActiveLayout = doc.Layouts.Item(saved_layout)
        except Exception:
            pass

        # Deduplicate by entity Handle (Pass 1 and Pass 2 may find the same entity)
        def _dedup(items):
            seen = set()
            out = []
            for r in items:
                try:
                    h = r[0].Handle
                except Exception:
                    h = id(r[0])
                if h not in seen:
                    seen.add(h)
                    out.append(r)
            return out

        # Priority: PaperSpace > ModelSpace (stamps in PaperSpace always visible in PDF)
        # Within each: exact name match > fallback (DRAWINGNUMBER+REVISION)
        if _ps_tbs:
            results = _dedup(_ps_tbs)
            print(f"    [DEBUG] Title blocks found in PaperSpace: {[r[3] for r in results]}")
        elif _ps_fallback:
            results = _dedup(_ps_fallback)
            print(f"    [DEBUG] Title blocks (fuzzy) found in PaperSpace: {[r[3] for r in results]}")
        elif _ms_tbs:
            results = _dedup(_ms_tbs)
            print(f"    [DEBUG] Title blocks found in ModelSpace: {[r[3] for r in results]}")
        elif _ms_fallback:
            results = _dedup(_ms_fallback)
            print(f"    [DEBUG] Title blocks (fuzzy) found in ModelSpace: {[r[3] for r in results]}")
        else:
            results = []

        return [(r[0], r[1], r[2], r[3]) for r in results] if results else []

    def _get_attrs_dict(self, block_ref) -> Dict:
        """Extract TAG -> attribute object dict from a block reference.

        Uses direct COM calls (no Dispatch wrapping, no _com_retry lambda).
        Dispatch() wrapping was found to BREAK attr.TextString on some DWGs
        (C-PLN-003) — the original COM proxy from GetAttributes() works correctly.
        """
        attrs = {}
        try:
            for attr in block_ref.GetAttributes():
                try:
                    attrs[attr.TagString.upper()] = attr
                except Exception:
                    pass
        except Exception:
            pass
        return attrs

    def _safe_get_text(self, attr) -> str:
        """Safely read attribute TextString with COM retry."""
        try:
            return attr.TextString or ''
        except Exception:
            try:
                # Re-wrap with Dispatch in case COM proxy lost type info
                attr2 = win32com.client.Dispatch(attr)
                return attr2.TextString or ''
            except Exception:
                try:
                    return self._com_retry(lambda: attr.TextString) or ''
                except Exception:
                    return ''

    def _safe_set_text(self, attr, value: str) -> bool:
        """Safely write attribute TextString with COM retry. Never raises.
        Returns True if the write succeeded, False otherwise."""
        try:
            attr.TextString = value
            return True
        except Exception as _e1:
            # "No database" = AutoCAD has no active drawing; retrying won't help
            # and would block for ~20s per attribute (5 retries × sleeps).
            if 'no database' in str(_e1).lower():
                logging.warning(f"_safe_set_text: No database — skip retry")
                return False
        try:
            # Re-wrap with Dispatch in case COM proxy lost type info
            attr2 = win32com.client.Dispatch(attr)
            attr2.TextString = value
            return True
        except Exception as _e2:
            if 'no database' in str(_e2).lower():
                logging.warning(f"_safe_set_text: No database — skip retry")
                return False
        try:
            self._com_retry(lambda: setattr(attr, 'TextString', value))
            return True
        except Exception as e:
            logging.warning(f"_safe_set_text failed (all 3 strategies): {e}")
            return False

    def _read_latest_ifr_row(self, attrs: Dict) -> Dict:
        """Find the highest non-empty revision row and return personnel info.

        Scans rows 6→1 looking for non-empty {N}REV tag.
        Returns {row, rev, drawn, check, engineer, qa, project, date, description}.
        """
        for row_num in range(self.REV_ROWS, 0, -1):
            rev_tag = f"{row_num}REV"
            if rev_tag in attrs:
                val = self._safe_get_text(attrs[rev_tag]).strip()
                if val:
                    info = {'row': row_num, 'rev': val}
                    for tag in self.PERSONNEL_TAGS:
                        full_tag = f"{row_num}{tag}"
                        if full_tag in attrs:
                            info[tag.lower()] = self._safe_get_text(attrs[full_tag]).strip()
                        else:
                            info[tag.lower()] = ''
                    date_tag = f"{row_num}DATE"
                    if date_tag in attrs:
                        info['date'] = self._safe_get_text(attrs[date_tag]).strip()
                    desc_tag = f"{row_num}DESCRIPTION"
                    if desc_tag in attrs:
                        info['description'] = self._safe_get_text(attrs[desc_tag]).strip()
                    # Per-FIELD backfill: a later revision row may leave a single
                    # personnel field blank (e.g. IFC row has DESIGNED='' while
                    # DRAWN/APPROVED are filled). Don't copy that blank into the
                    # AS BUILT row — backfill each empty field from the most recent
                    # EARLIER row that has a non-empty value for it.
                    for tag in self.PERSONNEL_TAGS:
                        key = tag.lower()
                        if info.get(key):
                            continue
                        for earlier in range(row_num - 1, 0, -1):
                            et = f"{earlier}{tag}"
                            if et in attrs:
                                ev = self._safe_get_text(attrs[et]).strip()
                                if ev:
                                    info[key] = ev
                                    break
                    return info
        return {}

    def _update_title_block(self, attrs: Dict, ifc_rev: int, personnel: Dict, date_str: str):
        """Update title block attributes for IFC conversion.

        When preserve_ifr=True:  keep IFR revision rows, add IFC row after them.
        When preserve_ifr=False: Rev0 clears all IFR rows, Rev1+ keeps existing IFC rows.
        """
        # Set main REVISION attribute
        if 'REVISION' in attrs:
            self._safe_set_text(attrs['REVISION'], str(ifc_rev))

        all_suffixes = ['REV', 'DATE', 'DESCRIPTION'] + self.PERSONNEL_TAGS

        if self.preserve_ifr:
            # Preserve IFR history: find last IFR row (skip existing IFC rows),
            # then place IFC row right after the last IFR row.
            last_ifr_row = 0
            existing_ifc_row = 0
            for row_num in range(1, self.REV_ROWS + 1):
                rev_tag = f"{row_num}REV"
                desc_tag = f"{row_num}DESCRIPTION"
                if rev_tag not in attrs:
                    continue
                val = self._safe_get_text(attrs[rev_tag]).strip()
                if not val:
                    continue
                # Check if this row is already an IFC row (from a previous run)
                desc_val = ''
                if desc_tag in attrs:
                    desc_val = self._safe_get_text(attrs[desc_tag]).strip().upper()
                if 'CONSTRUCTION' in desc_val:
                    existing_ifc_row = max(existing_ifc_row, row_num)
                else:
                    last_ifr_row = max(last_ifr_row, row_num)
            # If IFC row already exists, overwrite it; otherwise add after last IFR row
            if existing_ifc_row > 0:
                target_row = existing_ifc_row
            else:
                target_row = last_ifr_row + 1
        elif ifc_rev == 0:
            # Clear all IFR history rows (Tatua behavior)
            for row_num in range(1, self.REV_ROWS + 1):
                for suffix in all_suffixes:
                    tag = f"{row_num}{suffix}"
                    if tag in attrs:
                        self._safe_set_text(attrs[tag], '')
            target_row = 1
        else:
            # Rev1+: add to next row after existing IFC rows
            target_row = ifc_rev + 1

        if target_row > self.REV_ROWS:
            print(f"    Warning: IFC Rev{ifc_rev} needs row {target_row}, "
                  f"title block has {self.REV_ROWS} rows, using last row")
            target_row = self.REV_ROWS

        tag_prefix = str(target_row)
        if f'{tag_prefix}REV' in attrs:
            self._safe_set_text(attrs[f'{tag_prefix}REV'], str(ifc_rev))
        if f'{tag_prefix}DESCRIPTION' in attrs:
            self._safe_set_text(attrs[f'{tag_prefix}DESCRIPTION'], 'FOR CONSTRUCTION')
        if f'{tag_prefix}DATE' in attrs:
            self._safe_set_text(attrs[f'{tag_prefix}DATE'], date_str)

        # Copy personnel
        for tag in self.PERSONNEL_TAGS:
            full_tag = f"{tag_prefix}{tag}"
            if full_tag in attrs:
                self._safe_set_text(attrs[full_tag], personnel.get(tag.lower(), ''))

        # Clean up any duplicate IFC rows above target_row (from previous bug)
        if self.preserve_ifr:
            for row_num in range(target_row + 1, self.REV_ROWS + 1):
                desc_tag = f"{row_num}DESCRIPTION"
                if desc_tag in attrs:
                    desc_val = self._safe_get_text(attrs[desc_tag]).strip().upper()
                    if 'CONSTRUCTION' in desc_val:
                        for suffix in all_suffixes:
                            clear_tag = f"{row_num}{suffix}"
                            if clear_tag in attrs:
                                self._safe_set_text(attrs[clear_tag], '')

    def _publish_single_pdf(self, acad, dwg_path: Path,
                            pdf_path: Path) -> bool:
        """Export single DWG to PDF via PUBLISH + DSD (same pattern as PanelIFCManager).

        Closes all open docs, opens the DWG fresh, detects layout, builds DSD,
        runs -PUBLISH, waits for completion.
        """
        doc = None
        _temp_dir = None
        try:
            # PUBLISH SendCommand fails silently on non-ASCII paths (e.g. √)
            _use_temp = not str(dwg_path).isascii() or not str(pdf_path).isascii()
            actual_dwg = dwg_path
            actual_pdf = pdf_path
            if _use_temp:
                import tempfile
                _temp_dir = Path(tempfile.mkdtemp(prefix="publish_"))
                actual_dwg = _temp_dir / dwg_path.name
                actual_pdf = _temp_dir / pdf_path.name
                shutil.copy2(str(dwg_path), str(actual_dwg))
                print(f"    PUBLISH: Unicode path detected, using temp dir")

            # Close all open documents first (clean state)
            try:
                while acad.Documents.Count > 0:
                    acad.Documents.Item(0).Close(False)
                    time.sleep(1)
            except Exception:
                pass
            time.sleep(2)

            # Open the IFC DWG fresh
            for _retry in range(3):
                try:
                    doc = self._com_retry(
                        lambda p=str(actual_dwg): acad.Documents.Open(p))
                    break
                except Exception:
                    time.sleep(3)
            if doc is None:
                logging.warning(f"无法打开 {dwg_path.name} 用于 PUBLISH")
                return False

            # Wait for document to load
            for _wait in range(15):
                try:
                    _ = doc.ModelSpace.Count
                    break
                except Exception:
                    time.sleep(1)
            time.sleep(2)

            # Detect ALL non-Model layouts, sorted by TabOrder (correct page sequence)
            # Skip empty layouts (only default viewport, no real content) to avoid blank pages
            layout_info = []  # [(tab_order, name), ...]
            try:
                for layout in doc.Layouts:
                    if layout.Name.lower() != 'model':
                        try:
                            tab_order = layout.TabOrder
                        except Exception:
                            tab_order = 999
                        # Check if layout has real content (>= 2 entities means
                        # more than just the default viewport)
                        try:
                            block = layout.Block
                            entity_count = block.Count
                            if entity_count <= 1:
                                print(f"    跳过空布局: {layout.Name} ({entity_count} entities)")
                                continue
                        except Exception:
                            pass  # if can't check, include it
                        # Skip phantom layouts with no plot configuration
                        # (can appear after XREF bind adds entities to previously-empty layouts)
                        try:
                            config = layout.ConfigName
                            if not config or config.strip() == '' or config.strip().lower() == 'none':
                                print(f"    跳过无打印配置布局: {layout.Name}")
                                continue
                        except Exception:
                            pass
                        layout_info.append((tab_order, layout.Name))
            except Exception:
                pass
            layout_info.sort(key=lambda x: x[0])
            layout_names = [name for _, name in layout_info]
            if not layout_names:
                layout_names = ['Model']
            print(f"    PDF Layouts (sorted): {layout_names}")

            # Build DSD content with ALL layouts
            dwg_str = str(actual_dwg)
            pdf_str = str(actual_pdf)
            out_str = str(actual_pdf.parent)
            sheet_name = actual_dwg.stem

            dsd_lines = [
                '[DWF6Version]', 'Ver=1',
                '[DWF6MinorVersion]', 'MinorVer=1',
            ]
            for lname in layout_names:
                dsd_lines.extend([
                    f'[DWF6Sheet:{sheet_name}-{lname}]',
                    f'DWG={dwg_str}',
                    f'Layout={lname}',
                    'Setup=',
                    f'OriginalSheetPath={dwg_str}',
                    'Has Plot Port=0',
                    'Has3DDWF=0',
                ])
            dsd_lines.extend([
                '[Target]', 'Type=6',
                f'DWF={pdf_str}',
                f'OUT={out_str}',
                'PWD=',
                'PromptForDwfName=FALSE',
                '[PdfOptions]',
                'VectorResolution=600',
                'RasterResolution=400',
                '[SheetSetProperties]',
                'IsSheetSet=FALSE',
                'IsHomogeneous=FALSE',
                'SheetSet Storage File=',
                'AcadProfile=<<Default>>',
                'CategoryCount=0',
                '[AutoCAD Block Information]',
                'IncludeBlockInfo=0',
                'BlockTmplFilePath=',
            ])
            dsd_content = '\n'.join(dsd_lines)

            # Write DSD file
            to_long_path(pdf_path.parent).mkdir(parents=True, exist_ok=True)
            if _use_temp:
                _temp_dir.mkdir(parents=True, exist_ok=True)
            dsd_path = actual_dwg.parent / f"{sheet_name}.dsd"
            dsd_path.write_text(dsd_content, encoding='utf-8')

            # Suppress all dialogs for silent PUBLISH
            saved_filedia = doc.GetVariable("FILEDIA")
            saved_bgplot = doc.GetVariable("BACKGROUNDPLOT")
            doc.SetVariable("FILEDIA", 0)
            doc.SetVariable("BACKGROUNDPLOT", 0)

            # Run PUBLISH
            dsd_str = str(dsd_path)
            doc.SendCommand(f'-PUBLISH\n{dsd_str}\n')

            # Wait for PUBLISH to complete
            start = time.time()
            max_wait = 180
            last_log = start
            timed_out = True
            while time.time() - start < max_wait:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        timed_out = False
                        break
                except Exception:
                    timed_out = False
                    break
                elapsed = time.time() - start
                if time.time() - last_log > 30:
                    print(f"    PUBLISH 进行中... ({int(elapsed)}s)")
                    last_log = time.time()
                time.sleep(2)

            if timed_out:
                print(f"    PUBLISH 超时 ({max_wait}s)")

            # Restore system variables
            try:
                doc.SetVariable("FILEDIA", saved_filedia)
                doc.SetVariable("BACKGROUNDPLOT", saved_bgplot)
            except Exception:
                pass

            # Close document
            try:
                doc.Close(False)
                doc = None
            except Exception:
                pass

            # Cleanup DSD file
            try:
                dsd_path.unlink()
            except Exception:
                pass

            # Poll for PDF to appear and stabilize on disk
            # (CMDACTIVE=0 doesn't guarantee the plot engine has flushed)
            check_pdf = actual_pdf
            for _poll in range(30):
                if check_pdf.exists():
                    try:
                        sz1 = check_pdf.stat().st_size
                        if sz1 > 0:
                            time.sleep(2)
                            sz2 = check_pdf.stat().st_size
                            if sz1 == sz2:
                                break
                    except OSError:
                        pass
                time.sleep(2)

            # Move PDF from temp to target if needed
            if _use_temp and actual_pdf.exists():
                try:
                    shutil.move(str(actual_pdf), str(pdf_path))
                except Exception as e:
                    logging.warning(f"Failed to move PDF from temp: {e}")

            # Cleanup temp dir
            if _temp_dir and _temp_dir.exists():
                try:
                    shutil.rmtree(str(_temp_dir), ignore_errors=True)
                except Exception:
                    pass

            exists = pdf_path.exists()
            if not exists:
                print(f"    ⚠ PDF 未生成: {pdf_path.name}")
                print(f"      DWG={dwg_path}")
                print(f"      PDF={pdf_path}")
                if timed_out:
                    print(f"      原因: PUBLISH 超时 ({max_wait}s)")
                # Check if PUBLISH wrote PDF next to DWG instead of target
                alt_pdf = Path(str(actual_dwg)).parent / pdf_path.name
                if alt_pdf != pdf_path and alt_pdf.exists():
                    print(f"      发现 PDF 在 DWG 目录: {alt_pdf}")
                    try:
                        shutil.move(str(alt_pdf), str(pdf_path))
                        print(f"      已移动到目标目录")
                        exists = True
                    except Exception as e:
                        print(f"      移动失败: {e}")
            return exists

        except Exception as e:
            logging.warning(f"PDF PUBLISH failed: {e}")
            print(f"    ⚠ PDF PUBLISH 异常: {e}")
            if doc is not None:
                try:
                    doc.Close(False)
                except Exception:
                    pass
            if _temp_dir and _temp_dir.exists():
                try:
                    shutil.rmtree(str(_temp_dir), ignore_errors=True)
                except Exception:
                    pass
            return False

    def _build_ifc_filename(self, doc_id: str, description: str, ifc_rev: int) -> str:
        """Build IFC filename: {doc_id}_{description}_Rev{N}_IFC"""
        # Clean description — strip chars that are invalid in paths or break AutoCAD PUBLISH
        # () and & cause AutoCAD command-line parsing failures
        desc = re.sub(r'[<>:"/\\|?*()&]', '', description).strip()
        desc = re.sub(r'\s+', ' ', desc).replace(' ', '_')
        return f"{doc_id}_{desc}_Rev{ifc_rev}_IFC"

    def _wait_for_command(self, doc, timeout: int = 60):
        """Wait for AutoCAD command to finish."""
        start = time.time()
        time.sleep(0.5)
        while time.time() - start < timeout:
            try:
                if doc.GetVariable("CMDACTIVE") == 0:
                    return True
            except Exception:
                return False
            time.sleep(0.5)
        return False

    def convert_to_ifc(self, dwg_info: Dict) -> Dict:
        """Convert a single IFR DWG to IFC.

        Steps:
          1. Open DWG in AutoCAD
          2. Find title block, read latest IFR row
          3. Calculate IFC rev (0 if first, else existing + 1)
          4. Update title block
          5. SaveAs new IFC DWG
          6. Export PDF to IFC(Client)/
          7. Close without saving original

        Returns {success, dwg_path, pdf_path, ifc_rev, errors}.
        """
        result = {
            'success': False,
            'doc_id': dwg_info['doc_id'],
            'dwg_path': None,
            'pdf_path': None,
            'ifc_rev': None,
            'errors': [],
        }

        dwg_path = dwg_info['latest_ifr_dwg']
        doc_id = dwg_info['doc_id']
        description = dwg_info['description']

        # Check for lock files — try to remove stale ones, or reuse already-open doc
        lock1 = dwg_path.with_suffix('.dwl')
        lock2 = dwg_path.with_suffix('.dwl2')
        _reuse_open_doc = False  # Flag: DWG already open in AutoCAD
        if lock1.exists() or lock2.exists():
            # Try to delete stale locks first
            for _lock_attempt in range(3):
                any_locked = False
                for lf in (lock1, lock2):
                    if lf.exists():
                        try:
                            lf.unlink()
                            print(f"    锁文件已删除: {lf.name}")
                        except PermissionError:
                            any_locked = True
                if not any_locked:
                    break
                if _lock_attempt < 2:
                    print(f"    锁文件存在，等待后重试 ({_lock_attempt+1}/3)...")
                    time.sleep(3)
            # If still locked, check if DWG is already open in AutoCAD → reuse it
            if lock1.exists() or lock2.exists():
                try:
                    acad_check = self._get_acad()
                    for _di in range(acad_check.Documents.Count):
                        try:
                            _doc = acad_check.Documents.Item(_di)
                            if Path(_doc.FullName).resolve() == dwg_path.resolve():
                                _reuse_open_doc = True
                                print(f"    DWG 已在 AutoCAD 中打开，将直接使用")
                                break
                        except Exception:
                            continue
                except Exception:
                    pass
                if not _reuse_open_doc:
                    # Stale lock files (DWG not open in AutoCAD) — proceed anyway
                    print(f"    锁文件存在但 DWG 未在 AutoCAD 中打开，尝试强制打开")

        # Calculate IFC revision
        if dwg_info['existing_ifc_rev'] is not None:
            ifc_rev = dwg_info['existing_ifc_rev'] + 1
        else:
            ifc_rev = 0
        result['ifc_rev'] = ifc_rev

        if self.dry_run:
            ifc_name = self._build_ifc_filename(doc_id, description, ifc_rev)
            result['success'] = True
            result['dwg_path'] = str(dwg_info['folder'] / f"{ifc_name}.dwg")
            result['pdf_path'] = str(self.ifc_output / f"{ifc_name}.pdf")
            return result

        # --- Actual AutoCAD operations ---
        doc = None
        try:
            acad = self._get_acad()

            # For rev1+, try to open previous IFC DWG (preserves revision history)
            open_path = dwg_path
            if ifc_rev > 0:
                prev_name = self._build_ifc_filename(doc_id, description, ifc_rev - 1)
                prev_ifc = dwg_info['folder'] / f"{prev_name}.dwg"
                if prev_ifc.exists():
                    open_path = prev_ifc
                    print(f"    使用上一版 IFC DWG: {prev_ifc.name}")

            # Ensure Documents collection is accessible (avoid <unknown>.Open)
            for _doc_wait in range(10):
                try:
                    _ = acad.Documents.Count
                    break
                except Exception:
                    time.sleep(2)

            # Open the DWG — or reuse already-open document
            if _reuse_open_doc:
                # DWG is already open in AutoCAD (lock file confirmed it)
                try:
                    for _di in range(acad.Documents.Count):
                        _doc = acad.Documents.Item(_di)
                        if Path(_doc.FullName).resolve() == open_path.resolve():
                            doc = _doc
                            print(f"    已复用打开的文档: {doc.Name}")
                            break
                    if doc is None:
                        # Fallback: try normal open (AutoCAD may handle already-open gracefully)
                        doc = self._com_retry(
                            lambda p=str(open_path): acad.Documents.Open(p))
                except Exception as e:
                    result['errors'].append(f"无法获取已打开的 DWG: {e}")
                    return result
            else:
                try:
                    doc = self._com_retry(
                        lambda p=str(open_path): acad.Documents.Open(p))
                except Exception as e:
                    result['errors'].append(f"无法打开 DWG: {e}")
                    return result

            # Wait for document to be fully ready (not just ModelSpace accessible)
            for _wait in range(20):
                try:
                    _ = doc.ModelSpace.Count
                    # Also verify Layouts accessible (fails if doc not fully loaded)
                    _ = doc.Layouts.Count
                    break
                except Exception:
                    time.sleep(1)
            # Extra settle time for first document open in session
            time.sleep(2)

            # Find ALL title blocks — retry up to 3 times with increasing wait
            all_tbs = []
            for _tb_attempt in range(3):
                all_tbs = self._find_all_title_blocks(doc)
                if all_tbs:
                    break
                if _tb_attempt < 2:
                    print(f"    title block 搜索失败 (尝试 {_tb_attempt+1}/3)，等待后重试...")
                    time.sleep(3 * (_tb_attempt + 1))
            if not all_tbs:
                result['errors'].append(f"未找到 title block（尝试: {self.title_block_name}）")
                try:
                    doc.Close(False)
                except Exception:
                    pass
                return result
            print(f"    Found {len(all_tbs)} title block(s)")

            # Read personnel from first title block with valid IFR row
            personnel = {}
            for _tb_item in all_tbs:
                _attrs = _tb_item[1]
                personnel = self._read_latest_ifr_row(_attrs)
                if personnel:
                    break
            if not personnel and ifc_rev == 0:
                result['errors'].append("未找到有效的 IFR revision 行")
                try:
                    doc.Close(False)
                except Exception:
                    pass
                return result
            if not personnel:
                personnel = {}

            # Fix known typos (e.g. "Coulour" → "Colour")
            self._fix_known_typos(doc)

            # Scan for existing COLOUR stamp BEFORE removal (removal deletes IFR blocks)
            has_colour = self._scan_has_colour(doc)
            if has_colour:
                print(f"    COLOUR: 原DWG已有 COLOUR 印章，保留原样，仅画 FOR CONSTRUCTION")

            # Remove existing stamps ONCE (covers all spaces)
            self._remove_ifc_stamp(doc)

            # Update EVERY title block + add stamp near each one
            date_str = datetime.now().strftime('%d/%m/%y')
            for tb_idx, tb_item in enumerate(all_tbs, 1):
                block_ref = tb_item[0]
                attrs = tb_item[1]
                space = tb_item[2]
                layout_name = tb_item[3] if len(tb_item) > 3 else None
                print(f"    Sheet {tb_idx}/{len(all_tbs)}: 更新 title block + FOR CONSTRUCTION...")
                self._update_title_block(attrs, ifc_rev, personnel, date_str)
                if block_ref and space:
                    self._stamp_via_com_draw(doc, block_ref, space,
                                             has_colour=has_colour,
                                             layout_name=layout_name)
                else:
                    print(f"    ⚠ Sheet {tb_idx}: block_ref 或 space 为空，跳过印章")

            # Build output paths
            ifc_name = self._build_ifc_filename(doc_id, description, ifc_rev)
            ifc_dwg_path = dwg_info['folder'] / f"{ifc_name}.dwg"
            ifc_pdf_path = self.ifc_output / f"{ifc_name}.pdf"

            # Ensure IFC output directory exists
            to_long_path(self.ifc_output).mkdir(parents=True, exist_ok=True)

            # SaveAs new IFC DWG — use temp short path if full path > 240 chars
            # (AutoCAD COM does NOT support \\?\ long path prefix)
            final_dwg_path = ifc_dwg_path
            use_temp = len(str(ifc_dwg_path)) > 240
            if use_temp:
                import tempfile
                temp_dir = Path(tempfile.gettempdir()) / "IFC_TEMP"
                temp_dir.mkdir(parents=True, exist_ok=True)
                temp_dwg = temp_dir / f"{ifc_name}.dwg"
                save_path = temp_dwg
                print(f"    路径过长({len(str(ifc_dwg_path))}字符)，使用临时路径 SaveAs")

                # Bind all XREFs before SaveAs to temp — prevents missing content
                # when PUBLISH reopens DWG from temp dir (relative XREF paths break)
                xref_bound = 0
                try:
                    for bi in range(doc.Blocks.Count):
                        try:
                            blk = doc.Blocks.Item(bi)
                            if blk.IsXRef:
                                blk.Bind(False)  # Insert bind: keep original layer names
                                xref_bound += 1
                        except Exception:
                            pass
                    if xref_bound:
                        print(f"    XREF 绑定: {xref_bound} 个外部引用已绑定")
                        # Regen after XREF bind to stabilize document state.
                        # Sleep is OUTSIDE try so it always runs — 2s was too
                        # short; SaveAs immediately after bind triggers
                        # RPC_E_SERVERFAULT while AutoCAD is still processing.
                        try:
                            doc.Regen(1)  # acActiveViewport
                        except Exception:
                            pass
                        time.sleep(8)
                except Exception:
                    pass
            else:
                save_path = ifc_dwg_path

            # Clean up pre-existing file at save_path (from previous failed attempt)
            try:
                sp = Path(save_path)
                if sp.exists():
                    sp.unlink()
            except Exception:
                pass

            # SaveAs with multiple fallback strategies
            save_ok = False
            # Strategy 1: normal SaveAs to target path
            try:
                self._com_retry(lambda: doc.SaveAs(str(save_path)))
                save_ok = True
            except Exception as e1:
                print(f"    SaveAs 策略1失败({e1})，尝试备用方案...")

            # Strategy 2: Save() first (flush XREF bind), then SaveAs
            if not save_ok:
                try:
                    self._com_retry(lambda: doc.Save())
                    time.sleep(2)
                    self._com_retry(lambda: doc.SaveAs(str(save_path)))
                    save_ok = True
                except Exception as e2:
                    print(f"    SaveAs 策略2失败({e2})，尝试备用方案...")

            # Strategy 3: SaveAs with explicit DWG format (acNative=61)
            if not save_ok:
                try:
                    self._com_retry(lambda: doc.SaveAs(str(save_path), 61))
                    save_ok = True
                except Exception as e3:
                    print(f"    SaveAs 策略3失败({e3})，尝试备用方案...")

            # Strategy 4: If using temp, try direct long path as fallback
            if not save_ok and use_temp:
                try:
                    print(f"    尝试直接保存到长路径...")
                    save_path = ifc_dwg_path  # switch to original long path
                    use_temp = False  # disable temp move later
                    self._com_retry(lambda: doc.SaveAs(str(save_path)))
                    save_ok = True
                except Exception as e4:
                    print(f"    SaveAs 策略4失败({e4})，尝试 SendCommand...")

            # Strategy 5: Use SendCommand SAVEAS
            if not save_ok:
                try:
                    cmd_path = str(save_path).replace("\\", "\\\\")
                    doc.SendCommand(f'_SAVEAS\n\n{cmd_path}\n')
                    time.sleep(5)
                    if Path(save_path).exists():
                        save_ok = True
                    else:
                        print(f"    SendCommand SAVEAS 未生成文件")
                except Exception as e5:
                    print(f"    SaveAs 策略5失败({e5})")

            if not save_ok:
                # Strategy 6 (last resort): PUBLISH directly from open doc
                # Doc is still open with stamps + title block updated.
                # We can't SaveAs, but we can still produce the PDF.
                print(f"    SaveAs 全部失败 — 尝试直接从打开文档导出 PDF...")
                try:
                    # Save in-place (overwrite source) as last resort for DWG
                    self._com_retry(lambda: doc.Save())
                    print(f"    ⚠ DWG 已就地保存（覆盖源文件）")
                    save_path = Path(doc.FullName)
                    save_ok = True
                except Exception as e6:
                    print(f"    就地保存也失败({e6})，仍尝试导出 PDF...")
                    save_path = Path(doc.FullName)

                # Try PUBLISH from currently open doc path
                pdf_ok = self._publish_single_pdf(acad, save_path, ifc_pdf_path)
                try:
                    doc.Close(False)
                except Exception:
                    pass
                doc = None

                if pdf_ok:
                    result['success'] = True
                    result['pdf_path'] = str(ifc_pdf_path)
                    result['dwg_path'] = str(save_path)
                    result['errors'].append("⚠ SaveAs 失败，DWG 未另存为 IFC 文件名，但 PDF 已导出")
                else:
                    result['errors'].append("SaveAs 和 PDF 导出均失败")
                return result

            # Close document after SaveAs (clean state for PUBLISH)
            try:
                doc.Close(False)
            except Exception as _ce:
                logging.warning(f"doc.Close warning (non-fatal): {_ce}")
            doc = None
            time.sleep(2)

            # Export PDF via PUBLISH BEFORE moving DWG
            # (PUBLISH needs to reopen DWG — must use short path if temp was used)
            publish_dwg = save_path  # use the path AutoCAD actually saved to
            pdf_ok = self._publish_single_pdf(acad, publish_dwg, ifc_pdf_path)

            # NOW move DWG from temp to final path
            if use_temp:
                try:
                    shutil.move(str(save_path), str(to_long_path(final_dwg_path)))
                    ifc_dwg_path = final_dwg_path
                    print(f"    已移动到最终路径: {final_dwg_path.name}")
                except Exception as e:
                    ifc_dwg_path = save_path
                    print(f"    ⚠ 移动失败({e})，DWG 保留在临时目录: {save_path}")

            result['dwg_path'] = str(ifc_dwg_path)

            if pdf_ok:
                result['pdf_path'] = str(ifc_pdf_path)
                result['success'] = True
            else:
                result['errors'].append(
                    f"PDF PUBLISH 导出失败: {ifc_pdf_path.name} (DWG已保存)")
                result['success'] = False

        except Exception as e:
            result['errors'].append(f"转换异常: {e}")
            if doc is not None:
                try:
                    doc.Close(False)
                except Exception:
                    pass

        return result

    def batch_convert(self, doc_ids: Optional[List[str]] = None,
                      update_deliverables: bool = False) -> List[Dict]:
        """Convert multiple IFR DWGs to IFC.

        Args:
            doc_ids: If provided, only convert these doc-IDs. Otherwise convert all.
            update_deliverables: If True, auto-update deliverables Excel after conversion.

        Returns list of result dicts from convert_to_ifc().
        """
        scan = self.scan_native_folders()
        if doc_ids:
            scan = [s for s in scan if s['doc_id'] in doc_ids]

        if not scan:
            print("  没有找到需要转换的文件")
            return []

        results = []
        total = len(scan)
        for idx, dwg_info in enumerate(scan, 1):
            doc_id = dwg_info['doc_id']
            print(f"  [{idx}/{total}] {doc_id} (IFR Rev{dwg_info['latest_ifr_rev']} → "
                  f"IFC Rev{dwg_info['existing_ifc_rev'] + 1 if dwg_info['existing_ifc_rev'] is not None else 0})...")

            r = self.convert_to_ifc(dwg_info)
            results.append(r)

            if r['success']:
                status = "OK"
                extra = f"DWG={Path(r['dwg_path']).name}" if r['dwg_path'] else ''
                if r['pdf_path']:
                    extra += f", PDF={Path(r['pdf_path']).name}"
                if r['errors']:
                    extra += f" (警告: {'; '.join(r['errors'])})"
                print(f"    [{status}] {extra}")
            else:
                print(f"    [FAIL] {'; '.join(r['errors'])}")

        # Summary
        ok = sum(1 for r in results if r['success'])
        fail = total - ok
        print(f"\n  转换完成: 成功={ok}, 失败={fail}")

        # Auto-update deliverables if requested
        if update_deliverables and ok > 0:
            print("\n正在更新交付物状态...")
            self.update_ifc_deliverables(results)

        return results

    def update_ifc_deliverables(self, conversion_results: List[Dict]) -> Dict:
        """Update deliverables Excel with IFC revision info from conversion results.

        Uses exact revision numbers from the conversion rather than re-scanning,
        ensuring accuracy. Updates revision, status ('Approved IFC'), and date.

        Returns summary dict: {updated: int, skipped: int, errors: list}.
        """
        summary = {'updated': 0, 'skipped': 0, 'errors': []}

        # Filter to successful conversions
        successful = [r for r in conversion_results if r['success'] and r['ifc_rev'] is not None]
        if not successful:
            summary['errors'].append("没有成功的转换结果")
            return summary

        # Find deliverable Excel
        excel_path = self._dm.find_deliverable_excel()
        if not excel_path:
            summary['errors'].append("未找到交付物 Excel 文件")
            return summary

        if not OPENPYXL_AVAILABLE:
            summary['errors'].append("openpyxl 未安装")
            return summary

        wb = self._dm._load_workbook_with_retry(excel_path)
        if wb is None:
            summary['errors'].append(f"无法打开 Excel: {excel_path}")
            return summary

        ws = wb.active
        try:
            layout = self._dm.detect_layout(ws)
        except Exception as e:
            summary['errors'].append(f"Excel 格式检测失败: {e}")
            wb.close()
            return summary

        # Build doc_id → row mapping from Excel
        excel_items, _ = self._dm.read_excel_items(ws, layout)

        # Clear previous highlights
        self._dm.clear_previous_highlights(ws, layout)

        date_str = datetime.now().strftime('%d/%m/%y')
        updated_rows = []

        for result in successful:
            doc_id = result['doc_id']
            ifc_rev = result['ifc_rev']

            if doc_id not in excel_items:
                summary['skipped'] += 1
                print(f"    {doc_id}: 未在 Excel 中找到，跳过")
                continue

            row = excel_items[doc_id]['row']

            # Update revision (numeric IFC rev)
            ws.cell(row=row, column=layout.rev_col).value = str(ifc_rev)
            # Update status
            ws.cell(row=row, column=layout.status_col).value = 'Approved IFC'
            # Update date
            ws.cell(row=row, column=layout.date_col).value = date_str

            # Green fill on K-M
            ws.cell(row=row, column=layout.rev_col).fill = self._dm.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self._dm.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self._dm.STATUS_APPROVED_FILL
            # Marker highlights on J and N
            ws.cell(row=row, column=10).fill = self._dm.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self._dm.CHANGE_MARKER_FILL

            updated_rows.append(doc_id)
            summary['updated'] += 1
            print(f"    {doc_id}: Rev{ifc_rev} → Approved IFC ✓")

        if not updated_rows:
            wb.close()
            return summary

        # Update file revision and last updated date
        old_file_rev = ws[layout.file_rev_cell].value
        new_file_rev = self._dm._increment_file_revision(old_file_rev)
        old_str = str(old_file_rev).strip() if old_file_rev else ''
        prefix_match = re.match(r'^((?:Revision|Rev)\s*)', old_str, re.IGNORECASE)
        if prefix_match:
            ws[layout.file_rev_cell].value = f"{prefix_match.group(1)}{new_file_rev}"
        else:
            ws[layout.file_rev_cell].value = new_file_rev
        ws[layout.last_updated_cell].value = date_str

        if not self.dry_run:
            new_path = self._dm._get_new_filename(excel_path, new_file_rev)
            self._dm._supersede_file(excel_path)
            wb.save(str(to_long_path(new_path)))
            wb.close()
            self._dm._sync_deliverable(new_path)

            # Cleanup old versions
            ss_folder = excel_path.parent / "SS"
            if ss_folder.exists():
                self._dm.cleanup_ss_folder(ss_folder, max_versions=3)

            print(f"\n  交付物 Excel 已保存: {new_path.name}")
            print(f"  更新 {summary['updated']} 行, 跳过 {summary['skipped']} 行")
        else:
            wb.close()
            print(f"\n  [DRY-RUN] 将更新 {summary['updated']} 行")

        return summary


# =============================================================================
# Panel IFC Manager (multi-page panel designs)
# =============================================================================

# Regex: split filename into doc_no and page_number
# e.g. "TSF-EN-ELE-DRG-10-01" → ("TSF-EN-ELE-DRG-10", "01")
_RE_PANEL_PAGE = re.compile(r'^(.+)-(\d{2}|0[A-Za-z])$')

# Regex: match IFC PDF names like "TSF-EN-ELE-DRG-10_Rev0_IFC.pdf"
_RE_PANEL_IFC_PDF = re.compile(r'^(.+?)_Rev(\d+)_IFC\.pdf$', re.IGNORECASE)


class PanelIFCManager(IFCStampMixin):
    """IFC conversion for multi-page panel designs (one DWG per page).

    Unlike IFCManager which handles single-doc-per-DWG, this handles
    a flat folder of DWGs where multiple files belong to one document set
    (e.g. TSF-EN-ELE-DRG-10-00 through -10). Output is one multi-page
    PDF per document group via AutoCAD PUBLISH.

    Supports two title block update methods:
      - 'com': Direct COM API attribute modification (default)
      - 'utb': Lee-Mac Update Title Block AutoLISP plugin via CSV
    """

    # Max revision history rows to auto-detect (scan up to this)
    MAX_REV_ROWS = 6

    def __init__(self, source_folder, ifc_output=None, dry_run=False,
                 title_block_method='com', title_block_name=None,
                 utb_lsp_path=None):
        """
        Args:
            source_folder: Path to flat folder containing DWG files
            ifc_output: Path for PDF output (default: project's 4. IFC(Client))
            dry_run: Preview only, no changes
            title_block_method: 'com' or 'utb'
            title_block_name: Exact block name for COM mode (optional, uses fuzzy if None)
            utb_lsp_path: Path to Lee-Mac UTB .lsp file (required for 'utb' mode)
        """
        self.source_folder = Path(source_folder)
        self.ifc_output = Path(ifc_output) if ifc_output else self._detect_ifc_output()
        self.dry_run = dry_run
        self.title_block_method = title_block_method
        self.title_block_name = title_block_name
        self.utb_lsp_path = Path(utb_lsp_path) if utb_lsp_path else None
        self._acad = None

    def _detect_ifc_output(self) -> Path:
        """Walk up from source_folder to find project root, then return IFC(Client) path."""
        # Look for "1. Drawings" in parents
        for parent in self.source_folder.parents:
            ifc_dir = parent / "4. IFC(Client)"
            if ifc_dir.exists():
                return ifc_dir
            # Also check one level up from "1. Drawings"
            if parent.name == "1. Native" or parent.name.startswith("1."):
                drawings_dir = parent.parent
                ifc_dir = drawings_dir / "4. IFC(Client)"
                if ifc_dir.exists():
                    return ifc_dir
        # Fallback: sibling of source folder
        return self.source_folder.parent / "IFC_Output"

    def _get_acad(self):
        """Get or launch AutoCAD COM instance."""
        if not WIN32COM_AVAILABLE:
            raise RuntimeError("win32com 未安装，无法使用 AutoCAD COM API。请安装 pywin32。")
        if self._acad is not None:
            try:
                _ = self._acad.Visible
                return self._acad
            except Exception:
                self._acad = None
        try:
            self._acad = win32com.client.GetActiveObject("AutoCAD.Application")
            # NOTE: Do NOT wrap with Dispatch() — see IFCManager._get_acad comment
        except Exception:
            try:
                print("  正在启动 AutoCAD（可能需要30秒）...")
                self._acad = win32com.client.Dispatch("AutoCAD.Application")
                # Cold-start race: a freshly-launched AutoCAD returns from Dispatch
                # before it's ready; setting .Visible too soon raises "Visible can
                # not be set". Wait until the app responds, then set Visible best-effort.
                for _start_wait in range(30):
                    try:
                        _ = self._acad.Documents
                        break
                    except Exception:
                        time.sleep(1)
                try:
                    self._acad.Visible = True
                except Exception:
                    pass
                # Wait for AutoCAD to be ready
                for _wait in range(60):
                    try:
                        _ = self._acad.Documents
                        break
                    except Exception:
                        time.sleep(1)
            except Exception as e:
                raise RuntimeError(f"无法连接或启动 AutoCAD: {e}")
        # Wait for Documents collection to be ready
        for _doc_wait in range(15):
            try:
                _ = self._acad.Documents.Count
                break
            except Exception:
                time.sleep(1)
        # Suppress all notification dialogs
        try:
            self._acad.Preferences.System.BeepOnError = False
        except Exception:
            pass
        try:
            doc = self._acad.ActiveDocument
            if doc:
                doc.SetVariable("LAYERNOTIFY", 0)
                doc.SetVariable("LAYEREVALCTL", 0)
        except Exception:
            pass
        return self._acad

    # ── Scanning & Grouping ──────────────────────────────────────────────

    def scan_and_group(self) -> Dict[str, List[Dict]]:
        """Scan source folder DWGs and group by doc_no.

        Returns {doc_no: [{'page': '01', 'path': Path, 'filename': str}, ...]}
        sorted by page number within each group.
        """
        groups: Dict[str, List[Dict]] = {}
        for f in sorted(self.source_folder.iterdir()):
            if not f.is_file() or f.suffix.lower() != '.dwg':
                continue
            if f.name.startswith('~$'):
                continue
            m = _RE_PANEL_PAGE.match(f.stem)
            if not m:
                continue
            doc_no = m.group(1)
            page = m.group(2)
            if doc_no not in groups:
                groups[doc_no] = []
            groups[doc_no].append({
                'page': page,
                'path': f,
                'filename': f.name,
            })
        # Sort pages within each group
        for doc_no in groups:
            groups[doc_no].sort(key=lambda x: (
                # 00 first, then numeric, then 0A/0B etc at end
                0 if x['page'] == '00' else (2 if x['page'].startswith('0') and not x['page'].isdigit() else 1),
                x['page']
            ))
        return groups

    def _get_existing_ifc_rev(self, doc_no: str) -> int:
        """Check IFC(Client) folder for existing IFC PDFs of this doc_no.

        Returns next revision number (0 if no existing IFC found).
        """
        if not self.ifc_output.exists():
            return 0
        max_rev = -1
        for f in self.ifc_output.iterdir():
            if not f.is_file() or f.suffix.lower() != '.pdf':
                continue
            m = _RE_PANEL_IFC_PDF.match(f.name)
            if m and m.group(1) == doc_no:
                rev = int(m.group(2))
                if rev > max_rev:
                    max_rev = rev
        return max_rev + 1 if max_rev >= 0 else 0

    # ── Block Scanning (diagnostics) ────────────────────────────────────

    def scan_blocks(self, dwg_path: str) -> List[Dict]:
        """Scan a DWG file and return all block references with their attributes.

        Returns list of {layout, block_name, attributes: {tag: value}}.
        Used for diagnostics / remote debugging via bot.
        """
        import time as _time
        acad = self._get_acad()
        doc = acad.Documents.Open(str(dwg_path), True)  # ReadOnly
        _time.sleep(2)
        results = []
        try:
            for layout in doc.Layouts:
                layout_name = layout.Name
                block = layout.Block
                for i in range(block.Count):
                    entity = block.Item(i)
                    try:
                        if entity.EntityName != 'AcDbBlockReference':
                            continue
                    except Exception:
                        continue
                    try:
                        bname = entity.Name
                    except Exception:
                        continue
                    attrs = {}
                    try:
                        for attr in entity.GetAttributes():
                            attrs[attr.TagString] = attr.TextString
                    except Exception:
                        pass
                    results.append({
                        'layout': layout_name,
                        'block_name': bname,
                        'attributes': attrs,
                    })
        finally:
            doc.Close(False)
        return results

    # ── Title Block Operations (shared helpers) ──────────────────────────
    # NOTE: _com_retry inherited from IFCStampMixin

    def _find_title_block(self, doc):
        """Find title block in AutoCAD document.

        Returns (block_ref, attrs_dict, space). Search strategy:
          1. Exact name match (if title_block_name is set)
          2. Fuzzy: any block with DRAWINGNUMBER + REVISION tags
          3. Fuzzy: any block with >=5 attributes containing REV-like tags
        Searches PaperSpace first (title blocks are usually there).
        Logs all block names on failure to aid remote debugging.
        """
        # Search PaperSpace first — title blocks are almost always there
        spaces = []
        try:
            for layout in self._com_retry(lambda: doc.Layouts):
                if layout.Name.lower() != 'model':
                    spaces.append(self._com_retry(lambda: doc.PaperSpace))
                    break
        except Exception:
            pass
        try:
            ms = self._com_retry(lambda: doc.ModelSpace)
            if ms is not None:
                spaces.append(ms)
        except Exception:
            pass

        all_blocks_found = []
        best_candidate = None
        best_attr_count = 0

        for space in spaces:
            if space is None:
                continue
            count = self._com_retry(lambda: space.Count)
            if count is None:
                continue
            for i in range(count):
                try:
                    entity = self._com_retry(lambda idx=i: space.Item(idx))
                except Exception:
                    continue
                if entity is None:
                    continue
                try:
                    if entity.EntityName != 'AcDbBlockReference':
                        continue
                except Exception:
                    continue
                try:
                    block_name = entity.Name
                except Exception:
                    continue

                attrs = self._get_attrs_dict(entity)
                all_blocks_found.append((block_name, len(attrs), list(attrs.keys())[:10]))

                # Strategy 1: Exact match
                if self.title_block_name and block_name == self.title_block_name:
                    return entity, attrs, space

                # Strategy 2: Classic DRAWINGNUMBER + REVISION
                if 'DRAWINGNUMBER' in attrs and 'REVISION' in attrs:
                    return entity, attrs, space

                # Strategy 3: Any block with REV-like tags and many attributes
                rev_tags = [k for k in attrs if 'REV' in k.upper() or 'REVISION' in k.upper()]
                if rev_tags and len(attrs) >= 5 and len(attrs) > best_attr_count:
                    best_candidate = (entity, attrs, space)
                    best_attr_count = len(attrs)

        if best_candidate:
            print(f"    Title block 模糊匹配: {best_attr_count} 个属性")
            return best_candidate

        # Log all blocks found for remote debugging
        if all_blocks_found:
            print(f"    [DEBUG] 所有 block references ({len(all_blocks_found)}):")
            for bname, attr_count, sample_tags in all_blocks_found:
                print(f"      '{bname}' ({attr_count} attrs): {sample_tags}")
        else:
            print(f"    [DEBUG] 未找到任何 block reference")

        return None, {}, None

    def _get_attrs_dict(self, block_ref) -> Dict:
        """Extract TAG -> attribute object dict from a block reference."""
        attrs = {}
        try:
            raw_attrs = self._com_retry(lambda: block_ref.GetAttributes())
            if raw_attrs:
                for attr in raw_attrs:
                    tag = self._com_retry(lambda a=attr: a.TagString.upper())
                    if tag:
                        attrs[tag] = attr
        except Exception:
            pass
        return attrs

    def _detect_rev_rows(self, attrs: Dict) -> int:
        """Auto-detect how many revision history rows the title block has."""
        for n in range(self.MAX_REV_ROWS, 0, -1):
            if f"{n}REV" in attrs:
                return n
        return 0

    def _detect_personnel_tags(self, attrs: Dict) -> List[str]:
        """Auto-detect personnel TAG suffixes from row 1 attributes.

        Returns list like ['DRAWN', 'APPROVED', 'DESIGNED', 'PROJECT']
        or ['DRAWN', 'CHECK', 'ENGINEER', 'QA', 'PROJECT'].
        """
        known_suffixes = ['DRAWN', 'CHECK', 'ENGINEER', 'QA', 'PROJECT',
                          'APPROVED', 'DESIGNED', 'SUBJECT', 'DRAWINGNUMBER']
        found = []
        for suffix in known_suffixes:
            if f"1{suffix}" in attrs:
                found.append(suffix)
        return found

    def _read_latest_row(self, attrs: Dict) -> Dict:
        """Find the highest non-empty revision row and return all fields.

        Auto-detects row count and personnel tags.
        Returns {row, rev, drawn, approved, ...} with whatever tags exist.
        """
        rev_rows = self._detect_rev_rows(attrs)
        for row_num in range(rev_rows, 0, -1):
            rev_tag = f"{row_num}REV"
            if rev_tag in attrs:
                val = self._get_attr_text(attrs[rev_tag])
                if val:
                    info = {'row': row_num, 'rev': val}
                    # Collect all tags for this row
                    personnel_tags = self._detect_personnel_tags(attrs)
                    for tag in personnel_tags:
                        full_tag = f"{row_num}{tag}"
                        if full_tag in attrs:
                            info[tag.lower()] = self._get_attr_text(attrs[full_tag])
                        else:
                            info[tag.lower()] = ''
                    # Also get DATE and DESCRIPTION
                    for extra in ['DATE', 'DESCRIPTION']:
                        full_tag = f"{row_num}{extra}"
                        if full_tag in attrs:
                            info[extra.lower()] = self._get_attr_text(attrs[full_tag])
                    return info
        return {}

    def _set_attr_text(self, attr, value: str):
        """Set attribute TextString with COM retry protection."""
        self._com_retry(lambda: setattr(attr, 'TextString', value))

    def _get_attr_text(self, attr) -> str:
        """Get attribute TextString with COM retry protection."""
        result = self._com_retry(lambda: attr.TextString)
        return result.strip() if result else ''

    def _write_ifc_row(self, attrs: Dict, ifc_rev: int, personnel: Dict, date_str: str):
        """Update title block attributes for IFC conversion.

        - Rev 0 (first IFC): clear all IFR history, write rev0 to row 1.
        - Rev 1+ (subsequent IFC): keep existing IFC rows, add new rev to next row.
          Expects to be called on a DWG that already has previous IFC history
          (i.e. the previous IFC DWG, not the IFR source).
        Auto-detects row count and personnel tags.
        """
        # Set main REVISION attribute
        if 'REVISION' in attrs:
            self._set_attr_text(attrs['REVISION'], str(ifc_rev))

        # Auto-detect structure
        rev_rows = self._detect_rev_rows(attrs)
        personnel_tags = self._detect_personnel_tags(attrs)
        all_suffixes = ['REV', 'DATE', 'DESCRIPTION'] + personnel_tags

        if ifc_rev == 0:
            # First IFC: clear all IFR history rows, then write row 1
            for row_num in range(1, rev_rows + 1):
                for suffix in all_suffixes:
                    tag = f"{row_num}{suffix}"
                    if tag in attrs:
                        self._set_attr_text(attrs[tag], '')

        # Target row: rev0 → row 1, rev1 → row 2, rev2 → row 3, etc.
        target_row = ifc_rev + 1
        if target_row > rev_rows:
            print(f"    警告: IFC Rev{ifc_rev} 需要 row {target_row}，但 title block 只有 {rev_rows} 行")
            target_row = rev_rows  # 写到最后一行

        tag_prefix = str(target_row)
        if f'{tag_prefix}REV' in attrs:
            self._set_attr_text(attrs[f'{tag_prefix}REV'], str(ifc_rev))
        if f'{tag_prefix}DESCRIPTION' in attrs:
            self._set_attr_text(attrs[f'{tag_prefix}DESCRIPTION'], 'ISSUED FOR CONSTRUCTION')
        if f'{tag_prefix}DATE' in attrs:
            self._set_attr_text(attrs[f'{tag_prefix}DATE'], date_str)

        # Copy personnel
        for tag in personnel_tags:
            full_tag = f"{tag_prefix}{tag}"
            if full_tag in attrs and tag.lower() in personnel:
                self._set_attr_text(attrs[full_tag], personnel[tag.lower()])

    # ── IFC Stamp ──────────────────────────────────────────────────────────
    # NOTE: _remove_ifc_stamp and _add_ifc_stamp inherited from IFCStampMixin

    # ── COM Mode ─────────────────────────────────────────────────────────

    def _update_via_com(self, doc, ifc_rev: int, date_str: str,
                        skip_stamp: bool = False,
                        fallback_personnel: Optional[Dict] = None,
                        sheet_no: str = '', total_sheets: str = '') -> Dict:
        """Update title block via direct COM API attribute modification.

        Args:
            skip_stamp: If True, skip adding FOR CONSTRUCTION stamp (e.g. cover sheet).
            fallback_personnel: If this page has no revision rows, use personnel
                                from another page in the same group.
            sheet_no: Sheet number to fill into SHEET_NO attribute.
            total_sheets: Total sheet count to fill into TOTAL_SHEETS attribute.

        Returns {'success': bool, 'error': str, 'personnel': dict}.
        """
        block_ref, attrs, space = self._find_title_block(doc)
        if not attrs:
            return {'success': False, 'error': '未找到 title block', 'personnel': {}}

        personnel = self._read_latest_row(attrs)
        if not personnel:
            if fallback_personnel:
                print(f"    注意: 未找到 IFR revision 行，使用同组 personnel")
                personnel = fallback_personnel
            else:
                print(f"    注意: 未找到 IFR revision 行，使用空 personnel 继续")
                personnel = {}

        self._write_ifc_row(attrs, ifc_rev, personnel, date_str)

        # Fill sheet number / total sheets
        if sheet_no and 'SHEET_NO' in attrs:
            self._set_attr_text(attrs['SHEET_NO'], sheet_no)
        if total_sheets and 'TOTAL_SHEETS' in attrs:
            self._set_attr_text(attrs['TOTAL_SHEETS'], total_sheets)

        if not skip_stamp:
            self._add_ifc_stamp(doc, block_ref, space)
        return {'success': True, 'error': '', 'personnel': personnel}

    # ── UTB Mode ─────────────────────────────────────────────────────────

    def _generate_ifc_csv(self, pages: List[Dict], ifc_rev: int, date_str: str,
                          existing_csv: Optional[Path] = None) -> Path:
        """Generate IFC CSV for Lee-Mac UTB plugin.

        If existing_csv is provided, reads it and transforms IFR → IFC values.
        Otherwise, generates from scratch using COM to read current attributes.
        """
        import csv

        # Try to find existing CSV in source folder
        if existing_csv is None:
            for f in self.source_folder.iterdir():
                if f.suffix.lower() == '.csv' and 'RGSTR' in f.name.upper():
                    existing_csv = f
                    break

        output_csv = self.source_folder / '_IFC_TEMP.csv'

        if existing_csv and existing_csv.exists():
            # Read existing CSV and transform values
            with open(existing_csv, 'r', encoding='utf-8-sig') as fh:
                reader = csv.DictReader(fh)
                headers = reader.fieldnames
                rows = list(reader)

            page_filenames = {p['path'].stem for p in pages}

            new_rows = []
            for row in rows:
                filename = row.get('FILENAME', '').strip()
                if not filename or filename not in page_filenames:
                    continue
                # Transform IFR → IFC values
                row['1REV'] = str(ifc_rev)
                row['1DESCRIPTION'] = 'ISSUED FOR CONSTRUCTION'
                row['1DATE'] = date_str
                row['REVISION'] = str(ifc_rev)
                # Clear rows 2-4
                for n in range(2, 5):
                    for suffix in headers:
                        if suffix.startswith(f'{n}'):
                            row[suffix] = ''
                new_rows.append(row)

            with open(output_csv, 'w', encoding='utf-8', newline='') as fh:
                writer = csv.DictWriter(fh, fieldnames=headers)
                writer.writeheader()
                writer.writerows(new_rows)
        else:
            raise FileNotFoundError(
                f"未找到 drawing register CSV 文件。UTB 模式需要 CSV。\n"
                f"请确保 {self.source_folder} 中有 *RGSTR*.csv 文件。"
            )

        return output_csv

    def _update_via_utb(self, doc, csv_path: Path) -> Dict:
        """Update title block via Lee-Mac UTB AutoLISP plugin.

        Loads UTB.lsp, runs utb command which reads the CSV and updates attributes.
        Returns {'success': bool, 'error': str}.
        """
        if not self.utb_lsp_path or not self.utb_lsp_path.exists():
            return {'success': False, 'error': f'UTB .lsp 文件不存在: {self.utb_lsp_path}'}

        try:
            lsp_str = str(self.utb_lsp_path).replace('\\', '/')
            # Load the UTB lisp file
            doc.SendCommand(f'(load "{lsp_str}")\n')
            time.sleep(1)

            # Run the UTB command
            doc.SendCommand('utb\n')

            # Wait for completion
            start = time.time()
            while time.time() - start < 30:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        break
                except Exception:
                    break
                time.sleep(0.5)

            return {'success': True, 'error': ''}
        except Exception as e:
            return {'success': False, 'error': f'UTB 执行失败: {e}'}

    # ── Fallback Logic ───────────────────────────────────────────────────

    def _update_with_fallback(self, doc, ifc_rev: int, date_str: str,
                              csv_path: Optional[Path] = None,
                              skip_stamp: bool = False,
                              fallback_personnel: Optional[Dict] = None,
                              sheet_no: str = '', total_sheets: str = '') -> Dict:
        """Try primary method, auto-fallback on failure (no interactive prompt)."""
        if self.title_block_method == 'com':
            result = self._update_via_com(doc, ifc_rev, date_str,
                                          skip_stamp=skip_stamp,
                                          fallback_personnel=fallback_personnel,
                                          sheet_no=sheet_no,
                                          total_sheets=total_sheets)
            if not result['success'] and csv_path:
                print(f"    COM API 失败: {result['error']}，自动切换 UTB 模式")
                self.title_block_method = 'utb'
                result = self._update_via_utb(doc, csv_path)
        else:
            if csv_path is None:
                return {'success': False, 'error': 'UTB 模式需要 CSV 文件'}
            result = self._update_via_utb(doc, csv_path)
            if not result['success']:
                print(f"    UTB 失败: {result['error']}，自动切换 COM API 模式")
                self.title_block_method = 'com'
                result = self._update_via_com(doc, ifc_rev, date_str,
                                              skip_stamp=skip_stamp,
                                              fallback_personnel=fallback_personnel,
                                              sheet_no=sheet_no,
                                              total_sheets=total_sheets)
        return result

    # ── PDF Export (PUBLISH multi-page) ──────────────────────────────────

    def _detect_layout_name(self, dwg_path: Path) -> str:
        """Detect the first non-Model layout name from a DWG via COM.

        Falls back to 'Layout1' if detection fails.
        """
        try:
            acad = self._get_acad()
            doc = self._com_retry(lambda: acad.Documents.Open(str(dwg_path)))
            if doc is None:
                return 'Layout1'
            # Wait for load
            for _ in range(15):
                try:
                    _ = doc.ModelSpace.Count
                    break
                except Exception:
                    time.sleep(1)
            layout_name = 'Layout1'
            try:
                for layout in doc.Layouts:
                    if layout.Name.lower() != 'model':
                        layout_name = layout.Name
                        break
            except Exception:
                pass
            doc.Close(False)
            time.sleep(1)
            return layout_name
        except Exception:
            return 'Layout1'

    def _build_dsd(self, pages: List[Dict], pdf_path: Path, dwg_folder: Path,
                   layout_name: str = 'Layout1') -> str:
        """Build DSD (Drawing Set Description) file content for PUBLISH.

        Args:
            pages: List of page dicts with 'path' and 'page' keys
            pdf_path: Output PDF file path
            dwg_folder: Folder containing the DWGs to publish
            layout_name: Layout name to use for all sheets
        """
        lines = [
            '[DWF6Version]',
            'Ver=1',
            '[DWF6MinorVersion]',
            'MinorVer=1',
        ]

        for page in pages:
            ifc_dwg = dwg_folder / page['path'].name
            dwg_path = ifc_dwg if ifc_dwg.exists() else page['path']
            if not dwg_path.exists():
                continue  # Skip missing files to prevent PUBLISH from hanging
            sheet_name = page['path'].stem
            dwg_str = str(dwg_path)  # backslashes for Windows DSD
            lines.extend([
                f'[DWF6Sheet:{sheet_name}-{layout_name}]',
                f'DWG={dwg_str}',
                f'Layout={layout_name}',
                'Setup=',
                f'OriginalSheetPath={dwg_str}',
                'Has Plot Port=0',
                'Has3DDWF=0',
            ])

        # Global settings — backslash paths for Windows
        pdf_str = str(pdf_path)
        out_str = str(pdf_path.parent)
        lines.extend([
            '[Target]',
            'Type=6',
            f'DWF={pdf_str}',
            f'OUT={out_str}',
            'PWD=',
            'PromptForDwfName=FALSE',
            '[PdfOptions]',
            'VectorResolution=600',
            'RasterResolution=400',
            '[SheetSetProperties]',
            'IsSheetSet=FALSE',
            'IsHomogeneous=FALSE',
            'SheetSet Storage File=',
            'AcadProfile=<<Default>>',
            'CategoryCount=0',
            '[AutoCAD Block Information]',
            'IncludeBlockInfo=0',
            'BlockTmplFilePath=',
        ])

        return '\n'.join(lines)

    def _publish_group_pdf(self, doc_no: str, ifc_rev: int,
                           dwg_folder: Path, pages: List[Dict]) -> Dict:
        """Create multi-page PDF using PUBLISH command with DSD file.

        Returns {'success': bool, 'pdf_path': str, 'error': str}.
        """
        pdf_name = f"{doc_no}_Rev{ifc_rev}_IFC.pdf"
        pdf_path = self.ifc_output / pdf_name

        result = {'success': False, 'pdf_path': str(pdf_path), 'error': ''}

        if self.dry_run:
            result['success'] = True
            return result

        # Ensure output directory exists
        to_long_path(self.ifc_output).mkdir(parents=True, exist_ok=True)

        try:
            acad = self._get_acad()
            # Close any open documents first to avoid conflicts
            try:
                while acad.Documents.Count > 0:
                    acad.Documents.Item(0).Close(False)
                    time.sleep(1)
            except Exception:
                pass
            time.sleep(2)

            # Detect actual layout name from first DWG
            first_dwg = dwg_folder / pages[0]['path'].name
            if not first_dwg.exists():
                first_dwg = pages[0]['path']
            layout_name = self._detect_layout_name(first_dwg)
            print(f"    Layout 检测: '{layout_name}'")

            # Build DSD file with detected layout name
            dsd_content = self._build_dsd(pages, pdf_path, dwg_folder, layout_name)
            dsd_path = dwg_folder / f"{doc_no}_IFC.dsd"
            dsd_path.write_text(dsd_content, encoding='utf-8')

            # Need an open document to run PUBLISH
            doc = None
            for _retry in range(3):
                try:
                    doc = self._com_retry(
                        lambda p=str(first_dwg): acad.Documents.Open(p))
                    break
                except Exception:
                    time.sleep(3)
            if doc is None:
                result['error'] = f"无法打开 {first_dwg.name} 用于 PUBLISH"
                return result

            # Wait for document to load
            for _wait in range(30):
                try:
                    _ = doc.ModelSpace.Count
                    break
                except Exception:
                    time.sleep(1)
            time.sleep(2)

            # Suppress all dialogs for silent PUBLISH
            saved_filedia = doc.GetVariable("FILEDIA")
            saved_bgplot = doc.GetVariable("BACKGROUNDPLOT")
            doc.SetVariable("FILEDIA", 0)
            doc.SetVariable("BACKGROUNDPLOT", 0)

            dsd_str = str(dsd_path)
            doc.SendCommand(f'-PUBLISH\n{dsd_str}\n')

            # Wait for PUBLISH to complete (can be slow for many pages)
            start = time.time()
            max_wait = 300  # 5 minutes for large sets
            last_log = start
            timed_out = True
            while time.time() - start < max_wait:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        timed_out = False
                        break
                except Exception:
                    timed_out = False
                    break
                elapsed = time.time() - start
                if time.time() - last_log > 30:
                    print(f"    PUBLISH 进行中... ({int(elapsed)}s)")
                    last_log = time.time()
                time.sleep(2)

            if timed_out:
                print(f"    PUBLISH 超时 ({max_wait}s)，尝试关闭文档")

            # Restore system variables
            try:
                doc.SetVariable("FILEDIA", saved_filedia)
                doc.SetVariable("BACKGROUNDPLOT", saved_bgplot)
            except Exception:
                pass

            try:
                doc.Close(False)
            except Exception:
                pass

            # Poll for PDF to appear and stabilize on disk
            pdf_ok = False
            for _poll in range(30):
                if pdf_path.exists():
                    try:
                        sz1 = pdf_path.stat().st_size
                        if sz1 > 0:
                            time.sleep(2)
                            sz2 = pdf_path.stat().st_size
                            if sz1 == sz2:
                                pdf_ok = True
                                break
                    except OSError:
                        pass
                time.sleep(2)

            if pdf_ok or pdf_path.exists():
                result['success'] = True
            else:
                reason = "PUBLISH 超时" if timed_out else "PUBLISH 完成但 PDF 未生成"
                result['error'] = f"{reason}: {pdf_name}"

        except Exception as e:
            result['error'] = f"PUBLISH 失败: {e}"

        return result

    # ── Main Batch Conversion ────────────────────────────────────────────

    def batch_convert_panel(self, doc_no_filter: Optional[str] = None,
                            incremental: bool = True) -> List[Dict]:
        """Convert panel design DWGs to IFC.

        Args:
            doc_no_filter: If set, only convert this specific doc_no.
            incremental: If True, skip pages whose source DWG hasn't changed
                         since the last IFC export (compare file timestamps).

        Returns list of per-group result dicts.
        """
        groups = self.scan_and_group()
        if not groups:
            print("  未找到可分组的 DWG 文件")
            return []

        if doc_no_filter:
            groups = {k: v for k, v in groups.items() if k == doc_no_filter}
            if not groups:
                print(f"  未找到 doc_no={doc_no_filter} 的 DWG 文件")
                return []

        date_str = datetime.now().strftime('%d/%m/%y')
        results = []

        # Prepare IFC subfolder
        ifc_folder = self.source_folder / "IFC"
        if not self.dry_run:
            to_long_path(ifc_folder).mkdir(parents=True, exist_ok=True)

        # Prepare UTB CSV if needed
        csv_path = None
        if self.title_block_method == 'utb':
            all_pages = [p for pages in groups.values() for p in pages]
            # Determine IFC rev (use first group's rev for CSV generation)
            first_doc_no = next(iter(groups))
            first_rev = self._get_existing_ifc_rev(first_doc_no)
            try:
                csv_path = self._generate_ifc_csv(all_pages, first_rev, date_str)
                print(f"  已生成 IFC CSV: {csv_path.name}")
            except FileNotFoundError as e:
                print(f"  错误: {e}")
                return []

        for doc_no, pages in groups.items():
            ifc_rev = self._get_existing_ifc_rev(doc_no)

            group_result = {
                'doc_no': doc_no,
                'ifc_rev': ifc_rev,
                'pages': len(pages),
                'dwgs_updated': 0,
                'dwgs_failed': 0,
                'dwgs_skipped': 0,
                'pdf_result': None,
                'errors': [],
            }

            print(f"\n  [{doc_no}] {len(pages)} 页 -> IFC Rev{ifc_rev}")

            if self.dry_run:
                # Preview mode
                for page in pages:
                    print(f"    {page['filename']} (page {page['page']})")
                pdf_name = f"{doc_no}_Rev{ifc_rev}_IFC.pdf"
                print(f"    -> PDF: {pdf_name}")
                group_result['dwgs_updated'] = len(pages)
                group_result['pdf_result'] = {'success': True, 'pdf_path': str(self.ifc_output / pdf_name)}
                results.append(group_result)
                continue

            # --- Actual conversion ---
            acad = self._get_acad()
            group_personnel = {}  # collected from first page with personnel data

            # Ensure IFC output directory exists for SaveAs
            to_long_path(ifc_folder).mkdir(parents=True, exist_ok=True)

            skipped_pages = []  # pages skipped by incremental check (reuse existing IFC)

            for page in pages:
                source_path = page['path']
                prev_ifc_dwg = ifc_folder / source_path.name

                # Incremental: skip if source DWG unchanged since last IFC export
                if incremental and prev_ifc_dwg.exists():
                    src_mtime = source_path.stat().st_mtime
                    ifc_mtime = prev_ifc_dwg.stat().st_mtime
                    if src_mtime <= ifc_mtime:
                        print(f"    跳过(未变更): {source_path.name}")
                        skipped_pages.append(page)
                        group_result['dwgs_skipped'] += 1
                        continue

                # For rev1+, use previous IFC DWG (preserves rev history)
                if ifc_rev > 0 and prev_ifc_dwg.exists():
                    dwg_path = prev_ifc_dwg
                    print(f"    打开(IFC): {dwg_path.name}...", end=' ')
                else:
                    dwg_path = source_path
                    print(f"    打开: {dwg_path.name}...", end=' ')

                # Check lock files — try to clean stale locks first
                dwl1 = dwg_path.with_suffix('.dwl')
                dwl2 = dwg_path.with_suffix('.dwl2')
                if dwl1.exists() or dwl2.exists():
                    # Check if file is actually open in AutoCAD
                    stale = True
                    try:
                        open_docs = [acad.Documents.Item(i).FullName
                                     for i in range(acad.Documents.Count)]
                        if str(dwg_path) in open_docs:
                            stale = False
                    except Exception:
                        pass
                    if stale:
                        # Remove stale lock files
                        for lf in (dwl1, dwl2):
                            try:
                                lf.unlink(missing_ok=True)
                            except Exception:
                                pass
                        print("(清理残留锁文件) ", end='')
                    # If DWG is open in AutoCAD, reuse the open document
                    # (don't skip — user wants to succeed even with lock files)

                _panel_reuse_doc = False
                if dwl1.exists() or dwl2.exists():
                    # Lock files still present — check if we can reuse the open doc
                    try:
                        for _di in range(acad.Documents.Count):
                            _doc = acad.Documents.Item(_di)
                            if Path(_doc.FullName).resolve() == dwg_path.resolve():
                                _panel_reuse_doc = True
                                break
                    except Exception:
                        pass
                    if not _panel_reuse_doc:
                        print("(锁文件存在但 DWG 未打开，尝试强制打开) ", end='')

                try:
                    doc = None
                    if _panel_reuse_doc:
                        # Reuse already-open document
                        for _di in range(acad.Documents.Count):
                            _doc = acad.Documents.Item(_di)
                            if Path(_doc.FullName).resolve() == dwg_path.resolve():
                                doc = _doc
                                print(f"(复用已打开文档) ", end='')
                                break
                    if doc is None:
                        # Retry Open up to 5 times with increasing delay
                        for _retry in range(5):
                            try:
                                doc = self._com_retry(
                                    lambda p=str(dwg_path): acad.Documents.Open(p))
                                break
                            except Exception as open_err:
                                if _retry < 4:
                                    time.sleep(3 + _retry * 2)
                                else:
                                    raise open_err
                    # Wait for document to fully load
                    for _wait_load in range(30):
                        try:
                            _ = doc.ModelSpace.Count
                            break
                        except Exception:
                            time.sleep(1)
                    else:
                        raise RuntimeError("文档加载超时")
                    # Extra settle time before COM operations
                    time.sleep(2)
                except Exception as e:
                    print(f"打开失败: {e}")
                    group_result['dwgs_failed'] += 1
                    group_result['errors'].append(f"{dwg_path.name}: {e}")
                    continue

                # Fix known typos (e.g. "Coulour" → "Colour")
                self._fix_known_typos(doc)

                # Update title block (skip stamp for cover sheet page 00)
                is_cover = (page['page'] == '00')
                # Sheet numbering: page index within group (1-based)
                page_idx = pages.index(page) + 1
                tb_result = self._update_with_fallback(
                    doc, ifc_rev, date_str, csv_path,
                    skip_stamp=is_cover,
                    fallback_personnel=group_personnel or None,
                    sheet_no=str(page_idx),
                    total_sheets=str(len(pages)))

                # Collect personnel from first page that has it
                if tb_result.get('personnel') and not group_personnel:
                    group_personnel = tb_result['personnel']

                if tb_result['success']:
                    # SaveAs to IFC subfolder
                    ifc_dwg = ifc_folder / dwg_path.name
                    try:
                        doc.SaveAs(str(ifc_dwg))
                        doc.Close(False)
                        time.sleep(5)  # Let AutoCAD fully settle before next Open
                        group_result['dwgs_updated'] += 1
                        print("OK")
                    except Exception as e:
                        try:
                            doc.Close(False)
                        except Exception:
                            pass
                        time.sleep(2)
                        group_result['dwgs_failed'] += 1
                        group_result['errors'].append(f"{dwg_path.name}: SaveAs 失败 - {e}")
                        print(f"SaveAs 失败: {e}")
                else:
                    try:
                        doc.Close(False)
                    except Exception:
                        pass
                    time.sleep(2)
                    group_result['dwgs_failed'] += 1
                    group_result['errors'].append(f"{dwg_path.name}: {tb_result['error']}")
                    print(f"TB 更新失败: {tb_result['error']}")

            # Report skipped pages
            if skipped_pages:
                print(f"    增量跳过: {len(skipped_pages)} 页未变更")

            # Publish multi-page PDF — include both updated and reused IFC DWGs
            has_any_ok = group_result['dwgs_updated'] > 0 or group_result['dwgs_skipped'] > 0
            if has_any_ok:
                ok_pages = [p for p in pages if (ifc_folder / p['path'].name).exists()]
                if ok_pages:
                    print(f"    正在生成多页 PDF ({len(ok_pages)}/{len(pages)} 页)...")
                    pdf_result = self._publish_group_pdf(doc_no, ifc_rev, ifc_folder, ok_pages)
                    group_result['pdf_result'] = pdf_result
                    if pdf_result['success']:
                        print(f"    PDF: {Path(pdf_result['pdf_path']).name}")
                    else:
                        print(f"    PDF 失败: {pdf_result['error']}")
                        group_result['errors'].append(pdf_result['error'])
                else:
                    print(f"    IFC 文件夹中无可用 DWG，跳过 PDF 导出")
            else:
                print(f"    所有 DWG 更新失败，跳过 PDF 导出")

            results.append(group_result)

        # Summary
        total_ok = sum(r['dwgs_updated'] for r in results)
        total_fail = sum(r['dwgs_failed'] for r in results)
        total_skipped = sum(r.get('dwgs_skipped', 0) for r in results)
        pdf_ok = sum(1 for r in results if r.get('pdf_result', {}).get('success'))
        skip_msg = f", 跳过(未变更) {total_skipped}" if total_skipped else ""
        print(f"\n  完成: DWG 更新 {total_ok} 成功 / {total_fail} 失败{skip_msg}, PDF {pdf_ok}/{len(results)} 组")

        return results


# =============================================================================
# Approved IFC Manager — detect approved IFR files → IFC conversion
# =============================================================================

# Regex: filename contains "-Approved" (before or after extension-like parts)
_RE_APPROVED_SUFFIX = re.compile(r'-Approved', re.IGNORECASE)

# Doc-ID patterns (reuse from DeliverableManager for standalone extraction)
_RE_DOC_ID_GG = re.compile(r'^(GG\d{2}-[A-Z]-[A-Z]{3}-\d{3})', re.IGNORECASE)
_RE_DOC_ID_LMS = re.compile(r'^(\d{5}-[A-Z]{2}-\d{3})', re.IGNORECASE)
_RE_DOC_ID_GENERIC = re.compile(r'^([A-Z0-9][\w]+-[A-Z]+-[A-Z]*-?\d{3})', re.IGNORECASE)
_RE_DOC_ID_TSF = re.compile(r'^(TSF-[A-Z]{2}-[A-Z]{3}-\w+-\d{2})', re.IGNORECASE)


def _extract_doc_id_standalone(filename: str) -> Optional[str]:
    """Extract doc-ID from a filename without needing a class instance.

    Two-pass strategy:
      1. match() — doc-ID at start of filename (standard naming)
      2. search() — doc-ID anywhere in filename (non-standard naming)
    This ensures files with inaccurate names are still matched by FILE NO.
    """
    stem = Path(filename).stem
    # Strip -Approved suffix first for cleaner matching
    stem_clean = _RE_APPROVED_SUFFIX.sub('', stem)
    # Pass 1: standard — doc-ID at start of filename
    for pat in [_RE_DOC_ID_GG, _RE_DOC_ID_LMS, _RE_DOC_ID_TSF, _RE_DOC_ID_GENERIC]:
        m = pat.match(stem_clean)
        if m:
            return m.group(1)
    # Pass 2: non-standard — doc-ID anywhere in filename
    # Handles files where FILE NAME is inaccurate but FILE NO is present somewhere
    for pat in [_RE_DOC_ID_GG, _RE_DOC_ID_LMS, _RE_DOC_ID_TSF]:
        m = pat.search(stem_clean)
        if m:
            return m.group(1)
    # Fallback: take everything before first _Rev or _rev
    m = re.match(r'^(.+?)(?:[_\s]-?[Rr]ev)', stem_clean)
    if m:
        return m.group(1).rstrip('_- ')
    return None


class FileHealthChecker:
    """Scan project folders for anomalous files that need attention.

    Detects:
      - Files without valid doc-ID in doc-ID-named folders
      - Files whose doc-ID doesn't match the parent folder's doc-ID
      - Files missing revision suffix
      - Empty folders (no deliverable files)

    Suggests renames based on folder naming convention + SS/Superseded history.
    """

    SCAN_PATHS = [
        "Design/Engineering/2. Calcs & Reports/Reports/Electrical",
        "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
        "Design/Engineering/2. Calcs & Reports/Schedule",
        "Design/Engineering/1. Drawings/1. Native",
    ]

    SKIP_SUBFOLDERS = {'ss', 'superseded', 'superceded', 'approved to ifc', 'appendix',
                       'reference', 'reference information', 'bom', 'stk'}

    TARGET_EXTENSIONS = {'.pdf'}

    EXCLUDE_PREFIXES = ('~$', '.')

    _RE_REV_LETTER = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*([A-Z])(?=[_.\s]|$)', re.IGNORECASE)
    _RE_REV_NUMBER = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?=[_.\s]|$)', re.IGNORECASE)

    def __init__(self, project_path: Path, dry_run: bool = True):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.logger = logging.getLogger(self.__class__.__name__)

    def scan_anomalies(self) -> Dict[str, list]:
        """Scan project for file anomalies. Returns dict with categorized results."""
        results = {
            'anomalies': [],
            'empty_folders': [],
            'scan_paths_checked': 0,
            'folders_checked': 0,
            'files_checked': 0,
        }

        for rel_path in self.SCAN_PATHS:
            scan_dir = self.project_path / rel_path
            if not scan_dir.exists():
                continue
            results['scan_paths_checked'] += 1
            self._scan_directory(scan_dir, results)

        return results

    def _scan_directory(self, parent_dir: Path, results: Dict):
        """Scan a parent directory containing doc-ID-named subfolders."""
        try:
            items = list(parent_dir.iterdir())
        except (OSError, PermissionError):
            return

        for folder in items:
            if not folder.is_dir():
                continue
            if folder.name.lower() in self.SKIP_SUBFOLDERS:
                continue

            folder_doc_id = _extract_doc_id_standalone(folder.name)
            if not folder_doc_id:
                continue

            results['folders_checked'] += 1
            self._check_folder(folder, folder_doc_id, results)

    def _check_folder(self, folder: Path, folder_doc_id: str, results: Dict):
        """Check a single doc-ID folder for anomalous files."""
        try:
            files = [f for f in folder.iterdir()
                     if f.is_file()
                     and not f.name.startswith(self.EXCLUDE_PREFIXES)
                     and f.suffix.lower() in self.TARGET_EXTENSIONS]
        except (OSError, PermissionError):
            return

        if not files:
            results['empty_folders'].append({
                'folder': folder,
                'folder_doc_id': folder_doc_id,
            })
            return

        ss_folder = self._find_ss_folder(folder)
        ss_versions = self._collect_ss_versions(ss_folder, folder_doc_id)
        current_versions = self._collect_current_versions(folder, folder_doc_id)
        highest_rev = self._get_highest_revision(ss_versions + current_versions)

        for f in files:
            results['files_checked'] += 1

            file_doc_id = _extract_doc_id_standalone(f.name)
            file_rev = self._extract_revision(f.name)

            anomaly = None

            if not file_doc_id:
                anomaly = {
                    'type': 'no_doc_id',
                    'file': f,
                    'folder': folder,
                    'folder_doc_id': folder_doc_id,
                    'reason': '文件无 doc-ID，不符合文件夹命名规范',
                }
            elif file_doc_id.upper() != folder_doc_id.upper():
                anomaly = {
                    'type': 'wrong_doc_id',
                    'file': f,
                    'folder': folder,
                    'folder_doc_id': folder_doc_id,
                    'file_doc_id': file_doc_id,
                    'reason': f'文件 doc-ID ({file_doc_id}) 与文件夹 ({folder_doc_id}) 不匹配',
                }
            elif not file_rev:
                anomaly = {
                    'type': 'no_revision',
                    'file': f,
                    'folder': folder,
                    'folder_doc_id': folder_doc_id,
                    'reason': '文件缺少版本号后缀',
                }

            if anomaly:
                suggested = self._suggest_filename(
                    folder, folder_doc_id, f.suffix, highest_rev)
                anomaly['suggested_name'] = suggested['name']
                anomaly['suggested_rev'] = suggested['rev']
                anomaly['rev_basis'] = suggested['basis']
                results['anomalies'].append(anomaly)

    def _extract_revision(self, filename: str) -> Optional[str]:
        stem = Path(filename).stem
        m = self._RE_REV_LETTER.search(stem)
        if m:
            return m.group(1).upper()
        m = self._RE_REV_NUMBER.search(stem)
        if m:
            return m.group(1)
        return None

    def _find_ss_folder(self, folder: Path) -> Optional[Path]:
        try:
            for item in folder.iterdir():
                if item.is_dir() and item.name.lower() in ('ss', 'superseded', 'superceded'):
                    return item
        except (OSError, PermissionError):
            pass
        return None

    def _collect_ss_versions(self, ss_folder: Optional[Path], doc_id: str) -> List[str]:
        if not ss_folder or not ss_folder.exists():
            return []
        revs = []
        try:
            for f in ss_folder.iterdir():
                if not f.is_file():
                    continue
                fid = _extract_doc_id_standalone(f.name)
                if fid and fid.upper() == doc_id.upper():
                    rev = self._extract_revision(f.name)
                    if rev:
                        revs.append(rev)
        except (OSError, PermissionError):
            pass
        return revs

    def _collect_current_versions(self, folder: Path, doc_id: str) -> List[str]:
        revs = []
        try:
            for f in folder.iterdir():
                if not f.is_file():
                    continue
                fid = _extract_doc_id_standalone(f.name)
                if fid and fid.upper() == doc_id.upper():
                    rev = self._extract_revision(f.name)
                    if rev:
                        revs.append(rev)
        except (OSError, PermissionError):
            pass
        return revs

    @staticmethod
    def _get_highest_revision(revs: List[str]) -> Optional[str]:
        if not revs:
            return None
        letter_revs = [r for r in revs if r.isalpha()]
        number_revs = [r for r in revs if r.isdigit()]
        highest = None
        if number_revs:
            highest = str(max(int(r) for r in number_revs))
        if letter_revs:
            h_letter = max(letter_revs, key=lambda x: ord(x.upper()))
            if highest is None:
                highest = h_letter
            else:
                highest = h_letter
        return highest

    @staticmethod
    def _next_revision(rev: str) -> str:
        if rev.isdigit():
            return str(int(rev) + 1)
        if rev.isalpha() and len(rev) == 1:
            return chr(ord(rev.upper()) + 1)
        return rev

    def _suggest_filename(self, folder: Path, doc_id: str, ext: str,
                          highest_rev: Optional[str]) -> Dict[str, str]:
        folder_stem = folder.name
        m_rev = self._RE_REV_LETTER.search(folder_stem)
        if not m_rev:
            m_rev = self._RE_REV_NUMBER.search(folder_stem)

        if m_rev:
            base = folder_stem[:m_rev.start()]
        else:
            base = re.sub(r'[_\s-]?Rev\s*$', '', folder_stem, flags=re.IGNORECASE)

        if highest_rev:
            next_rev = self._next_revision(highest_rev)
            basis = f'SS/当前最高 Rev{highest_rev} → 下一版本 Rev{next_rev}'
        else:
            next_rev = 'A'
            basis = '无历史版本 → 默认 RevA'

        suggested = f"{base}_Rev{next_rev}{ext}"
        return {'name': suggested, 'rev': next_rev, 'basis': basis}

    def execute_rename(self, anomaly: Dict) -> bool:
        """Rename a single anomalous file to its suggested name."""
        src = Path(anomaly['file'])
        dest = src.parent / anomaly['suggested_name']
        if self.dry_run:
            self.logger.info(f"[DRY-RUN] 重命名: {src.name} → {anomaly['suggested_name']}")
            return True
        try:
            src_long = to_long_path(src)
            dest_long = to_long_path(dest)
            if dest_long.exists():
                self.logger.warning(f"目标文件已存在: {dest}")
                return False
            src_long.rename(dest_long)
            self.logger.info(f"已重命名: {src.name} → {anomaly['suggested_name']}")
            return True
        except Exception as e:
            self.logger.error(f"重命名失败: {src.name} → {e}")
            return False

    def execute_renames(self, anomalies: List[Dict]) -> Dict[str, int]:
        """Execute renames for a list of anomalies."""
        stats = {'renamed': 0, 'skipped': 0, 'failed': 0}
        for a in anomalies:
            if self.execute_rename(a):
                stats['renamed'] += 1
            else:
                stats['failed'] += 1
        return stats


class ApprovedIFCManager(IFCStampMixin):
    """Detect approved IFR files, archive them, locate Native DWGs, convert to IFC.

    Workflow:
      1. Scan 13. Client Sharepoint/1.IFR/{1.Report, 2.Drawing}/ for -Approved PDFs
      2. Move detected -Approved files to Approved to IFC/ subfolder
      3. Merge with existing Approved to IFC/ contents
      4. Extract doc-IDs → map to 1. Native/ DWG folders
      5. In each Native folder: identify latest IFR DWG, archive outdated DWGs to SS/
      6. Delegate IFC conversion to IFCManager.convert_to_ifc()
      7. Update deliverable Excel via IFCManager.update_ifc_deliverables()
    """

    # Subfolders under 13. Client Sharepoint/1.IFR/ to scan
    SHAREPOINT_CHILDREN = {
        '1.Report': 'Report',
        '2.Drawing': 'Drawing',
    }
    APPROVED_SUBFOLDER = 'Approved to IFC'

    # Path probing order for each key directory
    _SHAREPOINT_CANDIDATES = [
        "Design/Engineering/13. Client Sharepoint/1.IFR",
        "13. Client Sharepoint/1.IFR",
    ]
    _NATIVE_CANDIDATES = [
        "Design/Engineering/1. Drawings/1. Native",
        "1. Drawings/1. Native",
    ]
    _IFC_OUTPUT_CANDIDATES = [
        "Design/Engineering/1. Drawings/4. IFC(Client)",
        "1. Drawings/4. IFC(Client)",
    ]

    def __init__(self, project_path, dry_run=False, title_block_name=None):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.title_block_name = title_block_name
        self.logger = logging.getLogger(self.__class__.__name__)

        paths = self._detect_paths()
        self.sharepoint_ifr = paths['sharepoint_ifr']
        self.native_root = paths['native_root']
        self.ifc_output = paths['ifc_output']

    def _detect_paths(self) -> Dict[str, Path]:
        """Auto-detect key directory paths for this project."""
        result = {}

        # Sharepoint IFR
        for rel in self._SHAREPOINT_CANDIDATES:
            p = self.project_path / rel
            if p.exists():
                result['sharepoint_ifr'] = p
                break
        if 'sharepoint_ifr' not in result:
            raise FileNotFoundError(
                f"项目 {self.project_path.name} 未找到 Client Sharepoint IFR 目录。\n"
                f"尝试路径: {', '.join(self._SHAREPOINT_CANDIDATES)}"
            )

        # Native root
        for rel in self._NATIVE_CANDIDATES:
            p = self.project_path / rel
            if p.exists():
                result['native_root'] = p
                break
        if 'native_root' not in result:
            raise FileNotFoundError(
                f"项目 {self.project_path.name} 未找到 1. Native 目录"
            )

        # IFC output
        for rel in self._IFC_OUTPUT_CANDIDATES:
            p = self.project_path / rel
            if p.exists():
                result['ifc_output'] = p
                break
        if 'ifc_output' not in result:
            # Create default path based on native_root sibling
            result['ifc_output'] = result['native_root'].parent / "4. IFC(Client)"

        return result

    # ── Scanning ─────────────────────────────────────────────────────────

    def scan_approved_files(self) -> List[Dict]:
        """Scan Sharepoint IFR subfolders for approved PDFs.

        Merges:
          - Files in current directory with '-Approved' in name (needs_archive=True)
          - Files already in 'Approved to IFC/' subfolder (needs_archive=False)

        Returns list of {doc_id, filename, source_path, category, needs_archive}.
        Deduplicates by doc_id (prefers non-archived if both exist).
        """
        items = []  # (doc_id, dict)

        for child_name, category in self.SHAREPOINT_CHILDREN.items():
            child_dir = self.sharepoint_ifr / child_name
            if not child_dir.exists():
                continue

            # Method 1: scan current dir for -Approved files
            try:
                for f in child_dir.iterdir():
                    if not f.is_file():
                        continue
                    if f.suffix.lower() != '.pdf':
                        continue
                    if f.name.startswith('~$'):
                        continue
                    if not _RE_APPROVED_SUFFIX.search(f.stem):
                        continue
                    doc_id = _extract_doc_id_standalone(f.name)
                    if doc_id:
                        items.append((doc_id, {
                            'doc_id': doc_id,
                            'filename': f.name,
                            'source_path': f,
                            'category': category,
                            'needs_archive': True,
                        }))
                    else:
                        self.logger.warning(
                            f"无法提取 doc-ID: {f.name} — FILE NO 不在文件名中")
                        print(f"    ⚠ 跳过(无法提取 FILE NO): {f.name}")
            except (OSError, PermissionError) as e:
                self.logger.warning(f"扫描 {child_dir} 失败: {e}")

            # Method 2: scan Approved to IFC/ subfolder
            approved_dir = child_dir / self.APPROVED_SUBFOLDER
            if approved_dir.exists():
                try:
                    for f in approved_dir.iterdir():
                        if not f.is_file():
                            continue
                        if f.suffix.lower() != '.pdf':
                            continue
                        if f.name.startswith('~$'):
                            continue
                        doc_id = _extract_doc_id_standalone(f.name)
                        if doc_id:
                            items.append((doc_id, {
                                'doc_id': doc_id,
                                'filename': f.name,
                                'source_path': f,
                                'category': category,
                                'needs_archive': False,
                            }))
                        else:
                            self.logger.warning(
                                f"无法提取 doc-ID: {f.name} — FILE NO 不在文件名中")
                            print(f"    ⚠ 跳过(无法提取 FILE NO): {f.name}")
                except (OSError, PermissionError) as e:
                    self.logger.warning(f"扫描 {approved_dir} 失败: {e}")

        # Dedup by doc_id: prefer needs_archive=True (so we archive it), then keep first
        seen = {}
        for doc_id, item in items:
            if doc_id not in seen:
                seen[doc_id] = item
            elif item['needs_archive'] and not seen[doc_id]['needs_archive']:
                # New approved file supersedes already-archived copy
                seen[doc_id] = item

        return sorted(seen.values(), key=lambda x: x['doc_id'])

    def archive_approved_files(self, approved_items: List[Dict]) -> int:
        """Move -Approved files from current dir to Approved to IFC/ subfolder.

        Only moves items where needs_archive=True.
        Returns count of files moved.
        """
        moved = 0
        for item in approved_items:
            if not item['needs_archive']:
                continue

            source = item['source_path']
            approved_dir = source.parent / self.APPROVED_SUBFOLDER

            if self.dry_run:
                self.logger.info(f"[DRY-RUN] 归档: {source.name} → {self.APPROVED_SUBFOLDER}/")
                print(f"    [预览] 归档: {source.name} → {self.APPROVED_SUBFOLDER}/")
                moved += 1
                continue

            try:
                to_long_path(approved_dir).mkdir(parents=True, exist_ok=True)
                dest = approved_dir / source.name
                if dest.exists():
                    # File already in target; skip move
                    self.logger.info(f"已存在于归档目录，跳过: {source.name}")
                    continue
                shutil.move(str(to_long_path(source)), str(to_long_path(dest)))
                # Update item's source_path to new location
                item['source_path'] = dest
                item['needs_archive'] = False
                self.logger.info(f"归档: {source.name} → {self.APPROVED_SUBFOLDER}/")
                moved += 1
            except Exception as e:
                self.logger.error(f"归档失败: {source.name} - {e}")

        return moved

    # ── Native DWG Mapping ───────────────────────────────────────────────

    def _find_native_folder(self, doc_id: str) -> Optional[Path]:
        """Find the Native subfolder matching a doc-ID.

        Search strategy:
          1. Direct children of native_root (flat structure)
          2. One level deep: children of children (nested Category/DocFolder/)
        Uses _parse_folder_name() to extract doc-ID from folder names.
        """
        doc_id_upper = doc_id.upper()

        # Level 1: direct children
        try:
            for d in self.native_root.iterdir():
                if not d.is_dir():
                    continue
                if d.name.lower() in ('ss', 'superseded', 'superceded'):
                    continue
                folder_doc_id, _ = _parse_folder_name(d.name)
                if folder_doc_id and folder_doc_id.upper() == doc_id_upper:
                    return d
        except (OSError, PermissionError):
            pass

        # Level 2: nested one level deep (Category/DocFolder/)
        try:
            for category_dir in self.native_root.iterdir():
                if not category_dir.is_dir():
                    continue
                if category_dir.name.lower() in ('ss', 'superseded', 'superceded'):
                    continue
                # Skip if this looks like a doc folder itself (already checked above)
                for d in category_dir.iterdir():
                    if not d.is_dir():
                        continue
                    if d.name.lower() in ('ss', 'superseded', 'superceded'):
                        continue
                    folder_doc_id, _ = _parse_folder_name(d.name)
                    if folder_doc_id and folder_doc_id.upper() == doc_id_upper:
                        return d
        except (OSError, PermissionError):
            pass

        return None

    def _scan_native_dwgs(self, folder: Path, doc_id: str) -> Optional[Dict]:
        """Scan a Native folder for DWGs and identify the latest IFR revision.

        Returns {doc_id, folder, latest_ifr_dwg, latest_ifr_rev,
                 existing_ifc_rev, description, all_dwgs}
        or None if no valid IFR DWG found.
        """
        ifr_dwgs = []   # (path, revision_letter)
        ifc_dwgs = []   # (path, revision_number)
        other_dwgs = []  # paths

        try:
            for f in folder.iterdir():
                if not f.is_file() or f.suffix.lower() != '.dwg':
                    continue
                if f.name.startswith('~$'):
                    continue
                rev_type, revision = _classify_dwg(f.name)
                if rev_type == 'IFR':
                    ifr_dwgs.append((f, revision))
                elif rev_type == 'IFC':
                    try:
                        ifc_dwgs.append((f, int(revision)))
                    except ValueError:
                        ifc_dwgs.append((f, 0))
                else:
                    other_dwgs.append(f)
        except (OSError, PermissionError) as e:
            self.logger.warning(f"扫描 {folder} 失败: {e}")
            return None

        if not ifr_dwgs and not other_dwgs:
            return None

        if ifr_dwgs:
            # Find latest IFR DWG (highest letter revision)
            ifr_dwgs.sort(key=lambda x: x[1], reverse=True)
            latest_ifr, latest_rev = ifr_dwgs[0]
        else:
            # No IFR-classified DWGs — use most recently modified OTHER DWG as source
            # (handles DWGs without explicit revision suffix like "GG31-C-PLN-004-Road Pavement.dwg")
            other_dwgs_sorted = sorted(other_dwgs, key=lambda f: f.stat().st_mtime, reverse=True)
            latest_ifr = other_dwgs_sorted[0]
            latest_rev = ''
            other_dwgs = other_dwgs_sorted[1:]  # remaining others may be cleaned up

        # Check existing IFC revision from IFC output folder
        existing_ifc_rev = self._get_existing_ifc_rev(doc_id)

        # Extract description from folder name
        _, description = _parse_folder_name(folder.name)

        return {
            'doc_id': doc_id,
            'folder': folder,
            'latest_ifr_dwg': latest_ifr,
            'latest_ifr_rev': latest_rev,
            'existing_ifc_rev': existing_ifc_rev,
            'needs_ifc': True,
            'description': description,
            'all_ifr_dwgs': ifr_dwgs,
            'all_ifc_dwgs': ifc_dwgs,
            'other_dwgs': other_dwgs,
        }

    def _get_existing_ifc_rev(self, doc_id: str) -> Optional[int]:
        """Check IFC output folder for existing IFC files of a doc-ID.

        Files in 4. IFC(Client)/ are IFC by definition — treat any numeric
        revision (e.g., _Rev0) as IFC regardless of _IFC suffix.
        """
        if not self.ifc_output.exists():
            return None
        max_rev = None
        for f in self.ifc_output.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() not in ('.dwg', '.pdf'):
                continue
            if f.name.startswith('~$'):
                continue
            f_doc_id = _extract_doc_id_standalone(f.name)
            if f_doc_id and f_doc_id.upper() == doc_id.upper():
                rev_type, revision = _classify_dwg(f.name)
                if rev_type == 'IFC':
                    try:
                        rev_num = int(revision)
                        if max_rev is None or rev_num > max_rev:
                            max_rev = rev_num
                    except ValueError:
                        if max_rev is None:
                            max_rev = 0
                elif rev_type == 'IFR':
                    # In ifc_output folder, numeric revisions are IFC even
                    # without _IFC suffix (e.g., _Rev0.pdf = IFC Rev0)
                    try:
                        rev_num = int(revision)
                        if max_rev is None or rev_num > max_rev:
                            max_rev = rev_num
                    except ValueError:
                        pass  # letter revision in IFC folder — ignore
        return max_rev

    # ── Cleanup ──────────────────────────────────────────────────────────

    def _cleanup_native_folder(self, dwg_info: Dict) -> List[str]:
        """Move outdated DWGs in a Native folder to SS/.

        Keeps: latest IFR DWG + existing IFC DWGs.
        Moves: all other IFR DWGs + other DWGs to SS/.
        Returns list of moved filenames.
        """
        keep_paths = {dwg_info['latest_ifr_dwg']}
        # Also keep IFC DWGs
        for ifc_path, _ in dwg_info.get('all_ifc_dwgs', []):
            keep_paths.add(ifc_path)

        to_move = []
        for ifr_path, _ in dwg_info.get('all_ifr_dwgs', []):
            if ifr_path not in keep_paths:
                to_move.append(ifr_path)
        for other_path in dwg_info.get('other_dwgs', []):
            if other_path not in keep_paths:
                to_move.append(other_path)

        if not to_move:
            return []

        ss_folder = dwg_info['folder'] / 'SS'
        moved = []

        if self.dry_run:
            for f in to_move:
                print(f"    [预览] 移动旧版: {f.name} → SS/")
                moved.append(f.name)
            return moved

        try:
            to_long_path(ss_folder).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            self.logger.error(f"无法创建 SS 文件夹: {e}")
            return []

        for f in to_move:
            try:
                dest = ss_folder / f.name
                if dest.exists():
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = ss_folder / f"{f.stem}_{ts}{f.suffix}"
                shutil.move(str(to_long_path(f)), str(to_long_path(dest)))
                moved.append(f.name)
                self.logger.info(f"旧版归档: {f.name} → SS/")
            except Exception as e:
                self.logger.error(f"移动失败: {f.name} - {e}")

        return moved

    # ── Incremental Check ────────────────────────────────────────────────

    def _check_incremental(self, dwg_info: Dict) -> bool:
        """Check if IFC conversion can be skipped.

        Returns True if should SKIP.
        Only skip if a PDF with matching doc-ID exists in 4. IFC(Client)/.
        NO mtime comparison — unreliable with Dropbox/cloud sync.
        User can force re-conversion via force_doc_ids or bot 强制重转 button.
        """
        doc_id = dwg_info['doc_id']

        if self.ifc_output.exists():
            for f in self.ifc_output.iterdir():
                if not f.is_file():
                    continue
                if f.suffix.lower() != '.pdf':
                    continue
                if f.name.startswith('~$'):
                    continue
                f_doc_id = _extract_doc_id_standalone(f.name)
                if f_doc_id and f_doc_id.upper() == doc_id.upper():
                    return True

        return False

    # ── IFC Filename Normalization ────────────────────────────────────────

    def normalize_ifc_filenames(self) -> List[Tuple[str, str]]:
        """Rename files in 4. IFC(Client)/ to ensure _IFC suffix before extension.

        E.g., GG31-E-PLN-001_..._Rev0.pdf → GG31-E-PLN-001_..._Rev0_IFC.pdf
        Skips files that already have _IFC suffix.
        Returns list of (old_name, new_name) tuples.
        """
        renamed = []
        if not self.ifc_output.exists():
            return renamed

        _RE_NEEDS_IFC = re.compile(
            r'^(.+?_Rev\d+)(?!_IFC)(\.\w+)$', re.IGNORECASE)

        for f in self.ifc_output.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() not in ('.pdf', '.dwg'):
                continue
            if f.name.startswith('~$'):
                continue
            # Check if already has _IFC suffix
            stem = f.stem
            if stem.upper().endswith('_IFC'):
                continue
            m = _RE_NEEDS_IFC.match(f.name)
            if not m:
                continue
            new_name = f"{m.group(1)}_IFC{m.group(2)}"
            if new_name == f.name:
                continue
            if self.dry_run:
                renamed.append((f.name, new_name))
            else:
                try:
                    new_path = f.parent / new_name
                    f.rename(new_path)
                    renamed.append((f.name, new_name))
                    self.logger.info(f"IFC重命名: {f.name} → {new_name}")
                except Exception as e:
                    self.logger.error(f"重命名失败: {f.name} - {e}")
        return renamed

    # ── Batch Orchestrator ───────────────────────────────────────────────

    def batch_convert_approved(self, force_doc_ids: Optional[set] = None,
                                update_deliverables: bool = True) -> Dict:
        """Main entry point: scan approved → archive → locate → cleanup → convert.

        Args:
            force_doc_ids: If provided, skip incremental check for these doc-IDs.
            update_deliverables: If True, auto-update deliverable Excel after conversion.

        Returns result dict with all details for bot display.
        """
        result = {
            'approved_files': [],
            'archived_count': 0,
            'native_mapped': [],
            'native_not_found': [],
            'converted': [],
            'skipped': [],
            'cleaned_up': {},
            'errors': [],
            'renamed_ifc': [],
        }

        # Step 0: normalize IFC filenames (add _IFC suffix where missing)
        renamed = self.normalize_ifc_filenames()
        result['renamed_ifc'] = renamed
        if renamed:
            print(f"\n  [Step 0] Normalize IFC filenames (_IFC suffix): {len(renamed)} files")

        if force_doc_ids is None:
            force_doc_ids = set()

        # 1. Scan approved files
        print(f"\n  [Step 1] 扫描 Approved 文件...")
        approved = self.scan_approved_files()
        result['approved_files'] = approved
        print(f"    发现 {len(approved)} 个 Approved 文件")

        if not approved:
            print("    没有找到 Approved 文件")
            return result

        # 2. Archive non-archived files
        print(f"\n  [Step 2] 归档 Approved 文件到 {self.APPROVED_SUBFOLDER}/...")
        needs_archive = [a for a in approved if a['needs_archive']]
        if needs_archive:
            result['archived_count'] = self.archive_approved_files(approved)
            print(f"    归档 {result['archived_count']} 个文件")
        else:
            print(f"    所有文件已在归档目录中")

        # 3. Map doc-IDs to Native folders
        print(f"\n  [Step 3] 定位 Native DWG 文件夹...")
        unique_doc_ids = {a['doc_id'] for a in approved}
        for doc_id in sorted(unique_doc_ids):
            native_folder = self._find_native_folder(doc_id)
            if not native_folder:
                result['native_not_found'].append(doc_id)
                print(f"    {doc_id}: 未找到 Native 文件夹 (可能是报告/附表)")
                continue

            dwg_info = self._scan_native_dwgs(native_folder, doc_id)
            if not dwg_info:
                result['native_not_found'].append(doc_id)
                print(f"    {doc_id}: 文件夹 {native_folder.name} 中无有效 IFR DWG")
                continue

            result['native_mapped'].append(dwg_info)
            print(f"    {doc_id}: {dwg_info['latest_ifr_dwg'].name} "
                  f"(Rev{dwg_info['latest_ifr_rev']}, "
                  f"IFC Rev{dwg_info['existing_ifc_rev'] if dwg_info['existing_ifc_rev'] is not None else 'N/A'})")

        if not result['native_mapped']:
            print("\n  没有找到可转换的 Native DWG")
            return result

        # 4. Process each mapped DWG
        print(f"\n  [Step 4] IFC 转换 ({len(result['native_mapped'])} 个文件)...")

        # Create IFCManager for actual conversion (composition)
        # preserve_ifr=True: keep IFR revision history in title block
        ifc_mgr = IFCManager(
            self.project_path,
            dry_run=self.dry_run,
            title_block_name=self.title_block_name,
            native_root=self.native_root,
            ifc_output=self.ifc_output,
            preserve_ifr=True,
        )

        conversion_results = []
        for idx, dwg_info in enumerate(result['native_mapped'], 1):
            doc_id = dwg_info['doc_id']

            # Incremental check
            if doc_id not in force_doc_ids and self._check_incremental(dwg_info):
                result['skipped'].append(doc_id)
                print(f"    [{idx}/{len(result['native_mapped'])}] {doc_id}: "
                      f"跳过(IFC PDF 已存在且为最新)")
                continue

            # Cleanup old DWGs
            cleaned = self._cleanup_native_folder(dwg_info)
            if cleaned:
                result['cleaned_up'][doc_id] = cleaned
                if not self.dry_run:
                    print(f"    [{idx}/{len(result['native_mapped'])}] {doc_id}: "
                          f"清理 {len(cleaned)} 个旧版本")

            # Convert
            print(f"    [{idx}/{len(result['native_mapped'])}] {doc_id}: "
                  f"IFR Rev{dwg_info['latest_ifr_rev']} → "
                  f"IFC Rev{(dwg_info['existing_ifc_rev'] if dwg_info['existing_ifc_rev'] is not None else -1) + 1}...")

            conv_result = ifc_mgr.convert_to_ifc(dwg_info)
            conversion_results.append(conv_result)
            result['converted'].append(conv_result)

            # Recovery: if conversion failed, reset AutoCAD state for next file
            if not conv_result['success']:
                try:
                    acad = ifc_mgr._get_acad()
                    while acad.Documents.Count > 0:
                        acad.Documents.Item(0).Close(False)
                        time.sleep(1)
                except Exception:
                    # Force reconnect on next call
                    ifc_mgr._acad = None
                time.sleep(3)

            if conv_result['success']:
                print(f"      ✓ DWG: {Path(conv_result['dwg_path']).name if conv_result['dwg_path'] else 'N/A'}")
                if conv_result['pdf_path']:
                    print(f"      ✓ PDF: {Path(conv_result['pdf_path']).name}")
                if conv_result['errors']:
                    print(f"      ⚠ 警告: {'; '.join(conv_result['errors'])}")
            else:
                result['errors'].append(f"{doc_id}: {'; '.join(conv_result['errors'])}")
                print(f"      ✗ 失败: {'; '.join(conv_result['errors'])}")

        # 4b. Clean up old IFC revisions in IFC(Client) + sync to Client Sharepoint
        if conversion_results and any(r['success'] for r in conversion_results):
            print(f"\n  [Step 4b] 清理 IFC(Client) 旧版本...")
            try:
                tsmt_mgr = IFCTransmittalManager(self.project_path, dry_run=self.dry_run)
                grouped = tsmt_mgr.scan_ifc_files()
                duplicates = tsmt_mgr.identify_duplicates(grouped)
                moved = tsmt_mgr.move_old_to_ss(duplicates)
                result['ifc_version_cleanup'] = {'moved': moved, 'duplicates': len(duplicates)}
                if moved > 0:
                    for src, dest in duplicates:
                        print(f"      [->SS] {src.name}")
                    print(f"    已移动 {moved} 个旧版本到 SS/")
                else:
                    print(f"    无旧版本需要清理")
                # Sync IFC PDFs to Client Sharepoint/2.IFC + version cleanup
                tsmt_mgr._sync_ifc_files_to_sharepoint()
                sp_ifc_dir = self.project_path / "Design/Engineering/13. Client Sharepoint/2.IFC"
                if sp_ifc_dir.exists():
                    vm = VersionManager(str(self.project_path.parent), dry_run=self.dry_run)
                    vm_stats = vm.process_directory(sp_ifc_dir, show_details=False)
                    if vm_stats.get("moved", 0) > 0:
                        print(f"    Client Sharepoint/2.IFC: 已移动 {vm_stats['moved']} 个旧版本到 SS/")
            except Exception as e:
                self.logger.error(f"IFC版本清理失败: {e}")
                result['errors'].append(f"IFC版本清理失败: {e}")

        # 5. Update deliverables
        if update_deliverables and conversion_results:
            successful = [r for r in conversion_results if r['success']]
            if successful:
                print(f"\n  [Step 5] 更新交付物 Excel...")
                try:
                    deliv_result = ifc_mgr.update_ifc_deliverables(successful)
                    result['deliverable_update'] = deliv_result
                except Exception as e:
                    self.logger.error(f"交付物更新失败: {e}")
                    result['errors'].append(f"交付物更新失败: {e}")

        # Summary
        ok = sum(1 for r in result['converted'] if r['success'])
        fail = sum(1 for r in result['converted'] if not r['success'])
        skip = len(result['skipped'])
        not_found = len(result['native_not_found'])
        print(f"\n  完成: 成功={ok}, 失败={fail}, 跳过={skip}, 无 Native={not_found}")

        return result


# =============================================================================
# Issue Register Manager (NEW) — Three-Angle Responder Assignment
# =============================================================================
#
# Cross-project workflow for Design Review Comments Register / RFI Register
# Responder assignment. Applies three-angle review:
#   1. Role (position description) — fallback default
#   2. Email allocation (doc-prefix → team mapping, per-project overrides)
#   3. Title block ground truth (DRN/DES/CHK from latest Rev IFC PDF)
#
# Generates `{original}_updated.xlsx` with:
#   - Col M (Responder) auto-filled
#   - Light blue fill on auto-filled cells
#   - Yellow fill on cells where existing responder disagrees with allocation

@dataclass
class ResponderAssignment:
    """Assignment decision for a single register row."""
    row: int
    doc_id: str
    doc_desc: str
    comment_severity: str
    existing_responder: str
    dim1_role: str            # Role-based default (fallback)
    dim2_email: str           # Email allocation (per-project)
    dim3_titleblock: str      # Title-block ground truth (DES preferred over DRN)
    final_responder: str
    dim3_field: str           # Which field was used: DES/DRN/CHK
    conflict: bool            # True if final_responder ≠ existing_responder (non-empty)
    note: str                 # annotation like 'CP (AH PR)', 'RV (HOLD)', etc.


@dataclass
class IssueRegisterResult:
    """Result of an issue-register responder run."""
    project_name: str
    register_path: str
    output_path: str
    total_rows: int = 0
    open_rows: int = 0
    auto_filled: int = 0
    conflicts: int = 0
    skipped_closed: int = 0
    skipped_already_filled: int = 0
    titleblock_hits: int = 0
    titleblock_misses: int = 0
    assignments: List[ResponderAssignment] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


class IssueRegisterManager:
    """Assign responders to Issue Register / Design Review Comments / RFI Register rows
    using three-angle review (role + email allocation + title block).

    **Fully generic across all EPC projects. NO team/responder data is hardcoded
    in this module.** All project-specific allocation lives on D: drive only.
    Dropbox is never read or written for allocation config (company-visible).

    Allocation lookup (per-project, on D: drive):
        D:\\3.Career\\obsidian-vault\\04-Work-SOP\\Projects\\{code}-*\\Team-Allocation.md

    The md file contains a fenced ```json``` block with the runtime
    allocation config (single source of truth — human-readable notes and
    machine-readable config live in the same file to avoid sync drift).

    If the D: drive md (or its json fence) is absent, responder assignment
    falls back to title-block extraction (dimension 3) only. When both are
    absent, the cell is left blank for manual fill-in.
    """

    # Register file search paths (relative to project) — Dropbox, read-only register lookup
    REGISTER_SEARCH_PATHS = [
        "Design/Engineering/1. Drawings/3. IFR(Client)",
    ]

    # Team allocation lookup root — D: drive private vault ONLY.
    # Dropbox is company-visible and MUST NOT contain Team-Allocation.md.
    ALLOCATION_LOOKUP_ROOT = Path(r"D:\3.Career\obsidian-vault\04-Work-SOP\Projects")

    # IFC folder for title-block ground truth
    IFC_FOLDER = "Design/Engineering/1. Drawings/4. IFC(Client)"

    # Register filename patterns (lowercase match)
    REGISTER_PATTERNS = [
        "design review comments register",
        "rfi",
        "issue register",
    ]

    # Excel column map (1-based; standard client register layout)
    COL_STATUS = 1      # A — Open/Closed
    COL_DOC_ID = 2      # B — Doc Number
    COL_DESC = 3        # C — Description (sometimes)
    COL_COMMENT = 9     # I — Comment text
    COL_SEVERITY = 10   # J — Severity A/B/C/D
    COL_RESPONDER = 13  # M — Responder (this is what we fill)
    COL_CLOSEOUT_R = 18 # R — Closeout
    COL_CLOSEOUT_W = 23 # W — Closeout (post-workshop)

    # Visual markup
    FILL_AUTOFILL = PatternFill(start_color='FFADD8E6', end_color='FFADD8E6',
                                 fill_type='solid') if OPENPYXL_AVAILABLE else None  # light blue
    FILL_CONFLICT = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00',
                                 fill_type='solid') if OPENPYXL_AVAILABLE else None  # yellow

    # Team / responder data is NEVER hardcoded here.
    # All project-specific mappings load at runtime from the D: drive config.
    # These empty dicts exist only so code paths that iterate them stay valid
    # when no D: config is found.
    DEFAULT_ALLOCATION: Dict = {}
    SPECIAL_DOC_RULES: Dict = {}

    # Known management/approver stamps that don't count as ground-truth modifier.
    # This IS generic (cross-project company-wide), so it stays as code constant.
    GENERIC_STAMPS = {'ACE', 'AW', ''}  # e.g. AW = manager approval stamp

    def __init__(self, project_path: Path, dry_run: bool = False):
        self.project_path = project_path
        self.dry_run = dry_run
        self._project_allocation: Optional[Dict] = None

    # -------------------------------------------------------------------------
    # Register location
    # -------------------------------------------------------------------------

    def find_register(self) -> Optional[Path]:
        """Find the most recent Design Review Comments Register / RFI Excel.

        Preference: `Design Review Comments Register_Rev*` > `RFI_*`.
        Within matches, latest mtime wins.
        """
        candidates = []
        for rel_path in self.REGISTER_SEARCH_PATHS:
            folder = self.project_path / rel_path
            if not folder.exists():
                continue
            for f in folder.iterdir():
                if not f.is_file():
                    continue
                if f.suffix.lower() not in ('.xlsx', '.xlsm'):
                    continue
                if f.name.startswith('~$') or '_updated' in f.stem.lower():
                    continue
                name_low = f.name.lower()
                priority = None
                if 'design review comments register' in name_low:
                    priority = 1
                elif 'issue register' in name_low:
                    priority = 2
                elif 'rfi' in name_low:
                    priority = 3
                if priority is not None:
                    candidates.append((priority, -f.stat().st_mtime, f))
        if not candidates:
            return None
        candidates.sort()
        return candidates[0][2]

    # -------------------------------------------------------------------------
    # Project allocation override (Team-Allocation.md, json fence block)
    # -------------------------------------------------------------------------

    def _load_project_allocation(self, project_code: str = '') -> Dict:
        """Load per-project allocation from Team-Allocation.md on D: drive.

        Single source of truth: the project's ``Team-Allocation.md`` contains
        one fenced ```json``` code block with schema:

            {
                "allocation":    {"<prefix>": {"responder": "XX", "note": "..."}, ...},
                "special_rules": {"<doc-id>": {"responder": "XX", "note": "..."}, ...},
                "notes":         {"<responder>": "<annotation e.g. HOLD>"}
            }

        Lookup: ``ALLOCATION_LOOKUP_ROOT / '{project_code}-*' / 'Team-Allocation.md'``

        Dropbox project folders are NEVER read for allocation (company-visible).
        If the D: drive md (or its json fence) is absent, returns an empty
        allocation — the assignment then relies on title-block extraction
        (dimension 3) only.
        """
        if self._project_allocation is not None:
            return self._project_allocation

        merged = {'allocation': {}, 'special_rules': {}, 'notes': {}}

        md_path = self._find_d_drive_allocation(project_code)
        if md_path is None:
            if project_code:
                print(f"[IssueRegister] No D: drive Team-Allocation.md found for project "
                      f"'{project_code}' under {self.ALLOCATION_LOOKUP_ROOT}; "
                      f"falling back to title-block only.")
            self._project_allocation = merged
            return merged

        try:
            with open(md_path, 'r', encoding='utf-8') as f:
                md_text = f.read()
            proj = self._extract_allocation_json_from_md(md_text)
            if proj is None:
                print(f"[IssueRegister] WARN no json fence with allocation/special_rules "
                      f"in {md_path}; falling back to title-block only.")
            else:
                for k, v in (proj.get('allocation') or {}).items():
                    merged['allocation'][k.upper()] = v
                for k, v in (proj.get('special_rules') or {}).items():
                    merged['special_rules'][k.upper()] = v
                merged['notes'].update(proj.get('notes') or {})
                print(f"[IssueRegister] Loaded allocation from {md_path}")
        except Exception as e:
            print(f"[IssueRegister] WARN Team-Allocation.md load failed: {e}")

        self._project_allocation = merged
        return merged

    def _find_d_drive_allocation(self, project_code: str) -> Optional[Path]:
        """Locate ``{code}-*/Team-Allocation.md`` under ALLOCATION_LOOKUP_ROOT.

        Returns None if the root is missing or no matching project dir/file exists.
        """
        if not project_code:
            return None
        root = self.ALLOCATION_LOOKUP_ROOT
        if not root.exists():
            return None
        code_upper = project_code.upper()
        for sub in root.iterdir():
            if not sub.is_dir():
                continue
            name_upper = sub.name.upper()
            if (name_upper == code_upper
                    or name_upper.startswith(code_upper + '-')
                    or name_upper.startswith(code_upper + '_')):
                md_path = sub / "Team-Allocation.md"
                if md_path.exists():
                    return md_path
        return None

    @staticmethod
    def _extract_allocation_json_from_md(md_text: str) -> Optional[Dict]:
        """Scan md_text for fenced ```json``` code blocks.

        Returns the first parsed dict containing BOTH 'allocation' and
        'special_rules' keys (identifies the runtime-config block vs any
        unrelated json fence). Returns None if no matching block found.
        """
        pattern = re.compile(r"```json\s*\n(.*?)\n```", re.DOTALL | re.IGNORECASE)
        for match in pattern.finditer(md_text):
            raw = match.group(1)
            try:
                obj = json.loads(raw)
            except (json.JSONDecodeError, ValueError):
                continue
            if (isinstance(obj, dict)
                    and 'allocation' in obj
                    and 'special_rules' in obj):
                return obj
        return None

    @staticmethod
    def _detect_project_code(register_path: Path) -> str:
        """Extract project code from the register filename.

        Examples:
            '50023-Design Review Comments Register_Rev 11.0.xlsx' → '50023'
            'GG31-Design Review Comments Register_Rev 5.xlsx'    → 'GG31'
        """
        if not register_path:
            return ''
        m = re.match(r'^(\d{5}|GG\d{1,3})[\-_\s]', register_path.name, re.IGNORECASE)
        if m:
            return m.group(1).upper()
        return ''

    # -------------------------------------------------------------------------
    # Title-block ground truth (dimension 3)
    # -------------------------------------------------------------------------

    def build_titleblock_cache(self, wanted_doc_ids: Set[str]) -> Dict[str, Dict]:
        """For each doc-id, find the latest-rev IFC PDF and extract title-block fields.

        Returns: {doc_id: {'DRN': str, 'DES': str, 'CHK': str, 'APP': str,
                            'rev': str, 'pdf': str}}
        Missing doc-ids are simply absent from the dict.
        """
        cache: Dict[str, Dict] = {}
        ifc_folder = self.project_path / self.IFC_FOLDER
        if not ifc_folder.exists():
            return cache

        if not PDFPLUMBER_AVAILABLE:
            print("[IssueRegister] pdfplumber unavailable — dimension 3 skipped")
            return cache

        # Group PDFs by doc-id, keep the one with highest rev
        by_doc: Dict[str, List[Tuple[int, Path]]] = {}
        for pdf in ifc_folder.glob("*.pdf"):
            doc_id = _extract_doc_id_standalone(pdf.name)
            if not doc_id or doc_id not in wanted_doc_ids:
                continue
            rev = self._extract_rev_number(pdf.stem)
            by_doc.setdefault(doc_id, []).append((rev, pdf))

        for doc_id, entries in by_doc.items():
            entries.sort(reverse=True)  # highest rev first
            _, pdf = entries[0]
            try:
                fields = self._extract_titleblock_fields(pdf)
                if fields:
                    fields['pdf'] = pdf.name
                    cache[doc_id] = fields
            except Exception as e:
                print(f"[IssueRegister] TB extract fail {doc_id}: {e}")
        return cache

    @staticmethod
    def _extract_rev_number(stem: str) -> int:
        """Parse Rev<N> from filename. Returns -1 if not found."""
        m = re.search(r'[Rr]ev\.?\s*(\d+)', stem)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                pass
        # Letter rev → convert A=1, B=2, ...
        m = re.search(r'[Rr]ev\.?\s*([A-Z])', stem)
        if m:
            return ord(m.group(1).upper()) - ord('A') + 1
        return 0

    @staticmethod
    def _extract_titleblock_fields(pdf_path: Path) -> Optional[Dict[str, str]]:
        """Open the PDF, locate the revision history table on the last page (where
        AutoCAD title blocks usually live), and extract DRN/DES/CHK/APP from
        the latest (bottom-most) populated revision row.

        Uses pdfplumber table extraction. Works with AutoCAD PDFs that have
        table structure; falls back to text-based heuristic otherwise.
        """
        if not PDFPLUMBER_AVAILABLE:
            return None

        try:
            with pdfplumber.open(str(pdf_path)) as pdf:
                # Title block is usually on the last page of a multi-sheet drawing
                pages = list(pdf.pages)
                if not pages:
                    return None
                # Try last page first, then first
                for page in (pages[-1], pages[0] if len(pages) > 1 else None):
                    if page is None:
                        continue
                    fields = IssueRegisterManager._scan_page_for_tb(page)
                    if fields:
                        return fields
        except Exception:
            return None
        return None

    @staticmethod
    def _scan_page_for_tb(page) -> Optional[Dict[str, str]]:
        """Look for a table on the page with REV/DATE/DESCRIPTION and
        DRN/DES/CHK/APP column headers; return the latest non-empty row."""
        header_keys = {'DRN', 'DES', 'CHK', 'APP'}
        tables = []
        try:
            tables = page.extract_tables() or []
        except Exception:
            return None

        for tbl in tables:
            if not tbl or len(tbl) < 2:
                continue
            # Normalize cells
            norm = [[(c or '').strip().upper() for c in row] for row in tbl]
            # Find header row that contains at least 2 of our keys
            header_row_idx = None
            for i, row in enumerate(norm):
                hits = sum(1 for cell in row if any(k == cell or cell.startswith(k + ' ') for k in header_keys))
                if hits >= 2:
                    header_row_idx = i
                    break
            if header_row_idx is None:
                continue

            header = norm[header_row_idx]
            col_map = {}
            for ci, cell in enumerate(header):
                for k in ('DRN', 'DES', 'CHK', 'APP'):
                    if cell == k or cell.startswith(k + ' ') or cell.endswith(' ' + k):
                        col_map.setdefault(k, ci)

            if not col_map:
                continue

            # Find the latest populated data row (scan upward from the bottom)
            rev_col = None
            for ci, cell in enumerate(header):
                if cell in ('REV', 'REVISION') or cell.startswith('REV '):
                    rev_col = ci
                    break

            for row in reversed(norm[header_row_idx + 1:]):
                if rev_col is not None:
                    rev_cell = row[rev_col] if rev_col < len(row) else ''
                    if not rev_cell:
                        continue
                # Harvest fields
                fields: Dict[str, str] = {}
                for k, ci in col_map.items():
                    if ci < len(row) and row[ci]:
                        fields[k] = row[ci].strip()
                if fields.get('DES') or fields.get('DRN') or fields.get('CHK'):
                    if rev_col is not None and rev_col < len(row):
                        fields['rev'] = row[rev_col]
                    return fields
        return None

    # -------------------------------------------------------------------------
    # Three-angle assignment
    # -------------------------------------------------------------------------

    def _prefix_of(self, doc_id: str) -> str:
        """Extract the doc-type code from a doc-id. Examples:
            50023-EA-301 → EA
            GG31-E-SLD-001 → SLD
            Tatua-EL-001 → EL
        """
        if not doc_id:
            return ''
        # Prefer the 2-letter code right before the 3-digit number
        m = re.search(r'([A-Z]{2,3})-\d{3}', doc_id.upper())
        if m:
            return m.group(1)
        # Fallback: everything after first hyphen, take first letter token
        parts = doc_id.upper().split('-')
        for p in parts[1:]:
            if p.isalpha() and 1 <= len(p) <= 3:
                return p
        return ''

    def _dim1_role_default(self, doc_id: str, allocation: Dict) -> str:
        """Dimension 1 — prefix-level role baseline from project allocation.

        Returns the ``allocation[<prefix>].responder`` value (ignoring special_rules).
        Since team/responder data is no longer hardcoded, the role baseline now
        comes from the same D: drive JSON as dimension 2, just without the
        doc-id-level special_rules overlay.
        Returns '' if the allocation config is empty or no prefix match.
        """
        prefix = self._prefix_of(doc_id)
        info = allocation.get('allocation', {}).get(prefix)
        if info:
            return info.get('responder', '')
        return ''

    def _dim2_email_allocation(self, doc_id: str, allocation: Dict) -> Tuple[str, str]:
        """Responder per project email allocation. Returns (responder, note)."""
        # Special rule first (exact doc-id tail match)
        for rule_key, rule in allocation['special_rules'].items():
            # Rule key like 'EA-300' should match '50023-EA-300' or '50023-EA-300-XX'
            if rule_key in doc_id.upper():
                return rule['responder'], rule.get('note', '')
        prefix = self._prefix_of(doc_id)
        if prefix in allocation['allocation']:
            info = allocation['allocation'][prefix]
            return info['responder'], info.get('note', '')
        return '', ''

    def _dim3_titleblock(self, doc_id: str, tb_cache: Dict) -> Tuple[str, str]:
        """Return (responder, field_used) from title-block ground truth.
        Prefers DES > CHK > DRN. Filters generic stamps (ACE, AW, empty)."""
        fields = tb_cache.get(doc_id)
        if not fields:
            return '', ''
        for key in ('DES', 'CHK', 'DRN'):
            val = fields.get(key, '').strip().upper()
            if val and val not in self.GENERIC_STAMPS:
                # Strip any trailing digits/dots (e.g. "RV." → "RV")
                val = re.sub(r'[^\w].*$', '', val).strip()
                if val:
                    return val, key
        return '', ''

    def _resolve_responder(self, doc_id: str, allocation: Dict,
                            tb_cache: Dict) -> Tuple[str, str, str, str, str]:
        """Apply three-angle logic. Returns
        (final_responder, dim1, dim2, dim3, dim3_field)."""
        dim1 = self._dim1_role_default(doc_id, allocation)
        dim2, note = self._dim2_email_allocation(doc_id, allocation)
        dim3, dim3_field = self._dim3_titleblock(doc_id, tb_cache)

        # Conflict resolution:
        # - dim3 wins when present and differs
        # - special_rules (encoded in dim2) beat plain dim1
        if dim3:
            # If dim3 matches either dim1 or dim2 → confirmed, no conflict
            # If dim3 differs → dim3 takes precedence
            final = dim3
        elif dim2:
            final = dim2
        else:
            final = dim1 or ''

        return final, dim1, dim2, dim3, dim3_field

    # -------------------------------------------------------------------------
    # Main run
    # -------------------------------------------------------------------------

    def run(self) -> IssueRegisterResult:
        """Locate register → build TB cache → assign responders → write _updated.xlsx."""
        project_name = self.project_path.name
        result = IssueRegisterResult(
            project_name=project_name,
            register_path='',
            output_path='',
        )

        if not OPENPYXL_AVAILABLE:
            result.errors.append("openpyxl not available")
            return result

        register = self.find_register()
        if not register:
            result.errors.append("未找到 Design Review / RFI Register Excel")
            return result
        result.register_path = str(register)

        project_code = self._detect_project_code(register)
        allocation = self._load_project_allocation(project_code)

        # Pass 1: read rows, collect wanted doc-ids
        wb = openpyxl.load_workbook(str(register))
        ws = self._pick_master_sheet(wb)
        if ws is None:
            result.errors.append("未找到 Master Register 工作表")
            return result

        row_data: List[Tuple[int, str, str, str, str, str]] = []
        # tuple: (row, status, doc_id, existing_resp, closeout_r, closeout_w)
        wanted_doc_ids: Set[str] = set()

        # Auto-detect header row (scan rows 1-10 for 'Doc Number' or 'FILE NO')
        header_row = self._detect_header_row(ws)
        data_start = header_row + 1

        for r in range(data_start, ws.max_row + 1):
            status = self._cell_text(ws.cell(r, self.COL_STATUS))
            doc_id = self._cell_text(ws.cell(r, self.COL_DOC_ID))
            existing = self._cell_text(ws.cell(r, self.COL_RESPONDER))
            close_r = self._cell_text(ws.cell(r, self.COL_CLOSEOUT_R))
            close_w = self._cell_text(ws.cell(r, self.COL_CLOSEOUT_W))
            if not doc_id and not status and not existing:
                continue
            row_data.append((r, status, doc_id, existing, close_r, close_w))
            if doc_id:
                wanted_doc_ids.add(doc_id.upper())

        result.total_rows = len(row_data)

        # Pass 2: build title-block cache
        tb_cache = self.build_titleblock_cache(wanted_doc_ids)
        result.titleblock_hits = len(tb_cache)
        result.titleblock_misses = len(wanted_doc_ids) - len(tb_cache)

        # Pass 3: assign
        for (r, status, doc_id, existing, close_r, close_w) in row_data:
            if status.strip().lower() != 'open':
                result.skipped_closed += 1
                continue
            result.open_rows += 1
            if existing and close_w:
                # Already handled — skip
                result.skipped_already_filled += 1
                continue

            comment_sev = self._cell_text(ws.cell(r, self.COL_SEVERITY))
            doc_desc = self._cell_text(ws.cell(r, self.COL_DESC))

            final, dim1, dim2, dim3, dim3_field = self._resolve_responder(
                doc_id.upper(), allocation, tb_cache)

            _, note = self._dim2_email_allocation(doc_id.upper(), allocation)
            # Attach per-responder notes from project config
            extra = allocation['notes'].get(final, '')
            combined_note = note or extra

            final_with_note = final
            if combined_note:
                final_with_note = f"{final} ({combined_note})"

            conflict = bool(existing and existing.strip().upper() != final.strip().upper())

            assignment = ResponderAssignment(
                row=r,
                doc_id=doc_id,
                doc_desc=doc_desc,
                comment_severity=comment_sev,
                existing_responder=existing,
                dim1_role=dim1,
                dim2_email=dim2,
                dim3_titleblock=dim3,
                final_responder=final,
                dim3_field=dim3_field,
                conflict=conflict,
                note=combined_note,
            )
            result.assignments.append(assignment)

            # Count what would be (or is) changed — works for both dry-run and live
            if not existing:
                result.auto_filled += 1
                if not self.dry_run:
                    cell = ws.cell(r, self.COL_RESPONDER)
                    cell.value = final_with_note
                    cell.fill = self.FILL_AUTOFILL
            elif conflict:
                result.conflicts += 1
                if not self.dry_run:
                    cell = ws.cell(r, self.COL_RESPONDER)
                    cell.fill = self.FILL_CONFLICT

        # Save as _updated.xlsx
        if not self.dry_run:
            out_path = register.with_name(register.stem + "_updated" + register.suffix)
            try:
                wb.save(str(out_path))
                result.output_path = str(out_path)
            except Exception as e:
                result.errors.append(f"保存失败: {e}")
        else:
            result.output_path = f"(dry-run) {register.stem}_updated{register.suffix}"

        return result

    # -------------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------------

    @staticmethod
    def _pick_master_sheet(wb):
        """Pick the 'Master Register' sheet if present, else the first sheet."""
        for name in wb.sheetnames:
            if 'master' in name.lower() and 'register' in name.lower():
                return wb[name]
        return wb[wb.sheetnames[0]] if wb.sheetnames else None

    @staticmethod
    def _detect_header_row(ws) -> int:
        """Find the row containing 'Doc Number' / 'FILE NO' / 'Status' headers.
        Returns the row index (1-based). Defaults to 1."""
        for r in range(1, min(ws.max_row + 1, 15)):
            texts = []
            for c in range(1, min(ws.max_column + 1, 30)):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    texts.append(v.strip().lower())
            joined = ' | '.join(texts)
            if ('doc number' in joined or 'file no' in joined or 'document number' in joined) \
                    and 'status' in joined:
                return r
        return 1

    @staticmethod
    def _cell_text(cell) -> str:
        v = cell.value
        if v is None:
            return ''
        return str(v).strip()


# =============================================================================
# AS BUILT Manager (IFC → AS BUILT conversion)
# =============================================================================

# Regex to match IFC revision subfolders in 1. Native/ doc-ID folders
# Matches: "Rev.2 IFC", "Rev 2 - IFC", "Rev.1 - IFC", "Rev D - IFC", "IFC"
_RE_IFC_SUBFOLDER = re.compile(
    r'^Rev\.?\s*(\d+|[A-Z])\s*[-–]?\s*IFC$|^IFC$',
    re.IGNORECASE
)

# Regex to extract REV N from AS BUILT PDF filename
_RE_AB_REV_IN_FILENAME = re.compile(r'\bREV\s*(\d+)\b', re.IGNORECASE)


class AsBuiltManager(IFCManager):
    """Convert IFC DWGs to AS BUILT: stamp, title block update, PDF export.

    Workflow per doc-ID:
      1. Scan 1. Native/ for latest IFC DWG(s)
      2. Open DWG, remove old IFC stamp, add AS BUILT stamp
      3. Update title block: AS BUILT description, new revision row
      4. SaveAs new AB DWG to Rev.{N} - AB/ subfolder
      5. Export PDF to 5. As Built/ folder
      6. For multi-page doc-IDs, merge all pages into single PDF

    Inherits IFCManager (which inherits IFCStampMixin) — reuses COM helpers,
    title block finding, stamp drawing, PDF export infrastructure.
    """

    STAMP_TEXT = "{\\fArial Narrow|b1;AS BUILT}"

    # Stamp box style — user requirement: RED, THICK border, BOTH boxes uniform
    # (matches gold standard BLD-003/SLD-001). cw=2.0 matches the original pre-bot
    # stamp thickness; color=1 is AutoCAD red. Both AS BUILT and COLOUR boxes use
    # these same values via _stamp_via_com_draw (_cw + _stamp_color), guaranteeing
    # identical weight and color.
    _STAMP_CW = 2.0
    _STAMP_COLOR = 1  # AutoCAD red

    # AS BUILT drawings always carry the COLOUR stamp (matches gold standard
    # BLD-003/SLD-001). The source COLOUR stamp is often an inserted BLOCK that
    # _scan_has_colour (MText-only) misses; after geometry cleanup removes it,
    # force a fresh aligned COLOUR box rather than skipping on false overlap.
    _FORCE_COLOUR = True

    LMS_TITLE_BLOCKS = {"Coleamablly", "Riverina_tellhow"}

    PERSONNEL_TAGS = ['DESIGNED', 'DRAWN', 'CHECK', 'APPROVED',
                      'ENGINEER', 'QA', 'PROJECT']


    _NATIVE_CANDIDATES = [
        "Design/Engineering/1. Drawings/1. Native",
        "1. Drawings/1. Native",
    ]
    _AB_OUTPUT_CANDIDATES = [
        "Design/Engineering/1. Drawings/5. As Built/3. As Built Client",
        "Design/Engineering/1. Drawings/5. As Built",
        "1. Drawings/6.AS Built",
    ]

    # Extra roots scanned for AS BUILT IN ADDITION to native_root. Some native
    # drawing DWGs (e.g. the Civil/Structure foundation plans) do NOT live in
    # `1. Native/` — they sit inside their per-report doc-ID folder under
    # `2. Calcs & Reports/Reports/...`. For these the doc-ID + description come
    # from the DWG FILENAME, not the folder name: the folder is named by the
    # REPORT doc-ID but the DWG by the DRAWING doc-ID and they differ
    # (e.g. folder GG31-C-RPT-001 holds drawing GG31-C-PLN-006). Each entry is a
    # project-relative path; missing paths are silently skipped (cross-project).
    # Both folder layouts: `Design/Engineering/` prefix (Warnertown/LMS/Cole2) and
    # flat `1. Drawings/`-style projects (Tatua) — mirrors _NATIVE_CANDIDATES.
    # Non-existent paths are silently skipped, so listing both is safe and makes
    # the rule universal for any future project created via /new_project (which,
    # per SSOT mainv3.md rule 2b, files foundation/structural Civil drawings into
    # Reports/Civil & Structure rather than 1. Native).
    _EXTRA_DRAWING_SOURCES = [
        "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
        "Design/Engineering/2. Calcs & Reports/Reports/Electrical",
        "2. Calcs & Reports/Reports/Civil & Structure",
        "2. Calcs & Reports/Reports/Electrical",
    ]

    _SAME_DIR_AB_OUTPUTS = {
        "Design/Engineering/1. Drawings/5. As Built",
    }

    def __init__(self, project_path, dry_run=False, title_block_name=None,
                 native_root=None, ab_output=None):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.title_block_name = title_block_name or self.DEFAULT_TITLE_BLOCK
        self.preserve_ifr = True
        self._acad = None
        self._dm = DeliverableManager(self.project_path, dry_run=dry_run)
        self._native_root_override = Path(native_root) if native_root else None
        self._ab_output_override = Path(ab_output) if ab_output else None
        self._detected_native = self._NATIVE_CANDIDATES[0]
        self._detected_ab_output = self._AB_OUTPUT_CANDIDATES[0]
        if not native_root:
            for cand in self._NATIVE_CANDIDATES:
                if (self.project_path / cand).exists():
                    self._detected_native = cand
                    break
        if not ab_output:
            for cand in self._AB_OUTPUT_CANDIDATES:
                if (self.project_path / cand).exists():
                    self._detected_ab_output = cand
                    break
            if self._detected_ab_output in self._SAME_DIR_AB_OUTPUTS:
                ab_path = self.project_path / self._detected_ab_output
                for child in sorted(ab_path.iterdir()):
                    if child.is_dir() and child.name.lower().startswith('3.'):
                        self._detected_ab_output = (
                            f"{self._detected_ab_output}/{child.name}")
                        break
        self._save_in_source_dir = (
            self._detected_ab_output in self._SAME_DIR_AB_OUTPUTS)

    @property
    def native_root(self) -> Path:
        if self._native_root_override:
            return self._native_root_override
        return self.project_path / self._detected_native

    @property
    def ab_output(self) -> Path:
        if self._ab_output_override:
            return self._ab_output_override
        return self.project_path / self._detected_ab_output

    @property
    def ifc_output(self) -> Path:
        return self.ab_output

    # ── Scanning ─────────────────────────────────────────────────────────

    def scan_native_for_ab(self) -> List[Dict]:
        """Scan 1. Native/ for doc-ID folders with IFC DWGs ready for AS BUILT.

        Returns list of dicts:
          {doc_id, folder, ifc_source: {ifc_rev, dwg_paths, is_multi_page, source_dir},
           existing_ab_rev, description}
        """
        results = []
        native = self.native_root
        if not native.exists():
            print(f"  Native 目录不存在: {native}")
            return results

        existing_ab = self._scan_existing_ab()

        for item in sorted(native.iterdir()):
            if not item.is_dir():
                continue
            if item.name.lower() in ('ss', 'superseded', 'superceded',
                                      'lms template', 'not_named_drawings',
                                      'site plan_from development',
                                      'delivered to chint', 'new_drawings_not_named',
                                      'template', 'templates'):
                continue

            doc_id, description = _parse_folder_name(item.name)
            if not doc_id:
                continue
            if not re.search(r'-\d{2,3}', doc_id):
                continue

            ifc_source = self._find_latest_ifc_source(item)
            if ifc_source is None:
                continue

            ab_rev = existing_ab.get(doc_id)
            results.append({
                'doc_id': doc_id,
                'folder': item,
                'ifc_source': ifc_source,
                'existing_ab_rev': ab_rev,
                'description': description,
            })

        # Drawings that live inside report folders (Calcs & Reports) — not in
        # 1. Native/. Appended AFTER the native scan; deduped against doc-IDs the
        # native scan already produced so a drawing present in both is converted
        # once (native wins).
        seen = {r['doc_id'].upper() for r in results}
        results.extend(self._scan_report_drawings_for_ab(existing_ab, seen))

        return results

    def _scan_report_drawings_for_ab(self, existing_ab: Dict[str, int],
                                     seen_doc_ids: set) -> List[Dict]:
        """Scan `_EXTRA_DRAWING_SOURCES` report folders for native drawing DWGs.

        Unlike `1. Native/`, the folder here is named by the REPORT doc-ID while
        the drawing DWG inside carries the DRAWING doc-ID (e.g. folder
        GG31-C-RPT-001 holds GG31-C-PLN-006_..._Rev0_IFC.dwg). So doc-ID and
        description are parsed from the DWG FILENAME, not the folder. One report
        folder may hold drawings for >1 doc-ID → emitted as separate entries
        (grouped by drawing doc-ID; same doc-ID across sheets stays one entry).
        Output still routes through the normal subfolder-mode path: the AB DWG
        lands in a `Rev.N - AB/` subfolder of the report folder, the PDF in the
        shared AB output dir.
        """
        results: List[Dict] = []
        for rel in self._EXTRA_DRAWING_SOURCES:
            root = self.project_path / rel
            if not root.exists():
                continue
            for item in sorted(root.iterdir()):
                if not item.is_dir():
                    continue
                if item.name.lower() in ('ss', 'superseded', 'superceded',
                                          'appendix', 'template', 'templates'):
                    continue

                ifc_source = self._find_latest_ifc_source(item)
                if ifc_source is None:
                    continue

                # Group the chosen-rev DWGs by the DRAWING doc-ID (the folder's
                # doc-ID is the report's, which differs from the drawing's).
                groups: Dict[str, List[Path]] = {}
                names: Dict[str, str] = {}
                for p in ifc_source['dwg_paths']:
                    did = _extract_doc_id_standalone(p.name)
                    if not did:
                        print(f"  ⚠ 报告图纸 doc-ID 无法识别，跳过: {p.name}")
                        continue
                    key = did.upper()
                    groups.setdefault(key, []).append(p)
                    names.setdefault(key, did)

                for key, paths in groups.items():
                    doc_id = names[key]
                    if key in seen_doc_ids:
                        continue  # already covered by 1. Native scan
                    _, description = _parse_folder_name(paths[0].stem)
                    sub_source = dict(ifc_source)
                    sub_source['dwg_paths'] = sorted(paths, key=lambda p: p.name)
                    sub_source['is_multi_page'] = len(paths) > 1
                    results.append({
                        'doc_id': doc_id,
                        'folder': item,
                        'ifc_source': sub_source,
                        'existing_ab_rev': existing_ab.get(doc_id),
                        'description': description,
                    })
                    seen_doc_ids.add(key)

        return results

    def _find_latest_ifc_source(self, doc_folder: Path) -> Optional[Dict]:
        """Find latest IFC DWG(s) in a doc-ID folder.

        Handles:
          1. Flat: IFC DWGs directly in folder
          2. Nested: IFC DWGs in subfolders (Rev.N IFC, Rev N - IFC, etc.)

        Returns {ifc_rev, dwg_paths, is_multi_page, source_dir} or None.
        """
        # --- Flat IFC DWGs ---
        # Strict check: require 'IFC' explicitly in filename (not just numeric rev).
        # _classify_dwg treats ANY numeric rev as IFC, but in LMS many DWGs have
        # numeric revisions without being IFC files (e.g. "Rev 5 - POWER STATION...").
        flat_ifc = {}  # rev_num -> [paths]
        try:
            # os.scandir (dirent-based is_file) instead of Path.iterdir()+is_file():
            # Path.is_file() does a full os.stat() on the entry's FULL path, which
            # FAILS (FileNotFoundError → is_file()==False) when that path exceeds
            # Windows MAX_PATH (260). These report-folder DWGs sit at 260-290 chars
            # (deep Dropbox tree + long RPT folder + long description), so
            # iterdir()+is_file() silently HID them from the scan. scandir reads the
            # directory entry (dirent) without re-stat'ing the long path, so is_file
            # is correct regardless of length. (The long path is also why AutoCAD
            # COM Open fails — handled by _shortpath_open_target.) Cross-project: any
            # native DWG on a >260-char path used to be invisible to AS BUILT/IFC.
            for _entry in os.scandir(doc_folder):
                if not _entry.is_file():
                    continue
                f = Path(_entry.path)
                if f.suffix.lower() != '.dwg':
                    continue
                if f.name.startswith('~$'):
                    continue
                stem_upper = f.stem.upper()
                if 'IFC' not in stem_upper and '_AB' not in stem_upper:
                    continue
                if '_AB' in stem_upper:
                    continue
                if 'ASBUILT' in stem_upper or 'AS_BUILT' in stem_upper or 'AS BUILT' in stem_upper:
                    continue
                rev_type, revision = _classify_dwg(f.name)
                if rev_type == 'IFC':
                    try:
                        rev_num = int(revision)
                    except ValueError:
                        rev_num = 0
                    flat_ifc.setdefault(rev_num, []).append(f)
                elif '_IFC' in stem_upper:
                    if revision and revision[0].isalpha():
                        rev_num = ord(revision[0].upper()) - ord('A')
                    else:
                        rev_num = 0
                    flat_ifc.setdefault(rev_num, []).append(f)
        except (OSError, PermissionError):
            pass

        # --- Nested IFC subfolders ---
        nested_ifc = []  # (subfolder, rev_num, [dwg_paths])
        try:
            for d in doc_folder.iterdir():
                if not d.is_dir():
                    continue
                if d.name.lower() in ('ss', 'superseded', 'superceded'):
                    continue
                m = _RE_IFC_SUBFOLDER.match(d.name)
                if not m:
                    continue
                rev_str = m.group(1) if m.lastindex and m.group(1) else None
                if rev_str is None:
                    rev_num = 0
                elif rev_str.isdigit():
                    rev_num = int(rev_str)
                else:
                    rev_num = ord(rev_str.upper()) - ord('A')

                dwgs = []
                for _e in os.scandir(d):  # dirent-based — see flat-scan note re: MAX_PATH
                    if (_e.is_file() and _e.name.lower().endswith('.dwg')
                            and not _e.name.startswith('~$')):
                        dwgs.append(Path(_e.path))
                if dwgs:
                    nested_ifc.append((d, rev_num, sorted(dwgs, key=lambda p: p.name)))
        except (OSError, PermissionError):
            pass

        # Pick highest revision
        best_flat_rev = max(flat_ifc.keys(), default=-1)
        best_nested = max(nested_ifc, key=lambda x: x[1], default=None)
        best_nested_rev = best_nested[1] if best_nested else -1

        if best_flat_rev < 0 and best_nested_rev < 0:
            return None

        if best_flat_rev >= best_nested_rev and best_flat_rev >= 0:
            dwgs = sorted(flat_ifc[best_flat_rev], key=lambda p: p.name)
            return {
                'ifc_rev': best_flat_rev,
                'dwg_paths': dwgs,
                'is_multi_page': len(dwgs) > 1,
                'source_dir': doc_folder,
            }
        elif best_nested:
            return {
                'ifc_rev': best_nested[1],
                'dwg_paths': best_nested[2],
                'is_multi_page': len(best_nested[2]) > 1,
                'source_dir': best_nested[0],
            }

        return None

    def _scan_existing_ab(self) -> Dict[str, int]:
        """Scan 5. As Built/ for existing AS BUILT PDFs. Return {doc_id: max_rev}."""
        ab_revs = {}
        ab_dir = self.ab_output
        if not ab_dir.exists():
            return ab_revs
        for f in ab_dir.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() != '.pdf':
                continue
            stem_upper = f.stem.upper()
            if ('_AS BUILT' not in stem_upper and '_AS_BUILT' not in stem_upper
                    and '_ASBUILT' not in stem_upper):
                continue
            doc_id = _extract_doc_id_standalone(f.name)
            if not doc_id:
                continue
            m = _RE_AB_REV_IN_FILENAME.search(f.stem)
            if m:
                rev_num = int(m.group(1))
                if doc_id not in ab_revs or rev_num > ab_revs[doc_id]:
                    ab_revs[doc_id] = rev_num
        return ab_revs

    # ── AS BUILT native house-keeping ────────────────────────────────────

    # An AS BUILT subfolder inside a doc-ID folder, e.g. 'Rev.1 - AB', 'Rev1 AB',
    # 'Rev.A - AB'. The designated home for the converted AB DWG (subfolder mode).
    _RE_AB_SUBFOLDER = re.compile(r'^Rev\.?\s*(\d+|[A-Z])\s*[-–]?\s*AB$',
                                  re.IGNORECASE)

    @staticmethod
    def _is_ab_named(stem: str) -> bool:
        """True if a filename stem looks like an AS BUILT export (not IFC/IFR).

        Matches '..._AsBuilt', '..._AS BUILT', '..._AS_BUILT', or trailing '_AB'.
        Used to spot loose AB files that escaped the designated Rev.N - AB/ home.
        """
        u = stem.upper()
        return ('ASBUILT' in u or 'AS BUILT' in u or 'AS_BUILT' in u
                or u.endswith('_AB'))

    @staticmethod
    def _ab_rev_key(rev_str: Optional[str]) -> int:
        """Sort key for an AB rev token ('1','A',...). Lower = earlier/first."""
        if not rev_str:
            return 0
        rev_str = rev_str.strip()
        if rev_str.isdigit():
            return int(rev_str)
        return ord(rev_str.upper()) - ord('A')

    def _ab_superseded_dir(self, doc_folder: Path) -> Path:
        """Existing Superseded/SS folder inside a doc-ID folder, or default path."""
        try:
            for item in doc_folder.iterdir():
                if item.is_dir() and item.name.lower() in (
                        'superseded', 'superceded', 'ss'):
                    return item
        except (OSError, PermissionError):
            pass
        return doc_folder / 'Superseded'

    def _move_to_ss(self, src: Path, ss_dir: Path) -> bool:
        """Move a stray file/dir into ss_dir (reversible). Dropbox/long-path safe."""
        try:
            to_long_path(ss_dir).mkdir(parents=True, exist_ok=True)
            dest = ss_dir / src.name
            if to_long_path(dest).exists():
                stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                dest = ss_dir / f"{src.stem}_{stamp}{src.suffix}"
            shutil.move(str(to_long_path(src)), str(to_long_path(dest)))
            return True
        except Exception as e:
            logging.warning(f"[AB-cleanup] 移动失败 {src.name}: {e}")
            return False

    def cleanup_ab_native(self, report_only: bool = False) -> List[Dict]:
        """House-keep AS BUILT artefacts inside 1. Native/ doc-ID folders.

        Cross-project rule: a doc-ID keeps ONE live AS BUILT version — the first
        rev in its designated 'Rev.N - AB/' subfolder — and the doc-ID folder
        carries no loose AB exports. This collects the junk left over while the
        script was being refined:
          1. MULTIPLE 'Rev.N - AB' subfolders → keep the lowest rev, move the
             higher ones to the doc-ID folder's Superseded/.
          2. Loose AS BUILT files ('..._AsBuilt.dwg', '..._AsBuilt.dwg.dxf',
             '..._AB.dwg', etc.) sitting directly in the doc-ID folder (outside
             any 'Rev.N - AB/') → move to Superseded/.

        Deliberately conservative w.r.t. the user's "there may be special cases"
        note: a LONE AB subfolder is kept as-is whatever its rev (a special-case
        'For Construction' rev continuation is therefore never clobbered); only
        DUPLICATES are collapsed. Everything is MOVED (reversible), never deleted.
        Pure filesystem work — no AutoCAD/COM. Returns a list of action dicts.
        """
        actions: List[Dict] = []
        native = self.native_root
        if not native.exists():
            return actions

        for doc_folder in sorted(native.iterdir()):
            if not doc_folder.is_dir():
                continue
            if doc_folder.name.lower() in ('ss', 'superseded', 'superceded'):
                continue
            doc_id, _ = _parse_folder_name(doc_folder.name)
            if not doc_id or not re.search(r'-\d{2,3}', doc_id):
                continue

            try:
                children = list(doc_folder.iterdir())
            except (OSError, PermissionError):
                continue

            ss_dir = self._ab_superseded_dir(doc_folder)

            # 1. Collapse multiple 'Rev.N - AB' subfolders → keep the lowest rev.
            ab_subs = []  # (rev_key, path)
            for child in children:
                if not child.is_dir():
                    continue
                m = self._RE_AB_SUBFOLDER.match(child.name)
                if m:
                    ab_subs.append((self._ab_rev_key(m.group(1)), child))
            if len(ab_subs) > 1:
                ab_subs.sort(key=lambda t: t[0])
                keep = ab_subs[0][1]
                for _key, extra in ab_subs[1:]:
                    moved = True if report_only else self._move_to_ss(extra, ss_dir)
                    actions.append({'doc_id': doc_id, 'kind': 'extra_ab_rev',
                                    'path': str(extra), 'kept': keep.name,
                                    'moved': moved})

            # 2. Move loose AS BUILT files (outside any Rev.N - AB/ subfolder).
            for child in children:
                if not child.is_file() or child.name.startswith('~$'):
                    continue
                if not self._is_ab_named(child.stem):
                    continue
                moved = True if report_only else self._move_to_ss(child, ss_dir)
                actions.append({'doc_id': doc_id, 'kind': 'stray_ab_file',
                                'path': str(child), 'moved': moved})

        return actions

    def _ab_existing_pdf_qa_clean(self, dwg_info: Dict) -> bool:
        """Incremental-skip gate: True iff an existing AS BUILT PDF for this
        doc-ID is present AND passes QA.

        This is the 'skip only if QA-clean' policy: a clean prior export is NOT
        re-converted, but a missing OR QA-faulty PDF returns False → the doc-ID
        is re-exported and then re-QA'd. 'Export first, then QA' therefore still
        holds for everything that actually needs producing.
        """
        doc_id = dwg_info['doc_id']
        ab_dir = self.ab_output
        if not ab_dir.exists():
            return False
        target = None
        try:
            for f in ab_dir.iterdir():
                if not f.is_file() or f.suffix.lower() != '.pdf':
                    continue
                fid = _extract_doc_id_standalone(f.name)
                if fid and fid.upper() == doc_id.upper():
                    su = f.stem.upper()
                    if ('_AS BUILT' in su or '_AS_BUILT' in su
                            or '_ASBUILT' in su):
                        target = f
                        break
        except (OSError, PermissionError):
            return False
        if target is None:
            return False
        # QA the existing PDF. _qa_validate_ab_pdf returns [] when fitz is
        # unavailable; in that case we cannot verify, so this degrades to the
        # plain presence check (old incremental behaviour) rather than blocking.
        warnings = self._qa_validate_ab_pdf(target, doc_id, expected_pages=None)
        return not warnings

    # ── Filename builders ────────────────────────────────────────────────

    def _build_ab_pdf_filename(self, doc_id: str, description: str, ab_rev: int) -> str:
        """Build AS BUILT PDF filename: {doc_id} REV {N} {desc}_AS BUILT"""
        desc = re.sub(r'[<>:"/\\|?*()&]', '', description).strip()
        desc = re.sub(r'\s+', ' ', desc)
        return f"{doc_id} REV {ab_rev} {desc}_AS BUILT"

    def _build_ab_dwg_name(self, source_name: str, ab_rev: int) -> str:
        """Build AB DWG filename from IFC source: replace _IFC→_AB, update rev number."""
        stem = Path(source_name).stem
        ext = Path(source_name).suffix
        new_stem = re.sub(r'_IFC$', '_AB', stem, flags=re.IGNORECASE)
        if new_stem == stem:
            new_stem = stem + '_AB'
        new_stem = re.sub(r'([Rr]ev\.?\s*)(\d+)',
                         lambda m: m.group(1) + str(ab_rev), new_stem)
        return new_stem + ext

    # ── Title block update ───────────────────────────────────────────────

    def _update_title_block(self, attrs: Dict, ab_rev: int, personnel: Dict, date_str: str):
        """Update title block attributes for AS BUILT conversion.

        Keep all existing IFR + IFC revision rows, add AS BUILT row after them.
        Idempotent: if AS BUILT row already exists, overwrite in-place.
        """
        if 'REVISION' in attrs:
            self._safe_set_text(attrs['REVISION'], str(ab_rev))

        all_suffixes = ['REV', 'DATE', 'DESCRIPTION'] + self.PERSONNEL_TAGS

        # Detect the ACTUAL number of revision rows in THIS title block — do NOT
        # trust self.REV_ROWS (a class default of 6). The Coleambally `Coleamablly`
        # frame has only 4 rows; with the 6-default, a full 4-row TB computed
        # target_row=5, the '5 > 6' full-check stayed False, and the write targeted
        # nonexistent '5REV'/'5DESIGNED' tags → AS BUILT row SILENTLY not written
        # (root cause of "四行版本占满后 title block 没更新"). Scan a generous range
        # so we adapt whether the frame has 4, 6, or more rows.
        tb_rows = max((n for n in range(1, 13) if f"{n}REV" in attrs),
                      default=self.REV_ROWS)

        last_occupied_row = 0
        existing_ab_row = 0
        for row_num in range(1, tb_rows + 1):
            rev_tag = f"{row_num}REV"
            desc_tag = f"{row_num}DESCRIPTION"
            if rev_tag not in attrs:
                continue
            val = self._safe_get_text(attrs[rev_tag]).strip()
            if not val:
                continue
            desc_val = ''
            if desc_tag in attrs:
                desc_val = self._safe_get_text(attrs[desc_tag]).strip().upper()
            if 'AS BUILT' in desc_val or 'AS-BUILT' in desc_val:
                existing_ab_row = max(existing_ab_row, row_num)
            else:
                last_occupied_row = max(last_occupied_row, row_num)

        if existing_ab_row > 0:
            target_row = existing_ab_row
        else:
            target_row = last_occupied_row + 1

        if target_row > tb_rows:
            # Revision rows FULL → rolling history (bottom-to-top = old-to-new):
            # drop the oldest (row 1), shift rows 2..N DOWN into 1..N-1, then write
            # AS BUILT into the top row (tb_rows, newest). Preserves chronology;
            # only the oldest revision rolls off. (User-confirmed behavior.)
            print(f"    版本行已满({tb_rows}行) → 滚动:丢弃最旧Rev行1,"
                  f"行2..{tb_rows}整体下移,AS BUILT写入最上行{tb_rows}")
            for dest in range(1, tb_rows):
                src = dest + 1
                for suffix in all_suffixes:
                    d_tag = f"{dest}{suffix}"
                    s_tag = f"{src}{suffix}"
                    if d_tag in attrs:
                        sval = (self._safe_get_text(attrs[s_tag]).strip()
                                if s_tag in attrs else '')
                        self._safe_set_text(attrs[d_tag], sval)
            target_row = tb_rows

        # Clean up duplicate AB rows above target_row (idempotent safety)
        for row_num in range(1, target_row):
            desc_tag = f"{row_num}DESCRIPTION"
            if desc_tag in attrs:
                desc_val = self._safe_get_text(attrs[desc_tag]).strip().upper()
                if 'AS BUILT' in desc_val or 'AS-BUILT' in desc_val:
                    for suffix in all_suffixes:
                        tag = f"{row_num}{suffix}"
                        if tag in attrs:
                            self._safe_set_text(attrs[tag], '')

        # Write the AS BUILT row, tracking write outcomes for cheap QA (no COM
        # re-read — _safe_set_text returns success/failure). A non-empty value
        # that fails to write = a blank field in the output → QA warning.
        warnings = []
        tag_prefix = str(target_row)
        if f'{tag_prefix}REV' in attrs:
            if not self._safe_set_text(attrs[f'{tag_prefix}REV'], str(ab_rev)):
                warnings.append("AS BUILT 行 REV 写入失败")
        if f'{tag_prefix}DESCRIPTION' in attrs:
            if not self._safe_set_text(attrs[f'{tag_prefix}DESCRIPTION'], 'AS BUILT'):
                warnings.append("AS BUILT 行 DESCRIPTION 写入失败")
        if f'{tag_prefix}DATE' in attrs:
            if not self._safe_set_text(attrs[f'{tag_prefix}DATE'], date_str):
                warnings.append("AS BUILT 行 DATE 写入失败")

        _personnel_written = 0
        for tag in self.PERSONNEL_TAGS:
            full_tag = f"{tag_prefix}{tag}"
            if full_tag in attrs:
                val = personnel.get(tag.lower(), '')
                ok = self._safe_set_text(attrs[full_tag], val)
                if val and not ok:
                    warnings.append(f"AS BUILT 行 {tag} 写入失败 (应为 '{val}')")
                if val and ok:
                    _personnel_written += 1
        # All-empty personnel on the AS BUILT row = likely a personnel-read failure
        # (the source had none, or _read_latest_ifr_row backfill regressed).
        if _personnel_written == 0:
            warnings.append("AS BUILT 行人员字段全空 (检查源版本行/回退)")

        # Column alignment: some title-block definitions place the AS BUILT row's
        # attribute slots at a slightly different X than the rows above (Warnertown
        # PLN-010: row-5 PROJECT at X=286.6 vs rows 1-4 at X=281.8 → "GG31" sat
        # crooked, shifted right). Copy each written attribute's horizontal
        # position from the row directly below so every revision-table column
        # stays vertically straight. No-op for already-aligned columns.
        ref_row = target_row - 1
        if ref_row >= 1:
            for suffix in all_suffixes:
                t_tag = f"{target_row}{suffix}"
                r_tag = f"{ref_row}{suffix}"
                if t_tag in attrs and r_tag in attrs:
                    self._align_attr_x(attrs[t_tag], attrs[r_tag])
        return warnings

    def _align_attr_x(self, target_attr, ref_attr):
        """Copy ref_attr's horizontal position (InsertionPoint.X and
        TextAlignmentPoint.X) onto target_attr, preserving target's own Y/Z.

        Keeps a revision-table column visually straight when the block definition
        placed a row's attribute slot at a different X than the rows above.
        Best-effort + cosmetic: any COM failure is swallowed so it can never break
        a conversion. Update() re-evaluates the attribute so the move renders.
        """
        import pythoncom as _pc
        def _set_x(getter, setter):
            try:
                ref = self._com_retry(getter[0])
                tgt = self._com_retry(getter[1])
                if ref is None or tgt is None:
                    return
                if abs(float(ref[0]) - float(tgt[0])) < 1e-6:
                    return  # already aligned — skip the COM write
                v = win32com.client.VARIANT(
                    _pc.VT_ARRAY | _pc.VT_R8,
                    [float(ref[0]), float(tgt[1]), float(tgt[2])])
                self._com_retry(lambda: setter(v))
            except Exception:
                pass
        _set_x((lambda: ref_attr.InsertionPoint, lambda: target_attr.InsertionPoint),
               lambda v: setattr(target_attr, 'InsertionPoint', v))
        _set_x((lambda: ref_attr.TextAlignmentPoint, lambda: target_attr.TextAlignmentPoint),
               lambda v: setattr(target_attr, 'TextAlignmentPoint', v))
        try:
            self._com_retry(lambda: target_attr.Update())
        except Exception:
            pass

    # ── Stamp removal (LMS-enhanced) ────────────────────────────────────

    # 'PRINTED IN COLO' is the spelling-agnostic catch-all: matches British
    # "...PRINTED IN COLOUR" AND American "...PRINTED IN COLOR", plural OR singular
    # "DRAWING(S)". Warnertown SLD-001's old stamp reads "DRAWING TO BE PRINTED IN
    # COLOR" (singular, American) — the exact-plural-British phrase missed it.
    _QA_STAMP_PHRASES = {'FOR CONSTRUCTION', 'ISSUED FOR REVIEW', 'FOR REVIEW',
                         'DRAWINGS TO BE PRINTED IN COLOUR', 'PRINTED IN COLO',
                         'AS BUILT', 'AS-BUILT'}

    def _remove_ifc_stamp(self, doc):
        """Remove stamps from all spaces — parent + QA + geometry-based cleanup."""
        super()._remove_ifc_stamp(doc)
        self._remove_qa_layer_stamps(doc)
        self._remove_stamps_by_geometry(doc)

    def _remove_stamps_by_geometry(self, doc):
        """Remove stamp-like entities near title block by spatial detection.

        Catches stamps on any layer (not just IFC_STAMP/QA) — prevents
        overlapping frames when old stamp was drawn with different style.
        """
        all_tbs = self._find_all_title_blocks(doc)
        if not all_tbs:
            return

        for tb_item in all_tbs:
            block_ref = tb_item[0]
            layout_name = tb_item[3] if len(tb_item) > 3 else 'Model'

            try:
                min_pt, max_pt = block_ref.GetBoundingBox()
                tb_left = float(min_pt[0])
                tb_bottom = float(min_pt[1])
                tb_right = float(max_pt[0])
                tb_top = float(max_pt[1])
                tb_w = tb_right - tb_left
                tb_h = tb_top - tb_bottom
            except Exception:
                continue

            if tb_w <= 0 or tb_h <= 0:
                continue

            zone_left = tb_left + tb_w * 0.55
            zone_right = tb_right
            zone_bottom = tb_bottom
            zone_top = tb_bottom + tb_h * 0.40

            to_delete = []

            if layout_name and layout_name.lower() != 'model':
                try:
                    for layout in doc.Layouts:
                        if layout.Name == layout_name:
                            block = layout.Block
                            self._scan_block_for_stamp_entities(
                                block, zone_left, zone_right, zone_bottom, zone_top,
                                tb_w, tb_h, to_delete)
                            break
                except Exception:
                    pass
            else:
                import pythoncom as _pythoncom
                ss_name = f"_GeoClean_{int(time.time() * 1000) % 1_000_000}"
                try:
                    ss = doc.SelectionSets.Add(ss_name)
                    pt1 = win32com.client.VARIANT(
                        _pythoncom.VT_ARRAY | _pythoncom.VT_R8,
                        [zone_left, zone_bottom, 0.0])
                    pt2 = win32com.client.VARIANT(
                        _pythoncom.VT_ARRAY | _pythoncom.VT_R8,
                        [zone_right, zone_top, 0.0])
                    ss.Select(1, pt1, pt2)
                    for i in range(ss.Count):
                        try:
                            entity = ss.Item(i)
                            self._check_stamp_entity(
                                entity, zone_left, zone_right, zone_bottom, zone_top,
                                tb_w, tb_h, to_delete)
                        except Exception:
                            continue
                    ss.Delete()
                except Exception:
                    try:
                        doc.SelectionSets.Item(ss_name).Delete()
                    except Exception:
                        pass

            for entity in reversed(to_delete):
                try:
                    entity.Delete()
                except Exception:
                    pass

            if to_delete:
                print(f"    Geometry cleanup [{layout_name}]: "
                      f"{len(to_delete)} stamp entities removed")

    def _scan_block_for_stamp_entities(self, block, zone_left, zone_right,
                                        zone_bottom, zone_top, tb_w, tb_h,
                                        to_delete):
        """Scan a block's entities for stamp-like items in the given zone."""
        try:
            count = block.Count
        except Exception:
            return
        for i in range(count):
            try:
                entity = block.Item(i)
                self._check_stamp_entity(
                    entity, zone_left, zone_right, zone_bottom, zone_top,
                    tb_w, tb_h, to_delete)
            except Exception:
                continue

    def _check_stamp_entity(self, entity, zone_left, zone_right,
                             zone_bottom, zone_top, tb_w, tb_h, to_delete):
        """Check if a single entity is a stamp that should be removed."""

        try:
            ename = entity.EntityName
        except Exception:
            return

        if ename in ('AcDbPolyline', 'AcDbLwPolyline'):
            try:
                if not entity.Closed:
                    return
                mn, mx = entity.GetBoundingBox()
                el, eb = float(mn[0]), float(mn[1])
                er, et = float(mx[0]), float(mx[1])
                if not (el >= zone_left - 5 and er <= zone_right + 5 and
                        eb >= zone_bottom - 5 and et <= zone_top + 5):
                    return
                w = er - el
                h = et - eb
                w_ratio = w / tb_w
                h_ratio = h / tb_h
                if 0.02 < w_ratio < 0.20 and 0.005 < h_ratio < 0.08:
                    to_delete.append(entity)
            except Exception:
                pass

        elif ename in ('AcDbMText', 'AcDbText'):
            try:
                mn, mx = entity.GetBoundingBox()
                el, eb = float(mn[0]), float(mn[1])
                er, et = float(mx[0]), float(mx[1])
                if not (el >= zone_left - 5 and er <= zone_right + 5 and
                        eb >= zone_bottom - 5 and et <= zone_top + 5):
                    return
                text = self._com_retry(lambda e=entity: e.TextString) or ''
                plain = self._strip_mtext_formatting(text).upper().strip()
                if any(phrase in plain for phrase in self._QA_STAMP_PHRASES):
                    to_delete.append(entity)
            except Exception:
                pass

        elif ename == 'AcDbBlockReference':
            # COLOUR / FOR CONSTRUCTION stamps may be INSERTED BLOCKS (not
            # polylines). A stamp-sized block insert in the stamp zone is an
            # old pre-bot stamp — remove it so the bot can redraw aligned.
            # The title block itself is full-frame (w_ratio ~1.0) so won't match.
            try:
                mn, mx = entity.GetBoundingBox()
                el, eb = float(mn[0]), float(mn[1])
                er, et = float(mx[0]), float(mx[1])
                if not (el >= zone_left - 5 and er <= zone_right + 5 and
                        eb >= zone_bottom - 5 and et <= zone_top + 5):
                    return
                w = er - el
                h = et - eb
                w_ratio = w / tb_w
                h_ratio = h / tb_h
                if 0.02 < w_ratio < 0.20 and 0.005 < h_ratio < 0.08:
                    to_delete.append(entity)
            except Exception:
                pass

    def _remove_qa_layer_stamps(self, doc):
        """Remove stamp entities on QA layer (LMS projects use QA, not IFC_STAMP).

        Targets: closed polylines matching stamp box dimensions + stamp MText +
        block references with stamp-like names. Scans PaperSpace and ModelSpace.
        """
        _stamp_block_kw = ('IFR', 'IFC_STAMP', 'FOR_CONSTRUCTION', 'AS_BUILT', 'ASBUILT')

        def _is_qa_stamp(entity):
            """True if a QA-layer entity is a stamp box / text / stamp block."""
            try:
                ename = entity.EntityName
            except Exception:
                return False
            if ename in ('AcDbPolyline', 'AcDbLwPolyline'):
                try:
                    if not entity.Closed:
                        return False
                    mn, mx = entity.GetBoundingBox()
                    w = float(mx[0]) - float(mn[0]); h = float(mx[1]) - float(mn[1])
                    # cw>1.5: bot-drawn thick stamps; hand-drawn (cw=0) caught by geometry pass.
                    # Size thresholds relaxed (was w>300 h>30) — they're in the TB's OWN units,
                    # which vary with TB scale: Warnertown SLD-001's TB is ~1.8x, its QA COLOUR
                    # box is only 200x60 and was missed by w>300. Layer=='QA' already strongly
                    # filters to stamps, so a smaller box on QA is still a stamp.
                    return w > 100 and h > 20 and entity.ConstantWidth > 1.5
                except Exception:
                    return False
            if ename in ('AcDbMText', 'AcDbText'):
                try:
                    plain = self._strip_mtext_formatting(
                        self._com_retry(lambda e=entity: e.TextString) or '').upper().strip()
                    return any(p in plain for p in self._QA_STAMP_PHRASES)
                except Exception:
                    return False
            if ename == 'AcDbBlockReference':
                try:
                    return (any(k in entity.Name.upper() for k in _stamp_block_kw)
                            and len(entity.GetAttributes()) < 5)
                except Exception:
                    return False
            return False

        def _collect_from_block(block_space):
            """PaperSpace layout block — small, safe to iterate directly."""
            out = []
            try:
                cnt = block_space.Count
            except Exception:
                return out
            for i in range(cnt):
                try:
                    e = block_space.Item(i)
                    if e.Layer == 'QA' and _is_qa_stamp(e):
                        out.append(e)
                except Exception:
                    continue
            return out

        # PaperSpace layouts (per-layout try/except so one COM error doesn't skip all)
        try:
            layouts = list(doc.Layouts)
        except Exception:
            layouts = []
        for layout in layouts:
            try:
                if layout.Name.lower() == 'model':
                    continue
                victims = _collect_from_block(layout.Block)
                for e in reversed(victims):
                    try:
                        e.Delete()
                    except Exception:
                        pass
                if victims:
                    print(f"    QA stamp cleanup [{layout.Name}]: {len(victims)} entities removed")
            except Exception:
                continue

        # ModelSpace — use a SelectionSet filtered to layer 'QA' (spatial index).
        # NEVER iterate ModelSpace entity-by-entity: PLN-005 has 24854 MS entities;
        # the old full iteration was O(n) COM and pushed multi-layout drawings past
        # the timeout. The layer filter returns only the few QA entities, fast.
        import pythoncom as _pc
        try:
            ssn = f"_QAms_{int(time.time()*1000) % 1_000_000}"
            ss = doc.SelectionSets.Add(ssn)
            ft = win32com.client.VARIANT(_pc.VT_ARRAY | _pc.VT_I2, [8])      # 8 = layer
            fv = win32com.client.VARIANT(_pc.VT_ARRAY | _pc.VT_VARIANT, ["QA"])
            ss.Select(5, None, None, ft, fv)   # mode 5 = whole space, filtered
            victims = []
            for i in range(ss.Count):
                try:
                    e = ss.Item(i)
                    if _is_qa_stamp(e):
                        victims.append(e)
                except Exception:
                    continue
            ss.Delete()
            for e in reversed(victims):
                try:
                    e.Delete()
                except Exception:
                    pass
            if victims:
                print(f"    QA stamp cleanup [Model]: {len(victims)} entities removed")
        except Exception:
            pass

    # ── Stamp drawing ─────────────────────────────────────────────────────
    # Delegates to IFCStampMixin._stamp_via_com_draw so the AS BUILT box
    # has identical width / position / cw as the FOR CONSTRUCTION box that
    # was drawn during IFC conversion.  self.STAMP_TEXT is already
    # "AS BUILT", so the parent method prints the correct label.
    # COLOUR box also uses the same ratios → guaranteed alignment.

    # ── Single DWG conversion ────────────────────────────────────────────

    def _shortpath_open_target(self, dwg_path: Path):
        r"""Return (open_path:str, cleanup:callable) for AutoCAD COM Open.

        AutoCAD's COM `Documents.Open` cannot open paths longer than ~256 chars
        and does NOT accept the `\\?\` extended-length prefix (it errors
        "Invalid file name"). Report-folder DWGs routinely exceed MAX_PATH
        (deep Dropbox tree + long RPT folder name + long descriptive filename →
        260-290 chars), so 5 of 6 Civil drawings failed to open.

        Fix: expose the DWG's PARENT directory through a SHORT directory
        JUNCTION (`mklink /J`, no admin needed) in TEMP and open the DWG via the
        junction. Because a junction is a transparent alias to the real folder,
        relative XREF paths still resolve — unlike copying the lone DWG to a
        temp dir, which would orphan its XREFs. `rmdir` on the junction removes
        only the link, never the target files.

        Short/ASCII paths are returned unchanged (no junction). Cross-project:
        any project with a long native path benefits.
        """
        import subprocess, tempfile
        p = str(dwg_path)
        noop = (p, (lambda: None))
        if len(p) <= 240 and p.isascii():
            return noop
        try:
            link_root = (Path(tempfile.gettempdir())
                         / f"abj_{int(time.time() * 1000) % 1_000_000}")
            r = subprocess.run(
                ['cmd', '/c', 'mklink', '/J', str(link_root), str(dwg_path.parent)],
                capture_output=True, text=True, timeout=30)
            if not link_root.exists():
                print(f"    ⚠ 短路径 junction 创建失败: {r.stderr.strip() or r.stdout.strip()}")
                return noop
        except Exception as e:
            print(f"    ⚠ 短路径 junction 异常: {e}")
            return noop

        short = str(link_root / dwg_path.name)

        def _cleanup():
            try:
                subprocess.run(['cmd', '/c', 'rmdir', str(link_root)],
                               capture_output=True, timeout=30)
            except Exception:
                pass

        print(f"    长路径({len(p)}>256) → 经短 junction 打开: {link_root.name}")
        return (short, _cleanup)

    def convert_to_ab(self, dwg_info: Dict) -> Dict:
        """Convert a single IFC DWG to AS BUILT.

        Steps:
          1. Open IFC DWG in AutoCAD
          2. Find title block, read latest revision row (IFC personnel)
          3. Calculate AB rev (1 if first, else existing + 1)
          4. Update title block with AS BUILT description
          5. Add AS BUILT stamp (remove old IFC stamp first)
          6. SaveAs new AB DWG to Rev.{N} - AB/ subfolder
          7. Export PDF to 5. As Built/

        Returns {success, dwg_path, pdf_path, ab_rev, errors}.
        """
        result = {
            'success': False,
            'doc_id': dwg_info['doc_id'],
            'dwg_path': None,
            'pdf_path': None,
            'ab_rev': None,
            'errors': [],
        }

        ifc_source = dwg_info['ifc_source']
        dwg_path = ifc_source['dwg_paths'][0]
        doc_id = dwg_info['doc_id']
        description = dwg_info['description']

        # Calculate AB revision
        if dwg_info['existing_ab_rev'] is not None:
            ab_rev = dwg_info['existing_ab_rev'] + 1
        else:
            ab_rev = 1
        result['ab_rev'] = ab_rev

        # Build output paths
        if self._save_in_source_dir:
            ab_dwg_stem = re.sub(r'\s*_?IFC$', '_AsBuilt', dwg_path.stem,
                                 flags=re.IGNORECASE)
            ab_dwg_name = ab_dwg_stem + dwg_path.suffix
            ab_dwg_path = ifc_source['source_dir'] / ab_dwg_name
            ab_pdf_stem = re.sub(r'\s*_?IFC$', '_As Built', dwg_path.stem,
                                 flags=re.IGNORECASE)
            ab_pdf_path = self.ab_output / f"{ab_pdf_stem}.pdf"
            ab_subfolder = None
        else:
            ab_subfolder = dwg_info['folder'] / f"Rev.{ab_rev} - AB"
            ab_dwg_name = self._build_ab_dwg_name(dwg_path.name, ab_rev)
            ab_dwg_path = ab_subfolder / ab_dwg_name
            ab_pdf_name = self._build_ab_pdf_filename(doc_id, description, ab_rev)
            ab_pdf_path = self.ab_output / f"{ab_pdf_name}.pdf"

        if self.dry_run:
            result['success'] = True
            result['dwg_path'] = str(ab_dwg_path)
            result['pdf_path'] = str(ab_pdf_path)
            return result

        # --- Actual AutoCAD operations ---
        doc = None
        _jcleanup = lambda: None
        try:
            acad = self._get_acad()

            # AutoCAD COM cannot open paths >~256 chars (and rejects the \\?\
            # prefix). Long source paths — common for these report-folder DWGs
            # (deep Dropbox path + long RPT folder + long description) — are
            # opened through a short directory JUNCTION so XREFs still resolve.
            open_path, _jcleanup = self._shortpath_open_target(dwg_path)

            # Handle lock files
            lock1 = dwg_path.with_suffix('.dwl')
            lock2 = dwg_path.with_suffix('.dwl2')
            _reuse_open_doc = False
            if lock1.exists() or lock2.exists():
                for _lock_attempt in range(3):
                    any_locked = False
                    for lf in (lock1, lock2):
                        if lf.exists():
                            try:
                                lf.unlink()
                            except PermissionError:
                                any_locked = True
                    if not any_locked:
                        break
                    if _lock_attempt < 2:
                        time.sleep(3)
                if lock1.exists() or lock2.exists():
                    try:
                        for _di in range(acad.Documents.Count):
                            _doc = acad.Documents.Item(_di)
                            if Path(_doc.FullName).resolve() == dwg_path.resolve():
                                _reuse_open_doc = True
                                break
                    except Exception:
                        pass

            # Open the IFC DWG
            for _doc_wait in range(10):
                try:
                    _ = acad.Documents.Count
                    break
                except Exception:
                    time.sleep(2)

            if _reuse_open_doc:
                try:
                    for _di in range(acad.Documents.Count):
                        _doc = acad.Documents.Item(_di)
                        if Path(_doc.FullName).resolve() == dwg_path.resolve():
                            doc = _doc
                            break
                    if doc is None:
                        doc = self._com_retry(
                            lambda p=open_path: acad.Documents.Open(p))
                except Exception as e:
                    result['errors'].append(f"无法获取已打开的 DWG: {e}")
                    return result
            else:
                try:
                    doc = self._com_retry(
                        lambda p=open_path: acad.Documents.Open(p))
                except Exception as e:
                    result['errors'].append(f"无法打开 DWG: {e}")
                    return result

            # Wait for document ready
            for _wait in range(20):
                try:
                    _ = doc.ModelSpace.Count
                    _ = doc.Layouts.Count
                    break
                except Exception:
                    time.sleep(1)
            time.sleep(2)

            # Find ALL title blocks
            all_tbs = []
            for _tb_attempt in range(3):
                all_tbs = self._find_all_title_blocks(doc)
                if all_tbs:
                    break
                if _tb_attempt < 2:
                    print(f"    title block 搜索失败 (尝试 {_tb_attempt+1}/3)，等待后重试...")
                    time.sleep(3 * (_tb_attempt + 1))
            if not all_tbs:
                result['errors'].append(f"未找到 title block（尝试: {self.title_block_name}）")
                try:
                    doc.Close(False)
                except Exception:
                    pass
                return result
            print(f"    Found {len(all_tbs)} title block(s)")

            # Read personnel from latest revision row, fall back to root TB attrs
            personnel = {}
            _root_attrs = all_tbs[0][1] if all_tbs else {}
            for _tb_item in all_tbs:
                _attrs = _tb_item[1]
                personnel = self._read_latest_ifr_row(_attrs)
                if personnel:
                    _root_attrs = _attrs
                    break
            # If revision-row personnel fields are empty, read from root-level
            # title block attributes (e.g. DESIGNED, DRAWN — not row-prefixed).
            # Pre-bot IFC DWGs often leave revision row personnel empty.
            if not any(v for v in personnel.values() if isinstance(v, str) and v):
                for tag in self.PERSONNEL_TAGS:
                    if tag in _root_attrs and not personnel.get(tag.lower()):
                        val = self._safe_get_text(_root_attrs[tag]).strip()
                        if val:
                            personnel[tag.lower()] = val

            # Fix known typos
            self._fix_known_typos(doc)

            # Scan for existing COLOUR stamp BEFORE removal
            has_colour = self._scan_has_colour(doc)
            if has_colour:
                print(f"    COLOUR: 原DWG已有 COLOUR 印章，保留原样，仅画 AS BUILT")

            # Remove existing stamps ONCE (covers all spaces — removes IFC stamps too)
            self._remove_ifc_stamp(doc)

            # Update EVERY title block + add AS BUILT stamp near each one
            date_str = datetime.now().strftime('%d/%m/%y')
            _tb_qa_all = []   # title-block QA warnings, collected per sheet from writes
            for tb_idx, tb_item in enumerate(all_tbs, 1):
                block_ref = tb_item[0]
                attrs = tb_item[1]
                space = tb_item[2]
                layout_name = tb_item[3] if len(tb_item) > 3 else None
                print(f"    Sheet {tb_idx}/{len(all_tbs)}: 更新 title block + AS BUILT...")
                _tbw = self._update_title_block(attrs, ab_rev, personnel, date_str)
                for _w in (_tbw or []):
                    _tb_qa_all.append(f"Sheet {tb_idx}: {_w}")
                if block_ref and space:
                    self._stamp_via_com_draw(doc, block_ref, space,
                                             has_colour=has_colour,
                                             layout_name=layout_name)
                else:
                    print(f"    ⚠ Sheet {tb_idx}: block_ref 或 space 为空，跳过印章")

            # Title-block QA (same closed loop as stamps): warnings were collected
            # per sheet from the WRITE RESULTS during _update_title_block — cheap,
            # NO COM re-read (re-reading all sheets' attrs timed out 9-page files).
            if _tb_qa_all:
                for _w in _tb_qa_all:
                    print(f"    [TB-QA] {_w}")
                result['tb_qa_warnings'] = _tb_qa_all

            # Create output directories
            if ab_subfolder:
                to_long_path(ab_subfolder).mkdir(parents=True, exist_ok=True)
            to_long_path(self.ab_output).mkdir(parents=True, exist_ok=True)

            if not self._save_in_source_dir:
                # Bind XREFs before SaveAs — subfolder mode changes relative
                # path base. Without binding, PUBLISH reopens DWG and XREFs
                # fail to resolve → empty/missing PDF.
                xref_bound = 0
                try:
                    for bi in range(doc.Blocks.Count):
                        try:
                            blk = doc.Blocks.Item(bi)
                            if blk.IsXRef:
                                blk.Bind(False)
                                xref_bound += 1
                        except Exception:
                            pass
                    if xref_bound:
                        print(f"    XREF 绑定: {xref_bound} 个外部引用已绑定")
                        try:
                            doc.Regen(1)
                        except Exception:
                            pass
                        time.sleep(8)  # outside try — must always run
                        # Verify all XREFs resolved
                        _unresolved = 0
                        try:
                            for _bi2 in range(doc.Blocks.Count):
                                try:
                                    _blk2 = doc.Blocks.Item(_bi2)
                                    if _blk2.IsXRef:
                                        _unresolved += 1
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        if _unresolved:
                            print(f"    ⚠ XREF: {_unresolved} 个外部引用绑定后仍未解析")
                except Exception:
                    pass

            # SaveAs — use temp short path if > 240 chars
            final_dwg_path = ab_dwg_path
            use_temp = len(str(ab_dwg_path)) > 240
            if use_temp:
                import tempfile
                temp_dir = Path(tempfile.gettempdir()) / "AB_TEMP"
                temp_dir.mkdir(parents=True, exist_ok=True)
                temp_dwg = temp_dir / ab_dwg_name
                save_path = temp_dwg
                print(f"    路径过长({len(str(ab_dwg_path))}字符)，使用临时路径 SaveAs")
            else:
                save_path = ab_dwg_path

            # Clean up pre-existing file
            try:
                sp = Path(save_path)
                if sp.exists():
                    sp.unlink()
            except Exception:
                pass

            # SaveAs with fallback strategies (same as IFCManager)
            save_ok = False
            try:
                self._com_retry(lambda: doc.SaveAs(str(save_path)))
                save_ok = True
            except Exception as e1:
                print(f"    SaveAs 策略1失败({e1})，尝试备用方案...")

            if not save_ok:
                try:
                    self._com_retry(lambda: doc.Save())
                    time.sleep(2)
                    self._com_retry(lambda: doc.SaveAs(str(save_path)))
                    save_ok = True
                except Exception as e2:
                    print(f"    SaveAs 策略2失败({e2})，尝试备用方案...")

            if not save_ok:
                try:
                    self._com_retry(lambda: doc.SaveAs(str(save_path), 61))
                    save_ok = True
                except Exception as e3:
                    print(f"    SaveAs 策略3失败({e3})")

            if not save_ok:
                # Strategy 4 (last resort): in-place Save + PUBLISH from
                # current document. DWG won't have the AB filename but PDF
                # can still be exported. Triggered when XREF bind leaves
                # AutoCAD in RPC_E_SERVERFAULT state across all SaveAs retries.
                print(f"    SaveAs 全部失败 — 尝试就地 Save + 直接导出 PDF...")
                try:
                    self._com_retry(lambda: doc.Save())
                    time.sleep(3)
                    save_path = Path(doc.FullName)
                    save_ok = True
                    print(f"    ⚠ DWG 已就地保存（源文件），AB 文件名未另存")
                except Exception as e4:
                    print(f"    就地 Save 失败({e4})，仍尝试从当前路径导出 PDF...")
                    save_path = Path(doc.FullName)

                pdf_ok_s4 = self._publish_single_pdf(acad, save_path, ab_pdf_path)
                try:
                    doc.Close(False)
                except Exception:
                    pass
                doc = None
                if pdf_ok_s4:
                    result['pdf_path'] = str(ab_pdf_path)
                    result['success'] = True
                    result['errors'].append("⚠ SaveAs 失败，DWG 未另存为 AB 文件名，但 PDF 已导出")
                else:
                    result['errors'].append("SaveAs 和 PDF 导出均失败")
                return result

            # Close document — guarded: Close can throw "Open.Close" after
            # temp-path SaveAs while AutoCAD is still flushing. Continue to
            # PUBLISH regardless; the DWG is already saved.
            try:
                doc.Close(False)
            except Exception as _ce:
                logging.warning(f"doc.Close warning (non-fatal): {_ce}")
            doc = None
            time.sleep(2)

            # Export PDF via PUBLISH
            publish_dwg = save_path
            pdf_ok = self._publish_single_pdf(acad, publish_dwg, ab_pdf_path)

            # Move DWG from temp if needed
            if use_temp:
                try:
                    shutil.move(str(save_path), str(to_long_path(final_dwg_path)))
                    ab_dwg_path = final_dwg_path
                except Exception as e:
                    ab_dwg_path = save_path
                    print(f"    ⚠ 移动失败({e})，DWG 保留在临时目录")

            result['dwg_path'] = str(ab_dwg_path)

            if pdf_ok:
                result['pdf_path'] = str(ab_pdf_path)
                result['success'] = True
                # QA validation (expected_pages=None: single DWG may have many
                # layouts → many pages; phantom/missing-stamp checks cover defects)
                qa_warnings = self._qa_validate_ab_pdf(
                    ab_pdf_path, doc_id, expected_pages=None)
                if qa_warnings:
                    for _qw in qa_warnings:
                        print(f"    [QA] {_qw}")
                    result['qa_warnings'] = qa_warnings
            else:
                result['errors'].append(
                    f"PDF PUBLISH 导出失败: {ab_pdf_path.name} (DWG已保存)")
                result['success'] = False

        except Exception as e:
            result['errors'].append(f"转换异常: {e}")
            if doc is not None:
                try:
                    doc.Close(False)
                except Exception:
                    pass
        finally:
            _jcleanup()  # remove the short-path junction (link only, not target)

        return result

    # ── Multi-page conversion ────────────────────────────────────────────

    def convert_multi_to_ab(self, dwg_info: Dict) -> Dict:
        """Convert multi-page IFC DWGs to a single AS BUILT PDF.

        Adapted from PanelIFCManager: process each page DWG individually,
        then PUBLISH all pages together as one merged PDF.

        Returns {success, dwg_paths, pdf_path, ab_rev, errors}.
        """
        result = {
            'success': False,
            'doc_id': dwg_info['doc_id'],
            'dwg_paths': [],
            'pdf_path': None,
            'ab_rev': None,
            'errors': [],
        }

        ifc_source = dwg_info['ifc_source']
        page_dwgs = ifc_source['dwg_paths']
        doc_id = dwg_info['doc_id']
        description = dwg_info['description']

        # Calculate AB revision
        if dwg_info['existing_ab_rev'] is not None:
            ab_rev = dwg_info['existing_ab_rev'] + 1
        else:
            ab_rev = 1
        result['ab_rev'] = ab_rev

        # Working folder for AB DWGs
        if self._save_in_source_dir:
            ab_subfolder = None
            ab_dwg_dir = ifc_source['source_dir']
            first_stem = re.sub(r'\s*_?IFC$', '_As Built', page_dwgs[0].stem,
                                flags=re.IGNORECASE)
            ab_pdf_path = self.ab_output / f"{first_stem}.pdf"
        else:
            ab_subfolder = dwg_info['folder'] / f"Rev.{ab_rev} - AB"
            ab_dwg_dir = ab_subfolder
            ab_pdf_name = self._build_ab_pdf_filename(doc_id, description, ab_rev)
            ab_pdf_path = self.ab_output / f"{ab_pdf_name}.pdf"

        if self.dry_run:
            result['success'] = True
            if self._save_in_source_dir:
                result['dwg_paths'] = [
                    str(ab_dwg_dir / (re.sub(r'\s*_?IFC$', '_AsBuilt', p.stem,
                                             flags=re.IGNORECASE) + p.suffix))
                    for p in page_dwgs]
            else:
                result['dwg_paths'] = [str(ab_dwg_dir / self._build_ab_dwg_name(p.name, ab_rev))
                                       for p in page_dwgs]
            result['pdf_path'] = str(ab_pdf_path)
            return result

        # --- Process each page ---
        if ab_subfolder:
            to_long_path(ab_subfolder).mkdir(parents=True, exist_ok=True)
        to_long_path(self.ab_output).mkdir(parents=True, exist_ok=True)

        acad = self._get_acad()
        group_personnel = {}
        ok_pages = []  # pages successfully converted
        total_pages = len(page_dwgs)
        date_str = datetime.now().strftime('%d/%m/%y')

        for page_idx, page_dwg in enumerate(page_dwgs, 1):
            page_label = page_dwg.stem
            print(f"    [{page_idx}/{total_pages}] {page_label}")

            if self._save_in_source_dir:
                ab_dwg_stem = re.sub(r'\s*_?IFC$', '_AsBuilt', page_dwg.stem,
                                     flags=re.IGNORECASE)
                ab_dwg_name = ab_dwg_stem + page_dwg.suffix
            else:
                ab_dwg_name = self._build_ab_dwg_name(page_dwg.name, ab_rev)
            ab_dwg_path = ab_dwg_dir / ab_dwg_name

            doc = None
            # Long source path → open via short junction (XREF-safe). See
            # _shortpath_open_target. Cleaned up after the page is processed.
            _open_path, _jcleanup = self._shortpath_open_target(page_dwg)
            try:
                # Open page DWG
                doc = self._com_retry(
                    lambda p=_open_path: acad.Documents.Open(p))
                if doc is None:
                    result['errors'].append(f"无法打开 {page_dwg.name}")
                    _jcleanup()
                    continue

                # Wait for document ready
                for _wait in range(20):
                    try:
                        _ = doc.ModelSpace.Count
                        _ = doc.Layouts.Count
                        break
                    except Exception:
                        time.sleep(1)
                time.sleep(2)

                # Find title blocks
                all_tbs = self._find_all_title_blocks(doc)
                if not all_tbs:
                    for _tb_retry in range(2):
                        time.sleep(3)
                        all_tbs = self._find_all_title_blocks(doc)
                        if all_tbs:
                            break

                # Read personnel (first page with data → fallback for others)
                if all_tbs:
                    for _tb_item in all_tbs:
                        page_personnel = self._read_latest_ifr_row(_tb_item[1])
                        if page_personnel:
                            if not group_personnel:
                                group_personnel = page_personnel
                            break

                personnel = group_personnel if not self._read_latest_ifr_row(
                    all_tbs[0][1] if all_tbs else {}) else self._read_latest_ifr_row(
                    all_tbs[0][1]) if all_tbs else {}
                if not personnel:
                    personnel = group_personnel

                # Fix typos
                self._fix_known_typos(doc)

                # Scan COLOUR before removal
                has_colour = self._scan_has_colour(doc)

                # Remove old stamps
                self._remove_ifc_stamp(doc)

                # Update title blocks + stamp
                if all_tbs:
                    for tb_idx, tb_item in enumerate(all_tbs, 1):
                        block_ref, attrs, space = tb_item[0], tb_item[1], tb_item[2]
                        layout_name = tb_item[3] if len(tb_item) > 3 else None
                        self._update_title_block(attrs, ab_rev, personnel, date_str)
                        if block_ref and space:
                            self._stamp_via_com_draw(doc, block_ref, space,
                                                     has_colour=has_colour,
                                                     layout_name=layout_name)

                # Bind XREFs before SaveAs (only for subfolder mode)
                if not self._save_in_source_dir:
                    _xb = 0
                    try:
                        for bi in range(doc.Blocks.Count):
                            try:
                                blk = doc.Blocks.Item(bi)
                                if blk.IsXRef:
                                    blk.Bind(False)
                                    _xb += 1
                            except Exception:
                                pass
                        if _xb:
                            print(f"    XREF 绑定: {_xb} 个外部引用已绑定")
                            try:
                                doc.Regen(1)
                            except Exception:
                                pass
                            time.sleep(8)  # outside try — must always run
                            _unres = 0
                            try:
                                for _bi3 in range(doc.Blocks.Count):
                                    try:
                                        if doc.Blocks.Item(_bi3).IsXRef:
                                            _unres += 1
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                            if _unres:
                                print(f"    ⚠ XREF: {_unres} 个外部引用绑定后仍未解析")
                    except Exception:
                        pass

                # SaveAs to AB subfolder
                try:
                    sp = Path(ab_dwg_path)
                    if sp.exists():
                        sp.unlink()
                except Exception:
                    pass

                save_ok = False
                try:
                    self._com_retry(lambda: doc.SaveAs(str(ab_dwg_path)))
                    save_ok = True
                except Exception:
                    try:
                        self._com_retry(lambda: doc.SaveAs(str(ab_dwg_path), 61))
                        save_ok = True
                    except Exception as e:
                        result['errors'].append(f"{page_label}: SaveAs 失败 ({e})")

                try:
                    doc.Close(False)
                except Exception as _ce:
                    logging.warning(f"doc.Close warning (non-fatal): {_ce}")
                doc = None
                time.sleep(1)

                if save_ok:
                    ok_pages.append({
                        'path': page_dwg,
                        'ab_path': ab_dwg_path,
                        'page': page_label,
                    })
                    result['dwg_paths'].append(str(ab_dwg_path))

            except Exception as e:
                result['errors'].append(f"{page_label}: 异常 ({e})")
                if doc is not None:
                    try:
                        doc.Close(False)
                    except Exception:
                        pass
            finally:
                _jcleanup()  # remove this page's short-path junction (link only)

        if not ok_pages:
            result['errors'].append("所有页面转换失败")
            return result

        # --- PUBLISH merged PDF ---
        print(f"    合并 {len(ok_pages)} 页为 PDF...")
        pdf_result = self._publish_ab_group_pdf(
            doc_id, ab_rev, ab_subfolder, ok_pages, ab_pdf_path)

        if pdf_result['success']:
            result['pdf_path'] = pdf_result['pdf_path']
            result['success'] = True
            # QA validation
            qa_warnings = self._qa_validate_ab_pdf(
                ab_pdf_path, doc_id,
                expected_pages=len(ok_pages))
            if qa_warnings:
                for _qw in qa_warnings:
                    print(f"    [QA] {_qw}")
                result['qa_warnings'] = qa_warnings
        else:
            result['errors'].append(pdf_result.get('error', 'PDF 合并失败'))
            result['success'] = False

        return result

    def _publish_ab_group_pdf(self, doc_id: str, ab_rev: int,
                               dwg_folder: Path, pages: List[Dict],
                               pdf_path: Path) -> Dict:
        """Create multi-page AS BUILT PDF via PUBLISH + DSD.

        Adapted from PanelIFCManager._publish_group_pdf().
        """
        result = {'success': False, 'pdf_path': str(pdf_path), 'error': ''}

        if self.dry_run:
            result['success'] = True
            return result

        to_long_path(pdf_path.parent).mkdir(parents=True, exist_ok=True)

        try:
            acad = self._get_acad()
            # Close all open documents
            try:
                while acad.Documents.Count > 0:
                    acad.Documents.Item(0).Close(False)
                    time.sleep(1)
            except Exception:
                pass
            time.sleep(2)

            # Detect layout name from first page
            first_dwg = pages[0]['ab_path']
            layout_name = self._detect_ab_layout_name(first_dwg)
            print(f"    Layout 检测: '{layout_name}'")

            # Build DSD
            dsd_lines = ['[DWF6Version]', 'Ver=1', '[DWF6MinorVersion]', 'MinorVer=1']
            for page in pages:
                dwg_path = page['ab_path']
                if not dwg_path.exists():
                    continue
                sheet_name = dwg_path.stem
                dwg_str = str(dwg_path)
                dsd_lines.extend([
                    f'[DWF6Sheet:{sheet_name}-{layout_name}]',
                    f'DWG={dwg_str}',
                    f'Layout={layout_name}',
                    'Setup=',
                    f'OriginalSheetPath={dwg_str}',
                    'Has Plot Port=0',
                    'Has3DDWF=0',
                ])
            pdf_str = str(pdf_path)
            out_str = str(pdf_path.parent)
            dsd_lines.extend([
                '[Target]', 'Type=6',
                f'DWF={pdf_str}',
                f'OUT={out_str}',
                'PWD=',
                'PromptForDwfName=FALSE',
                '[PdfOptions]',
                'VectorResolution=600',
                'RasterResolution=400',
                '[SheetSetProperties]',
                'IsSheetSet=FALSE',
                'IsHomogeneous=FALSE',
                'SheetSet Storage File=',
                'AcadProfile=<<Default>>',
                'CategoryCount=0',
                '[AutoCAD Block Information]',
                'IncludeBlockInfo=0',
                'BlockTmplFilePath=',
            ])
            dsd_content = '\n'.join(dsd_lines)
            dsd_path = dwg_folder / f"{doc_id}_AB.dsd"
            dsd_path.write_text(dsd_content, encoding='utf-8')

            # Open first DWG to run PUBLISH
            doc = None
            for _retry in range(3):
                try:
                    doc = self._com_retry(
                        lambda p=str(first_dwg): acad.Documents.Open(p))
                    break
                except Exception:
                    time.sleep(3)
            if doc is None:
                result['error'] = f"无法打开 {first_dwg.name} 用于 PUBLISH"
                return result

            for _wait in range(30):
                try:
                    _ = doc.ModelSpace.Count
                    break
                except Exception:
                    time.sleep(1)
            time.sleep(2)

            # PUBLISH
            saved_filedia = doc.GetVariable("FILEDIA")
            saved_bgplot = doc.GetVariable("BACKGROUNDPLOT")
            doc.SetVariable("FILEDIA", 0)
            doc.SetVariable("BACKGROUNDPLOT", 0)

            dsd_str = str(dsd_path)
            doc.SendCommand(f'-PUBLISH\n{dsd_str}\n')

            # Wait for completion
            start = time.time()
            max_wait = 300
            last_log = start
            timed_out = True
            while time.time() - start < max_wait:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        timed_out = False
                        break
                except Exception:
                    timed_out = False
                    break
                if time.time() - last_log > 30:
                    print(f"    PUBLISH 进行中... ({int(time.time() - start)}s)")
                    last_log = time.time()
                time.sleep(2)

            if timed_out:
                print(f"    PUBLISH 超时 ({max_wait}s)")

            try:
                doc.SetVariable("FILEDIA", saved_filedia)
                doc.SetVariable("BACKGROUNDPLOT", saved_bgplot)
            except Exception:
                pass

            try:
                doc.Close(False)
            except Exception:
                pass

            # Cleanup DSD
            try:
                dsd_path.unlink()
            except Exception:
                pass

            # Poll for PDF to appear and stabilize on disk
            pdf_ok = False
            for _poll in range(30):
                if pdf_path.exists():
                    try:
                        sz1 = pdf_path.stat().st_size
                        if sz1 > 0:
                            time.sleep(2)
                            sz2 = pdf_path.stat().st_size
                            if sz1 == sz2:
                                pdf_ok = True
                                break
                    except OSError:
                        pass
                time.sleep(2)

            if pdf_ok or pdf_path.exists():
                result['success'] = True
            else:
                reason = "PUBLISH 超时" if timed_out else "PUBLISH 完成但 PDF 未生成"
                result['error'] = f"{reason}: {pdf_path.name}"

        except Exception as e:
            result['error'] = f"PUBLISH 失败: {e}"

        return result

    def _detect_ab_layout_name(self, dwg_path: Path) -> str:
        """Detect first non-Model layout name from a DWG. Falls back to 'Layout1'."""
        try:
            acad = self._get_acad()
            doc = self._com_retry(lambda: acad.Documents.Open(str(dwg_path)))
            if doc is None:
                return 'Layout1'
            for _ in range(15):
                try:
                    _ = doc.ModelSpace.Count
                    break
                except Exception:
                    time.sleep(1)
            layout_name = 'Layout1'
            try:
                for layout in doc.Layouts:
                    if layout.Name.lower() != 'model':
                        layout_name = layout.Name
                        break
            except Exception:
                pass
            doc.Close(False)
            time.sleep(1)
            return layout_name
        except Exception:
            return 'Layout1'

    # ── Incremental check ────────────────────────────────────────────────

    def _check_ab_incremental(self, dwg_info: Dict) -> bool:
        """Check if this doc-ID already has an AS BUILT PDF. Returns True to skip."""
        doc_id = dwg_info['doc_id']
        ab_dir = self.ab_output
        if not ab_dir.exists():
            return False
        for f in ab_dir.iterdir():
            if not f.is_file() or f.suffix.lower() != '.pdf':
                continue
            file_doc_id = _extract_doc_id_standalone(f.name)
            if file_doc_id and file_doc_id.upper() == doc_id.upper():
                s_upper = f.stem.upper()
                if '_AS BUILT' in s_upper or '_AS_BUILT' in s_upper or '_ASBUILT' in s_upper:
                    return True
        return False

    # ── Post-conversion QA validation ────────────────────────────────────

    def _qa_validate_ab_pdf(self, pdf_path: Path, doc_id: str,
                             expected_pages: int = None) -> List[str]:
        """Post-conversion QA validation of an AS BUILT PDF.

        Returns list of warning strings. Empty list = all checks passed.
        """
        warnings = []
        if not pdf_path.exists():
            return [f'PDF not found: {pdf_path.name}']

        try:
            import fitz
        except ImportError:
            return []

        try:
            pdf_doc = fitz.open(str(pdf_path))
            page_count = pdf_doc.page_count

            if expected_pages is not None and page_count != expected_pages:
                warnings.append(
                    f'页数不匹配: PDF={page_count}, 预期={expected_pages}')

            page_dims = []
            for pi in range(page_count):
                page = pdf_doc[pi]
                page_text = page.get_text().upper()
                pw, ph = page.rect.width, page.rect.height
                page_dims.append((round(pw), round(ph)))

                # Duplicate COLOUR stamps
                import re as _re_qa
                colour_count = len(_re_qa.findall(
                    r'PRINTED\s+IN\s+COLOU?R', page_text))
                if colour_count > 1:
                    warnings.append(
                        f'Page {pi+1}: COLOUR 印章出现 {colour_count} 次 (重复)')

                # Duplicate AS BUILT stamps
                ab_count = len(_re_qa.findall(r'AS\s*BUILT', page_text))
                if ab_count > 1:
                    warnings.append(
                        f'Page {pi+1}: AS BUILT 印章出现 {ab_count} 次 (重复)')

                # Leftover IFC stamps
                if _re_qa.search(r'FOR\s+CONSTRUCTION', page_text):
                    warnings.append(
                        f'Page {pi+1}: "FOR CONSTRUCTION" 残留 — IFC 印章未清除')

                # Leftover IFR stamps
                if _re_qa.search(r'ISSUED\s+FOR\s+REVIEW', page_text):
                    warnings.append(
                        f'Page {pi+1}: "ISSUED FOR REVIEW" 残留 — IFR 印章未清除')

                # Missing AS BUILT stamp
                if not _re_qa.search(r'AS\s*BUILT', page_text):
                    warnings.append(
                        f'Page {pi+1}: 缺少 AS BUILT 印章')

                # Stamp position — must be in bottom-right quadrant (>50% x, >50% y)
                # Use fitz word extraction for bounding box of "AS BUILT" text
                try:
                    words = page.get_text("words")
                    ab_boxes = [w for w in words if 'BUILT' in w[4].upper()]
                    for wb in ab_boxes:
                        wx_centre = (wb[0] + wb[2]) / 2
                        wy_centre = (wb[1] + wb[3]) / 2
                        if wx_centre < pw * 0.5 or wy_centre < ph * 0.5:
                            warnings.append(
                                f'Page {pi+1}: 印章位置异常 — 不在底右区域 '
                                f'({wx_centre/pw:.0%} x, {wy_centre/ph:.0%} y) '
                                f'[需人工检查]')
                            break
                except Exception:
                    pass

            # Phantom page detection (different dimensions from majority)
            if page_count > 1:
                from collections import Counter
                dim_counts = Counter(page_dims)
                majority_dim = dim_counts.most_common(1)[0][0]
                for pi, dims in enumerate(page_dims):
                    if dims != majority_dim:
                        warnings.append(
                            f'Page {pi+1}: 尺寸异常 ({dims[0]}x{dims[1]} '
                            f'vs 多数页 {majority_dim[0]}x{majority_dim[1]}) '
                            f'— 可能是 Model tab [需人工检查]')

            # Stamp box alignment: COLOUR and AS BUILT rects must share same left edge
            # Uses get_drawings() to detect large filled rects in the stamp zone.
            # Misalignment is retryable — current code (4e010d2) should produce aligned output.
            for pi in range(page_count):
                page = pdf_doc[pi]
                pw, ph = page.rect.width, page.rect.height
                try:
                    # Accept both filled and stroked (thin-border) rects — AS BUILT
                    # boxes use cw=0 (thin stroke), so fill is None. Match by
                    # zone + size + rectangularity instead of requiring fill.
                    stamp_boxes = [
                        p['rect'] for p in page.get_drawings()
                        if p['rect'].x0 > pw * 0.60
                        and p['rect'].y0 > ph * 0.60
                        and p['rect'].width  > pw * 0.08
                        and p['rect'].height > ph * 0.02
                        and p['rect'].width  < pw * 0.30
                        and p['rect'].height < ph * 0.10
                    ]
                    if len(stamp_boxes) >= 2:
                        lefts  = [r.x0 for r in stamp_boxes]
                        rights = [r.x0 + r.width for r in stamp_boxes]
                        l_spread = max(lefts)  - min(lefts)
                        r_spread = max(rights) - min(rights)
                        if l_spread > 20 or r_spread > 20:
                            warnings.append(
                                f'Page {pi+1}: 印章框未对齐 '
                                f'(左边差={l_spread:.0f}pt, 右边差={r_spread:.0f}pt)')
                except Exception:
                    pass

            pdf_doc.close()
        except Exception as e:
            warnings.append(f'QA 扫描异常: {e}')

        return warnings

    # ── QA retry wrapper ─────────────────────────────────────────────────

    # Warnings containing these substrings → escalate to user immediately (no retry)
    _QA_ESCALATE_KEYWORDS = ('印章位置异常', '尺寸异常', '页数不匹配', '需人工检查')

    def _convert_with_qa_retry(self, dwg_info: Dict,
                                max_retries: int = 3) -> Dict:
        """Convert a single DWG with QA check and auto-retry closed loop.

        Retry policy:
          - Auto-retryable: stamp missing, FOR CONSTRUCTION残留, IFR残留, duplicate stamp
          - Escalate immediately: position anomaly, page count mismatch, dimension anomaly
          - Max 3 attempts; only correct output reaches the caller.
        """
        doc_id = dwg_info['doc_id']
        ifc_source = dwg_info['ifc_source']
        is_multi = ifc_source['is_multi_page']
        # Do NOT enforce a strict page count: a single source DWG can legitimately
        # have multiple PaperSpace layouts → a multi-page PDF. We can't know the
        # layout count from the PDF, so passing expected=1 caused false "页数不匹配"
        # escalations on every multi-layout drawing. Phantom Model-tab pages are
        # still caught by the dimension check, and dropped pages by the missing-
        # AS-BUILT check, inside _qa_validate_ab_pdf.
        expected_pages = None

        for attempt in range(1, max_retries + 1):
            # Convert
            if is_multi:
                result = self.convert_multi_to_ab(dwg_info)
            else:
                result = self.convert_to_ab(dwg_info)

            if not result['success']:
                if attempt < max_retries:
                    logging.warning(f"[QA-RETRY] {doc_id} 转换失败 "
                                    f"({attempt}/{max_retries}): "
                                    f"{result.get('errors')} — 重试中...")
                    self._acad = None
                    time.sleep(5)
                    continue
                return result  # max retries exhausted on conversion failure

            # QA check (PDF-level: stamps, leftovers, phantom pages, alignment).
            pdf_path = Path(result['pdf_path'])
            qa_warns = self._qa_validate_ab_pdf(
                pdf_path, doc_id, expected_pages=expected_pages)

            # Title-block QA rides ALONGSIDE as a non-blocking WARN — it is
            # DETERMINISTIC (re-converting re-reads the SAME source → an identical
            # personnel/rev result), so it must NOT drive the retry/escalate loop.
            # Retrying an empty-personnel source is futile thrash that would FAIL
            # an otherwise gold-standard deliverable PDF (root cause of the mass
            # Coleambally2 'personnel 全空' FAILs). Surface it for review; ship the
            # PDF. Only the PDF-QA defects below are retryable/escalatable.
            tb_qa = result.get('tb_qa_warnings', [])
            if tb_qa:
                _existing = result.setdefault('qa_warnings', [])
                for _w in tb_qa:
                    if _w not in _existing:
                        _existing.append(_w)

            if not qa_warns:
                return result  # ✅ PDF QA PASS — return (tb_qa, if any, as WARN)

            # Categorise
            must_escalate = [w for w in qa_warns
                             if any(kw in w for kw in self._QA_ESCALATE_KEYWORDS)]
            retryable = [w for w in qa_warns if w not in must_escalate]

            if must_escalate:
                result['success'] = False
                result['errors'].append(
                    f"QA失败(需人工介入): {'; '.join(must_escalate)}")
                if retryable:
                    result['errors'].append(
                        f"另有可重试问题: {'; '.join(retryable)}")
                return result

            # All failures are retryable
            if attempt < max_retries:
                logging.warning(f"[QA-RETRY] {doc_id} QA失败 "
                                f"({attempt}/{max_retries}): {retryable} — 重试...")
                self._acad = None
                time.sleep(5)
                continue

            # Max retries exhausted with retryable failures
            result['success'] = False
            result['errors'].append(
                f"QA失败({max_retries}次重试后仍未通过): {'; '.join(qa_warns)}")
            return result

        return result  # fallback (should not reach here)

    # ── Batch conversion ─────────────────────────────────────────────────

    def batch_convert(self, doc_ids: Optional[List[str]] = None,
                      force_doc_ids: Optional[set] = None,
                      force_rev_1: bool = True,
                      cleanup_native: bool = True) -> List[Dict]:
        """Convert IFC DWGs to AS BUILT (export FIRST, then QA — every project).

        Policy:
          - cleanup_native (default True): house-keep 1. Native/ FIRST — collapse
            duplicate 'Rev.N - AB' subfolders to the first rev + move loose AB
            exports to Superseded/. See cleanup_ab_native().
          - force_rev_1 (default True): pin output to the FIRST AB rev and
            OVERWRITE it — the default is one live AB version per doc-ID, never
            an accumulating Rev.2/Rev.3. Pass force_rev_1=False for the
            (caller-driven) special case where a genuine new AB rev must
            continue the For-Construction numbering.
          - Incremental skip is QA-GATED ('skip only if QA-clean'): a doc-ID is
            skipped ONLY when its existing AB PDF passes QA. A missing OR
            QA-faulty PDF is re-exported then re-QA'd — so 'export first, then
            QA' holds for everything that actually needs producing.

        Args:
            doc_ids: If provided, only convert these doc-IDs.
            force_doc_ids: Always re-convert these (bypass the QA-clean skip).
        """
        mode_label = "同目录" if self._save_in_source_dir else "子文件夹"
        print(f"  AS BUILT 模式: {mode_label} | Native: {self._detected_native} | "
              f"Output: {self._detected_ab_output}")

        # Native house-keeping FIRST (filesystem only, reversible → Superseded/).
        if cleanup_native and not self.dry_run:
            acts = self.cleanup_ab_native()
            if acts:
                print(f"  [AB-清理] Native 整理: {len(acts)} 项移入 Superseded/")
                for a in acts[:12]:
                    print(f"    - {a['kind']}: {Path(a['path']).name}")

        scan = self.scan_native_for_ab()
        if doc_ids:
            doc_id_set = {d.upper() for d in doc_ids}
            scan = [s for s in scan if s['doc_id'].upper() in doc_id_set]
        force_set = {d.upper() for d in (force_doc_ids or set())}

        if force_rev_1:
            # Pin output to the FIRST AB rev (overwrite). Does NOT force-convert
            # everything — the QA-clean gate below still skips clean output.
            for item in scan:
                item['existing_ab_rev'] = None
            print("  [force_rev_1] 输出锁定 REV 1（已有 PDF 将被覆盖；QA 通过者仍跳过）")

        if not scan:
            print("  没有找到需要转换的 IFC 文件")
            return []

        results = []
        total = len(scan)
        for idx, dwg_info in enumerate(scan, 1):
            doc_id = dwg_info['doc_id']
            ifc_source = dwg_info['ifc_source']
            is_multi = ifc_source['is_multi_page']
            page_count = len(ifc_source['dwg_paths'])
            ab_rev = (dwg_info['existing_ab_rev'] or 0) + 1

            # QA-gated incremental skip: skip ONLY when the existing AB PDF is
            # already QA-clean. Missing/faulty → fall through to export + QA.
            if doc_id.upper() not in force_set and self._ab_existing_pdf_qa_clean(dwg_info):
                print(f"  [{idx}/{total}] {doc_id} — 已有 AS BUILT PDF 且 QA 通过，跳过")
                continue

            type_label = f"multi({page_count}p)" if is_multi else "single"
            print(f"  [{idx}/{total}] {doc_id} ({type_label}, IFC Rev{ifc_source['ifc_rev']} -> "
                  f"AB Rev{ab_rev})...")

            r = self._convert_with_qa_retry(dwg_info)
            results.append(r)

            if r['success']:
                status = "OK"
                extra_parts = []
                if r.get('dwg_path'):
                    extra_parts.append(f"DWG={Path(r['dwg_path']).name}")
                elif r.get('dwg_paths'):
                    extra_parts.append(f"DWG x{len(r['dwg_paths'])}")
                if r.get('pdf_path'):
                    extra_parts.append(f"PDF={Path(r['pdf_path']).name}")
                if r.get('errors'):
                    extra_parts.append(f"警告: {'; '.join(r['errors'])}")
                print(f"    [{status}] {', '.join(extra_parts)}")
            else:
                print(f"    [FAIL] {'; '.join(r.get('errors', ['未知错误']))}")
                self._acad = None
                time.sleep(3)

        # Summary
        ok = sum(1 for r in results if r['success'])
        fail = len(results) - ok
        qa_issues = sum(1 for r in results if r.get('qa_warnings'))
        print(f"\n  AS BUILT done: ok={ok}, fail={fail}")
        if qa_issues:
            print(f"  ⚠ QA 警告: {qa_issues} 个文件有质量问题，请检查日志")

        # Post-batch self-check: scan ALL PDFs in output dir for stamp issues
        if not self.dry_run and ok > 0:
            print(f"\n  [POST-QA] 对 {self.ab_output.name}/ 做批量自检...")
            batch_qa = self._run_post_batch_qa()
            if batch_qa:
                fail_count = sum(1 for v in batch_qa.values() if v)
                print(f"  [POST-QA] {len(batch_qa)} 个 PDF 扫描完毕 — "
                      f"{fail_count} 个有问题, {len(batch_qa)-fail_count} 个通过")
                for fname, issues in batch_qa.items():
                    if issues:
                        print(f"    ✗ {fname}: {'; '.join(issues[:3])}")
            for r in results:
                if r.get('success') and r.get('pdf_path'):
                    pname = Path(r['pdf_path']).name
                    r['post_qa'] = batch_qa.get(pname, [])

        return results

    def _run_post_batch_qa(self) -> Dict[str, List[str]]:
        """Scan all PDFs in ab_output for stamp issues. Returns {filename: [issues]}."""
        try:
            import fitz as _fitz
            import re as _re
        except ImportError:
            return {}

        results: Dict[str, List[str]] = {}
        ab_dir = self.ab_output
        if not ab_dir.exists():
            return results

        _ZONE_X, _ZONE_Y = 0.60, 0.60
        _MIN_W, _MIN_H    = 0.08, 0.02
        _ALIGN_TOL        = 20.0

        for pdf_path in sorted(ab_dir.glob("*.pdf")):
            issues: List[str] = []
            try:
                doc = _fitz.open(str(pdf_path))
                for pi in range(doc.page_count):
                    page = doc[pi]
                    pw, ph = page.rect.width, page.rect.height
                    text = page.get_text().upper()

                    # FOR CONSTRUCTION leftover
                    if _re.search(r'FOR\s+CONSTRUCTION', text):
                        issues.append(f"p{pi+1}: FOR CONSTRUCTION 残留")
                    # Duplicate AS BUILT
                    if len(_re.findall(r'AS\s*BUILT', text)) > 1:
                        issues.append(f"p{pi+1}: AS BUILT 重复")
                    # Missing AS BUILT
                    if not _re.search(r'AS\s*BUILT', text):
                        issues.append(f"p{pi+1}: 缺 AS BUILT")
                    # RECT alignment — accept thin stroked rects (cw=0), not just filled
                    rects = [p['rect'] for p in page.get_drawings()
                             if p['rect'].x0 > pw*_ZONE_X
                             and p['rect'].y0 > ph*_ZONE_Y
                             and p['rect'].width  > pw*_MIN_W
                             and p['rect'].height > ph*_MIN_H
                             and p['rect'].width  < pw*0.30
                             and p['rect'].height < ph*0.10]
                    if len(rects) >= 2:
                        lefts = [r.x0 for r in rects]
                        spread = max(lefts) - min(lefts)
                        if spread > _ALIGN_TOL:
                            issues.append(f"p{pi+1}: 印章框未对齐({spread:.0f}pt)")
                doc.close()
            except Exception as e:
                issues.append(f"扫描异常: {e}")
            results[pdf_path.name] = issues

        return results


# =============================================================================
# Pipeline Orchestrator (NEW in v7)
# =============================================================================

class PipelineOrchestrator:
    """Sequences the 6 stages per project: Health Check → IFR Sync → Version Mgmt → IFC Transmittal → Sharepoint Sync → Deliverable."""

    def __init__(self, config: ConfigManager, dry_run: bool = False,
                 stages: Optional[List[str]] = None,
                 filter_doc_ids: Optional[Set[str]] = None):
        self.config = config
        self.dry_run = dry_run
        self.stages = stages or ['health_check', 'ifr_sync', 'version_mgmt', 'ifc_transmittal', 'sharepoint_sync', 'deliverable']
        self.filter_doc_ids = filter_doc_ids

    def run_pipeline(self, project_path: Path, project_validation: ProjectValidation) -> Dict:
        """Run the full pipeline for a single project."""
        results = {
            'project_name': project_validation.project_name,
            'health_check': None,
            'ifr_sync': None,
            'version_mgmt': None,
            'ifc_transmittal': None,
            'sharepoint_sync': None,
            'deliverable': None,
            'success': True,
        }

        UIHelper.print_separator("=", 72)
        print(f"\n  管线处理: {project_validation.project_name}")
        UIHelper.print_separator("=", 72)

        # Pre-load Deliverable Excel doc-IDs for report validation
        # This allows Stage 1 (_mirror_reports) to validate files against
        # the Deliverable before Stage 4 (cross_check) runs.
        excel_doc_ids: Set[str] = set()
        try:
            dm_preload = DeliverableManager(project_path, dry_run=True)
            excel_doc_ids = dm_preload.preload_doc_ids()
            if excel_doc_ids:
                print(f"  预加载交付物 Excel: {len(excel_doc_ids)} 个 doc-ID")
        except Exception:
            pass  # Non-fatal — fallback to prefix-only matching

        # Stage 0: File Health Check
        if 'health_check' in self.stages:
            print(f"\n  [Stage 0/6] 文件健康检查...")
            try:
                checker = FileHealthChecker(project_path, dry_run=self.dry_run)
                check_result = checker.scan_anomalies()
                anomalies = check_result.get('anomalies', [])
                empty = check_result.get('empty_folders', [])
                # Apply doc-ID filter if provided
                if self.filter_doc_ids is not None:
                    anomalies = [a for a in anomalies
                                 if a.get('folder_doc_id', '').upper() in self.filter_doc_ids]
                    check_result['anomalies'] = anomalies
                if anomalies:
                    if not self.dry_run:
                        rename_stats = checker.execute_renames(anomalies)
                        UIHelper.print_success(
                            f"文件健康检查: 发现 {len(anomalies)} 个异常, "
                            f"重命名={rename_stats['renamed']}, 失败={rename_stats['failed']}")
                    else:
                        UIHelper.print_success(
                            f"文件健康检查: 发现 {len(anomalies)} 个异常文件")
                else:
                    UIHelper.print_success(
                        f"文件健康检查: 检查 {check_result['folders_checked']} 个文件夹, 无���常")
                results['health_check'] = {
                    'folders_checked': check_result['folders_checked'],
                    'files_checked': check_result['files_checked'],
                    'anomalies': anomalies,
                    'empty_folders': empty,
                    'renamed': rename_stats['renamed'] if not self.dry_run and anomalies else 0,
                    'failed': rename_stats['failed'] if not self.dry_run and anomalies else 0,
                }
            except Exception as e:
                UIHelper.print_error(f"文件健康检查异常: {e}")
                results['health_check'] = {'error': str(e)}

        # Stage 1: IFR Sync
        if 'ifr_sync' in self.stages:
            print(f"\n  [Stage 1/6] IFR 同步...")
            try:
                automation = IFRAutomation(
                    root_path=str(project_path.parent),
                    config=self.config,
                    dry_run=self.dry_run,
                    interactive=True,
                    excel_doc_ids=excel_doc_ids
                )
                sync_result = automation.process_project(project_validation)
                results['ifr_sync'] = {
                    'success': sync_result.success,
                    'folders': sync_result.folders_created,
                    'drawings': sync_result.drawings_copied,
                    'reports': sync_result.reports_copied,
                }
                if sync_result.success:
                    UIHelper.print_success(
                        f"IFR同步完成: 文件夹={sync_result.folders_created}, "
                        f"图纸={sync_result.drawings_copied}, 报告={sync_result.reports_copied}")
                else:
                    UIHelper.print_error(f"IFR同步失败: {sync_result.errors}")
                    results['success'] = False
            except Exception as e:
                UIHelper.print_error(f"IFR同步异常: {e}")
                results['ifr_sync'] = {'success': False, 'error': str(e)}
                results['success'] = False

        # Stage 2: Version Management
        if 'version_mgmt' in self.stages:
            print(f"\n  [Stage 2/6] 版本管理...")
            try:
                vm = VersionManager(str(project_path.parent), dry_run=self.dry_run)
                vm_stats = vm.process_project(project_path, show_details=True)
                results['version_mgmt'] = vm_stats
                UIHelper.print_success(
                    f"版本管理完成: 扫描={vm_stats['scanned']}, 移动={vm_stats['moved']}")
            except Exception as e:
                UIHelper.print_error(f"版本管理异常: {e}")
                results['version_mgmt'] = {'error': str(e)}
                results['success'] = False

        # Stage 3: IFC Transmittal (scan + dedup, pass ifc_map to Stage 5)
        if 'ifc_transmittal' in self.stages:
            ifc_dir = project_path / "Design/Engineering/1. Drawings/4. IFC(Client)"
            if ifc_dir.exists():
                print(f"\n  [Stage 3/6] IFC Transmittal 管理...")
                try:
                    ifc_cfg = self.config.config if hasattr(self.config, 'config') else {}
                    mgr = IFCTransmittalManager(project_path, config=ifc_cfg, dry_run=self.dry_run)
                    ifc_result = mgr.run_for_pipeline()
                    results['ifc_transmittal'] = ifc_result
                    parts = [f"IFC文件={ifc_result['total_ifc_files']}"]
                    if ifc_result['duplicates_archived']:
                        parts.append(f"去重={ifc_result['duplicates_archived']}")
                    parts.append(f"doc-ID={len(ifc_result.get('ifc_map', {}))}")
                    UIHelper.print_success(f"IFC Transmittal完成: {', '.join(parts)}")
                except Exception as e:
                    UIHelper.print_error(f"IFC Transmittal异常: {e}")
                    results['ifc_transmittal'] = {'error': str(e)}
            else:
                results['ifc_transmittal'] = {'skipped': True, 'reason': 'IFC(Client)不存在'}

        # Stage 4: Sharepoint Sync (archive approved + sync to Client Sharepoint)
        if 'sharepoint_sync' in self.stages:
            print(f"\n  [Stage 4/6] Client Sharepoint 同步...")
            try:
                automation = IFRAutomation(
                    root_path=str(project_path.parent),
                    config=self.config,
                    dry_run=self.dry_run,
                    interactive=True
                )
                sp_result = automation.sync_to_sharepoint(project_path)
                results['sharepoint_sync'] = sp_result
                parts = [f"复制={sp_result['copied']}"]
                if sp_result.get('archived', 0) > 0:
                    parts.append(f"归档={sp_result['archived']}")
                parts.append(f"跳过={sp_result['skipped']}")
                parts.append(f"已审批跳过={sp_result['skipped_approved']}")
                UIHelper.print_success(f"Sharepoint同步完成: {', '.join(parts)}")
            except Exception as e:
                UIHelper.print_error(f"Sharepoint同步异常: {e}")
                results['sharepoint_sync'] = {'error': str(e)}
                results['success'] = False

        # Stage 5: Deliverable Cross-Check
        if 'deliverable' in self.stages:
            print(f"\n  [Stage 5/6] 交付物检查...")
            try:
                dm = DeliverableManager(project_path, dry_run=self.dry_run)
                excel_path = dm.find_deliverable_excel()
                if excel_path:
                    ifc_map = None
                    ifc_data = results.get('ifc_transmittal')
                    if ifc_data and isinstance(ifc_data, dict):
                        ifc_map = ifc_data.get('ifc_map')
                    check_result = dm.cross_check(excel_path, external_ifc_map=ifc_map)
                    self._print_cross_check_summary(check_result)
                    if not self.dry_run and (check_result.items_in_folders_not_excel or
                                              check_result.revision_mismatches or
                                              check_result.doc_id_corrections or
                                              check_result.status_updates):
                        check_result = dm.apply_updates(excel_path, check_result)
                        UIHelper.print_success(
                            f"交付物更新完成: 新增={check_result.rows_inserted}, "
                            f"更新={check_result.rows_updated}")
                    results['deliverable'] = {
                        'new_items': len(check_result.items_in_folders_not_excel),
                        'rev_mismatches': len(check_result.revision_mismatches),
                        'doc_id_corrections': len(check_result.doc_id_corrections),
                        'status_updates': len(check_result.status_updates),
                        'naming_warnings': len(check_result.naming_warnings),
                        'inserted': check_result.rows_inserted,
                        'updated': check_result.rows_updated,
                    }
                else:
                    UIHelper.print_warning("未找到交付物 Excel 文件")
                    results['deliverable'] = {'skipped': True, 'reason': '未找到Excel'}
            except Exception as e:
                UIHelper.print_error(f"交付物检查异常: {e}")
                results['deliverable'] = {'error': str(e)}
                results['success'] = False

        return results

    def _print_cross_check_summary(self, result: DeliverableCrossCheckResult):
        """Print cross-check results summary."""
        if result.errors:
            for err in result.errors:
                UIHelper.print_error(err)
            return

        if result.items_in_folders_not_excel:
            print(f"\n    新增文件 (在文件夹中但不在Excel中): {len(result.items_in_folders_not_excel)}")
            for item in result.items_in_folders_not_excel[:10]:
                print(f"      + {item['doc_id']} ({item['filename']})")
            if len(result.items_in_folders_not_excel) > 10:
                print(f"      ... 还有 {len(result.items_in_folders_not_excel) - 10} 个")

        if result.items_in_excel_not_folders:
            print(f"\n    仅在Excel中 (信息): {len(result.items_in_excel_not_folders)}")
            for doc_id in result.items_in_excel_not_folders[:5]:
                print(f"      ? {doc_id}")

        if result.doc_id_corrections:
            print(f"\n    FILE NO 修正 (按描述匹配): {len(result.doc_id_corrections)}")
            for corr in result.doc_id_corrections:
                print(f"      Row {corr['row']}: {corr['old_doc_id']} -> {corr['doc_id']} ({corr['description']})")

        if result.revision_mismatches:
            print(f"\n    版本不一致: {len(result.revision_mismatches)}")
            for mm in result.revision_mismatches[:10]:
                print(f"      {mm['doc_id']}: Excel={mm['excel_rev']} -> 文件夹={mm['folder_rev']}")

        if result.status_updates:
            print(f"\n    IFC 状态更新: {len(result.status_updates)}")
            for su in result.status_updates[:10]:
                print(f"      {su['doc_id']}: {su['old_status'] or '(空)'} -> {su['new_status']}")

        if result.naming_warnings:
            print(f"\n    命名规范警告: {len(result.naming_warnings)}")
            for nw in result.naming_warnings[:10]:
                print(f"      {nw['doc_id']}: {', '.join(nw['changes'])}")
                print(f"        建议: {nw['suggested']}")

        if (not result.items_in_folders_not_excel and not result.revision_mismatches
                and not result.doc_id_corrections and not result.status_updates):
            UIHelper.print_success("交付物清单与文件夹一致，无需更新")


# =============================================================================
# Interactive Mode (Unified)
# =============================================================================

class UnifiedInteractiveMode:
    """Unified interactive mode combining IFR Sync + Version Mgmt + Sharepoint Sync + Deliverable."""

    def __init__(self, config: ConfigManager):
        self.config = config
        self.ui = UIHelper()
        self.validator = ProjectValidator(config)
        self.scanner = ProjectScanner(config, self.validator)
        self.safety_checker = SafetyChecker(config)
        self.report_generator = ValidationReportGenerator()
        self.root_path: Optional[Path] = None

    def run(self):
        """Main interactive loop."""
        self.ui.clear_screen()
        self.show_welcome()

        while True:
            choice = self.show_main_menu()

            if choice == '0':
                self.ui.print_info("感谢使用，再见！")
                break
            elif choice == '1':
                self.full_pipeline_mode()
            elif choice == '2':
                self.scan_and_process_projects()
            elif choice == '3':
                self.version_mgmt_mode()
            elif choice == '4':
                self.deliverable_mode()
            elif choice == '5':
                self.native_dwg_menu()
            elif choice == '6':
                self.folder_relocate_menu()
            elif choice == '7':
                self.ifc_copy_mode()
            elif choice == '8':
                self.validate_only_mode()
            elif choice == '9':
                self.modify_config()
            elif choice == '10':
                self.ifc_convert_mode()
            elif choice == '11':
                self.panel_ifc_convert_mode()
            else:
                self.ui.print_warning("无效选项，请重新选择")

    def show_welcome(self):
        """Display welcome screen."""
        self.ui.print_header(
            "工程文档自动化管线 v8.0",
            "IFR Sync + Version Mgmt + Sharepoint Sync + Deliverable + Panel IFC"
        )

        print("当前配置信息：")
        root = self.config.get("default_root_path", "")
        if root:
            display = root[:50] + "..." if len(root) > 50 else root
            print(f"  [DIR] 根目录: {display}")
        else:
            print(f"  [DIR] 根目录: 未设置")

        print(f"  日志级别: {self.config.get('log_level', 'INFO')}")
        print(f"  [v] 自动备份: {'已启用' if self.config.get('auto_backup', True) else '已禁用'}")

        # Show filters
        filters = self.config.get("project_filters", {})
        whitelist = filters.get("whitelist", [])
        blacklist = filters.get("blacklist", [])
        if whitelist:
            print(f"  白名单: {', '.join(whitelist[:3])}{'...' if len(whitelist) > 3 else ''}")
        if blacklist:
            print(f"  黑名单: {', '.join(blacklist[:3])}{'...' if len(blacklist) > 3 else ''}")

        print()

    def show_main_menu(self) -> str:
        """Display main menu."""
        self.ui.print_separator()

        options = [
            ("1", "完整流程 (IFR Sync + Version Mgmt + Sharepoint Sync + Deliverable)"),
            ("2", "仅 IFR 同步"),
            ("3", "仅版本管理 (PDF)"),
            ("4", "仅交付物检查 & 更新"),
            ("5", "Native/Reports/Schedule 版本管理"),
            ("6", "文件夹归位"),
            ("7", "复制 IFC(Client) 到目标"),
            ("8", "仅验证项目结构"),
            ("9", "配置 / 日志"),
            ("10", "IFR → IFC 转换 (AutoCAD)"),
            ("11", "Panel IFC 批量转换 (多页 DWG)"),
            ("0", "退出"),
        ]

        return self.ui.print_menu(options, "请输入选项 [0-11]")

    def get_root_path(self) -> Optional[Path]:
        """Get root path from config or user input."""
        if self.root_path:
            return self.root_path

        config_root = self.config.get("default_root_path")
        if config_root and Path(config_root).exists():
            print(f"\n当前根目录: {config_root}")
            if self.ui.confirm("使用此目录？"):
                self.root_path = Path(config_root)
                return self.root_path

        print("\n请输入项目根目录路径：")
        path_str = input("> ").strip().strip('"')

        if path_str and Path(path_str).exists():
            self.root_path = Path(path_str)
            self.config.set("default_root_path", str(self.root_path))
            return self.root_path
        else:
            self.ui.print_error("路径不存在！")
            return None

    # =========================================================================
    # V7: New Pipeline / Deliverable / Version Mgmt modes
    # =========================================================================

    def full_pipeline_mode(self):
        """Run the full pipeline (IFR Sync + Version Mgmt + Sharepoint Sync + Deliverable) per project."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)
        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        all_projects = self.display_projects_with_status(projects_by_region)
        selected = self.get_project_selection(projects_by_region)
        if not selected:
            return

        # Ask for dry-run
        dry_run = not self.ui.confirm("\n执行实际操作？(选N为预览模式)", default=False)

        pipeline = PipelineOrchestrator(self.config, dry_run=dry_run)

        all_results = []
        for idx, project in enumerate(selected, 1):
            print(f"\n{'='*72}")
            print(f"管线处理项目 {idx}/{len(selected)}: {project.project_name}")
            print(f"{'='*72}")

            result = pipeline.run_pipeline(Path(project.project_path), project)
            all_results.append(result)

            if result['success']:
                self.ui.print_success(f"项目 {project.project_name} 管线完成")
            else:
                self.ui.print_warning(f"项目 {project.project_name} 部分失败")

        # Summary
        self.ui.print_header("管线执行完成", f"{'预览模式' if dry_run else '已执行'}")
        success_count = sum(1 for r in all_results if r['success'])
        print(f"  成功: {success_count}/{len(all_results)}")
        input("\n按 Enter 返回主菜单...")

    def version_mgmt_mode(self):
        """PDF version management mode (from version_manager_v4)."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        vm = VersionManager(str(root))
        projects = vm.scan_for_projects()

        if not projects:
            self.ui.print_warning("未找到任何项目！")
            return

        print(f"\n找到 {len(projects)} 个项目:")
        for i, project in enumerate(projects, 1):
            print(f"  [{i}] {project.name}")
        print(f"\n  [0] 返回主菜单")
        self.ui.print_separator()

        choice = input(f"\n请选择项目 [0-{len(projects)}，输入 'all' 处理全部]: ").strip()

        if choice == '0':
            return
        elif choice.lower() == 'all':
            selected_projects = projects
        else:
            try:
                idx = int(choice)
                if 1 <= idx <= len(projects):
                    selected_projects = [projects[idx - 1]]
                else:
                    return
            except ValueError:
                return

        # Preview or execute
        print("\n选择操作:")
        print("  [1] 执行移动")
        print("  [2] 仅预览 (dry-run)")
        print("  [0] 取消返回")
        op = input("\n请选择: ").strip()

        if op == '0':
            return

        dry_run = op != '1'
        if not dry_run:
            confirm = input("\n确认执行? [Y/n]: ").strip().lower()
            if confirm not in ('', 'y', 'yes'):
                print("已取消")
                return

        vm.dry_run = dry_run
        total_stats = {"scanned": 0, "groups": 0, "moved": 0}

        for project in selected_projects:
            print(f"\n处理: {project.name}")
            self.ui.print_separator()
            stats = vm.process_project(project)
            total_stats["scanned"] += stats["scanned"]
            total_stats["groups"] += stats["groups"]
            total_stats["moved"] += stats["moved"]

        print(f"\n完成: 扫描={total_stats['scanned']}, 组={total_stats['groups']}, 移动={total_stats['moved']}")
        input("\n按 Enter 返回主菜单...")

    def deliverable_mode(self):
        """Deliverable cross-check and update mode."""
        if not OPENPYXL_AVAILABLE:
            self.ui.print_error("需要安装 openpyxl: pip install openpyxl")
            input("\n按 Enter 返回...")
            return

        root = self.get_root_path()
        if not root:
            return

        # Scan for projects with deliverable Excel files
        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)
        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        all_projects = [p for projects in projects_by_region.values() for p in projects]

        # Find which projects have deliverable Excel files
        dlv_projects = []
        for p in all_projects:
            dm = DeliverableManager(Path(p.project_path))
            excel = dm.find_deliverable_excel()
            if excel:
                dlv_projects.append((p, excel))

        if not dlv_projects:
            self.ui.print_warning("没有找到包含交付物 Excel 的项目！")
            input("\n按 Enter 返回...")
            return

        print(f"\n找到 {len(dlv_projects)} 个包含交付物 Excel 的项目:")
        for i, (p, excel) in enumerate(dlv_projects, 1):
            print(f"  [{i}] {p.project_name}")
            print(f"      Excel: {excel.name}")

        print(f"\n  [0] 返回主菜单")
        self.ui.print_separator()

        choice = input(f"\n请选择项目 [0-{len(dlv_projects)}]: ").strip()
        if choice == '0':
            return

        try:
            idx = int(choice)
            if not (1 <= idx <= len(dlv_projects)):
                return
        except ValueError:
            return

        project, excel_path = dlv_projects[idx - 1]
        project_path = Path(project.project_path)

        # Cross-check
        print(f"\n正在检查: {project.project_name}")
        dm = DeliverableManager(project_path)

        check_result = dm.cross_check(excel_path)

        if check_result.errors:
            for err in check_result.errors:
                self.ui.print_error(err)
            input("\n按 Enter 返回...")
            return

        # Display results
        print(f"\n{'='*60}")
        print(f"  交付物检查结果: {project.project_name}")
        print(f"{'='*60}")

        if check_result.items_in_folders_not_excel:
            print(f"\n  新增文件 (在文件夹中但不在Excel): {len(check_result.items_in_folders_not_excel)}")
            for item in check_result.items_in_folders_not_excel:
                print(f"    + {item['doc_id']} | {item['description'][:40]} | Rev {item['revision'] or '?'}")

        if check_result.items_in_excel_not_folders:
            print(f"\n  仅在Excel中 (信息): {len(check_result.items_in_excel_not_folders)}")
            for doc_id in check_result.items_in_excel_not_folders[:10]:
                print(f"    ? {doc_id}")

        if check_result.revision_mismatches:
            print(f"\n  版本不一致: {len(check_result.revision_mismatches)}")
            for mm in check_result.revision_mismatches:
                print(f"    {mm['doc_id']}: Excel=Rev{mm['excel_rev']} -> 文件夹=Rev{mm['folder_rev']}")

        if check_result.status_updates:
            print(f"\n  IFC 状态更新: {len(check_result.status_updates)}")
            for su in check_result.status_updates:
                print(f"    {su['doc_id']}: {su['old_status'] or '(空)'} -> {su['new_status']}")

        if check_result.naming_warnings:
            print(f"\n  命名规范警告: {len(check_result.naming_warnings)}")
            for nw in check_result.naming_warnings[:10]:
                print(f"    {nw['doc_id']}: {', '.join(nw['changes'])}")
                print(f"      当前: {nw['filename']}")
                print(f"      建议: {nw['suggested']}")

        if (not check_result.items_in_folders_not_excel and not check_result.revision_mismatches
                and not check_result.status_updates and not check_result.doc_id_corrections):
            self.ui.print_success("交付物清单与文件夹一致，无需更新")
            input("\n按 Enter 返回...")
            return

        # Ask to apply
        self.ui.print_separator()
        print("\n选择操作:")
        print("  [1] 应用更新到 Excel")
        print("  [2] 预览模式 (dry-run)")
        print("  [0] 返回 (不执行)")
        op = input("\n请选择: ").strip()

        if op == '0':
            return

        dry_run = op != '1'
        if not dry_run:
            if not self.ui.confirm("确认更新Excel文件？"):
                return

        dm.dry_run = dry_run
        updated = dm.apply_updates(excel_path, check_result)

        if dry_run:
            print(f"\n[预览] 将新增 {updated.rows_inserted} 行, 更新 {updated.rows_updated} 行")
            print(f"[预览] 文件版本将更新为: {updated.new_file_rev}")
        else:
            self.ui.print_success(
                f"完成: 新增={updated.rows_inserted}, 更新={updated.rows_updated}, "
                f"新版本={updated.new_file_rev}")
            if updated.excel_path:
                print(f"  新文件: {updated.excel_path}")

        input("\n按 Enter 返回主菜单...")

    def native_dwg_menu(self):
        """Native/Reports/Schedule version management sub-menu."""
        root = self.get_root_path()
        if not root:
            return

        vm = VersionManager(str(root))
        projects = vm.scan_for_projects()

        if not projects:
            self.ui.print_warning("未找到任何项目")
            input("\n按 Enter 返回...")
            return

        print("\n选择项目:")
        for i, project in enumerate(projects, 1):
            print(f"  [{i}] {project.name}")
        print(f"\n  [0] 返回主菜单")

        choice = input(f"\n请选择项目 [0-{len(projects)}]: ").strip()
        try:
            idx = int(choice)
            if idx == 0:
                return
            if 1 <= idx <= len(projects):
                project = projects[idx - 1]
            else:
                return
        except ValueError:
            return

        print(f"\n  项目: {project.name}")
        native_mgr = NativeVersionManager(str(project), dry_run=True)

        # Scope selection
        print("\n  扫描范围:")
        print("    [1] 全部 (Native + Reports + Schedule)")
        print("    [2] 仅 Native")
        print("    [3] 仅 Reports")
        print("    [4] 仅 Schedule")
        scope_choice = input("\n  请选择 [1-4, 默认=1]: ").strip() or '1'
        scope_map = {'1': 'all', '2': 'native', '3': 'reports', '4': 'schedule'}
        scope = scope_map.get(scope_choice, 'all')

        folder_filter = input("\n  筛选文件夹 (留空=全部): ").strip() or None

        results = native_mgr.process_all(folder_filter, scope)
        if not results:
            print("\n  没有找到文档文件夹")
            input("\n按 Enter 返回...")
            return

        total_actions = sum(len(r.actions) for _, r in results)
        folders_with_files = sum(1 for _, r in results if r.dwg_files)

        print(f"\n  [预览] 扫描 {len(results)} 个文件夹, "
              f"{folders_with_files} 个有文件, {total_actions} 个动作")
        self.ui.print_separator()

        for group_name, group_results in groupby(results, key=lambda x: x[0]):
            group_list = list(group_results)
            group_folders_with_files = [r for _, r in group_list if r.dwg_files]
            print(f"\n  === {group_name} === ({len(group_folders_with_files)} folders)")

            for _, result in group_list:
                if not result.dwg_files:
                    continue
                print(f"    {result.folder_name}")
                print(f"      doc_id: {result.doc_id}  |  {len(result.dwg_files)} files")
                for kept in result.kept_ifr_all:
                    print(f"      [保留 IFR] {kept.filename}")
                for kept in result.kept_ifc_all:
                    print(f"      [保留 IFC] {kept.filename}")
                for action in result.actions:
                    if action.action == 'rename':
                        print(f"      [重命名] {action.source.name} -> {action.dest.name}")
                    else:
                        print(f"      [->{action.dest.parent.name}/] {action.source.name}")

        if total_actions == 0:
            print("\n  [v] 所有文件已是最新标准格式")
            input("\n按 Enter 返回...")
            return

        renames = sum(1 for _, r in results for a in r.actions if a.action == 'rename')
        moves = sum(1 for _, r in results for a in r.actions if a.action == 'move_to_ss')
        print(f"\n  汇总: {renames} 个重命名, {moves} 个移动到 SS/")

        print("\n  选择操作:")
        print("    [1] 执行")
        print("    [0] 返回 (不执行)")
        execute_choice = input("\n  请选择: ").strip()
        if execute_choice == '1':
            confirm = input("\n  确认执行? 输入 'YES': ").strip()
            if confirm == 'YES':
                native_mgr.dry_run = False
                stats = native_mgr.execute_actions(results)
                print(f"\n  完成: {stats['renamed']} 重命名, {stats['moved']} 移动, "
                      f"{stats['ss_created']} SS/ 创建, {stats['errors']} 错误")
            else:
                print("\n  已取消")
        input("\n按 Enter 返回主菜单...")

    def folder_relocate_menu(self):
        """Folder relocation sub-menu."""
        root = self.get_root_path()
        if not root:
            return

        vm = VersionManager(str(root))
        projects = vm.scan_for_projects()

        if not projects:
            self.ui.print_warning("未找到任何项目")
            input("\n按 Enter 返回...")
            return

        print("\n选择项目:")
        for i, project in enumerate(projects, 1):
            print(f"  [{i}] {project.name}")
        print(f"\n  [0] 返回主菜单")

        choice = input(f"\n请选择项目 [0-{len(projects)}]: ").strip()
        try:
            idx = int(choice)
            if idx == 0:
                return
            if 1 <= idx <= len(projects):
                project = projects[idx - 1]
            else:
                return
        except ValueError:
            return

        print(f"\n  项目: {project.name}")
        relocator = FolderRelocator(project, dry_run=True)

        if not relocator.drawings_root.exists():
            print(f"\n  [!] 1. Drawings/ 不存在")
            input("\n按 Enter 返回...")
            return

        relocations = relocator.scan()
        if not relocations:
            self.ui.print_success("未检测到错位的文件夹，结构正确")
            input("\n按 Enter 返回...")
            return

        actionable = [r for r in relocations if r.action != 'warn']
        warnings = [r for r in relocations if r.action == 'warn']

        if actionable:
            print(f"\n  检测到 {len(actionable)} 个错位文件夹:")
            for i, rel in enumerate(actionable, 1):
                action_str = '移动' if rel.action == 'move' else '合并'
                print(f"    [{i}] [{action_str}] {rel.source.name}")
                print(f"         原因: {rel.reason}")

        if warnings:
            print(f"\n  警告 ({len(warnings)}):")
            for rel in warnings:
                print(f"    [!] {rel.reason}")

        if not actionable:
            input("\n按 Enter 返回...")
            return

        print("\n  选择操作:")
        print("    [1] 执行归位")
        print("    [0] 返回")
        if input("\n  请选择: ").strip() == '1':
            if input("\n  确认执行? 输入 'YES': ").strip() == 'YES':
                relocator.dry_run = False
                stats = relocator.execute(actionable)
                print(f"\n  完成: {stats['moved']} 移动, {stats['merged']} 合并, {stats['errors']} 错误")
            else:
                print("  已取消")
        input("\n按 Enter 返回主菜单...")

    def scan_and_process_projects(self):
        """Scan projects and process with confirmation (IFR sync only)."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)

        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        # Display projects with status
        self.display_projects_with_status(projects_by_region)

        # Get user selection
        selected = self.get_project_selection(projects_by_region)
        if not selected:
            return

        # Select operation
        operation = self.select_operation()
        if operation is None:
            return

        # IFC copy has its own flow
        if operation == 'ifc_copy':
            self.process_ifc_copy(selected)
            return

        # Process projects with confirmation
        self.process_projects_with_confirmation(selected, operation)

    def display_projects_with_status(self, projects_by_region: Dict[str, List[ProjectValidation]]):
        """Display projects grouped by region with status indicators."""
        total = sum(len(projects) for projects in projects_by_region.values())

        self.ui.print_header(f"检测到 {total} 个项目目录", "")

        all_projects = []
        project_idx = 1

        for region, projects in sorted(projects_by_region.items()):
            print(f"\n[DIR] 区域: {region}")
            self.ui.print_separator("-", 72)

            for project in projects:
                all_projects.append(project)

                # Status
                status_icon, status_text = self.ui.status_icon(
                    "ready" if project.recommended_action == "process" else
                    "confirm" if project.recommended_action == "confirm" else "legacy"
                )

                # Action
                action_text = self.ui.action_icon(project.recommended_action)

                # Source dirs count
                active_sources = sum(1 for s in project.source_dirs_found if s.exists and s.file_count > 0)

                # V4: IFC status
                if project.has_existing_ifc_client:
                    ifc_status = f"[v] {project.ifc_file_count} 个文件"
                else:
                    ifc_status = "[x] 不存在"

                print(f"\n  [{project_idx}] {project.project_name}")
                print(f"      结构状态: {status_icon} {status_text}")
                print(f"      建议操作: {action_text}")
                print(f"      源目录: {active_sources} 个有效")
                print(f"      IFC(Client): {ifc_status}")
                print(f"      最后修改: {project.last_modified}")

                if project.warning_message:
                    if COLORAMA_AVAILABLE:
                        print(f"      {Fore.YELLOW}警告: {project.warning_message}{Style.RESET_ALL}")
                    else:
                        print(f"      警告: {project.warning_message}")

                project_idx += 1

        # Summary
        print()
        self.ui.print_separator("=", 72)

        ready_projects = [p for projects in projects_by_region.values()
                        for p in projects if p.recommended_action == "process"]
        confirm_projects = [p for projects in projects_by_region.values()
                          for p in projects if p.recommended_action == "confirm"]
        skip_projects = [p for projects in projects_by_region.values()
                        for p in projects if p.recommended_action == "skip"]

        ready_nums = [str(i+1) for i, p in enumerate(all_projects) if p.recommended_action == "process"]
        confirm_nums = [str(i+1) for i, p in enumerate(all_projects) if p.recommended_action == "confirm"]
        skip_nums = [str(i+1) for i, p in enumerate(all_projects) if p.recommended_action == "skip"]

        print(f"\n自动推荐处理: 项目 #{', #'.join(ready_nums) if ready_nums else '无'} ({len(ready_projects)}个)")
        print(f"需要确认的项目: 项目 #{', #'.join(confirm_nums) if confirm_nums else '无'} ({len(confirm_projects)}个)")
        print(f"建议跳过的项目: 项目 #{', #'.join(skip_nums) if skip_nums else '无'} ({len(skip_projects)}个)")

        # V4: IFC summary
        ifc_projects = [p for projects in projects_by_region.values()
                       for p in projects if p.has_existing_ifc_client]
        if ifc_projects:
            ifc_nums = [str(i+1) for i, p in enumerate(all_projects) if p.has_existing_ifc_client]
            print(f"含 IFC(Client): 项目 #{', #'.join(ifc_nums)} ({len(ifc_projects)}个)")

        self.ui.print_separator("=", 72)

        return all_projects

    def get_project_selection(self, projects_by_region: Dict[str, List[ProjectValidation]]) -> List[ProjectValidation]:
        """Get user's project selection."""
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        ready_indices = [i for i, p in enumerate(all_projects) if p.recommended_action == "process"]

        options = [
            ("1", f"只处理推荐的项目 ({len(ready_indices)}个)"),
            ("2", "手动选择项目 (输入编号)"),
            ("3", "查看项目详细信息"),
            ("4", "处理所有项目（包括需要确认的）"),
            ("5", "生成验证报告"),
            ("0", "返回主菜单"),
        ]

        choice = self.ui.print_menu(options, "请选择")

        if choice == '0':
            return []
        elif choice == '1':
            return [all_projects[i] for i in ready_indices]
        elif choice == '2':
            print("\n输入项目编号（多选用逗号分隔，如: 1,3,4 或 1-5 或 'all'）：")
            selection = input("> ").strip()
            indices = self._parse_selection(selection, len(all_projects))
            return [all_projects[i-1] for i in indices]
        elif choice == '3':
            self._show_project_details(all_projects)
            return self.get_project_selection(projects_by_region)
        elif choice == '4':
            return [p for p in all_projects if p.recommended_action != "skip"]
        elif choice == '5':
            self._generate_and_save_report(projects_by_region)
            return self.get_project_selection(projects_by_region)

        return []

    def _parse_selection(self, selection: str, max_num: int) -> List[int]:
        """Parse user selection string."""
        indices = []
        selection = selection.strip().lower()

        if selection == 'all':
            return list(range(1, max_num + 1))

        parts = selection.replace(' ', '').split(',')
        for part in parts:
            if '-' in part:
                try:
                    start, end = part.split('-')
                    indices.extend(range(int(start), min(int(end) + 1, max_num + 1)))
                except:
                    pass
            else:
                try:
                    num = int(part)
                    if 1 <= num <= max_num:
                        indices.append(num)
                except:
                    pass

        return sorted(set(indices))

    def _show_project_details(self, projects: List[ProjectValidation]):
        """Show detailed information for a project."""
        print("\n输入项目编号查看详情（0 返回）：")
        choice = input("> ").strip()

        try:
            idx = int(choice)
            if idx == 0:
                return
            if 1 <= idx <= len(projects):
                p = projects[idx - 1]
                print()
                self.ui.print_separator("=", 72)
                print(f"[i] 项目: {p.project_name}")
                print(f"[DIR] 路径: {p.project_path}")
                print(f"[i] 区域: {p.region}")
                print(f"[i] 最后修改: {p.last_modified} ({p.days_since_modified} 天前)")
                print(f"\n[DIR] 源目录状态:")

                for source in p.source_dirs_found:
                    status = "[v]" if source.exists and source.file_count > 0 else "[x]"
                    if COLORAMA_AVAILABLE:
                        color = Fore.GREEN if source.exists and source.file_count > 0 else Fore.RED
                        print(f"   {color}{status}{Style.RESET_ALL} {source.name}", end="")
                    else:
                        print(f"   {status} {source.name}", end="")

                    if source.exists:
                        print(f" ({source.file_count} 个文件)")
                    else:
                        print(" (不存在)")

                # V4: Show IFC(Client) status
                print(f"\n[DIR] IFC(Client) 状态:")
                if p.has_existing_ifc_client:
                    ifc_path = Path(p.project_path) / self.validator.ifc_client_rel_path
                    if COLORAMA_AVAILABLE:
                        print(f"   {Fore.GREEN}[v]{Style.RESET_ALL} 存在 ({p.ifc_file_count} 个文件)")
                    else:
                        print(f"   [v] 存在 ({p.ifc_file_count} 个文件)")
                    print(f"   路径: {ifc_path}")
                else:
                    if COLORAMA_AVAILABLE:
                        print(f"   {Fore.RED}[x]{Style.RESET_ALL} 不存在")
                    else:
                        print(f"   [x] 不存在")

                print(f"\n[i] 预计操作:")
                print(f"   - 创建 {p.folders_to_create} 个文件夹")
                print(f"   - 复制 {p.drawings_count} 个图纸")
                print(f"   - 复制 {p.reports_count} 个报告")

                if p.warning_message:
                    print(f"\n[!] 警告: {p.warning_message}")

                self.ui.print_separator("=", 72)
                input("\n按 Enter 返回...")
        except:
            pass

    def _generate_and_save_report(self, projects_by_region: Dict[str, List[ProjectValidation]]):
        """Generate and save validation report."""
        report = self.report_generator.generate_report(
            projects_by_region,
            str(self.root_path)
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = Path(__file__).parent / "logs" / f"validation_report_{timestamp}.txt"
        report_path.parent.mkdir(exist_ok=True)

        self.report_generator.save_report(report, report_path)
        self.ui.print_success(f"报告已保存: {report_path}")

        if self.ui.confirm("是否打开报告？"):
            if os.name == 'nt':
                os.startfile(report_path)
            else:
                os.system(f'open "{report_path}"')

    def select_operation(self) -> Optional[str]:
        """Let user select operation type."""
        print("\n选择要执行的操作：")
        options = [
            ("1", "仅创建文件夹"),
            ("2", "仅镜像文件"),
            ("3", "全部执行 (推荐)"),
            ("4", "预览模式 (dry-run)"),
            ("5", "复制 IFC(Client) 到目标文件夹"),
            ("0", "返回"),
        ]

        choice = self.ui.print_menu(options, "请选择操作")

        operation_map = {
            '1': 'create_folders',
            '2': 'mirror_files',
            '3': 'full',
            '4': 'dry_run',
            '5': 'ifc_copy',
            '0': None
        }

        return operation_map.get(choice)

    # =========================================================================
    # V4: IFC(Client) Copy Mode
    # =========================================================================

    def ifc_copy_mode(self):
        """Standalone IFC(Client) copy mode from main menu."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)

        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        # Filter to only show projects that have IFC(Client)
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        ifc_projects = [p for p in all_projects if p.has_existing_ifc_client]

        if not ifc_projects:
            self.ui.print_warning("没有找到包含 4. IFC(Client) 文件夹的项目！")
            input("\n按 Enter 返回...")
            return

        # Display IFC projects
        self.ui.print_header(
            f"IFC(Client) 复制 - 找到 {len(ifc_projects)} 个项目",
            "选择要复制 IFC(Client) 的项目"
        )

        for idx, p in enumerate(ifc_projects, 1):
            print(f"  [{idx}] {p.project_name}")
            print(f"      IFC(Client): {p.ifc_file_count} 个文件")
            print(f"      路径: {p.project_path}")
            print()

        self.ui.print_separator()

        # Select projects
        print("\n输入项目编号（多选用逗号分隔，如: 1,3,4 或 1-5 或 'all'，0 返回）：")
        selection = input("> ").strip()

        if selection == '0':
            return

        indices = self._parse_selection(selection, len(ifc_projects))
        if not indices:
            self.ui.print_warning("未选择任何项目")
            return

        selected = [ifc_projects[i-1] for i in indices]
        self.process_ifc_copy(selected)

    def process_ifc_copy(self, projects: List[ProjectValidation]):
        """Process IFC(Client) copy for selected projects.

        For each project, ask user for target folder, then copy.
        """
        # Filter projects that actually have IFC(Client)
        valid_projects = []
        for p in projects:
            if p.has_existing_ifc_client:
                valid_projects.append(p)
            else:
                self.ui.print_warning(f"跳过 {p.project_name}: 没有 4. IFC(Client) 文件夹")

        if not valid_projects:
            self.ui.print_warning("没有可复制的项目（所有项目均缺少 IFC(Client) 文件夹）")
            input("\n按 Enter 返回...")
            return

        results = []

        for idx, project in enumerate(valid_projects, 1):
            print(f"\n{'='*72}")
            print(f"IFC(Client) 复制 - 项目 {idx}/{len(valid_projects)}")
            print(f"{'='*72}\n")

            print(f"[i] 项目: {project.project_name}")
            ifc_source = Path(project.project_path) / self.validator.ifc_client_rel_path
            print(f"[DIR] IFC 源目录: {ifc_source}")
            print(f"[i] 文件数量: {project.ifc_file_count} 个")

            # Ask for target path
            print(f"\n请输入目标文件夹路径（可拖拽文件夹，输入 0 跳过此项目）：")
            target_str = input("> ").strip().strip('"')

            if not target_str or target_str == '0':
                results.append(ProcessResult(
                    project_name=project.project_name,
                    project_path=project.project_path,
                    success=False,
                    skipped=True,
                    skip_reason="用户跳过"
                ))
                print("[->] 跳过此项目")
                continue

            target_path = Path(target_str)

            # Validate target path (parent must exist, or we create it)
            if not target_path.parent.exists():
                self.ui.print_error(f"目标路径的父目录不存在: {target_path.parent}")
                results.append(ProcessResult(
                    project_name=project.project_name,
                    project_path=project.project_path,
                    success=False,
                    errors=[f"目标路径父目录不存在: {target_path.parent}"]
                ))
                continue

            # Show confirmation
            print(f"\n[i] 操作确认:")
            print(f"   源: {ifc_source}")
            print(f"   目标: {target_path}")
            print(f"   文件数: {project.ifc_file_count}")
            print(f"   方式: 复制 (copy & paste)")

            if target_path.exists():
                self.ui.print_warning("目标文件夹已存在，同名文件将根据大小/日期决定是否覆盖")

            if not self.ui.confirm("\n确认开始复制？"):
                results.append(ProcessResult(
                    project_name=project.project_name,
                    project_path=project.project_path,
                    success=False,
                    skipped=True,
                    skip_reason="用户跳过"
                ))
                print("[->] 跳过此项目")
                continue

            # Execute copy
            automation = IFRAutomation(
                root_path=str(Path(project.project_path).parent),
                config=self.config,
                interactive=True
            )

            print("\n正在复制...")
            ifc_result = automation.copy_ifc_to_target(
                Path(project.project_path),
                target_path
            )

            result = ProcessResult(
                project_name=project.project_name,
                project_path=project.project_path,
                success=len(ifc_result["errors"]) == 0,
                ifc_files_copied=ifc_result["copied"],
                ifc_files_skipped=ifc_result["skipped"],
                ifc_target_path=str(target_path),
                errors=[str(e) for e in ifc_result["errors"]]
            )
            results.append(result)

            if result.success:
                self.ui.print_success(
                    f"完成: {project.project_name} -> "
                    f"复制 {ifc_result['copied']} 个文件, "
                    f"跳过 {ifc_result['skipped']} 个 (已存在)"
                )
            else:
                self.ui.print_warning(
                    f"部分完成: {project.project_name} -> "
                    f"复制 {ifc_result['copied']} 个, "
                    f"跳过 {ifc_result['skipped']} 个, "
                    f"错误 {len(ifc_result['errors'])} 个"
                )
                for err in ifc_result["errors"][:5]:
                    print(f"   错误: {err}")

        # Show IFC copy summary
        self.show_ifc_copy_summary(results)

    def show_ifc_copy_summary(self, results: List[ProcessResult]):
        """Show IFC copy completion summary."""
        print()
        self.ui.print_separator("=", 72)
        self.ui.print_header("IFC(Client) 复制完成", "")

        success = sum(1 for r in results if r.success)
        skipped = sum(1 for r in results if r.skipped)
        failed = len(results) - success - skipped

        total_copied = sum(r.ifc_files_copied for r in results)
        total_skipped_files = sum(r.ifc_files_skipped for r in results)

        print(f"  项目统计:")
        print(f"    成功: {success}")
        print(f"    跳过: {skipped}")
        print(f"    失败: {failed}")

        print(f"\n  文件统计:")
        print(f"    复制: {total_copied} 个文件")
        print(f"    跳过 (已存在): {total_skipped_files} 个文件")

        if success > 0:
            print(f"\n  复制详情:")
            for r in results:
                if r.success:
                    print(f"    {r.project_name}")
                    print(f"      -> {r.ifc_target_path}")
                    print(f"      复制: {r.ifc_files_copied}, 跳过: {r.ifc_files_skipped}")

        self.ui.print_separator()
        input("\n按 Enter 返回...")

    def ifc_convert_mode(self):
        """IFR → IFC 转换模式 (AutoCAD COM)."""
        if not WIN32COM_AVAILABLE:
            self.ui.print_error("需要安装 pywin32 才能使用此功能 (pip install pywin32)")
            input("\n按 Enter 返回...")
            return

        root = self.get_root_path()
        if not root:
            return

        # Scan projects
        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)
        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        all_projects = [p for projects in projects_by_region.values() for p in projects]

        # Filter to projects that have Native drawings folder
        native_projects = []
        for p in all_projects:
            native_dir = Path(p.project_path) / IFCManager.NATIVE_ROOT
            if native_dir.exists():
                native_projects.append(p)

        if not native_projects:
            self.ui.print_warning("没有找到包含 1. Native 文件夹的项目！")
            input("\n按 Enter 返回...")
            return

        # Display and select project
        self.ui.print_header(
            f"IFR → IFC 转换 - 找到 {len(native_projects)} 个项目",
            "选择要转换的项目"
        )

        for idx, p in enumerate(native_projects, 1):
            print(f"  [{idx}] {p.project_name}")
            print(f"      路径: {p.project_path}")

        print(f"\n  [0] 返回主菜单")
        self.ui.print_separator()

        choice = input(f"\n请选择项目 [0-{len(native_projects)}]: ").strip()
        if choice == '0':
            return

        try:
            idx = int(choice)
            if not (1 <= idx <= len(native_projects)):
                return
        except ValueError:
            return

        project = native_projects[idx - 1]
        project_path = Path(project.project_path)

        # Scan native folders for IFR DWGs
        print(f"\n正在扫描 Native 文件夹: {project.project_name}")
        mgr = IFCManager(project_path, dry_run=True)
        scan_results = mgr.scan_native_folders()

        if not scan_results:
            self.ui.print_warning("未找到可转换的 IFR DWG 文件")
            input("\n按 Enter 返回...")
            return

        # Display scan results
        print(f"\n{'='*60}")
        print(f"  找到 {len(scan_results)} 个 doc-ID 的 IFR DWG 文件")
        print(f"{'='*60}")

        for i, info in enumerate(scan_results, 1):
            ifc_status = f"已有 IFC Rev{info['existing_ifc_rev']}" if info['existing_ifc_rev'] is not None else "无 IFC"
            print(f"  [{i:3d}] {info['doc_id']}")
            print(f"        IFR: {info['latest_ifr_dwg'].name} (Rev{info['latest_ifr_rev']})")
            print(f"        IFC: {ifc_status}")

        self.ui.print_separator()

        # Preview (dry-run)
        print(f"\n  预览转换结果（dry-run）:")
        preview_results = []
        for info in scan_results:
            r = mgr.convert_to_ifc(info)
            preview_results.append(r)
            ifc_rev = r['ifc_rev']
            dwg_name = Path(r['dwg_path']).name if r['dwg_path'] else '?'
            pdf_name = Path(r['pdf_path']).name if r['pdf_path'] else '?'
            print(f"    {info['doc_id']}: → IFC Rev{ifc_rev}")
            print(f"      DWG: {dwg_name}")
            print(f"      PDF: {pdf_name}")

        self.ui.print_separator()

        # Confirm
        print("\n选择操作:")
        print("  [1] 执行转换（批量全部转）")
        print("  [2] 预览模式（仅显示，不执行）")
        print("  [0] 返回（不执行）")

        action = input("\n请选择 [0-2]: ").strip()

        if action == '0':
            return
        elif action == '2':
            self.ui.print_success("预览完成，未执行任何操作")
            input("\n按 Enter 返回...")
            return

        if action != '1':
            return

        # Execute conversion
        print(f"\n正在启动 AutoCAD 转换...")
        real_mgr = IFCManager(project_path, dry_run=False)
        results = real_mgr.batch_convert()

        # Summary
        ok_count = sum(1 for r in results if r['success'])
        fail_count = len(results) - ok_count

        print(f"\n{'='*60}")
        print(f"  IFC 转换完成")
        print(f"  成功: {ok_count}, 失败: {fail_count}")
        print(f"{'='*60}")

        if fail_count > 0:
            print("\n  失败详情:")
            for r in results:
                if not r['success']:
                    print(f"    {r['doc_id']}: {'; '.join(r['errors'])}")

        # Trigger deliverable sync using IFCManager's dedicated method
        if ok_count > 0:
            print("\n正在更新交付物状态...")
            real_mgr.update_ifc_deliverables(results)

        input("\n按 Enter 返回...")

    def panel_ifc_convert_mode(self):
        """Panel IFC 批量转换模式 (多页 DWG → 合并 PDF)."""
        if not WIN32COM_AVAILABLE:
            self.ui.print_error("需要安装 pywin32 才能使用此功能 (pip install pywin32)")
            input("\n按 Enter 返回...")
            return

        self.ui.print_header(
            "Panel IFC 批量转换",
            "多页 DWG Title Block 更新 + 合并 PDF 导出"
        )

        # 1. Get source folder
        print("请输入 Panel Design 文件夹路径 (包含多个 DWG 的平面目录):")
        print("  示例: .../1. Native/STS1&2 Panel Design/CAD/Project-TSF")
        folder_str = input("\n> ").strip().strip('"')
        if not folder_str or not Path(folder_str).exists():
            self.ui.print_warning("路径不存在")
            input("\n按 Enter 返回...")
            return

        source_folder = Path(folder_str)

        # 2. Scan and group
        print(f"\n正在扫描: {source_folder.name}")
        preview_mgr = PanelIFCManager(source_folder, dry_run=True)
        groups = preview_mgr.scan_and_group()

        if not groups:
            self.ui.print_warning("未找到可分组的 DWG 文件 (需要 {DocNO}-{PageNum}.dwg 格式)")
            input("\n按 Enter 返回...")
            return

        # Display groups
        print(f"\n{'='*60}")
        print(f"  找到 {len(groups)} 个文档组:")
        print(f"{'='*60}")

        for doc_no, pages in groups.items():
            ifc_rev = preview_mgr._get_existing_ifc_rev(doc_no)
            ifc_label = f"Rev{ifc_rev}" if ifc_rev == 0 else f"Rev{ifc_rev} (已有 Rev{ifc_rev-1})"
            print(f"\n  [{doc_no}] {len(pages)} 页 -> IFC {ifc_label}")
            for p in pages:
                page_type = ""
                if p['page'] == '00':
                    page_type = " (Cover)"
                elif p['page'].startswith('0') and not p['page'].isdigit():
                    page_type = " (TOC/Appendix)"
                print(f"    {p['filename']}{page_type}")
            pdf_name = f"{doc_no}_Rev{ifc_rev}_IFC.pdf"
            print(f"    -> PDF: {pdf_name}")

        self.ui.print_separator()

        # 3. Title block update method — default COM API (auto, no user prompt)
        tb_method = 'com'
        utb_lsp_path = None

        # 4. Dry-run preview
        print(f"\n{'='*60}")
        print(f"  预览 (dry-run) — 使用 COM API 模式")
        print(f"{'='*60}")

        dry_mgr = PanelIFCManager(
            source_folder, dry_run=True,
            title_block_method=tb_method,
            utb_lsp_path=utb_lsp_path,
        )
        dry_mgr.batch_convert_panel()

        # 5. Confirm
        print("\n选择操作:")
        print("  [1] 执行转换")
        print("  [0] 返回（不执行）")
        action = input("\n请选择 [0-1]: ").strip()

        if action != '1':
            return

        # 6. Execute
        print(f"\n正在执行 IFC 转换...")
        real_mgr = PanelIFCManager(
            source_folder, dry_run=False,
            title_block_method=tb_method,
            utb_lsp_path=utb_lsp_path,
        )
        results = real_mgr.batch_convert_panel()

        # 7. Summary
        total_groups = len(results)
        ok_groups = sum(1 for r in results if r.get('pdf_result', {}).get('success'))
        total_dwgs = sum(r['dwgs_updated'] for r in results)

        print(f"\n{'='*60}")
        print(f"  Panel IFC 转换完成")
        print(f"  文档组: {ok_groups}/{total_groups} 成功")
        print(f"  DWG 更新: {total_dwgs} 个")
        print(f"{'='*60}")

        if any(r['errors'] for r in results):
            print("\n  错误详情:")
            for r in results:
                for err in r['errors']:
                    print(f"    [{r['doc_no']}] {err}")

        input("\n按 Enter 返回...")

    def process_projects_with_confirmation(self, projects: List[ProjectValidation], operation: str):
        """Process projects with per-project confirmation."""
        dry_run = operation == 'dry_run'
        create_folders_only = operation == 'create_folders'
        mirror_files_only = operation == 'mirror_files'

        results = []

        for idx, project in enumerate(projects, 1):
            print(f"\n{'='*72}")
            print(f"正在处理项目 {idx}/{len(projects)}")
            print(f"{'='*72}\n")

            # Display project details
            print(f"[i] 项目: {project.project_name}")
            print(f"[DIR] 路径: {project.project_path}")
            print(f"[i] 最后修改: {project.last_modified}")

            # Safety check
            safety = self.safety_checker.check_before_process(project)

            if not safety["safe"]:
                print(f"\n[!] 安全检查警告:")
                for warning in safety["warnings"]:
                    if COLORAMA_AVAILABLE:
                        print(f"   {Fore.YELLOW}- {warning}{Style.RESET_ALL}")
                    else:
                        print(f"   - {warning}")

            # Show operation preview
            print(f"\n[i] 操作预览:")
            if not mirror_files_only:
                print(f"   - 创建文件夹: {project.folders_to_create} 个")
            if not create_folders_only:
                print(f"   - 复制图纸: {project.drawings_count} 个文件")
                print(f"   - 复制报告: {project.reports_count} 个文件")

            if dry_run:
                print(f"\n   [i] 预览模式 - 不会执行实际操作")

            # Confirm if needed
            if project.recommended_action != "process" or not safety["safe"]:
                if not self.ui.confirm("\n是否继续处理此项目？", default=False):
                    results.append(ProcessResult(
                        project_name=project.project_name,
                        project_path=project.project_path,
                        success=False,
                        skipped=True,
                        skip_reason="用户跳过"
                    ))
                    print("[->] 跳过此项目")
                    continue
            else:
                if not self.ui.confirm("\n开始处理？"):
                    results.append(ProcessResult(
                        project_name=project.project_name,
                        project_path=project.project_path,
                        success=False,
                        skipped=True,
                        skip_reason="用户跳过"
                    ))
                    print("[->] 跳过此项目")
                    continue

            # Execute
            automation = IFRAutomation(
                root_path=str(Path(project.project_path).parent),
                config=self.config,
                dry_run=dry_run,
                create_folders_only=create_folders_only,
                mirror_files_only=mirror_files_only,
                interactive=True
            )

            result = automation.process_project(project)
            results.append(result)

            if result.success:
                self.ui.print_success(f"完成: {project.project_name}")
                print(f"   文件夹: {result.folders_created}, 图纸: {result.drawings_copied}, 报告: {result.reports_copied}")
                self.config.add_recent_project(project.project_path)
            else:
                self.ui.print_error(f"失败: {project.project_name}")
                for error in result.errors:
                    print(f"   错误: {error}")

        # Show summary
        self.show_completion_summary(results, dry_run)

    def show_completion_summary(self, results: List[ProcessResult], dry_run: bool):
        """Show processing summary."""
        print()
        self.ui.print_separator("=", 72)

        mode = "预览模式" if dry_run else "执行完成"
        self.ui.print_header(f"操作{mode}", "")

        success = sum(1 for r in results if r.success)
        skipped = sum(1 for r in results if r.skipped)
        failed = len(results) - success - skipped

        total_folders = sum(r.folders_created for r in results)
        total_drawings = sum(r.drawings_copied for r in results)
        total_reports = sum(r.reports_copied for r in results)

        print(f"  项目统计:")
        print(f"    成功: {success}")
        print(f"    跳过: {skipped}")
        print(f"    失败: {failed}")

        print(f"\n  操作统计:")
        print(f"    文件夹创建: {total_folders}")
        print(f"    图纸复制: {total_drawings}")
        print(f"    报告复制: {total_reports}")

        self.ui.print_separator()

        options = [
            ("1", "打开日志文件"),
            ("2", "返回主菜单"),
            ("0", "退出"),
        ]

        choice = self.ui.print_menu(options, "请选择")

        if choice == '1':
            self.view_logs()
        elif choice == '0':
            sys.exit(0)

    def validate_only_mode(self):
        """Validate projects without processing."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描和验证项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)

        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        # Display results
        self.display_projects_with_status(projects_by_region)

        # Generate report option
        if self.ui.confirm("\n是否生成验证报告？"):
            self._generate_and_save_report(projects_by_region)

    def input_single_project(self):
        """Handle single project input."""
        print("\n请输入项目路径：")
        print("（可以直接拖拽文件夹到此窗口）")

        path_str = input("> ").strip().strip('"')

        if not path_str:
            return

        project_path = Path(path_str)

        if not project_path.exists():
            self.ui.print_error("路径不存在！")
            return

        # Validate
        validation = self.validator.validate_project(project_path)

        print()
        self.ui.print_separator()
        print(f"[i] 项目: {validation.project_name}")
        print(f"[i] 状态: {self.ui.status_icon(validation.recommended_action)[1]}")
        print(f"[i] 建议: {self.ui.action_icon(validation.recommended_action)}")

        if validation.has_existing_ifc_client:
            print(f"[i] IFC(Client): {validation.ifc_file_count} 个文件")

        if validation.warning_message:
            self.ui.print_warning(validation.warning_message)

        self.ui.print_separator()

        if not self.ui.confirm("\n是否处理此项目？"):
            return

        operation = self.select_operation()
        if operation is None:
            return

        if operation == 'ifc_copy':
            self.process_ifc_copy([validation])
        else:
            self.process_projects_with_confirmation([validation], operation)

    def modify_config(self):
        """Modify configuration and view logs."""
        while True:
            print("\n当前配置：")
            print(f"  1. 默认根目录: {self.config.get('default_root_path', '未设置')}")
            print(f"  2. 日志级别: {self.config.get('log_level', 'INFO')}")
            print(f"  3. 白名单: {self.config.get('project_filters.whitelist', [])}")
            print(f"  4. 黑名单: {self.config.get('project_filters.blacklist', [])}")
            print(f"  5. 最少源目录数: {self.config.get('validation_rules.min_source_dirs', 1)}")
            print(f"  6. 最大未修改天数: {self.config.get('validation_rules.max_days_since_modified', 180)}")
            print(f"  7. IFC(Client) 路径: {self.config.get('ifc_client_path', ProjectValidator.IFC_CLIENT_REL_PATH)}")
            print(f"  8. 查看日志")
            print(f"  0. 返回主菜单")

            choice = input("\n选择配置项 [0-8]: ").strip()

            if choice == '0':
                break
            elif choice == '1':
                new_path = input("输入新的根目录路径: ").strip().strip('"')
                if new_path and Path(new_path).exists():
                    self.config.set("default_root_path", new_path)
                    self.root_path = Path(new_path)
                    self.ui.print_success("已更新根目录")
                else:
                    self.ui.print_error("路径无效")
            elif choice == '2':
                levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR']
                print("可选级别: " + ", ".join(levels))
                level = input("输入日志级别: ").strip().upper()
                if level in levels:
                    self.config.set("log_level", level)
                    self.ui.print_success("已更新日志级别")
            elif choice == '3':
                print("输入白名单项目（逗号分隔，支持通配符如 NSW*）：")
                whitelist = input("> ").strip()
                if whitelist:
                    items = [x.strip() for x in whitelist.split(',')]
                    self.config.set("project_filters.whitelist", items)
                else:
                    self.config.set("project_filters.whitelist", [])
                self.ui.print_success("已更新白名单")
            elif choice == '4':
                print("输入黑名单项目（逗号分隔，支持通配符如 Old_*）：")
                blacklist = input("> ").strip()
                if blacklist:
                    items = [x.strip() for x in blacklist.split(',')]
                    self.config.set("project_filters.blacklist", items)
                else:
                    self.config.set("project_filters.blacklist", [])
                self.ui.print_success("已更新黑名单")
            elif choice == '5':
                try:
                    num = int(input("输入最少源目录数 (1-4): ").strip())
                    if 1 <= num <= 4:
                        self.config.set("validation_rules.min_source_dirs", num)
                        self.ui.print_success("已更新")
                except:
                    self.ui.print_error("无效输入")
            elif choice == '6':
                try:
                    num = int(input("输入最大未修改天数: ").strip())
                    if num > 0:
                        self.config.set("validation_rules.max_days_since_modified", num)
                        self.ui.print_success("已更新")
                except:
                    self.ui.print_error("无效输入")
            elif choice == '7':
                print(f"当前 IFC(Client) 路径: {self.config.get('ifc_client_path', ProjectValidator.IFC_CLIENT_REL_PATH)}")
                print("输入新的相对路径（相对于项目根目录，留空保持默认）：")
                new_path = input("> ").strip()
                if new_path:
                    self.config.set("ifc_client_path", new_path)
                    self.validator = ProjectValidator(self.config)
                    self.scanner = ProjectScanner(self.config, self.validator)
                    self.ui.print_success("已更新 IFC(Client) 路径")
            elif choice == '8':
                self.view_logs()

    def view_logs(self):
        """View log files."""
        log_dir = Path(__file__).parent / "logs"

        if not log_dir.exists():
            self.ui.print_warning("日志目录不存在")
            return

        log_files = sorted(log_dir.glob("ifr_automation_*.log"), reverse=True)

        if not log_files:
            self.ui.print_warning("没有找到日志文件")
            return

        print("\n最近的日志文件：")
        for i, log_file in enumerate(log_files[:5], 1):
            mtime = datetime.fromtimestamp(log_file.stat().st_mtime)
            print(f"  {i}. {log_file.name} ({mtime.strftime('%Y-%m-%d %H:%M')})")

        print("\n选择要查看的日志（输入序号，或 0 返回）：")
        choice = input("> ").strip()

        try:
            idx = int(choice)
            if 1 <= idx <= len(log_files[:5]):
                log_file = log_files[idx - 1]
                if os.name == 'nt':
                    os.startfile(log_file)
                else:
                    os.system(f'open "{log_file}"')
        except:
            pass


# =============================================================================
# Command Line Interface
# =============================================================================

def main():
    """Main entry point."""

    # Interactive mode if no arguments
    if len(sys.argv) == 1:
        config = ConfigManager()
        interactive = UnifiedInteractiveMode(config)
        interactive.run()
        return

    # Command line mode
    parser = argparse.ArgumentParser(
        description="Engineering Pipeline v7.0 - IFR Sync + Version Mgmt + Sharepoint Sync + Deliverable",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode
  python ifr_automation_v7.py

  # Full pipeline (IFR sync + version mgmt + deliverable)
  python ifr_automation_v7.py --pipeline --root "..."

  # Specific stages only
  python ifr_automation_v7.py --stages ifr_sync version_mgmt --root "..."

  # Deliverable cross-check only (dry-run)
  python ifr_automation_v7.py --deliverable-only --dry-run --root "..."

  # Deliverable check only (report, no update)
  python ifr_automation_v7.py --deliverable-check-only --root "..."

  # IFR sync only
  python ifr_automation_v7.py --root "..." --auto-safe-only

  # Version management (PDF)
  python ifr_automation_v7.py --version-mgmt --root "..."

  # Native/Reports/Schedule version management
  python ifr_automation_v7.py --native --root "..." --project "GG-31*"
        """
    )

    parser.add_argument('--root', help='Root directory containing project folders')
    parser.add_argument('--dry-run', action='store_true', help='Preview mode (no changes)')
    parser.add_argument('--create-folders', action='store_true', help='Only create folder structure')
    parser.add_argument('--mirror-files', action='store_true', help='Only mirror files')
    parser.add_argument('--project', help='Process specific project (partial match)')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'], default='INFO')
    parser.add_argument('--output-json', action='store_true', help='Generate JSON report')

    # V3 arguments
    parser.add_argument('--validate-only', action='store_true', help='Only validate, no processing')
    parser.add_argument('--whitelist', nargs='+', help='Only process matching projects')
    parser.add_argument('--blacklist', nargs='+', help='Exclude matching projects')
    parser.add_argument('--auto-safe-only', action='store_true', help='Auto-process only safe projects')
    parser.add_argument('--yes', '-y', action='store_true', dest='yes_to_all', help='Yes to all prompts')
    parser.add_argument('--export-report', type=str, metavar='PATH', help='Export validation report')

    # V4 arguments
    parser.add_argument('--ifc-copy-target', type=str, metavar='PATH',
                       help='Copy IFC(Client) to target folder')

    # V7 arguments
    parser.add_argument('--pipeline', action='store_true',
                       help='Run full pipeline (IFR sync + version mgmt + deliverable)')
    parser.add_argument('--stages', nargs='+',
                       choices=['ifr_sync', 'version_mgmt', 'sharepoint_sync', 'deliverable'],
                       help='Run specific pipeline stages')
    parser.add_argument('--deliverable-only', action='store_true',
                       help='Run deliverable cross-check and update')
    parser.add_argument('--deliverable-check-only', action='store_true',
                       help='Run deliverable cross-check (report only, no update)')
    parser.add_argument('--version-mgmt', action='store_true',
                       help='Run PDF version management')
    parser.add_argument('--native', action='store_true',
                       help='Run Native/Reports/Schedule version management')
    parser.add_argument('--scope', choices=['native', 'reports', 'schedule', 'all'],
                       default='all', help='Scope for --native mode')
    parser.add_argument('--folder', type=str, help='Folder filter for --native mode')
    parser.add_argument('--execute', action='store_true',
                       help='Execute operations (for --native/--version-mgmt)')

    args = parser.parse_args()

    # Load config
    config = ConfigManager()

    if args.whitelist:
        config.set("project_filters.whitelist", args.whitelist)
    if args.blacklist:
        config.set("project_filters.blacklist", args.blacklist)

    root_path = args.root or config.get("default_root_path")
    if not root_path:
        print("Error: No root path specified. Use --root or set in config.")
        sys.exit(1)
    if not Path(root_path).exists():
        print(f"Error: Root path does not exist: {root_path}")
        sys.exit(1)

    # V7: Native mode (from version_manager_v4)
    if args.native:
        project_path = args.project
        if not project_path:
            # Try to find project from recent
            cfg_file = Path(__file__).parent / 'config.json'
            if cfg_file.exists():
                try:
                    with open(cfg_file, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    recent = cfg.get('recent_projects', [])
                    if recent:
                        project_path = recent[0]
                except Exception:
                    pass
        if not project_path:
            print("Error: --native requires --project")
            sys.exit(1)
        project_path = Path(project_path)
        if not project_path.exists():
            print(f"Error: Path not found: {project_path}")
            sys.exit(1)

        dry_run = not args.execute
        mgr = NativeVersionManager(str(project_path), dry_run=dry_run)
        results = mgr.process_all(folder_filter=args.folder, scope=args.scope)
        if not results:
            print("No document folders found")
            sys.exit(0)
        total_actions = sum(len(r.actions) for _, r in results)
        for group_name, group_results in groupby(results, key=lambda x: x[0]):
            group_list = list(group_results)
            print(f"\n  === {group_name} ===")
            for _, result in group_list:
                if not result.dwg_files:
                    continue
                print(f"    {result.folder_name} ({len(result.dwg_files)} files)")
                for action in result.actions:
                    if action.action == 'rename':
                        print(f"      [重命名] {action.source.name} -> {action.dest.name}")
                    else:
                        print(f"      [->SS] {action.source.name}")
        if not dry_run and total_actions > 0:
            stats = mgr.execute_actions(results)
            print(f"\n  Done: {stats['renamed']} renamed, {stats['moved']} moved")
        elif total_actions > 0:
            print(f"\n  Use --execute to apply {total_actions} actions")
        sys.exit(0)

    # V7: Version management mode
    if args.version_mgmt:
        vm = VersionManager(root_path, dry_run=not args.execute)
        projects = vm.scan_for_projects()
        if args.project:
            projects = [p for p in projects if args.project.lower() in p.name.lower()]
        for project in projects:
            print(f"\nProcessing: {project.name}")
            vm.process_project(project)
        sys.exit(0)

    # V7: Deliverable modes
    if args.deliverable_only or args.deliverable_check_only:
        automation = IFRAutomation(root_path=root_path, config=config, interactive=True)
        projects_by_region = automation.scan_projects()
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        if args.project:
            all_projects = [p for p in all_projects if args.project.lower() in p.project_name.lower()]
        tg_lines = []
        for project in all_projects:
            dm = DeliverableManager(Path(project.project_path), dry_run=args.dry_run)
            excel = dm.find_deliverable_excel()
            if not excel:
                print(f"[SKIP] {project.project_name}: No deliverable Excel found")
                continue
            check = dm.cross_check(excel)
            print(f"\n{project.project_name}: "
                  f"new={len(check.items_in_folders_not_excel)}, "
                  f"rev_mismatch={len(check.revision_mismatches)}")
            if check.errors:
                for e in check.errors:
                    print(f"  ERROR: {e}")
            if args.deliverable_only and not args.dry_run:
                if check.items_in_folders_not_excel or check.revision_mismatches:
                    updated = dm.apply_updates(excel, check)
                    print(f"  Updated: inserted={updated.rows_inserted}, "
                          f"updated={updated.rows_updated}, new_rev={updated.new_file_rev}")
                    tg_lines.append(
                        f"📋 <b>{project.project_name}</b>\n"
                        f"  新增={len(check.items_in_folders_not_excel)}, "
                        f"更新={len(check.revision_mismatches)}, "
                        f"inserted={updated.rows_inserted}, updated={updated.rows_updated}")
        if tg_lines:
            send_telegram_notification("📋 <b>Deliverable Update</b>\n\n" + "\n\n".join(tg_lines))
        sys.exit(0)

    # V7: Pipeline mode
    if args.pipeline or args.stages:
        stages = args.stages or ['ifr_sync', 'version_mgmt', 'sharepoint_sync', 'deliverable']
        automation = IFRAutomation(root_path=root_path, config=config, interactive=True)
        projects_by_region = automation.scan_projects()
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        if args.auto_safe_only:
            all_projects = [p for p in all_projects if p.recommended_action == "process"]
        if args.project:
            all_projects = [p for p in all_projects if args.project.lower() in p.project_name.lower()]
        pipeline = PipelineOrchestrator(config, dry_run=args.dry_run, stages=stages)
        all_results = []
        for project in all_projects:
            result = pipeline.run_pipeline(Path(project.project_path), project)
            all_results.append(result)
            status = "OK" if result['success'] else "FAIL"
            print(f"[{status}] {project.project_name}")
        # Send Telegram notification
        if all_results and not args.dry_run:
            tg_lines = [format_pipeline_result(r) for r in all_results]
            send_telegram_notification("\n\n".join(tg_lines))
        sys.exit(0)

    # Default: IFR sync mode (v6 behavior)
    automation = IFRAutomation(
        root_path=root_path, config=config,
        dry_run=args.dry_run, create_folders_only=args.create_folders,
        mirror_files_only=args.mirror_files, validate_only=args.validate_only,
        yes_to_all=args.yes_to_all, auto_safe_only=args.auto_safe_only,
        log_level=args.log_level, output_json=args.output_json
    )

    projects_by_region = automation.scan_projects()
    if not projects_by_region:
        print("No projects found!")
        sys.exit(0)

    if args.export_report or args.validate_only:
        report = automation.report_generator.generate_report(projects_by_region, root_path)
        print(report)
        if args.export_report:
            automation.report_generator.save_report(report, Path(args.export_report))
        if args.validate_only:
            sys.exit(0)

    all_projects = [p for projects in projects_by_region.values() for p in projects]
    if args.auto_safe_only:
        all_projects = [p for p in all_projects if p.recommended_action == "process"]
    if args.project:
        all_projects = [p for p in all_projects if args.project.lower() in p.project_name.lower()]

    if args.ifc_copy_target:
        target = Path(args.ifc_copy_target)
        for project in all_projects:
            if not project.has_existing_ifc_client:
                continue
            ifc_result = automation.copy_ifc_to_target(Path(project.project_path), target)
            status = "OK" if len(ifc_result["errors"]) == 0 else "WARN"
            print(f"[{status}] {project.project_name}: copied={ifc_result['copied']}")
        sys.exit(0)

    for project in all_projects:
        if not args.yes_to_all and project.recommended_action != "process":
            response = input(f"Process {project.project_name}? [y/N]: ").strip().lower()
            if response not in ('y', 'yes'):
                continue
        result = automation.process_project(project)
        status = "OK" if result.success else "FAIL"
        print(f"[{status}] {project.project_name}: folders={result.folders_created}, "
              f"drawings={result.drawings_copied}, reports={result.reports_copied}")


def keep_window_open():
    """Keep console window open."""
    print()
    input("按 Enter 键退出...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n用户中断操作")
    except Exception as e:
        print(f"\n发生错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if len(sys.argv) == 1:
            keep_window_open()
