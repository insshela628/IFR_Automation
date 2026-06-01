#!/usr/bin/env python3
"""
Engineering Pipeline - IFR Sync + Version Management + Deliverable Cross-Check
工程文档自动化管线工具 v7.0

v7.0 New Features (merged pipeline):
    - Merged ifr_automation_v6 + version_manager_v4 into single script
    - Deliverable cross-check: auto-detect Excel layout, compare files vs Excel,
      auto-insert new items, auto-update revisions, highlight changes
    - Pipeline orchestrator: sequential IFR Sync → Version Mgmt → Deliverable per project
    - Unified interactive menu with 10 options
    - CLI: --pipeline, --deliverable-only, --deliverable-check-only, --stages

v6.0 Features:
    - IFC/IFR file classification: prevents IFC files from IFR(Client) directories
v5.0 Features:
    - Folder-based file classification fallback (any naming convention)
v4.0 Features:
    - IFC(Client) folder copy to external target
    - Version manager: PDF old version cleanup, Native/Reports/Schedule management
    - Folder relocator: detect/fix misplaced structural folders
v3.0 Features:
    - Smart project validation, hierarchical scanning, multi-source directories
    - Whitelist/Blacklist, safety checker, comprehensive validation reports

Usage:
    # Interactive mode (recommended)
    python ifr_automation_v7.py

    # Full pipeline for all projects
    python ifr_automation_v7.py --pipeline

    # Deliverable cross-check only
    python ifr_automation_v7.py --deliverable-only

    # Specific stages
    python ifr_automation_v7.py --stages ifr_sync version_mgmt deliverable

    # Validate only (no changes)
    python ifr_automation_v7.py --validate-only

Author: Generated with Claude Code
Date: 2026-03-11
Version: 7.0
"""

import os
import sys
import argparse
import logging
import json
import re
import shutil
import fnmatch
import time
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional, Set
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from itertools import groupby

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

try:
    from colorama import init, Fore, Style, Back
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import requests as _requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False


# =============================================================================
# Telegram Notification Helper
# =============================================================================

def send_telegram_notification(message: str, parse_mode: str = "HTML") -> bool:
    """Send a notification message to the configured Telegram chat.

    Reads TELEGRAM_TOKEN and TELEGRAM_CHAT_ID from the .env file at:
      D:/1. SOP/SOP_Project Status&Tasks√/V3 Manul&Auto CSV to excel/Automatic Export√/.env

    Returns True if sent successfully, False otherwise (silently fails).
    """
    if not REQUESTS_AVAILABLE:
        return False

    env_path = Path(r"D:\1. SOP\SOP_Project Status&Tasks√\V3 Manul&Auto CSV to excel\Automatic Export√\.env")
    if not env_path.exists():
        return False

    token = chat_id = None
    try:
        for line in env_path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if line.startswith('TELEGRAM_TOKEN='):
                token = line.split('=', 1)[1].strip()
            elif line.startswith('TELEGRAM_CHAT_ID='):
                chat_id = line.split('=', 1)[1].strip()
    except Exception:
        return False

    if not token or not chat_id:
        return False

    try:
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = _requests.post(url, json={
            "chat_id": chat_id,
            "text": message,
            "parse_mode": parse_mode,
        }, timeout=10)
        return resp.status_code == 200
    except Exception:
        return False


def format_pipeline_result(results: Dict) -> str:
    """Format a pipeline result dict into a Telegram-friendly message."""
    project = results.get('project_name', 'Unknown')
    success = results.get('success', False)
    icon = "✅" if success else "❌"

    lines = [f"{icon} <b>Pipeline: {project}</b>"]

    ifr = results.get('ifr_sync')
    if ifr:
        if ifr.get('success'):
            lines.append(f"  📄 IFR Sync: 文件夹={ifr.get('folders',0)}, "
                         f"图纸={ifr.get('drawings',0)}, 报告={ifr.get('reports',0)}")
        elif 'error' in ifr:
            lines.append(f"  ❌ IFR Sync: {ifr['error']}")

    vm = results.get('version_mgmt')
    if vm and not vm.get('error'):
        lines.append(f"  🗂 Version Mgmt: 扫描={vm.get('scanned',0)}, 移动={vm.get('moved',0)}")
    elif vm and vm.get('error'):
        lines.append(f"  ❌ Version Mgmt: {vm['error']}")

    dlv = results.get('deliverable')
    if dlv and not dlv.get('error') and not dlv.get('skipped'):
        lines.append(f"  📋 Deliverable: 新增={dlv.get('new_items',0)}, "
                     f"更新={dlv.get('rev_mismatches',0)}, "
                     f"插入={dlv.get('inserted',0)}, 更新={dlv.get('updated',0)}")
    elif dlv and dlv.get('skipped'):
        lines.append(f"  ⚠️ Deliverable: {dlv.get('reason','skipped')}")
    elif dlv and dlv.get('error'):
        lines.append(f"  ❌ Deliverable: {dlv['error']}")

    return "\n".join(lines)


# =============================================================================
# Utility Functions
# =============================================================================

def to_long_path(path: Path) -> Path:
    """Convert path to long path format on Windows to support paths > 260 chars."""
    if os.name == 'nt':
        path_str = str(path.resolve())
        if not path_str.startswith('\\\\?\\'):
            path_str = '\\\\?\\' + path_str
        return Path(path_str)
    return path


# IFC/IFR classification patterns
_RE_IFC_FILE = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?:[_\s-]?IFC)?(?=[_.\s]|$)', re.IGNORECASE)
_RE_IFR_FILE = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*([A-Z])(?=[_.\s]|$)', re.IGNORECASE)


def _is_ifc_file(filename: str) -> bool:
    """判断文件是否为 IFC 文件（数字版本号）"""
    stem = Path(filename).stem
    if _RE_IFC_FILE.search(stem):
        return True
    if 'IFC' in stem.upper():
        return True
    return False


def _is_ifr_file(filename: str) -> bool:
    """判断文件是否为 IFR 文件（字母版本号）"""
    stem = Path(filename).stem
    return bool(_RE_IFR_FILE.search(stem))


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class SourceDirectory:
    """Information about a source directory."""
    name: str
    path: str
    exists: bool = False
    file_count: int = 0
    files: List[str] = field(default_factory=list)
    priority: int = 1


@dataclass
class ProjectValidation:
    """Validation result for a project."""
    valid: bool
    project_name: str
    project_path: str
    region: str = ""
    missing_dirs: List[str] = field(default_factory=list)
    source_dirs_found: List[SourceDirectory] = field(default_factory=list)
    has_legacy_structure: bool = False
    has_existing_ifr_client: bool = False
    has_existing_ifc_client: bool = False
    ifc_file_count: int = 0
    last_modified: str = ""
    days_since_modified: int = 0
    recommended_action: str = "skip"  # process/confirm/skip
    warning_message: str = ""
    drawings_count: int = 0
    reports_count: int = 0
    folders_to_create: int = 0

    def to_dict(self) -> Dict:
        result = asdict(self)
        result['source_dirs_found'] = [asdict(s) for s in self.source_dirs_found]
        return result


@dataclass
class ProcessResult:
    """Result of processing a single project."""
    project_name: str
    project_path: str
    success: bool
    folders_created: int = 0
    drawings_copied: int = 0
    drawings_skipped: int = 0
    reports_copied: int = 0
    reports_skipped: int = 0
    ifc_files_copied: int = 0
    ifc_files_skipped: int = 0
    ifc_target_path: str = ""
    errors: List[str] = field(default_factory=list)
    skipped: bool = False
    skip_reason: str = ""


# =============================================================================
# Configuration Management
# =============================================================================

class ConfigManager:
    """Manages configuration file for IFR automation."""

    DEFAULT_CONFIG = {
        "default_root_path": "",
        "recent_projects": [],
        "auto_backup": True,
        "skip_existing_files": True,
        "show_progress_bar": True,
        "log_level": "INFO",
        "color_output": True,
        "max_recent_projects": 10,
        "last_operation": None,
        "last_run_time": None,

        # V3: Project filters
        "project_filters": {
            "whitelist": [],
            "blacklist": [],
            "auto_process_only_validated": True,
            "require_confirmation_for_legacy": True
        },

        # V3: Drawing sources configuration
        "drawing_sources": [
            {
                "path": "Design/Engineering/1. Drawings/2. IFR_internal",
                "enabled": True,
                "priority": 1,
                "description": "Internal drawings directory (primary source)"
            },
            {
                "path": "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
                "enabled": True,
                "priority": 2,
                "description": "Civil reports directory (some drawings here)"
            }
        ],

        # V3: Validation rules
        "validation_rules": {
            "require_ifr_internal": False,
            "require_at_least_one_source": True,
            "min_source_dirs": 1,
            "max_days_since_modified": 180  # 6 months
        },

        # V4: IFC(Client) source path (relative to project root)
        "ifc_client_path": "Design/Engineering/1. Drawings/4. IFC(Client)"
    }

    def __init__(self, config_path: Optional[Path] = None):
        if config_path is None:
            self.config_path = Path(__file__).parent / "config.json"
        else:
            self.config_path = Path(config_path)
        self.config = self.load()

    def load(self) -> Dict:
        if self.config_path.exists():
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                config = self._deep_merge(self.DEFAULT_CONFIG.copy(), loaded)
                return config
            except Exception as e:
                print(f"Warning: Could not load config: {e}")
                return self.DEFAULT_CONFIG.copy()
        else:
            return self.DEFAULT_CONFIG.copy()

    def _deep_merge(self, base: Dict, override: Dict) -> Dict:
        """Deep merge two dictionaries."""
        result = base.copy()
        for key, value in override.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = self._deep_merge(result[key], value)
            else:
                result[key] = value
        return result

    def save(self):
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Warning: Could not save config: {e}")

    def get(self, key: str, default=None):
        keys = key.split('.')
        value = self.config
        for k in keys:
            if isinstance(value, dict) and k in value:
                value = value[k]
            else:
                return default
        return value

    def set(self, key: str, value):
        keys = key.split('.')
        config = self.config
        for k in keys[:-1]:
            if k not in config:
                config[k] = {}
            config = config[k]
        config[keys[-1]] = value
        self.save()

    def add_recent_project(self, project_path: str):
        recent = self.config.get("recent_projects", [])
        if project_path in recent:
            recent.remove(project_path)
        recent.insert(0, project_path)
        max_recent = self.config.get("max_recent_projects", 10)
        self.config["recent_projects"] = recent[:max_recent]
        self.save()


# =============================================================================
# UI Utilities
# =============================================================================

class UIHelper:
    """Helper class for console UI elements."""

    @staticmethod
    def clear_screen():
        os.system('cls' if os.name == 'nt' else 'clear')

    @staticmethod
    def colorize(text: str, color: str, bold: bool = False) -> str:
        if not COLORAMA_AVAILABLE:
            return text
        color_map = {
            "green": Fore.GREEN,
            "red": Fore.RED,
            "yellow": Fore.YELLOW,
            "blue": Fore.BLUE,
            "cyan": Fore.CYAN,
            "magenta": Fore.MAGENTA,
            "white": Fore.WHITE,
        }
        prefix = color_map.get(color, "")
        if bold:
            prefix = Style.BRIGHT + prefix
        return f"{prefix}{text}{Style.RESET_ALL}"

    @staticmethod
    def print_header(title: str, subtitle: str = ""):
        width = 72
        print()
        if COLORAMA_AVAILABLE:
            print(Fore.CYAN + "+" + "=" * width + "+")
            print(Fore.CYAN + "|" + Style.BRIGHT + Fore.WHITE + title.center(width) + Style.RESET_ALL + Fore.CYAN + "|")
            if subtitle:
                print(Fore.CYAN + "|" + Fore.WHITE + subtitle.center(width) + Fore.CYAN + "|")
            print(Fore.CYAN + "+" + "=" * width + "+" + Style.RESET_ALL)
        else:
            print("+" + "=" * width + "+")
            print("|" + title.center(width) + "|")
            if subtitle:
                print("|" + subtitle.center(width) + "|")
            print("+" + "=" * width + "+")
        print()

    @staticmethod
    def print_separator(char: str = "-", width: int = 72):
        if COLORAMA_AVAILABLE:
            print(Fore.CYAN + char * width + Style.RESET_ALL)
        else:
            print(char * width)

    @staticmethod
    def print_menu(options: List[Tuple[str, str]], prompt: str = "请输入选项") -> str:
        print()
        for key, label in options:
            if COLORAMA_AVAILABLE:
                print(f"  {Fore.YELLOW}[{key}]{Style.RESET_ALL} {label}")
            else:
                print(f"  [{key}] {label}")
        print()
        UIHelper.print_separator()
        return input(f"\n{prompt}: ").strip()

    @staticmethod
    def print_success(message: str):
        symbol = "[OK]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.GREEN}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def print_warning(message: str):
        symbol = "[!]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.YELLOW}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def print_error(message: str):
        symbol = "[X]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.RED}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def print_info(message: str):
        symbol = "[i]"
        if COLORAMA_AVAILABLE:
            print(f"{Fore.CYAN}{symbol}{Style.RESET_ALL} {message}")
        else:
            print(f"{symbol} {message}")

    @staticmethod
    def confirm(message: str, default: bool = True) -> bool:
        suffix = " [Y/n]: " if default else " [y/N]: "
        response = input(message + suffix).strip().lower()
        if not response:
            return default
        return response in ('y', 'yes')

    @staticmethod
    def status_icon(status: str) -> str:
        """Get status icon with color."""
        icons = {
            "ready": (Fore.GREEN + "[v]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[v]", "完整"),
            "confirm": (Fore.YELLOW + "[!]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[!]", "需确认"),
            "legacy": (Fore.RED + "[x]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[x]", "旧结构"),
            "skip": (Fore.RED + "[-]" + Style.RESET_ALL if COLORAMA_AVAILABLE else "[-]", "跳过"),
        }
        return icons.get(status, ("[?]", "未知"))

    @staticmethod
    def action_icon(action: str) -> str:
        """Get action recommendation icon."""
        if COLORAMA_AVAILABLE:
            icons = {
                "process": Fore.GREEN + "[SAFE]" + Style.RESET_ALL + " 可以安全处理",
                "confirm": Fore.YELLOW + "[WARN]" + Style.RESET_ALL + " 需要人工确认",
                "skip": Fore.RED + "[SKIP]" + Style.RESET_ALL + " 建议跳过",
            }
        else:
            icons = {
                "process": "[SAFE] 可以安全处理",
                "confirm": "[WARN] 需要人工确认",
                "skip": "[SKIP] 建议跳过",
            }
        return icons.get(action, "[?] 未知")


# =============================================================================
# Project Validator
# =============================================================================

class ProjectValidator:
    """Project validator - ensures only valid project structures are processed."""

    # Required directory structure
    REQUIRED_STRUCTURE = [
        "Design/Engineering/1. Drawings",
        "Design/Engineering/2. Calcs & Reports"
    ]

    # Source directories to check
    SOURCE_DIRECTORIES = [
        {"path": "Design/Engineering/1. Drawings/2. IFR_internal", "name": "IFR_internal", "priority": 1},
        {"path": "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure", "name": "Civil Reports", "priority": 2},
        {"path": "Design/Engineering/2. Calcs & Reports/Reports/Electrical", "name": "Electrical Reports", "priority": 2},
        {"path": "Design/Engineering/2. Calcs & Reports/Schedule", "name": "Schedule", "priority": 3},
    ]

    # Drawing patterns
    DRAWING_PATTERNS = [
        r'-C-PLN-\d+', r'-C-GEN-\d+', r'-C-SEC-\d+',
        r'-E-PLN-\d+', r'-E-SLD-\d+', r'-E-BLD-\d+',
        r'-E-CFG-\d+', r'-E-GAD-\d+',
    ]

    # Report patterns
    REPORT_PATTERNS = [
        r'-C-RPT-', r'-E-RPT-', r'-E-SCH-',
    ]

    # V4: IFC(Client) path
    IFC_CLIENT_REL_PATH = "Design/Engineering/1. Drawings/4. IFC(Client)"

    def __init__(self, config: ConfigManager):
        self.config = config
        self.validation_rules = config.get("validation_rules", {})
        self.ifc_client_rel_path = config.get("ifc_client_path", self.IFC_CLIENT_REL_PATH)

    def validate_project(self, project_path: Path, region: str = "") -> ProjectValidation:
        """Validate a project's structure and return detailed validation result."""
        validation = ProjectValidation(
            valid=False,
            project_name=project_path.name,
            project_path=str(project_path),
            region=region
        )

        # Check basic structure
        missing_dirs = []
        for required in self.REQUIRED_STRUCTURE:
            if not (project_path / required).exists():
                missing_dirs.append(required)

        validation.missing_dirs = missing_dirs

        if missing_dirs:
            validation.has_legacy_structure = True
            validation.warning_message = f"缺少必要目录: {', '.join(missing_dirs)}"
            validation.recommended_action = "skip"
            return validation

        # Check source directories
        source_dirs_found = []
        total_drawings = 0
        total_reports = 0

        for source in self.SOURCE_DIRECTORIES:
            source_path = project_path / source["path"]
            source_dir = SourceDirectory(
                name=source["name"],
                path=source["path"],
                exists=source_path.exists(),
                priority=source["priority"]
            )

            if source_path.exists():
                # Count files (use rglob to search subdirectories)
                pdf_files = list(source_path.rglob("*.pdf"))
                source_dir.file_count = len(pdf_files)
                source_dir.files = [f.name for f in pdf_files]

                # Categorize files using folder-based fallback
                is_drawing_source = "1. Drawings" in source["path"]
                is_report_source = "2. Calcs & Reports" in source["path"]
                for pdf_file in pdf_files:
                    name = pdf_file.name
                    if self._is_drawing(name):
                        total_drawings += 1
                    elif self._is_report(name):
                        total_reports += 1
                    elif is_drawing_source:
                        total_drawings += 1
                    elif is_report_source:
                        # Skip files in non-deliverable subdirs
                        try:
                            rel_parts = pdf_file.relative_to(source_path).parts[:-1]
                            in_excluded = any(
                                part.lower() in ('ss', 'superseded', '_export', 'data')
                                or part.lower().startswith(('appendix', 'stk'))
                                for part in rel_parts
                            )
                        except Exception:
                            rel_parts = ()
                            in_excluded = False
                        if in_excluded:
                            pass
                        elif rel_parts:
                            # Parent prefix check for subdirectory files
                            parent_name = rel_parts[0]
                            parent_prefix = parent_name.split('_')[0]
                            if '_' in parent_name and name.startswith(parent_prefix):
                                total_reports += 1
                        else:
                            total_reports += 1

            source_dirs_found.append(source_dir)

        validation.source_dirs_found = source_dirs_found
        validation.drawings_count = total_drawings
        validation.reports_count = total_reports

        # Check for existing IFR(Client)
        ifr_client = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)"
        validation.has_existing_ifr_client = ifr_client.exists()

        # V4: Check for existing IFC(Client)
        ifc_client = project_path / self.ifc_client_rel_path
        validation.has_existing_ifc_client = ifc_client.exists()
        if ifc_client.exists():
            try:
                validation.ifc_file_count = sum(1 for _ in ifc_client.rglob("*") if _.is_file())
            except Exception:
                validation.ifc_file_count = 0

        # Calculate folders to create
        if not validation.has_existing_ifr_client:
            validation.folders_to_create = 7  # Main + 3 subs + 3 SS folders
        else:
            # Count missing subfolders
            subfolders = ["1.Drawing/SS", "2.Reports/SS", "3.Deliverables/SS"]
            validation.folders_to_create = sum(
                1 for sf in subfolders if not (ifr_client / sf).exists()
            )

        # Get last modified time
        try:
            mtime = datetime.fromtimestamp(project_path.stat().st_mtime)
            validation.last_modified = mtime.strftime("%Y-%m-%d")
            validation.days_since_modified = (datetime.now() - mtime).days
        except:
            validation.last_modified = "未知"
            validation.days_since_modified = 999

        # Determine recommendation
        active_sources = sum(1 for s in source_dirs_found if s.exists and s.file_count > 0)
        min_sources = self.validation_rules.get("min_source_dirs", 1)
        max_days = self.validation_rules.get("max_days_since_modified", 180)

        if active_sources >= min_sources:
            if validation.days_since_modified > max_days:
                validation.valid = True
                validation.recommended_action = "confirm"
                validation.warning_message = f"项目超过 {max_days} 天未修改，可能是旧项目"
            elif validation.has_existing_ifr_client:
                validation.valid = True
                validation.recommended_action = "confirm"
                validation.warning_message = "目标目录已存在，可能会覆盖现有内容"
            else:
                validation.valid = True
                validation.recommended_action = "process"
        else:
            validation.has_legacy_structure = True
            validation.recommended_action = "skip"
            validation.warning_message = f"有效源目录不足 (找到 {active_sources}, 需要 {min_sources})"

        return validation

    def _is_drawing(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.DRAWING_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                # Exclude reports
                for exclude in self.REPORT_PATTERNS:
                    if re.search(exclude, filename, re.IGNORECASE):
                        return False
                return True
        return False

    def _is_report(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.REPORT_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return True
        return False


# =============================================================================
# Safety Checker
# =============================================================================

class SafetyChecker:
    """Safety checker - prevents accidental operations."""

    def __init__(self, config: ConfigManager):
        self.config = config

    def check_before_process(self, validation: ProjectValidation) -> Dict:
        """Perform safety checks before processing."""
        warnings = []
        checks = {
            "has_recent_activity": validation.days_since_modified < 180,
            "has_existing_ifr_client": validation.has_existing_ifr_client,
            "source_dirs_not_empty": validation.drawings_count > 0 or validation.reports_count > 0,
            "structure_complete": not validation.has_legacy_structure
        }

        if validation.has_existing_ifr_client:
            warnings.append("目标目录已存在文件，可能会覆盖现有内容")

        if not checks["has_recent_activity"]:
            warnings.append(f"项目超过 {validation.days_since_modified} 天未修改，可能是旧项目")

        if not checks["source_dirs_not_empty"]:
            warnings.append("源目录中没有找到任何文件")

        if not checks["structure_complete"]:
            warnings.append("项目结构不完整")

        return {
            "safe": len(warnings) == 0,
            "warnings": warnings,
            "checks": checks
        }


# =============================================================================
# Project Scanner
# =============================================================================

class ProjectScanner:
    """Hierarchical project scanner."""

    # Known region prefixes
    REGION_PATTERNS = [
        r'^(\d+)\.',      # 1.NSW, 2.SA
        r'^([A-Z]{2,3})', # NSW, SA, VIC
    ]

    def __init__(self, config: ConfigManager, validator: ProjectValidator):
        self.config = config
        self.validator = validator

    def scan_hierarchical(self, root_path: Path) -> Dict[str, List[ProjectValidation]]:
        """Scan projects hierarchically by region."""
        projects_by_region = defaultdict(list)

        if not root_path.exists():
            return projects_by_region

        # Get whitelist and blacklist
        filters = self.config.get("project_filters", {})
        whitelist = filters.get("whitelist", [])
        blacklist = filters.get("blacklist", [])

        # First level: look for regions or direct projects
        for item in root_path.iterdir():
            if not item.is_dir():
                continue

            # Check if this is a region folder
            if self._is_region_folder(item):
                region_name = item.name
                # Scan projects within region
                for project_dir in item.iterdir():
                    if not project_dir.is_dir():
                        continue

                    # Apply whitelist/blacklist
                    if not self._passes_filter(project_dir.name, whitelist, blacklist):
                        continue

                    # Check if it's a valid project
                    if self._is_project_folder(project_dir):
                        validation = self.validator.validate_project(project_dir, region_name)
                        projects_by_region[region_name].append(validation)

            # Check if this is a direct project folder
            elif self._is_project_folder(item):
                if not self._passes_filter(item.name, whitelist, blacklist):
                    continue

                validation = self.validator.validate_project(item, "Root")
                projects_by_region["Root"].append(validation)

        # Include additional (non-standard structure) projects from config
        additional = self.config.get("additional_projects", [])
        already_paths = {v.project_path for vals in projects_by_region.values() for v in vals}
        for entry in additional:
            proj_path = Path(entry["path"])
            if str(proj_path) in already_paths or not proj_path.exists():
                continue
            region = entry.get("region", "Other")
            validation = ProjectValidation(
                valid=True,
                project_name=proj_path.name,
                project_path=str(proj_path),
                region=region,
            )
            validation.source_dirs_found = []
            # Auto-detect drawings root for non-standard structure
            drawings_root = entry.get("drawings_root", "1. Drawings")
            drw_path = proj_path / drawings_root
            if drw_path.exists():
                sd = SourceDirectory(
                    name=drawings_root,
                    path=drawings_root,
                    exists=True,
                    priority=1,
                )
                pdf_files = list(drw_path.rglob("*.pdf"))
                sd.file_count = len(pdf_files)
                sd.files = [f.name for f in pdf_files[:50]]
                validation.source_dirs_found.append(sd)
            projects_by_region[region].append(validation)

        # Sort projects within each region
        for region in projects_by_region:
            projects_by_region[region].sort(key=lambda x: x.project_name)

        return dict(projects_by_region)

    def _is_region_folder(self, path: Path) -> bool:
        """Check if folder is a region container (like 1.NSW, 2.SA)."""
        name = path.name

        # Check if matches region patterns
        for pattern in self.REGION_PATTERNS:
            if re.match(pattern, name):
                # Check if it contains project-like subfolders
                for subdir in path.iterdir():
                    if subdir.is_dir() and self._is_project_folder(subdir):
                        return True

        return False

    def _is_project_folder(self, path: Path) -> bool:
        """Check if folder is a project (has Design/Engineering structure)."""
        design_path = path / "Design" / "Engineering"
        return design_path.exists()

    def _passes_filter(self, name: str, whitelist: List[str], blacklist: List[str]) -> bool:
        """Check if project name passes whitelist/blacklist filters."""
        # If whitelist is set, only include matching projects
        if whitelist:
            if not any(fnmatch.fnmatch(name, pattern) for pattern in whitelist):
                return False

        # Exclude blacklisted projects
        if blacklist:
            if any(fnmatch.fnmatch(name, pattern) for pattern in blacklist):
                return False

        return True


# =============================================================================
# Drawing Collector
# =============================================================================

class DrawingCollector:
    """Collects drawings from multiple source directories."""

    DRAWING_PATTERNS = [
        r'-C-PLN-\d+', r'-C-GEN-\d+', r'-C-SEC-\d+',
        r'-E-PLN-\d+', r'-E-SLD-\d+', r'-E-BLD-\d+',
        r'-E-CFG-\d+', r'-E-GAD-\d+',
    ]

    EXCLUDE_PATTERNS = [
        r'-C-RPT-', r'-E-RPT-', r'-E-SCH-',
    ]

    def __init__(self, config: ConfigManager):
        self.config = config
        self.sources = config.get("drawing_sources", [])

    def collect_drawings(self, project_path: Path) -> Dict:
        """Collect drawings from all configured source directories."""
        result = {
            "files": [],
            "total_count": 0,
            "by_source": {},
            "duplicates_removed": 0
        }

        seen_files = set()

        for source in self.sources:
            if not source.get("enabled", True):
                continue

            source_path = project_path / source["path"]
            if not source_path.exists():
                continue

            source_name = Path(source["path"]).name
            source_count = 0
            is_drawing_source = "1. Drawings" in source["path"]

            try:
                pdf_files = list(source_path.glob("*.pdf"))
            except Exception as e:
                continue

            for pdf_file in pdf_files:
                if not self._should_collect_as_drawing(pdf_file.name, is_drawing_source):
                    continue

                # Check for duplicates
                if pdf_file.name in seen_files:
                    result["duplicates_removed"] += 1
                    continue

                seen_files.add(pdf_file.name)

                try:
                    # Use long path for stat operations
                    pdf_file_long = to_long_path(pdf_file)
                    file_info = {
                        "filename": pdf_file.name,
                        "source_path": str(pdf_file_long),
                        "source_dir": source_name,
                        "size": pdf_file_long.stat().st_size,
                        "modified": datetime.fromtimestamp(pdf_file_long.stat().st_mtime).strftime("%Y-%m-%d")
                    }
                    result["files"].append(file_info)
                    source_count += 1
                except Exception:
                    # Fallback without long path
                    try:
                        file_info = {
                            "filename": pdf_file.name,
                            "source_path": str(pdf_file),
                            "source_dir": source_name,
                            "size": 0,
                            "modified": "unknown"
                        }
                        result["files"].append(file_info)
                        source_count += 1
                    except Exception:
                        pass

            result["by_source"][source_name] = source_count

        result["total_count"] = len(result["files"])
        return result

    def _should_collect_as_drawing(self, filename: str, is_drawing_source: bool) -> bool:
        """Check if file should be collected as a drawing.

        Uses pattern matching first, falls back to folder-based classification
        for projects with unrecognized naming conventions (e.g. 50023-xx, TFS-xx).
        """
        if not filename.lower().endswith('.pdf'):
            return False
        # Pattern match: explicitly a drawing
        if self._is_drawing(filename):
            return True
        # Pattern match: explicitly a report → exclude
        for pattern in self.EXCLUDE_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return False
        # No pattern match: trust folder structure
        return is_drawing_source

    def _is_drawing(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.DRAWING_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                for exclude in self.EXCLUDE_PATTERNS:
                    if re.search(exclude, filename, re.IGNORECASE):
                        return False
                return True
        return False


# =============================================================================
# Report Generator
# =============================================================================

class ValidationReportGenerator:
    """Generates validation reports."""

    def generate_report(self, projects_by_region: Dict[str, List[ProjectValidation]],
                       root_path: str) -> str:
        """Generate a detailed validation report."""
        lines = []
        lines.append("=" * 70)
        lines.append("项目验证报告")
        lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"扫描根目录: {root_path}")
        lines.append("=" * 70)
        lines.append("")

        # Summary
        total = sum(len(projects) for projects in projects_by_region.values())
        ready = sum(1 for projects in projects_by_region.values()
                   for p in projects if p.recommended_action == "process")
        confirm = sum(1 for projects in projects_by_region.values()
                     for p in projects if p.recommended_action == "confirm")
        skip = sum(1 for projects in projects_by_region.values()
                  for p in projects if p.recommended_action == "skip")

        lines.append(f"总计: {total} 个项目")
        lines.append(f"  [v] 可处理: {ready} 个")
        lines.append(f"  [!] 需确认: {confirm} 个")
        lines.append(f"  [x] 建议跳过: {skip} 个")
        lines.append("")

        # Ready projects
        if ready > 0:
            lines.append("-" * 70)
            lines.append(f"[v] 可以安全处理的项目 ({ready})")
            lines.append("-" * 70)
            lines.append("")

            idx = 1
            for region, projects in projects_by_region.items():
                for p in projects:
                    if p.recommended_action == "process":
                        lines.extend(self._format_project(idx, p))
                        idx += 1

        # Confirm projects
        if confirm > 0:
            lines.append("-" * 70)
            lines.append(f"[!] 需要人工确认的项目 ({confirm})")
            lines.append("-" * 70)
            lines.append("")

            idx = 1
            for region, projects in projects_by_region.items():
                for p in projects:
                    if p.recommended_action == "confirm":
                        lines.extend(self._format_project(idx, p))
                        idx += 1

        # Skip projects
        if skip > 0:
            lines.append("-" * 70)
            lines.append(f"[x] 建议跳过的项目 ({skip})")
            lines.append("-" * 70)
            lines.append("")

            idx = 1
            for region, projects in projects_by_region.items():
                for p in projects:
                    if p.recommended_action == "skip":
                        lines.extend(self._format_project(idx, p))
                        idx += 1

        return "\n".join(lines)

    def _format_project(self, idx: int, p: ProjectValidation) -> List[str]:
        lines = []
        lines.append(f"{idx}. {p.project_name}")
        lines.append(f"   路径: {p.project_path}")
        lines.append(f"   区域: {p.region}")
        lines.append(f"   状态: {'完整结构' if p.valid else '不完整结构'}")

        if p.warning_message:
            lines.append(f"   警告: {p.warning_message}")

        lines.append("   源目录:")
        for source in p.source_dirs_found:
            status = "[v]" if source.exists and source.file_count > 0 else "[x]"
            count = f"({source.file_count} 个文件)" if source.exists else "(不存在)"
            lines.append(f"     {status} {source.name} {count}")

        # V4: Show IFC(Client) status
        if p.has_existing_ifc_client:
            lines.append(f"   IFC(Client): [v] 存在 ({p.ifc_file_count} 个文件)")
        else:
            lines.append(f"   IFC(Client): [x] 不存在")

        if p.recommended_action == "process":
            lines.append("   预计操作:")
            lines.append(f"     - 创建 {p.folders_to_create} 个文件夹")
            lines.append(f"     - 复制 {p.drawings_count} 个图纸")
            lines.append(f"     - 复制 {p.reports_count} 个报告")

        lines.append("")
        return lines

    def save_report(self, content: str, path: Path):
        """Save report to file."""
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)


# =============================================================================
# Main IFR Automation Class
# =============================================================================

class IFRAutomation:
    """Main IFR automation class with v4 features."""

    IFR_STRUCTURE = {
        "1.Drawing": ["SS"],
        "2.Reports": ["SS"],
        "3.Deliverables": ["SS"]
    }

    DRAWING_PATTERNS = [
        r'-C-PLN-\d+', r'-C-GEN-\d+', r'-C-SEC-\d+',
        r'-E-PLN-\d+', r'-E-SLD-\d+', r'-E-BLD-\d+',
        r'-E-CFG-\d+', r'-E-GAD-\d+',
    ]

    REPORT_PATTERNS = [
        r'-C-RPT-', r'-E-RPT-', r'-E-SCH-',
    ]

    # V4: IFC(Client) source path (relative to project root)
    IFC_CLIENT_REL_PATH = "Design/Engineering/1. Drawings/4. IFC(Client)"

    def __init__(self, root_path: str, config: ConfigManager,
                 dry_run: bool = False, create_folders_only: bool = False,
                 mirror_files_only: bool = False, validate_only: bool = False,
                 yes_to_all: bool = False, auto_safe_only: bool = False,
                 log_level: str = "INFO", output_json: bool = False,
                 interactive: bool = False):

        self.root_path = Path(root_path).resolve()
        self.config = config
        self.dry_run = dry_run
        self.create_folders_only = create_folders_only
        self.mirror_files_only = mirror_files_only
        self.validate_only = validate_only
        self.yes_to_all = yes_to_all
        self.auto_safe_only = auto_safe_only
        self.log_level = log_level
        self.output_json = output_json
        self.interactive = interactive

        self.validator = ProjectValidator(config)
        self.scanner = ProjectScanner(config, self.validator)
        self.safety_checker = SafetyChecker(config)
        self.drawing_collector = DrawingCollector(config)
        self.report_generator = ValidationReportGenerator()

        self.ifc_client_rel_path = config.get("ifc_client_path", self.IFC_CLIENT_REL_PATH)

        self.results: List[ProcessResult] = []
        self._setup_logging()

    def _setup_logging(self):
        log_dir = Path(__file__).parent / "logs"
        log_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = log_dir / f"ifr_automation_{timestamp}.log"

        log_format = '%(asctime)s - %(levelname)s - %(message)s'

        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter(log_format))

        console_handler = logging.StreamHandler(sys.stdout)
        if self.interactive:
            console_handler.setLevel(logging.WARNING)
        else:
            console_handler.setLevel(getattr(logging, self.log_level.upper()))
        console_handler.setFormatter(logging.Formatter('%(levelname)s - %(message)s'))

        self.logger = logging.getLogger(f'IFRAutomation_{timestamp}')
        self.logger.setLevel(logging.DEBUG)
        self.logger.handlers.clear()
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)

        self.log_file = log_file

    def scan_projects(self) -> Dict[str, List[ProjectValidation]]:
        """Scan all projects hierarchically."""
        return self.scanner.scan_hierarchical(self.root_path)

    def process_project(self, validation: ProjectValidation) -> ProcessResult:
        """Process a single validated project."""
        project_path = Path(validation.project_path)
        result = ProcessResult(
            project_name=validation.project_name,
            project_path=validation.project_path,
            success=False
        )

        try:
            # Create folder structure
            if not self.mirror_files_only:
                folders_created = self._create_folder_structure(project_path)
                result.folders_created = folders_created

            # Mirror files
            if not self.create_folders_only:
                drawings = self._mirror_drawings(project_path)
                result.drawings_copied = drawings["copied"]
                result.drawings_skipped = drawings["skipped"]

                reports = self._mirror_reports(project_path)
                result.reports_copied = reports["copied"]
                result.reports_skipped = reports["skipped"]

            result.success = True

        except Exception as e:
            result.errors.append(str(e))
            self.logger.error(f"Error processing {validation.project_name}: {e}")

        return result

    def copy_ifc_to_target(self, project_path: Path, target_path: Path) -> Dict[str, int]:
        """Copy the entire 4. IFC(Client) folder to the target path.

        Copies all files and subdirectories preserving folder structure.
        Uses copy (not move). Skips files that already exist with same size/date.

        Returns dict with copied/skipped/error counts.
        """
        result = {"copied": 0, "skipped": 0, "errors": []}

        ifc_source = project_path / self.ifc_client_rel_path
        ifc_source_long = self._to_long_path(ifc_source)

        if not ifc_source_long.exists():
            result["errors"].append(f"IFC(Client) 源目录不存在: {ifc_source}")
            return result

        target_long = self._to_long_path(target_path)

        # Create target directory if it doesn't exist
        if not self.dry_run:
            target_long.mkdir(parents=True, exist_ok=True)

        # Walk through all files in source and copy preserving structure
        try:
            source_files = list(ifc_source_long.rglob("*"))
        except Exception as e:
            result["errors"].append(f"无法扫描 IFC(Client) 目录: {e}")
            return result

        for source_item in source_files:
            # Calculate relative path from ifc_source
            try:
                rel_path = source_item.relative_to(ifc_source_long)
            except ValueError:
                # Fallback: try without long path prefix
                try:
                    rel_path = source_item.relative_to(ifc_source)
                except ValueError:
                    continue

            dest_item = target_long / rel_path

            if source_item.is_dir():
                # Create subdirectory
                if not self.dry_run:
                    dest_item.mkdir(parents=True, exist_ok=True)
                continue

            # It's a file - copy it
            try:
                # Ensure parent directory exists
                if not self.dry_run:
                    dest_item.parent.mkdir(parents=True, exist_ok=True)

                if self._should_copy(source_item, dest_item):
                    if not self.dry_run:
                        shutil.copy2(str(source_item), str(dest_item))
                    result["copied"] += 1
                    self.logger.info(f"IFC 复制: {rel_path}")
                else:
                    result["skipped"] += 1
            except OSError as e:
                self.logger.warning(f"无法复制 IFC 文件 (路径过长?): {rel_path} - {e}")
                result["errors"].append(f"{rel_path}: {e}")
            except Exception as e:
                self.logger.warning(f"复制 IFC 文件出错: {rel_path} - {e}")
                result["errors"].append(f"{rel_path}: {e}")

        return result

    def _create_folder_structure(self, project_path: Path) -> int:
        """Create IFR folder structure."""
        created_count = 0
        ifr_client = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)"

        # Use long path format on Windows
        ifr_client_long = self._to_long_path(ifr_client)

        if not ifr_client_long.exists():
            if not self.dry_run:
                ifr_client_long.mkdir(parents=True, exist_ok=True)
            created_count += 1

        for main_folder, subfolders in self.IFR_STRUCTURE.items():
            main_path = self._to_long_path(ifr_client / main_folder)
            if not main_path.exists():
                if not self.dry_run:
                    main_path.mkdir(parents=True, exist_ok=True)
                created_count += 1

            for subfolder in subfolders:
                sub_path = self._to_long_path(ifr_client / main_folder / subfolder)
                if not sub_path.exists():
                    if not self.dry_run:
                        sub_path.mkdir(parents=True, exist_ok=True)
                    created_count += 1

        return created_count

    def _mirror_drawings(self, project_path: Path) -> Dict[str, int]:
        """Mirror drawings from source directories."""
        result = {"copied": 0, "skipped": 0, "errors": []}
        dest_dir = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)/1.Drawing"

        if not dest_dir.exists():
            return result

        drawings = self.drawing_collector.collect_drawings(project_path)

        for file_info in drawings["files"]:
            # IFR(Client) 目录只接受 IFR 文件，排除 IFC 文件
            if _is_ifc_file(file_info["filename"]) and not _is_ifr_file(file_info["filename"]):
                continue

            source = Path(file_info["source_path"])
            dest = dest_dir / file_info["filename"]

            try:
                # Use long path format on Windows
                source_long = self._to_long_path(source)
                dest_long = self._to_long_path(dest)

                if self._should_copy(source_long, dest_long):
                    if not self.dry_run:
                        shutil.copy2(str(source_long), str(dest_long))
                    result["copied"] += 1
                else:
                    result["skipped"] += 1
            except OSError as e:
                self.logger.warning(f"无法复制图纸 (路径过长?): {file_info['filename']} - {e}")
                result["errors"].append(f"{file_info['filename']}: {e}")
            except Exception as e:
                self.logger.warning(f"复制图纸出错: {file_info['filename']} - {e}")
                result["errors"].append(f"{file_info['filename']}: {e}")

        return result

    def _to_long_path(self, path: Path) -> Path:
        """Convert path to long path format on Windows to support paths > 260 chars."""
        return to_long_path(path)

    def _mirror_reports(self, project_path: Path) -> Dict[str, int]:
        """Mirror reports from source directories."""
        result = {"copied": 0, "skipped": 0, "errors": []}
        dest_dir = project_path / "Design/Engineering/1. Drawings/3. IFR(Client)/2.Reports"

        if not dest_dir.exists():
            return result

        source_dirs = [
            project_path / "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
            project_path / "Design/Engineering/2. Calcs & Reports/Reports/Electrical",
            project_path / "Design/Engineering/2. Calcs & Reports/Schedule",
            project_path / "Design/Engineering/1. Drawings/2. IFR_internal",
        ]

        seen_files = set()

        for source_dir in source_dirs:
            if not source_dir.exists():
                continue

            is_report_source = "2. Calcs & Reports" in str(source_dir)

            try:
                pdf_files = list(source_dir.rglob("*.pdf"))
            except Exception as e:
                self.logger.warning(f"无法扫描目录 {source_dir}: {e}")
                continue

            for pdf_file in pdf_files:
                # Check if file is in a non-deliverable subdirectory
                try:
                    rel_parts = pdf_file.relative_to(source_dir).parts[:-1]
                    # Always exclude SS/superseded
                    in_ss = any(
                        part.lower() in ('ss', 'superseded', '_export', 'data')
                        for part in rel_parts
                    )
                    if in_ss:
                        in_excluded_subdir = True
                    else:
                        # Appendix: exclude unless file has a valid doc-ID
                        in_appendix = any(
                            part.lower().startswith(('appendix', 'stk'))
                            for part in rel_parts
                        )
                        if in_appendix:
                            from pathlib import Path as _P
                            _doc_id_pat = re.match(r'^(GG\d{2}-[A-Z]-[A-Z]{3}-\d{3})', pdf_file.name, re.IGNORECASE)
                            in_excluded_subdir = not bool(_doc_id_pat)
                        else:
                            in_excluded_subdir = False
                except Exception:
                    in_excluded_subdir = False

                if not self._should_collect_as_report(pdf_file.name, is_report_source,
                                                       in_excluded_subdir, rel_parts):
                    continue

                # IFR(Client) 目录只接受 IFR 文件，排除 IFC 文件
                if _is_ifc_file(pdf_file.name) and not _is_ifr_file(pdf_file.name):
                    continue

                if pdf_file.name in seen_files:
                    continue
                seen_files.add(pdf_file.name)

                dest = dest_dir / pdf_file.name

                try:
                    # Use long path format on Windows
                    source_long = self._to_long_path(pdf_file)
                    dest_long = self._to_long_path(dest)

                    if self._should_copy(source_long, dest_long):
                        if not self.dry_run:
                            shutil.copy2(str(source_long), str(dest_long))
                        result["copied"] += 1
                    else:
                        result["skipped"] += 1
                except OSError as e:
                    # Handle long path errors gracefully
                    self.logger.warning(f"无法复制文件 (路径过长?): {pdf_file.name} - {e}")
                    result["errors"].append(f"{pdf_file.name}: {e}")
                except Exception as e:
                    self.logger.warning(f"复制文件出错: {pdf_file.name} - {e}")
                    result["errors"].append(f"{pdf_file.name}: {e}")

        return result

    def _should_collect_as_report(self, filename: str, is_report_source: bool,
                                    in_excluded_subdir: bool = False,
                                    rel_parts: tuple = ()) -> bool:
        """Check if file should be collected as a report.

        Uses pattern matching first, falls back to folder-based classification
        for projects with unrecognized naming conventions.
        Files in excluded subdirs (Appendix, SS, STK, etc.) only included
        if they explicitly match a report pattern.
        In fallback mode, files in subdirectories must match parent folder's
        doc-ID prefix to avoid picking up third-party reference documents.
        """
        if not filename.lower().endswith('.pdf'):
            return False
        # Pattern match: explicitly a report → always include
        if self._is_report(filename):
            return True
        # Pattern match: explicitly a drawing → exclude
        for pattern in self.DRAWING_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return False
        # No pattern match: trust folder structure, but skip excluded subdirs
        if not is_report_source or in_excluded_subdir:
            return False
        # For files in subdirectories: check parent folder prefix match
        if rel_parts:
            parent_name = rel_parts[0]
            parent_prefix = parent_name.split('_')[0]
            # Only include if parent has a doc-ID style name and filename matches
            if '_' in parent_name:
                if not filename.startswith(parent_prefix):
                    return False
            else:
                # Non-standard folder (e.g. "Geotech Report") → skip in fallback
                return False
        return True

    def _is_report(self, filename: str) -> bool:
        if not filename.lower().endswith('.pdf'):
            return False
        for pattern in self.REPORT_PATTERNS:
            if re.search(pattern, filename, re.IGNORECASE):
                return True
        return False

    def _should_copy(self, source: Path, dest: Path) -> bool:
        if not dest.exists():
            return True
        try:
            src_stat = source.stat()
            dst_stat = dest.stat()
            if src_stat.st_size != dst_stat.st_size:
                return True
            if src_stat.st_mtime > dst_stat.st_mtime:
                return True
            return False
        except:
            return True


# =============================================================================
# Version Manager — PDF Mode (merged from version_manager_v4.py)
# =============================================================================

class VersionManager:
    """Manage file versions by moving older versions to SS folder."""

    VERSION_PATTERNS = [
        r'[_\s]Rev\.?\s*([A-Z0-9]+)',
        r'[_\s]R([A-Z])(?=[_\.\s]|$)',
        r'[_\s]R(\d+)(?=[_\.\s]|$)',
        r'[_\s]V(\d+)',
        r'[_\s]v(\d+)',
        r'-([A-Z])(?=[_\.\s]|$)',
    ]

    TARGET_SUBDIRS = [
        r"Design\Engineering\1. Drawings\2. IFR_internal",
        r"Design\Engineering\1. Drawings\3. IFR(Client)\1.Drawing",
        r"Design\Engineering\1. Drawings\3. IFR(Client)\2.Reports",
        r"Design\Engineering\1. Drawings\3. IFR(Client)\3.Deliverables",
    ]

    PROJECT_PATTERNS = [
        r'^GG-?\d+',
        r'^NSW\d+',
        r'^[A-Z]{2,4}-\d+',
    ]

    def __init__(self, root_path: str, dry_run: bool = False):
        self.root_path = Path(root_path)
        self.dry_run = dry_run
        self.total_stats = {"scanned": 0, "groups": 0, "moved": 0}

    def extract_base_name_and_version(self, filename: str) -> Tuple[str, Optional[str]]:
        name_without_ext = Path(filename).stem
        extension = Path(filename).suffix
        for pattern in self.VERSION_PATTERNS:
            match = re.search(pattern, name_without_ext, re.IGNORECASE)
            if match:
                version = match.group(1).upper()
                base_name = re.sub(pattern, '', name_without_ext, flags=re.IGNORECASE)
                base_name = re.sub(r'[_\s]+$', '', base_name)
                return (base_name + extension, version)
        return (filename, None)

    def scan_for_projects(self) -> List[Path]:
        projects = []
        if not self.root_path.exists():
            return projects

        def scan_level(path: Path, depth: int = 0, max_depth: int = 3):
            if depth > max_depth:
                return
            try:
                for item in path.iterdir():
                    if not item.is_dir():
                        continue
                    try:
                        ifr_path = item / "Design" / "Engineering" / "1. Drawings"
                        if ifr_path.exists():
                            for pattern in self.PROJECT_PATTERNS:
                                if re.match(pattern, item.name, re.IGNORECASE):
                                    projects.append(item)
                                    break
                            else:
                                if (ifr_path / "2. IFR_internal").exists() or (ifr_path / "3. IFR(Client)").exists():
                                    projects.append(item)
                        else:
                            scan_level(item, depth + 1, max_depth)
                    except (OSError, PermissionError):
                        continue
            except (OSError, PermissionError):
                pass

        scan_level(self.root_path)
        projects = sorted(set(projects), key=lambda p: p.name)
        return projects

    def scan_files_pdf(self, target_dir: Path) -> Dict[str, List[Tuple[Path, str, datetime]]]:
        if not target_dir.exists():
            return {}
        file_groups: Dict[str, List[Tuple[Path, str, datetime]]] = defaultdict(list)
        try:
            pdf_files = list(target_dir.glob("*.pdf"))
        except (OSError, PermissionError):
            return file_groups
        for file_path in pdf_files:
            try:
                if file_path.is_file():
                    base_name, version = self.extract_base_name_and_version(file_path.name)
                    pdf_long = to_long_path(file_path)
                    modified_time = datetime.fromtimestamp(pdf_long.stat().st_mtime)
                    version_str = version if version else "NO_VERSION"
                    file_groups[base_name].append((file_path, version_str, modified_time))
            except (OSError, PermissionError):
                continue
        return file_groups

    def identify_old_versions(self, file_groups: Dict, ss_folder: Path) -> List[Tuple[Path, Path, str]]:
        files_to_move = []
        for base_name, files in file_groups.items():
            if len(files) <= 1:
                continue
            files_sorted = sorted(files, key=lambda x: x[2], reverse=True)
            newest_file = files_sorted[0]
            for file_path, version, modified_time in files_sorted[1:]:
                dest_path = ss_folder / file_path.name
                reason = f"较旧版本 (最新: {newest_file[0].name})"
                files_to_move.append((file_path, dest_path, reason))
        return files_to_move

    def identify_ifc_in_ifr_client(self, target_dir: Path, ss_folder: Path) -> List[Tuple[Path, Path, str]]:
        """识别 IFR(Client) 目录中的 IFC 文件，标记移动到 SS。"""
        files_to_move = []
        if "IFR(Client)" not in str(target_dir):
            return files_to_move
        try:
            pdf_files = list(target_dir.glob("*.pdf"))
        except (OSError, PermissionError):
            return files_to_move
        for pdf_file in pdf_files:
            stem = pdf_file.stem
            is_ifc = bool(_RE_IFC_FILE.search(stem))
            is_ifr = bool(_RE_IFR_FILE.search(stem))
            if is_ifc and not is_ifr:
                dest_path = ss_folder / pdf_file.name
                reason = "IFC文件不应在IFR(Client)目录中"
                files_to_move.append((pdf_file, dest_path, reason))
        return files_to_move

    def move_files(self, files_to_move: List[Tuple[Path, Path, str]], ss_folder: Path) -> int:
        if not files_to_move:
            return 0
        if not self.dry_run:
            try:
                ss_folder.mkdir(exist_ok=True)
            except (OSError, PermissionError):
                try:
                    to_long_path(ss_folder).mkdir(exist_ok=True)
                except Exception as e:
                    print(f"      [!] 无法创建 SS 文件夹: {e}")
                    return 0
        moved_count = 0
        for source, dest, reason in files_to_move:
            if self.dry_run:
                print(f"      [预览] {source.name} -> SS/")
            else:
                try:
                    if dest.exists():
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        new_name = f"{dest.stem}_{timestamp}{dest.suffix}"
                        dest = ss_folder / new_name
                    shutil.move(str(to_long_path(source)), str(to_long_path(dest)))
                    print(f"      [v] 已移动: {source.name} -> SS/")
                    moved_count += 1
                except Exception as e:
                    print(f"      [!] 移动失败: {source.name} - {e}")

        # Cleanup old timestamped versions in SS (keep max 3 per base file)
        if not self.dry_run and ss_folder.exists():
            self._cleanup_ss_folder(ss_folder, max_versions=3)

        return moved_count

    @staticmethod
    def _cleanup_ss_folder(ss_path: Path, max_versions: int = 3):
        """Remove old timestamped copies in SS, keeping only the most recent N per base file."""
        ts_pattern = re.compile(r'^(.+?)_(\d{8}_\d{6})(\.\w+)$')
        groups = defaultdict(list)
        for f in ss_path.iterdir():
            if not f.is_file():
                continue
            m = ts_pattern.match(f.name)
            if m:
                base = m.group(1) + m.group(3)
                groups[base].append(f)
        for base, files in groups.items():
            if len(files) <= max_versions:
                continue
            files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            for old_file in files[max_versions:]:
                try:
                    old_file.unlink()
                except Exception:
                    pass

    def analyze_directory(self, target_dir: Path) -> Dict:
        result = {
            "exists": target_dir.exists(), "total_files": 0,
            "multi_version_groups": {}, "files_to_move": 0,
            "ifc_files": [], "ifc_files_count": 0
        }
        if not target_dir.exists():
            return result
        file_groups = self.scan_files_pdf(target_dir)
        result["total_files"] = sum(len(files) for files in file_groups.values())
        multi_version_groups = {k: v for k, v in file_groups.items() if len(v) > 1}
        result["multi_version_groups"] = multi_version_groups
        result["files_to_move"] = sum(len(v) - 1 for v in multi_version_groups.values())
        ss_folder = target_dir / "SS"
        ifc_files = self.identify_ifc_in_ifr_client(target_dir, ss_folder)
        result["ifc_files"] = ifc_files
        result["ifc_files_count"] = len(ifc_files)
        result["files_to_move"] += len(ifc_files)
        return result

    def process_directory(self, target_dir: Path, show_details: bool = True) -> Dict[str, int]:
        stats = {"scanned": 0, "groups": 0, "moved": 0}
        if not target_dir.exists():
            if show_details:
                print(f"    [!] 目录不存在，跳过")
            return stats
        ss_folder = target_dir / "SS"
        file_groups = self.scan_files_pdf(target_dir)
        total_files = sum(len(files) for files in file_groups.values())
        stats["scanned"] = total_files
        if show_details:
            print(f"    找到 {total_files} 个 PDF 文件")
        multi_version_groups = {k: v for k, v in file_groups.items() if len(v) > 1}
        stats["groups"] = len(multi_version_groups)
        ifc_files = self.identify_ifc_in_ifr_client(target_dir, ss_folder)
        if not multi_version_groups and not ifc_files:
            if show_details:
                print(f"    [v] 没有重复版本文件")
            return stats
        if multi_version_groups and show_details:
            print(f"    发现 {len(multi_version_groups)} 组多版本文件:")
            for base_name, files in multi_version_groups.items():
                files_sorted = sorted(files, key=lambda x: x[2], reverse=True)
                print(f"\n      基础文件: {base_name}")
                for i, (file_path, version, modified_time) in enumerate(files_sorted):
                    status = "[保留]" if i == 0 else "[->SS]"
                    time_str = modified_time.strftime('%Y-%m-%d %H:%M')
                    print(f"        {status} {file_path.name} (版本:{version}, {time_str})")
        files_to_move = self.identify_old_versions(file_groups, ss_folder)
        if files_to_move:
            if show_details:
                print(f"\n    移动 {len(files_to_move)} 个旧版本到 SS 文件夹:")
            moved_count = self.move_files(files_to_move, ss_folder)
            stats["moved"] = moved_count
        if ifc_files:
            if show_details:
                print(f"\n    发现 {len(ifc_files)} 个 IFC 文件在 IFR(Client) 目录中:")
                for src, dst, reason in ifc_files:
                    print(f"        [->SS] {src.name} ({reason})")
            moved_ifc = self.move_files(ifc_files, ss_folder)
            stats["moved"] += moved_ifc
        return stats

    def process_project(self, project_path: Path, show_details: bool = True) -> Dict[str, int]:
        project_stats = {"scanned": 0, "groups": 0, "moved": 0}
        for subdir in self.TARGET_SUBDIRS:
            target_dir = project_path / subdir
            if show_details:
                short_path = subdir.split("\\")[-1] if "\\" in subdir else subdir
                print(f"\n  [{short_path}]")
            stats = self.process_directory(target_dir, show_details)
            project_stats["scanned"] += stats["scanned"]
            project_stats["groups"] += stats["groups"]
            project_stats["moved"] += stats["moved"]
        return project_stats


# =============================================================================
# Native Version Manager (merged from version_manager_v4.py)
# =============================================================================

# Scan root configs
SCAN_ROOTS = [
    ("Design/Engineering/1. Drawings/1. Native", 1, "1. Native"),
    ("Design/Engineering/2. Calcs & Reports/Reports", 2, "Reports"),
    ("Design/Engineering/2. Calcs & Reports/Schedule", 1, "Schedule"),
]

_SKIP_FOLDER_NAMES = {
    'ss', 'superseded', 'superceded',
    'bom', 'appendix', 'approved',
}
_SKIP_FOLDER_PREFIXES = ('sy supply',)
_IGNORE_EXT = {'.dwl', '.dwl2', '.err', '.log', '.tmp', '.lnk', '.db', '.ini'}


@dataclass
class DwgFile:
    """A versioned file with parsed metadata."""
    path: Path
    filename: str
    rev_type: str        # 'IFR', 'IFC', or 'OTHER'
    revision: str
    mtime: datetime
    bak_path: Optional[Path] = None

    @property
    def rev_display(self) -> str:
        if self.rev_type == 'IFR':
            return f"Rev{self.revision}"
        elif self.rev_type == 'IFC':
            return f"Rev{self.revision}_IFC"
        return "no-rev"


@dataclass
class NativeFileAction:
    source: Path
    dest: Path
    action: str   # 'rename', 'move_to_ss'
    reason: str


@dataclass
class FolderRelocation:
    source: Path
    dest: Path
    reason: str
    action: str   # 'move', 'merge', 'warn'


@dataclass
class NativeFolderResult:
    folder_name: str
    folder_path: Path
    doc_id: str
    description: str
    dwg_files: List[DwgFile] = field(default_factory=list)
    actions: List[NativeFileAction] = field(default_factory=list)
    kept_ifr: Optional[DwgFile] = None
    kept_ifc: Optional[DwgFile] = None


def _classify_dwg(filename: str) -> Tuple[str, str]:
    """Classify a filename as IFR, IFC, or OTHER."""
    stem = Path(filename).stem
    m = _RE_IFC_FILE.search(stem)
    if m:
        return ('IFC', m.group(1))
    m = _RE_IFR_FILE.search(stem)
    if m:
        return ('IFR', m.group(1).upper())
    return ('OTHER', '')


def _parse_folder_name(folder_name: str) -> Tuple[str, str]:
    m = re.match(r'^([A-Z0-9][\w-]+-\d{3})[_\s](.+?)(?:[_\s]Rev.*)?$', folder_name, re.IGNORECASE)
    if m:
        return (m.group(1), m.group(2).strip())
    if '_' in folder_name:
        parts = folder_name.split('_', 1)
        return (parts[0], parts[1].rstrip().rstrip('_Rev').strip())
    return (folder_name, '')


def _make_standard_dwg_name(doc_id: str, description: str, rev_type: str,
                            revision: str, ext: str = '.dwg') -> str:
    if rev_type == 'IFC':
        return f"{doc_id}_{description}_Rev{revision}_IFC{ext}"
    return f"{doc_id}_{description}_Rev{revision.upper()}{ext}"


class NativeVersionManager:
    """Manage file versions across Native, Reports, and Schedule folders."""

    def __init__(self, project_path: str, dry_run: bool = True):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.results: List[Tuple[str, NativeFolderResult]] = []

    @property
    def native_root(self) -> Path:
        return self.project_path / "Design" / "Engineering" / "1. Drawings" / "1. Native"

    def find_doc_folders(self, folder_filter: Optional[str] = None,
                         scope: str = 'all') -> List[Tuple[str, Path]]:
        results = []
        for rel_path, depth, display_name in SCAN_ROOTS:
            if scope != 'all':
                if scope == 'native' and 'Native' not in rel_path:
                    continue
                if scope == 'reports' and 'Reports' not in rel_path:
                    continue
                if scope == 'schedule' and 'Schedule' not in rel_path:
                    continue
            root = self.project_path / rel_path
            if not root.exists():
                continue
            if depth == 1:
                candidates = self._list_doc_dirs(root)
            elif depth == 2:
                candidates = []
                for category_dir in sorted(root.iterdir()):
                    if category_dir.is_dir() and not self._should_skip(category_dir.name):
                        candidates.extend(self._list_doc_dirs(category_dir))
            else:
                candidates = []
            for folder in candidates:
                if folder_filter and folder_filter.lower() not in folder.name.lower():
                    continue
                results.append((display_name, folder))
        return results

    def _list_doc_dirs(self, parent: Path) -> List[Path]:
        dirs = []
        try:
            for item in sorted(parent.iterdir(), key=lambda p: p.name):
                if item.is_dir() and not self._should_skip(item.name):
                    dirs.append(item)
        except (OSError, PermissionError):
            pass
        return dirs

    @staticmethod
    def _should_skip(name: str) -> bool:
        if name.startswith('.'):
            return True
        low = name.lower()
        if low in _SKIP_FOLDER_NAMES:
            return True
        for prefix in _SKIP_FOLDER_PREFIXES:
            if low.startswith(prefix):
                return True
        return False

    @staticmethod
    def _find_or_create_ss_folder(folder: Path, create: bool = False) -> Path:
        try:
            for item in folder.iterdir():
                if item.is_dir() and item.name.lower() in ('ss', 'superseded', 'superceded'):
                    return item
        except (OSError, PermissionError):
            pass
        ss_path = folder / 'SS'
        if create:
            to_long_path(ss_path).mkdir(exist_ok=True)
        return ss_path

    def scan_files(self, folder: Path) -> List[DwgFile]:
        files = []
        bak_map: Dict[str, Path] = {}
        ss_names = {'ss', 'superseded', 'superceded'}
        try:
            all_items = list(folder.iterdir())
        except (OSError, PermissionError):
            return []
        for f in all_items:
            if f.is_file() and f.suffix.lower() == '.bak':
                bak_map[f.stem.lower()] = f
        for f in all_items:
            if not f.is_file():
                continue
            if f.parent.name.lower() in ss_names:
                continue
            ext = f.suffix.lower()
            if ext in _IGNORE_EXT or ext == '.bak':
                continue
            rev_type, revision = _classify_dwg(f.name)
            try:
                mtime = datetime.fromtimestamp(to_long_path(f).stat().st_mtime)
            except (OSError, PermissionError):
                mtime = datetime.min
            files.append(DwgFile(
                path=f, filename=f.name,
                rev_type=rev_type, revision=revision,
                mtime=mtime, bak_path=bak_map.get(f.stem.lower()),
            ))
        return files

    def process_folder(self, folder: Path) -> NativeFolderResult:
        doc_id, description = _parse_folder_name(folder.name)
        result = NativeFolderResult(
            folder_name=folder.name, folder_path=folder,
            doc_id=doc_id or folder.name, description=description or '',
        )
        files = self.scan_files(folder)
        result.dwg_files = files
        if not files:
            return result
        ifr_files = [f for f in files if f.rev_type == 'IFR']
        ifc_files = [f for f in files if f.rev_type == 'IFC']
        other_files = [f for f in files if f.rev_type == 'OTHER']
        ss_folder = self._find_or_create_ss_folder(folder)
        if ifr_files:
            ifr_sorted = sorted(ifr_files, key=lambda f: f.mtime, reverse=True)
            result.kept_ifr = ifr_sorted[0]
            for old in ifr_sorted[1:]:
                self._plan_move(result, old, ss_folder, "older IFR revision")
        if ifc_files:
            ifc_sorted = sorted(ifc_files, key=lambda f: f.mtime, reverse=True)
            result.kept_ifc = ifc_sorted[0]
            for old in ifc_sorted[1:]:
                self._plan_move(result, old, ss_folder, "older IFC revision")
        for other in other_files:
            if len(files) > 1:
                self._plan_move(result, other, ss_folder, "unversioned/legacy")
        if result.kept_ifr and doc_id and description:
            self._plan_rename(result, result.kept_ifr, doc_id, description)
        if result.kept_ifc and doc_id and description:
            self._plan_rename(result, result.kept_ifc, doc_id, description)
        return result

    def _plan_move(self, result, dwg, ss_folder, reason):
        result.actions.append(NativeFileAction(
            source=dwg.path, dest=ss_folder / dwg.filename,
            action='move_to_ss', reason=reason
        ))
        if dwg.bak_path and dwg.bak_path.exists():
            result.actions.append(NativeFileAction(
                source=dwg.bak_path, dest=ss_folder / dwg.bak_path.name,
                action='move_to_ss', reason=f".bak follows {dwg.filename}"
            ))

    def _plan_rename(self, result, dwg, doc_id, description):
        ext = dwg.path.suffix
        standard = _make_standard_dwg_name(doc_id, description, dwg.rev_type,
                                           dwg.revision, ext)
        if dwg.filename != standard:
            result.actions.append(NativeFileAction(
                source=dwg.path, dest=dwg.path.parent / standard,
                action='rename', reason=f"标准化 -> {standard}"
            ))
            if dwg.bak_path and dwg.bak_path.exists():
                bak_std = _make_standard_dwg_name(doc_id, description,
                                                  dwg.rev_type, dwg.revision, '.bak')
                if dwg.bak_path.name != bak_std:
                    result.actions.append(NativeFileAction(
                        source=dwg.bak_path, dest=dwg.bak_path.parent / bak_std,
                        action='rename', reason=f".bak 标准化 -> {bak_std}"
                    ))

    def process_all(self, folder_filter=None, scope='all'):
        doc_folders = self.find_doc_folders(folder_filter, scope)
        results = []
        for group_name, folder in doc_folders:
            result = self.process_folder(folder)
            results.append((group_name, result))
        self.results = results
        return results

    def execute_actions(self, results):
        stats = {'renamed': 0, 'moved': 0, 'errors': 0, 'ss_created': 0}
        for _group_name, result in results:
            if not result.actions:
                continue
            moves = [a for a in result.actions if a.action == 'move_to_ss']
            if moves:
                ss_folder = self._find_or_create_ss_folder(result.folder_path, create=True)
                if not ss_folder.exists():
                    try:
                        to_long_path(ss_folder).mkdir(exist_ok=True)
                        stats['ss_created'] += 1
                    except Exception as e:
                        print(f"      [!] 无法创建 SS: {e}")
                        stats['errors'] += 1
                        continue
                for action in moves:
                    action.dest = ss_folder / action.source.name
            for action in result.actions:
                try:
                    src = str(to_long_path(action.source))
                    dst_path = action.dest
                    if to_long_path(dst_path).exists():
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        dst_path = dst_path.parent / f"{dst_path.stem}_{ts}{dst_path.suffix}"
                    dst = str(to_long_path(dst_path))
                    shutil.move(src, dst)
                    if action.action == 'rename':
                        stats['renamed'] += 1
                    else:
                        stats['moved'] += 1
                except Exception as e:
                    print(f"      [!] 失败 {action.source.name}: {e}")
                    stats['errors'] += 1
        return stats


# =============================================================================
# Folder Relocator (merged from version_manager_v4.py)
# =============================================================================

_L1_PATTERNS = [
    (re.compile(r'^1[\.\s]*Native', re.IGNORECASE),        '1. Native'),
    (re.compile(r'^2[\.\s]*IFR[_\s]*internal', re.IGNORECASE), '2. IFR_internal'),
    (re.compile(r'^3[\.\s]*IFR\s*\(Client\)', re.IGNORECASE), '3. IFR(Client)'),
    (re.compile(r'^4[\.\s]*IFC\s*\(Client\)', re.IGNORECASE), '4. IFC(Client)'),
]

_L2_PATTERNS = [
    (re.compile(r'^1[\.\s]*Drawing', re.IGNORECASE),       '1.Drawing'),
    (re.compile(r'^2[\.\s]*Reports?', re.IGNORECASE),      '2.Reports'),
    (re.compile(r'^3[\.\s]*Deliverables?', re.IGNORECASE), '3.Deliverables'),
]

_FLAT_L1 = {'1. Native', '2. IFR_internal', '4. IFC(Client)'}
_IFR_CLIENT_CHILDREN = {'1.Drawing', '2.Reports', '3.Deliverables'}


class FolderRelocator:
    """Detect and relocate misplaced structural folders under 1. Drawings/."""

    def __init__(self, project_path: Path, dry_run: bool = True):
        self.project_path = project_path
        self.drawings_root = project_path / "Design" / "Engineering" / "1. Drawings"
        self.dry_run = dry_run

    @staticmethod
    def _match_l1(name: str) -> Optional[str]:
        for pat, canonical in _L1_PATTERNS:
            if pat.match(name):
                return canonical
        return None

    @staticmethod
    def _match_l2(name: str) -> Optional[str]:
        for pat, canonical in _L2_PATTERNS:
            if pat.match(name):
                return canonical
        return None

    def _ifr_client_path(self) -> Path:
        try:
            for item in self.drawings_root.iterdir():
                if item.is_dir() and self._match_l1(item.name) == '3. IFR(Client)':
                    return item
        except (OSError, PermissionError):
            pass
        return self.drawings_root / '3. IFR(Client)'

    def scan(self) -> List[FolderRelocation]:
        relocations: List[FolderRelocation] = []
        if not self.drawings_root.exists():
            return relocations
        ifr_client = self._ifr_client_path()
        try:
            depth0_items = sorted(self.drawings_root.iterdir(), key=lambda p: p.name)
        except (OSError, PermissionError):
            return relocations
        for item in depth0_items:
            if not item.is_dir():
                continue
            l1_match = self._match_l1(item.name)
            l2_match = self._match_l2(item.name)
            if l2_match:
                dest = ifr_client / l2_match
                relocations.append(FolderRelocation(
                    source=item, dest=dest,
                    reason=f"Level-2 文件夹 '{item.name}' 不应直接在 1. Drawings/ 下",
                    action='merge' if dest.exists() else 'move',
                ))
                continue
            if not l1_match:
                continue
            try:
                depth1_items = sorted(item.iterdir(), key=lambda p: p.name)
            except (OSError, PermissionError):
                continue
            for child in depth1_items:
                if not child.is_dir():
                    continue
                child_l1 = self._match_l1(child.name)
                child_l2 = self._match_l2(child.name)
                if child_l1:
                    dest = self.drawings_root / child_l1
                    relocations.append(FolderRelocation(
                        source=child, dest=dest,
                        reason=f"Level-1 文件夹 '{child.name}' 被误放在 '{item.name}/' 内",
                        action='merge' if dest.exists() else 'move',
                    ))
                elif child_l2 and l1_match in _FLAT_L1:
                    dest = ifr_client / child_l2
                    relocations.append(FolderRelocation(
                        source=child, dest=dest,
                        reason=f"'{child.name}' 不应在 flat 文件夹 '{item.name}/' 内",
                        action='merge' if dest.exists() else 'move',
                    ))
                elif child_l2 and l1_match == '3. IFR(Client)':
                    self._scan_deep(child, item, ifr_client, relocations)
        if ifr_client.exists():
            existing_children = set()
            try:
                for c in ifr_client.iterdir():
                    if c.is_dir():
                        m = self._match_l2(c.name)
                        if m:
                            existing_children.add(m)
            except (OSError, PermissionError):
                pass
            for expected in _IFR_CLIENT_CHILDREN:
                if expected not in existing_children:
                    relocations.append(FolderRelocation(
                        source=ifr_client / expected,
                        dest=ifr_client / expected,
                        reason=f"3. IFR(Client)/ 缺少子文件夹 '{expected}'",
                        action='warn',
                    ))
        return relocations

    def _scan_deep(self, folder, parent_l1, ifr_client, relocations):
        try:
            children = sorted(folder.iterdir(), key=lambda p: p.name)
        except (OSError, PermissionError):
            return
        for child in children:
            if not child.is_dir():
                continue
            child_l1 = self._match_l1(child.name)
            child_l2 = self._match_l2(child.name)
            if child_l1:
                dest = self.drawings_root / child_l1
                relocations.append(FolderRelocation(
                    source=child, dest=dest,
                    reason=f"Level-1 '{child.name}' 嵌套在 '{folder.name}/' 内 (深层错位)",
                    action='merge' if dest.exists() else 'move',
                ))
            elif child_l2:
                dest = ifr_client / child_l2
                relocations.append(FolderRelocation(
                    source=child, dest=dest,
                    reason=f"Level-2 '{child.name}' 嵌套在 '{folder.name}/' 内 (深层错位)",
                    action='merge' if dest.exists() else 'move',
                ))

    def execute(self, relocations):
        stats = {'moved': 0, 'merged': 0, 'warnings': 0, 'errors': 0}
        for rel in relocations:
            if rel.action == 'warn':
                stats['warnings'] += 1
                continue
            try:
                src = to_long_path(rel.source)
                dst = to_long_path(rel.dest)
                if rel.action == 'move':
                    dst.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(src), str(dst))
                    print(f"  [v] 已移动: {rel.source.name} -> {rel.dest}")
                    stats['moved'] += 1
                elif rel.action == 'merge':
                    dst.mkdir(parents=True, exist_ok=True)
                    moved_children = 0
                    try:
                        for child in src.iterdir():
                            child_dst = dst / child.name
                            if child_dst.exists():
                                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                                if child.is_file():
                                    child_dst = dst / f"{child.stem}_{ts}{child.suffix}"
                                else:
                                    child_dst = dst / f"{child.name}_{ts}"
                            shutil.move(str(child), str(child_dst))
                            moved_children += 1
                    except Exception as e:
                        print(f"  [!] 合并部分失败 {rel.source.name}: {e}")
                        stats['errors'] += 1
                    try:
                        if src.exists() and not any(src.iterdir()):
                            src.rmdir()
                    except (OSError, PermissionError):
                        pass
                    print(f"  [v] 已合并: {rel.source.name} -> {rel.dest} ({moved_children} 项)")
                    stats['merged'] += 1
            except Exception as e:
                print(f"  [!] 失败: {rel.source.name} - {e}")
                stats['errors'] += 1
        return stats


# =============================================================================
# Deliverable Manager (NEW in v7)
# =============================================================================

class FormatChangeWarning(Exception):
    """Raised when deliverable Excel layout cannot be auto-detected."""
    pass


@dataclass
class DeliverableLayout:
    """Detected Excel layout for a deliverable file."""
    header_row: int           # Row number where headers are found
    rev_col: int              # Column index for Revision (1-based)
    date_col: int             # Column index for Submission Date
    status_col: int           # Column index for Status
    doc_id_col: int           # Column index for Doc ID (usually B=2)
    desc_col: int             # Column index for Description (usually C=3)
    file_rev_cell: str        # Cell address for file revision (e.g. 'K1')
    last_updated_cell: str    # Cell address for last updated date (e.g. 'C79')


@dataclass
class DeliverableCrossCheckResult:
    """Result of deliverable cross-check."""
    project_name: str
    excel_path: str
    items_in_folders_not_excel: List[Dict] = field(default_factory=list)
    items_in_excel_not_folders: List[str] = field(default_factory=list)
    revision_mismatches: List[Dict] = field(default_factory=list)
    doc_id_corrections: List[Dict] = field(default_factory=list)  # rows where FILE NO was corrected
    status_updates: List[Dict] = field(default_factory=list)  # IFC status updates
    naming_warnings: List[Dict] = field(default_factory=list)  # filename normalization warnings
    rows_inserted: int = 0
    rows_updated: int = 0
    new_file_rev: str = ""
    errors: List[str] = field(default_factory=list)


class DeliverableManager:
    """Cross-check and update deliverable Excel files against project folders."""

    # Doc ID extraction patterns
    _RE_DOC_ID_GG = re.compile(r'^(GG\d{2}-[A-Z]-[A-Z]{3}-\d{3})', re.IGNORECASE)
    _RE_DOC_ID_LMS = re.compile(r'^(\d{5}-[A-Z]{2}-\d{3})', re.IGNORECASE)
    _RE_DOC_ID_GENERIC = re.compile(r'^([A-Z0-9][\w]+-[A-Z]+-[A-Z]*-?\d{3})', re.IGNORECASE)

    # Revision extraction from filename
    _RE_REV_LETTER = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*([A-Z])(?=[_.\s]|$)', re.IGNORECASE)
    _RE_REV_NUMBER = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?=[_.\s]|$)', re.IGNORECASE)

    # Header pattern to detect column layout
    HEADER_KEYWORDS = ['revision', 'submission date', 'status']

    HIGHLIGHT_FILL = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99',
                                 fill_type='solid') if OPENPYXL_AVAILABLE else None
    CHANGE_MARKER_FILL = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99',
                                     fill_type='solid') if OPENPYXL_AVAILABLE else None
    STATUS_APPROVED_FILL = PatternFill(start_color='FF00B050', end_color='FF00B050',
                                       fill_type='solid') if OPENPYXL_AVAILABLE else None
    STATUS_SUBMITTED_FILL = PatternFill(start_color='FFFFC000', end_color='FFFFC000',
                                        fill_type='solid') if OPENPYXL_AVAILABLE else None

    # IFC folder path (relative to project)
    IFC_FOLDER = "Design/Engineering/1. Drawings/4. IFC(Client)"

    # Source folders to scan for deliverable items
    SOURCE_FOLDERS = [
        "Design/Engineering/1. Drawings/2. IFR_internal",
        "Design/Engineering/2. Calcs & Reports/Reports/Electrical",
        "Design/Engineering/2. Calcs & Reports/Reports/Civil & Structure",
        "Design/Engineering/2. Calcs & Reports/Schedule",
    ]

    # Deliverable folder path (relative to project)
    DELIVERABLE_REL_PATH = "Design/Engineering/1. Drawings/3. IFR(Client)/3.Deliverables"

    # Sync target paths
    PRIMARY_SYNC = "Design/Engineering/1. Drawings/3. IFR(Client)/3.Deliverables"
    SECONDARY_SYNC = "Design/Engineering/13. Client Sharepoint/1.IFR/3.Deliverables"

    def __init__(self, project_path: Path, dry_run: bool = False):
        self.project_path = project_path
        self.dry_run = dry_run

    # All possible locations for the deliverable Excel (searched in order)
    DELIVERABLE_SEARCH_PATHS = [
        "Design/Engineering/8. Deliverables",
        "Design/Engineering/1. Drawings/3. IFR(Client)/3.Deliverables",
        "Design/Engineering/13. Client Sharepoint/1.IFR/3.Deliverables",
    ]

    def find_deliverable_excel(self) -> Optional[Path]:
        """Find the deliverable Excel file across known locations.

        Searches: 8. Deliverables (master) → 3.Deliverables → 13. Client Sharepoint.
        """
        for rel_path in self.DELIVERABLE_SEARCH_PATHS:
            dlv_path = self.project_path / rel_path
            found = self._find_excel_in_folder(dlv_path)
            if found:
                return found
        return None

    def _find_excel_in_folder(self, folder: Path) -> Optional[Path]:
        """Find a deliverable Excel in a specific folder."""
        if not folder.exists():
            return None
        # Priority 1: filename contains 'dlv' or 'deliverable'
        for f in folder.iterdir():
            if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm') and not f.name.startswith('~$'):
                if 'dlv' in f.name.lower() or 'deliverable' in f.name.lower():
                    return f
        # Priority 2: any xlsx that isn't a temp file
        for f in folder.iterdir():
            if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm') and not f.name.startswith('~$'):
                return f
        return None

    def detect_layout(self, ws) -> DeliverableLayout:
        """Auto-detect Excel column layout by scanning rows 1-20 for header pattern."""
        for row_idx in range(1, 21):
            for col_idx in range(1, 20):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val and isinstance(cell_val, str) and 'revision' in cell_val.lower():
                    # Check if next two columns match pattern
                    next1 = ws.cell(row=row_idx, column=col_idx + 1).value
                    next2 = ws.cell(row=row_idx, column=col_idx + 2).value
                    if (next1 and isinstance(next1, str) and 'date' in next1.lower() and
                        next2 and isinstance(next2, str) and 'status' in next2.lower()):
                        # Found the header pattern
                        layout = DeliverableLayout(
                            header_row=row_idx,
                            rev_col=col_idx,
                            date_col=col_idx + 1,
                            status_col=col_idx + 2,
                            doc_id_col=2,  # Column B
                            desc_col=3,    # Column C
                            file_rev_cell=f"{openpyxl.utils.get_column_letter(col_idx)}1",
                            last_updated_cell=self._detect_last_updated_cell(ws, col_idx),
                        )
                        return layout
        raise FormatChangeWarning(
            f"无法在前20行检测到 'Revision | Submission Date | Status' 列头模式"
        )

    def _detect_last_updated_cell(self, ws, rev_col: int) -> str:
        """Try to find the 'last updated' date cell by scanning for the label."""
        # Search rows 1-10 for "Last Updated" label, return the cell to its right
        for row in range(1, 11):
            for col in range(1, 20):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and 'last updated' in val.lower():
                    date_col = col + 1
                    return f'{openpyxl.utils.get_column_letter(date_col)}{row}'
        # Fallback: search near top for date-like values
        for row in range(1, 10):
            for col in range(1, 20):
                val = ws.cell(row=row, column=col).value
                if isinstance(val, datetime):
                    return f'{openpyxl.utils.get_column_letter(col)}{row}'
        return 'L2'  # Reasonable default

    def extract_doc_id(self, filename: str) -> Optional[str]:
        """Extract doc ID from filename."""
        stem = Path(filename).stem
        for pattern in [self._RE_DOC_ID_GG, self._RE_DOC_ID_LMS, self._RE_DOC_ID_GENERIC]:
            m = pattern.match(stem)
            if m:
                return m.group(1)
        # Fallback: take everything before first _Rev or _rev
        m = re.match(r'^(.+?)(?:[_\s]-?[Rr]ev)', stem)
        if m:
            return m.group(1).rstrip('_- ')
        return None

    def extract_revision(self, filename: str) -> Optional[str]:
        """Extract revision from filename."""
        stem = Path(filename).stem
        m = self._RE_REV_LETTER.search(stem)
        if m:
            return m.group(1).upper()
        m = self._RE_REV_NUMBER.search(stem)
        if m:
            return m.group(1)
        return None

    def extract_description(self, filename: str) -> str:
        """Extract description from filename (between doc_id and revision)."""
        stem = Path(filename).stem
        doc_id = self.extract_doc_id(filename)
        if not doc_id:
            return stem
        remainder = stem[len(doc_id):].lstrip('_- ')
        # Remove revision suffix
        remainder = re.sub(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*[A-Z0-9]+.*$', '', remainder, flags=re.IGNORECASE)
        return remainder.strip('_- ') or stem

    def scan_source_folders(self) -> Dict[str, Dict]:
        """Scan source folders and extract doc IDs with their revisions."""
        items = {}  # doc_id -> {revision, filename, source, description}
        for rel_path in self.SOURCE_FOLDERS:
            source_dir = self.project_path / rel_path
            if not source_dir.exists():
                continue
            try:
                for f in source_dir.rglob("*.pdf"):
                    if f.name.startswith('~$'):
                        continue
                    # Skip files in excluded subfolders
                    try:
                        rel_parts = f.relative_to(source_dir).parts
                        skip_exact = ('ss', 'superseded', 'superceded')
                        skip_prefix = ('appendix', 'reference', 'app ')
                        # ALWAYS skip SS/superseded — even inside appendix folders
                        if any(p.lower() in skip_exact for p in rel_parts[:-1]):
                            continue
                        # In appendix subfolders, only include files with a valid doc-ID
                        in_appendix = any(
                            any(p.lower().startswith(pfx) for pfx in skip_prefix)
                            for p in rel_parts[:-1]
                        )
                        if in_appendix and not self.extract_doc_id(f.name):
                            continue
                    except ValueError:
                        pass
                    # Skip IFC files (only IFR goes into deliverables)
                    if _is_ifc_file(f.name) and not _is_ifr_file(f.name):
                        continue
                    doc_id = self.extract_doc_id(f.name)
                    if not doc_id:
                        continue
                    rev = self.extract_revision(f.name)
                    # Keep the highest revision per doc_id
                    if doc_id in items:
                        existing_rev = items[doc_id]['revision']
                        if rev and existing_rev and self._compare_revisions(rev, existing_rev) > 0:
                            items[doc_id] = {
                                'revision': rev, 'filename': f.name,
                                'source': rel_path, 'description': self.extract_description(f.name)
                            }
                    else:
                        items[doc_id] = {
                            'revision': rev, 'filename': f.name,
                            'source': rel_path, 'description': self.extract_description(f.name)
                        }
            except (OSError, PermissionError):
                continue
        return items

    @staticmethod
    def _normalize_desc(desc: str) -> str:
        """Normalize description for fuzzy matching: lowercase, strip punctuation/spaces."""
        return re.sub(r'[^a-z0-9]', '', desc.lower())

    @staticmethod
    def _desc_similarity(a: str, b: str) -> float:
        """Sequence-based similarity ratio using difflib-style matching."""
        if not a or not b:
            return 0.0
        if a == b:
            return 1.0
        # Use difflib SequenceMatcher for accurate similarity
        from difflib import SequenceMatcher
        return SequenceMatcher(None, a, b).ratio()

    def _find_match_by_description(self, folder_doc_id: str, folder_desc: str,
                                    all_excel_rows: List[Dict],
                                    matched_rows: Set[int]) -> Optional[Dict]:
        """Try to find an Excel row matching by description when doc-ID doesn't match.
        Also requires that the doc-ID category prefix matches (e.g. both are E-BLD).
        Returns the row_info dict if a match is found, else None."""
        norm_folder = self._normalize_desc(folder_desc)
        if not norm_folder or len(norm_folder) < 5:
            return None
        # Extract category prefix (e.g. "E-BLD" from "GG31-E-BLD-003")
        folder_cat = re.sub(r'^[A-Z0-9]+-', '', folder_doc_id, count=1)  # strip project prefix
        folder_cat = re.sub(r'-\d+$', '', folder_cat)  # strip trailing number
        best_match = None
        best_ratio = 0.0
        for row_info in all_excel_rows:
            if row_info['row'] in matched_rows:
                continue
            excel_id = row_info['doc_id']
            # Check category prefix matches
            excel_cat = re.sub(r'^[A-Z0-9]+-', '', excel_id, count=1)
            excel_cat = re.sub(r'-\d+$', '', excel_cat)
            if folder_cat.upper() != excel_cat.upper():
                continue
            norm_excel = self._normalize_desc(row_info.get('description', ''))
            if not norm_excel:
                continue
            # Exact match
            if norm_folder == norm_excel:
                return row_info
            # Similarity check (handles minor differences like "communication" vs "communications")
            ratio = self._desc_similarity(norm_folder, norm_excel)
            if ratio > best_ratio and ratio >= 0.85:
                best_ratio = ratio
                best_match = row_info
        return best_match

    def _compare_revisions(self, rev_a: str, rev_b: str) -> int:
        """Compare two revision strings. Returns >0 if a>b, <0 if a<b, 0 if equal."""
        if rev_a == rev_b:
            return 0
        # Try numeric comparison
        try:
            return int(rev_a) - int(rev_b)
        except ValueError:
            pass
        # Letter comparison
        if rev_a.isalpha() and rev_b.isalpha():
            return ord(rev_a.upper()) - ord(rev_b.upper())
        # Mixed: letters come after numbers in IFR convention
        if rev_a.isalpha() and rev_b.isdigit():
            return 1
        if rev_a.isdigit() and rev_b.isalpha():
            return -1
        return 0

    def read_excel_items(self, ws, layout: DeliverableLayout) -> Tuple[Dict[str, Dict], List[Dict]]:
        """Read doc IDs and revisions from Excel.
        Returns (items_dict, all_rows_list).
        items_dict: doc_id -> info (last row wins for duplicates, used for revision matching)
        all_rows_list: every row's data (used for description matching with duplicates)
        """
        items = {}
        all_rows = []
        for row in range(layout.header_row + 1, ws.max_row + 1):
            doc_id_val = ws.cell(row=row, column=layout.doc_id_col).value
            if not doc_id_val or not isinstance(doc_id_val, str):
                continue
            doc_id = doc_id_val.strip()
            if not doc_id:
                continue
            rev_val = ws.cell(row=row, column=layout.rev_col).value
            status_val = ws.cell(row=row, column=layout.status_col).value
            desc_val = ws.cell(row=row, column=layout.desc_col).value
            row_info = {
                'doc_id': doc_id,
                'row': row,
                'revision': str(rev_val).strip() if rev_val else '',
                'status': str(status_val).strip() if status_val else '',
                'description': str(desc_val).strip() if desc_val else '',
            }
            items[doc_id] = row_info
            all_rows.append(row_info)
        return items, all_rows

    def _is_yellow_fill(self, cell) -> bool:
        """Check if a cell has yellow highlight fill."""
        if not cell.fill or not cell.fill.fgColor:
            return False
        color = cell.fill.fgColor
        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb)
            # Check for yellow-ish colors (FFFF99, FFFFFF00, etc)
            return rgb in ('FFFFFF99', '00FFFF99', 'FFFFFF00', '00FFFF00')
        return False

    def clear_previous_highlights(self, ws, layout: DeliverableLayout):
        """Clear change marker highlights (J=10 and N=14) and status fills (K-M) from previous runs."""
        no_fill = PatternFill(fill_type=None)
        clear_cols = [10, 14, layout.rev_col, layout.date_col, layout.status_col]
        for row in range(layout.header_row + 1, ws.max_row + 1):
            for col in clear_cols:
                cell = ws.cell(row=row, column=col)
                if self._is_yellow_fill(cell) or self._is_status_fill(cell):
                    cell.fill = no_fill

    def _is_status_fill(self, cell) -> bool:
        """Check if a cell has green or amber status fill."""
        if not cell.fill or not cell.fill.fgColor:
            return False
        color = cell.fill.fgColor
        if color.type == 'rgb' and color.rgb:
            rgb = str(color.rgb)
            return rgb in ('FF00B050', '00FF00B050', 'FFFFC000', '00FFC000')
        return False

    def scan_ifc_folder(self) -> Dict[str, Dict]:
        """Scan IFC(Client) folder and return {doc_id: {'revision': int, 'filename': str}}.

        For each doc_id, keeps only the highest numeric revision found.
        """
        ifc_info: Dict[str, Dict] = {}
        ifc_dir = self.project_path / self.IFC_FOLDER
        if not ifc_dir.exists():
            return ifc_info
        try:
            for f in ifc_dir.rglob("*.pdf"):
                if f.name.startswith('~$'):
                    continue
                doc_id = self.extract_doc_id(f.name)
                if not doc_id:
                    continue
                # Extract numeric revision
                m = self._RE_REV_NUMBER.search(f.name)
                rev_num = int(m.group(1)) if m else 0
                if doc_id not in ifc_info or rev_num > ifc_info[doc_id]['revision']:
                    ifc_info[doc_id] = {'revision': rev_num, 'filename': f.name}
        except (OSError, PermissionError):
            pass
        return ifc_info

    def normalize_filename(self, filename: str, deliverable_desc: Optional[str] = None) -> Tuple[str, List[str]]:
        """Normalize filename conventions. Returns (normalized_name, list_of_changes).
        Rules:
        1. After FILE NO, separator must be '_' (not '-' or ' ')
        2. Revision format: '_Rev' + letter/number, no space (e.g. _RevA, _RevC)
        3. Description should match deliverable Excel column C (optional)
        """
        stem = Path(filename).stem
        ext = Path(filename).suffix
        changes = []
        doc_id = self.extract_doc_id(filename)
        if not doc_id:
            return filename, changes

        remainder = stem[len(doc_id):]

        # Rule 1: separator after doc-ID must be '_'
        if remainder and remainder[0] in ('-', ' '):
            old_sep = remainder[0]
            remainder = '_' + remainder[1:]
            changes.append(f"separator '{old_sep}' → '_' after FILE NO")

        # Rule 2: normalize revision format
        # _rA → _RevA, _Rev A → _RevA, _rev0 → _Rev0, Rev C → _RevC
        def rev_normalizer(m):
            prefix = m.group(1)   # separator (_, -, space or empty)
            rev_part = m.group(2) # 'Rev', 'rev', 'r', 'R'
            space = m.group(3)    # optional space
            letter = m.group(4)   # version letter/number
            normalized = f'_Rev{letter.upper()}' if letter.isalpha() else f'_Rev{letter}'
            original = m.group(0)
            if original != normalized:
                changes.append(f"revision '{original.strip()}' → '{normalized.strip()}'")
            return normalized

        remainder = re.sub(
            r'([_\s-])([Rr]ev|[Rr])\.?(\s?)([A-Z0-9])(?=[_.\s]|$)',
            rev_normalizer,
            remainder,
            flags=re.IGNORECASE
        )

        # Rule 3: description matching (optional)
        # Replace abbreviated or wrong descriptions with deliverable Excel column C
        if deliverable_desc:
            current_desc = self.extract_description(doc_id + remainder + ext)
            norm_current = self._normalize_desc(current_desc)
            norm_expected = self._normalize_desc(deliverable_desc)
            if norm_current and norm_expected and norm_current != norm_expected:
                ratio = self._desc_similarity(norm_current, norm_expected)
                if ratio < 0.7:
                    # Description is significantly different — suggest replacement
                    # Replace the description portion in remainder
                    # remainder = _description_RevX
                    rev_match = re.search(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*[A-Z0-9](?=[_.\s]|$)', remainder)
                    if rev_match:
                        rev_suffix = remainder[rev_match.start():]
                        remainder = f'_{deliverable_desc}{rev_suffix}'
                    else:
                        remainder = f'_{deliverable_desc}'
                    changes.append(f"description '{current_desc}' -> '{deliverable_desc}'")

        new_name = doc_id + remainder + ext
        return new_name, changes

    def cleanup_ss_folder(self, ss_path: Path, max_versions: int = 3):
        """Remove old timestamped copies in SS folder, keeping only the most recent N per base file."""
        if not ss_path.exists():
            return
        # Group files by base name (without timestamp suffix)
        ts_pattern = re.compile(r'^(.+?)_(\d{8}_\d{6})(\.\w+)$')
        groups = defaultdict(list)
        for f in ss_path.iterdir():
            if not f.is_file():
                continue
            m = ts_pattern.match(f.name)
            if m:
                base = m.group(1) + m.group(3)
                groups[base].append(f)

        for base, files in groups.items():
            if len(files) <= max_versions:
                continue
            # Sort by modification time descending, keep newest max_versions
            files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            for old_file in files[max_versions:]:
                try:
                    old_file.unlink()
                except Exception:
                    pass

    def cross_check(self, excel_path: Path) -> DeliverableCrossCheckResult:
        """Perform full cross-check between source folders and deliverable Excel."""
        if not OPENPYXL_AVAILABLE:
            result = DeliverableCrossCheckResult(
                project_name=self.project_path.name,
                excel_path=str(excel_path)
            )
            result.errors.append("openpyxl 未安装。请运行: pip install openpyxl")
            return result

        result = DeliverableCrossCheckResult(
            project_name=self.project_path.name,
            excel_path=str(excel_path)
        )

        # Load workbook with retry for Dropbox locks
        wb = self._load_workbook_with_retry(excel_path)
        if wb is None:
            result.errors.append(f"无法打开 Excel 文件: {excel_path}")
            return result

        ws = wb.active
        try:
            layout = self.detect_layout(ws)
        except FormatChangeWarning as e:
            result.errors.append(str(e))
            wb.close()
            return result

        # Scan source folders
        folder_items = self.scan_source_folders()

        # Read Excel items
        excel_items, all_excel_rows = self.read_excel_items(ws, layout)

        # Dimension 1: Item presence (with smart matching by description)
        matched_folder_ids = set()  # folder doc_ids that matched an Excel row
        matched_excel_rows = set()  # excel row numbers that matched a folder item

        # First pass: exact doc-ID matches
        for doc_id in folder_items:
            if doc_id in excel_items:
                matched_folder_ids.add(doc_id)
                matched_excel_rows.add(excel_items[doc_id]['row'])

        # Second pass: try description matching for unmatched folder items
        # Uses all_excel_rows to handle duplicate doc-IDs in Excel
        for doc_id, info in folder_items.items():
            if doc_id in matched_folder_ids:
                continue
            match_row_info = self._find_match_by_description(
                doc_id, info['description'], all_excel_rows, matched_excel_rows
            )
            if match_row_info:
                # Found a match by description — the Excel FILE NO is wrong
                matched_folder_ids.add(doc_id)
                matched_excel_rows.add(match_row_info['row'])
                excel_info = match_row_info
                result.doc_id_corrections.append({
                    'doc_id': doc_id,              # correct doc-ID from file
                    'old_doc_id': excel_info['doc_id'],  # wrong doc-ID in Excel
                    'row': excel_info['row'],
                    'description': info['description'],
                    'filename': info['filename'],
                })
                # Also check revision for this matched row
                folder_rev = info['revision']
                excel_rev = excel_info['revision']
                if folder_rev and excel_rev and self._compare_revisions(folder_rev, excel_rev) > 0:
                    result.revision_mismatches.append({
                        'doc_id': doc_id,
                        'row': excel_info['row'],
                        'excel_rev': excel_rev,
                        'folder_rev': folder_rev,
                        'filename': info['filename'],
                    })
            else:
                result.items_in_folders_not_excel.append({
                    'doc_id': doc_id,
                    'revision': info['revision'],
                    'filename': info['filename'],
                    'description': info['description'],
                    'source': info['source'],
                })

        for doc_id, info in excel_items.items():
            if info['row'] in matched_excel_rows:
                continue
            if doc_id not in folder_items:
                if info['status'].lower() not in ('n/a', 'reserved', ''):
                    result.items_in_excel_not_folders.append(doc_id)

        # Dimension 2: Revision comparison (for exact doc-ID matches)
        for doc_id, folder_info in folder_items.items():
            if doc_id not in excel_items:
                continue
            excel_info = excel_items[doc_id]
            folder_rev = folder_info['revision']
            excel_rev = excel_info['revision']
            if not folder_rev or not excel_rev:
                continue
            if excel_info['status'].lower() in ('n/a', 'reserved'):
                continue
            if self._compare_revisions(folder_rev, excel_rev) > 0:
                result.revision_mismatches.append({
                    'doc_id': doc_id,
                    'row': excel_info['row'],
                    'excel_rev': excel_rev,
                    'folder_rev': folder_rev,
                    'filename': folder_info['filename'],
                })

        # IFC status tracking: scan IFC folder and flag items as "Approved IFC"
        ifc_info = self.scan_ifc_folder()
        for doc_id, info in ifc_info.items():
            if doc_id in excel_items:
                excel_info = excel_items[doc_id]
                current_status = excel_info.get('status', '').strip()
                if current_status.lower() != 'approved ifc':
                    result.status_updates.append({
                        'doc_id': doc_id,
                        'row': excel_info['row'],
                        'old_status': current_status,
                        'new_status': 'Approved IFC',
                        'ifc_rev': info['revision'],
                    })

        # Naming warnings: check filename conventions for matched items
        for doc_id, info in folder_items.items():
            excel_desc = None
            if doc_id in excel_items:
                excel_desc = excel_items[doc_id].get('description')
            normalized, changes = self.normalize_filename(info['filename'], excel_desc)
            if changes:
                result.naming_warnings.append({
                    'doc_id': doc_id,
                    'filename': info['filename'],
                    'suggested': normalized,
                    'changes': changes,
                    'source': info['source'],
                })

        wb.close()
        return result

    def apply_updates(self, excel_path: Path, check_result: DeliverableCrossCheckResult) -> DeliverableCrossCheckResult:
        """Apply cross-check results to the Excel file."""
        if not OPENPYXL_AVAILABLE:
            check_result.errors.append("openpyxl 未安装")
            return check_result

        if (not check_result.items_in_folders_not_excel and
                not check_result.revision_mismatches and
                not check_result.doc_id_corrections and
                not check_result.status_updates):
            return check_result

        wb = self._load_workbook_with_retry(excel_path)
        if wb is None:
            check_result.errors.append(f"无法打开 Excel 文件: {excel_path}")
            return check_result

        ws = wb.active
        try:
            layout = self.detect_layout(ws)
        except FormatChangeWarning as e:
            check_result.errors.append(str(e))
            wb.close()
            return check_result

        # Clear previous highlights
        self.clear_previous_highlights(ws, layout)

        # Apply doc-ID corrections (matched by description, FILE NO was wrong)
        for correction in check_result.doc_id_corrections:
            row = correction['row']
            ws.cell(row=row, column=layout.doc_id_col).value = correction['doc_id']
            # Marker highlights on J (col 10) and N (col 14)
            ws.cell(row=row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self.CHANGE_MARKER_FILL
            # Status-based fill on K-M
            ws.cell(row=row, column=layout.rev_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self.STATUS_SUBMITTED_FILL
            check_result.rows_updated += 1

        # Apply revision updates
        for mismatch in check_result.revision_mismatches:
            row = mismatch['row']
            ws.cell(row=row, column=layout.rev_col).value = mismatch['folder_rev']
            ws.cell(row=row, column=layout.status_col).value = 'Submitted'
            # Marker highlights on J and N
            ws.cell(row=row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self.CHANGE_MARKER_FILL
            # Status-based fill on K-M (amber for Submitted)
            ws.cell(row=row, column=layout.rev_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self.STATUS_SUBMITTED_FILL
            check_result.rows_updated += 1

        # Apply IFC status updates (revision, date, and status)
        for update in check_result.status_updates:
            row = update['row']
            ws.cell(row=row, column=layout.status_col).value = 'Approved IFC'
            # Write IFC numeric revision if available
            ifc_rev = update.get('ifc_rev')
            if ifc_rev is not None:
                ws.cell(row=row, column=layout.rev_col).value = str(ifc_rev)
            # Write submission date
            ws.cell(row=row, column=layout.date_col).value = datetime.now().strftime('%Y-%m-%d')
            # Green fill on K-M for approved
            ws.cell(row=row, column=layout.rev_col).fill = self.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self.STATUS_APPROVED_FILL
            # Marker highlights on J and N
            ws.cell(row=row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self.CHANGE_MARKER_FILL
            check_result.rows_updated += 1

        # Insert new items
        for item in check_result.items_in_folders_not_excel:
            insert_row = ws.max_row + 1
            ws.cell(row=insert_row, column=layout.doc_id_col).value = item['doc_id']
            ws.cell(row=insert_row, column=layout.desc_col).value = item['description']
            if item['revision']:
                ws.cell(row=insert_row, column=layout.rev_col).value = item['revision']
            # Marker highlights on J and N
            ws.cell(row=insert_row, column=10).fill = self.CHANGE_MARKER_FILL
            ws.cell(row=insert_row, column=14).fill = self.CHANGE_MARKER_FILL
            # Status-based fill on K-M
            ws.cell(row=insert_row, column=layout.rev_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=insert_row, column=layout.date_col).fill = self.STATUS_SUBMITTED_FILL
            ws.cell(row=insert_row, column=layout.status_col).fill = self.STATUS_SUBMITTED_FILL
            check_result.rows_inserted += 1

        # Update file revision and last updated date
        old_file_rev = ws[layout.file_rev_cell].value
        new_file_rev = self._increment_file_revision(old_file_rev)
        # Preserve prefix format (e.g. "Revision 1.9" → "Revision 2.0")
        old_str = str(old_file_rev).strip() if old_file_rev else ''
        prefix_match = re.match(r'^((?:Revision|Rev)\s*)', old_str, re.IGNORECASE)
        if prefix_match:
            ws[layout.file_rev_cell].value = f"{prefix_match.group(1)}{new_file_rev}"
        else:
            ws[layout.file_rev_cell].value = new_file_rev
        ws[layout.last_updated_cell].value = datetime.now().strftime('%Y-%m-%d')
        check_result.new_file_rev = str(new_file_rev)

        if not self.dry_run:
            # Save to a new filename with updated revision
            new_path = self._get_new_filename(excel_path, new_file_rev)
            old_path = excel_path

            # Supersede old version
            self._supersede_file(old_path)

            # Save new version
            wb.save(str(to_long_path(new_path)))
            wb.close()

            # Sync to target locations
            self._sync_deliverable(new_path)

            # Cleanup old versions in SS folders (keep max 3)
            ss_folder = old_path.parent / "SS"
            if ss_folder.exists():
                self.cleanup_ss_folder(ss_folder, max_versions=3)

            check_result.excel_path = str(new_path)
        else:
            wb.close()

        return check_result

    def _load_workbook_with_retry(self, path: Path, max_retries: int = 3):
        """Load workbook with retry for Dropbox locks."""
        for attempt in range(max_retries):
            try:
                return openpyxl.load_workbook(str(to_long_path(path)))
            except PermissionError:
                if attempt < max_retries - 1:
                    wait = 2 ** attempt
                    print(f"    [!] 文件被锁定，{wait}秒后重试... (尝试 {attempt + 1}/{max_retries})")
                    time.sleep(wait)
                else:
                    return None
            except Exception:
                return None
        return None

    def _increment_file_revision(self, current_rev) -> str:
        """Increment file revision: 1.9→2.0, 14→15, 'Revision 1.9'→'2.0', etc."""
        if current_rev is None:
            return '1.0'
        rev_str = str(current_rev).strip()
        # Strip prefix like "Revision " or "Rev "
        rev_str = re.sub(r'^(?:Revision|Rev)\s*', '', rev_str, flags=re.IGNORECASE).strip()
        if '.' in rev_str:
            try:
                parts = rev_str.split('.')
                major = int(parts[0])
                minor = int(parts[1])
                if minor >= 9:
                    return f"{major + 1}.0"
                else:
                    return f"{major}.{minor + 1}"
            except (ValueError, IndexError):
                return rev_str
        else:
            try:
                return str(int(rev_str) + 1)
            except ValueError:
                return rev_str

    def _get_new_filename(self, old_path: Path, new_rev: str) -> Path:
        """Generate new filename with updated revision."""
        stem = old_path.stem
        ext = old_path.suffix
        # Match _rev or _Revision_ followed by version number (e.g., _rev1.9, _Revision_2.0)
        new_stem = re.sub(
            r'(_rev(?:ision_?)?)[\d.]+',
            rf'\g<1>{new_rev}',
            stem,
            flags=re.IGNORECASE,
            count=1
        )
        if new_stem == stem:
            # No revision pattern found, append it
            new_stem = f"{stem}_rev{new_rev}"
        return old_path.parent / (new_stem + ext)

    def _supersede_file(self, file_path: Path):
        """Move file to SUPERSEDED subfolder (copy-then-delete for Dropbox)."""
        ss_folder = file_path.parent / "SS"
        if not ss_folder.exists():
            # Check for existing superseded folder
            for item in file_path.parent.iterdir():
                if item.is_dir() and item.name.lower() in ('ss', 'superseded', 'superceded'):
                    ss_folder = item
                    break
            else:
                ss_folder.mkdir(exist_ok=True)

        dest = ss_folder / file_path.name
        if dest.exists():
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = ss_folder / f"{file_path.stem}_{ts}{file_path.suffix}"

        try:
            # Copy then delete (Dropbox-safe)
            shutil.copy2(str(to_long_path(file_path)), str(to_long_path(dest)))
            to_long_path(file_path).unlink()
        except Exception as e:
            print(f"    [!] 无法supersede文件: {file_path.name} - {e}")

    def _sync_deliverable(self, new_file: Path):
        """Sync new deliverable file to target locations."""
        # Primary sync target (already in the right place if file is in 3.Deliverables)
        primary_target = self.project_path / self.PRIMARY_SYNC
        if primary_target.exists() and new_file.parent != primary_target:
            dest = primary_target / new_file.name
            try:
                # Supersede old versions in target
                self._supersede_old_versions_in_folder(primary_target, new_file.name)
                shutil.copy2(str(to_long_path(new_file)), str(to_long_path(dest)))
            except Exception as e:
                print(f"    [!] 主同步失败: {e}")

        # Secondary sync target (only if exists)
        secondary_target = self.project_path / self.SECONDARY_SYNC
        if secondary_target.exists():
            dest = secondary_target / new_file.name
            try:
                self._supersede_old_versions_in_folder(secondary_target, new_file.name)
                shutil.copy2(str(to_long_path(new_file)), str(to_long_path(dest)))
                print(f"    [v] 已同步到 Client Sharepoint: {new_file.name}")
            except Exception as e:
                print(f"    [!] 次要同步失败: {e}")

    def _supersede_old_versions_in_folder(self, folder: Path, new_filename: str):
        """Move old versions of the deliverable in a folder to SS."""
        # Find the base name pattern (without revision)
        base_pattern = re.sub(r'(?:_rev|_Revision_?)[\d.]+', '', Path(new_filename).stem,
                             flags=re.IGNORECASE)
        ss_folder = folder / "SS"
        for f in folder.iterdir():
            if not f.is_file():
                continue
            if f.name == new_filename:
                continue
            f_base = re.sub(r'(?:_rev|_Revision_?)[\d.]+', '', f.stem, flags=re.IGNORECASE)
            if f_base.lower() == base_pattern.lower() and f.suffix.lower() == Path(new_filename).suffix.lower():
                if not ss_folder.exists():
                    ss_folder.mkdir(exist_ok=True)
                dest = ss_folder / f.name
                try:
                    shutil.copy2(str(to_long_path(f)), str(to_long_path(dest)))
                    to_long_path(f).unlink()
                except Exception:
                    pass


# =============================================================================
# IFC Transmittal Manager - Version Management & Transmittal Generation
# =============================================================================

@dataclass
class IFCResult:
    """Result of IFC management run."""
    total_ifc_files: int = 0
    duplicates_archived: int = 0
    new_files_since_last: int = 0
    deliverable_updates: int = 0
    transmittal_path: Optional[str] = None
    transmittal_number: int = 0
    errors: List[str] = field(default_factory=list)
    file_list: List[Dict] = field(default_factory=list)  # {name, action, doc_id, revision}


class IFCTransmittalManager:
    """Manage IFC file versions, update deliverable Excel, generate transmittals.

    Workflow:
      1. Scan 4. IFC(Client)/ for PDFs, group by doc-ID
      2. Identify duplicates → move old versions to SS/
      3. Collect new/updated files since last run
      4. Update deliverable Excel column L with IFC revision numbers
      5. Generate transmittal Excel with header info from config
      6. Save transmittal to 8. Deliverables/Transmittal/ + sync targets
      7. Save run state for incremental tracking
    """

    IFC_CLIENT_REL = "Design/Engineering/1. Drawings/4. IFC(Client)"
    TRANSMITTAL_REL = "Design/Engineering/8. Deliverables/Transmittal"

    # Reuse patterns from module level
    RE_DOC_ID = re.compile(r'^(\d{5}-[A-Z]{2}-\d{3})', re.IGNORECASE)
    RE_IFC_REV = re.compile(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*(\d+)(?:[_\s-]?IFC)?(?=[_.\s]|$)', re.IGNORECASE)

    # Transmittal styling
    HEADER_FONT = openpyxl.styles.Font(bold=True, size=14) if OPENPYXL_AVAILABLE else None
    SUBHEADER_FONT = openpyxl.styles.Font(bold=True, size=11) if OPENPYXL_AVAILABLE else None
    NORMAL_FONT = openpyxl.styles.Font(size=10) if OPENPYXL_AVAILABLE else None
    COL_HEADER_FONT = openpyxl.styles.Font(bold=True, size=10) if OPENPYXL_AVAILABLE else None
    COL_HEADER_FILL = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2',
                                   fill_type='solid') if OPENPYXL_AVAILABLE else None
    THIN_BORDER = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin'),
    ) if OPENPYXL_AVAILABLE else None
    IFC_APPROVED_FILL = PatternFill(start_color='FF00B050', end_color='FF00B050',
                                     fill_type='solid') if OPENPYXL_AVAILABLE else None

    def __init__(self, project_path, config=None, dry_run=False):
        self.project_path = Path(project_path)
        self.config = config or {}
        self.dry_run = dry_run
        self.logger = logging.getLogger(self.__class__.__name__)

        # Config-driven paths
        ifc_cfg = self.config.get("ifc_transmittal", {})
        self.transmittal_rel = ifc_cfg.get("transmittal_path", self.TRANSMITTAL_REL)
        self.sync_targets = ifc_cfg.get("sync_targets", [])
        self.header_config = ifc_cfg.get("transmittal_header", {})
        self.state_file_name = ifc_cfg.get("state_file", "ifc_state.json")

        self.ifc_dir = self.project_path / self.IFC_CLIENT_REL
        self.transmittal_dir = self.project_path / self.transmittal_rel
        self.state_file = self.transmittal_dir / self.state_file_name

    # ── 2a. Version Management ──

    def scan_ifc_files(self) -> Dict[str, List[Path]]:
        """Scan 4. IFC(Client)/ for PDFs, group by doc-ID. Excludes SS/."""
        grouped = defaultdict(list)
        if not self.ifc_dir.exists():
            self.logger.warning(f"IFC directory not found: {self.ifc_dir}")
            return grouped
        for f in self.ifc_dir.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() != '.pdf':
                continue
            if f.name.startswith('~$'):
                continue
            doc_id = self._extract_doc_id(f.name)
            if doc_id:
                grouped[doc_id].append(f)
        return dict(grouped)

    def identify_duplicates(self, grouped: Dict[str, List[Path]]) -> List[Tuple[Path, Path]]:
        """For doc-IDs with >1 file, identify old versions to archive.

        Returns list of (source, destination) tuples for files to move to SS/.
        """
        ss_folder = self.ifc_dir / "SS"
        duplicates = []
        for doc_id, files in grouped.items():
            if len(files) <= 1:
                continue
            # Sort by revision number (desc), then by mtime (desc) for tie-breaking
            scored = []
            for f in files:
                rev = self._extract_revision_number(f.name)
                mtime = f.stat().st_mtime
                scored.append((f, rev, mtime))
            scored.sort(key=lambda x: (x[1], x[2]), reverse=True)
            # Keep first (highest rev, newest), archive the rest
            for f, rev, mtime in scored[1:]:
                dest = ss_folder / f.name
                duplicates.append((f, dest))
        return duplicates

    def move_old_to_ss(self, duplicates: List[Tuple[Path, Path]]) -> int:
        """Move old versions to SS/ subfolder."""
        if not duplicates:
            return 0
        ss_folder = self.ifc_dir / "SS"
        if not self.dry_run:
            try:
                ss_folder.mkdir(exist_ok=True)
            except (OSError, PermissionError):
                try:
                    to_long_path(ss_folder).mkdir(exist_ok=True)
                except Exception as e:
                    self.logger.error(f"Cannot create SS folder: {e}")
                    return 0
        moved = 0
        for source, dest in duplicates:
            if self.dry_run:
                self.logger.info(f"[DRY-RUN] Would move: {source.name} -> SS/")
                moved += 1
            else:
                try:
                    if dest.exists():
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        dest = ss_folder / f"{dest.stem}_{timestamp}{dest.suffix}"
                    shutil.move(str(to_long_path(source)), str(to_long_path(dest)))
                    self.logger.info(f"Archived: {source.name} -> SS/")
                    moved += 1
                except Exception as e:
                    self.logger.error(f"Move failed: {source.name} - {e}")
        return moved

    # ── 2b. Deliverable Update ──

    def update_deliverable_ifc_status(self, ifc_files: Dict[str, Dict]) -> int:
        """Update deliverable Excel column L with IFC revision numbers.

        Args:
            ifc_files: {doc_id: {revision: int, filename: str}}

        Returns:
            Number of rows updated.
        """
        if not ifc_files:
            return 0

        dm = DeliverableManager(self.project_path, dry_run=self.dry_run)
        excel_path = dm.find_deliverable_excel()
        if not excel_path:
            self.logger.warning("Deliverable Excel not found")
            return 0

        if self.dry_run:
            # In dry-run, just count matches
            try:
                wb = openpyxl.load_workbook(str(to_long_path(excel_path)))
                ws = wb.active
                layout = dm.detect_layout(ws)
                count = 0
                for row in range(layout.header_row + 1, ws.max_row + 1):
                    cell_val = ws.cell(row=row, column=layout.doc_id_col).value
                    if cell_val and str(cell_val).strip() in ifc_files:
                        count += 1
                wb.close()
                return count
            except Exception as e:
                self.logger.error(f"Dry-run deliverable scan failed: {e}")
                return 0

        # Actual update
        try:
            wb = openpyxl.load_workbook(str(to_long_path(excel_path)))
            ws = wb.active
            layout = dm.detect_layout(ws)
            updated = 0

            # Column L = column 12
            ifc_rev_col = 12

            for row in range(layout.header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=layout.doc_id_col).value
                if not cell_val:
                    continue
                doc_id = str(cell_val).strip()
                if doc_id not in ifc_files:
                    continue

                rev_num = ifc_files[doc_id]["revision"]
                # Write IFC revision to column L
                ws.cell(row=row, column=ifc_rev_col, value=rev_num)
                # Set status to "Approved IFC" with green fill
                ws.cell(row=row, column=layout.status_col, value="Approved IFC")
                ws.cell(row=row, column=layout.status_col).fill = self.IFC_APPROVED_FILL
                updated += 1

            if updated > 0:
                # Version the deliverable: move old to SS, save new
                self._version_and_save_deliverable(wb, excel_path)
            else:
                wb.close()

            return updated
        except Exception as e:
            self.logger.error(f"Deliverable update failed: {e}")
            return 0

    def _version_and_save_deliverable(self, wb, excel_path: Path):
        """Save workbook, versioning old file to SS/."""
        ss_folder = excel_path.parent / "SS"
        ss_folder.mkdir(exist_ok=True)

        # Move current file to SS
        if excel_path.exists():
            ss_dest = ss_folder / excel_path.name
            if ss_dest.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                ss_dest = ss_folder / f"{excel_path.stem}_{timestamp}{excel_path.suffix}"
            try:
                shutil.copy2(str(to_long_path(excel_path)), str(to_long_path(ss_dest)))
            except Exception as e:
                self.logger.warning(f"Failed to backup deliverable: {e}")

        # Save updated workbook
        try:
            wb.save(str(to_long_path(excel_path)))
            self.logger.info(f"Deliverable updated: {excel_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to save deliverable: {e}")
        finally:
            wb.close()

    # ── 2c. Transmittal Generation ──

    def get_last_run_state(self) -> Dict:
        """Read state from ifc_state.json."""
        if not self.state_file.exists():
            return {}
        try:
            with open(str(to_long_path(self.state_file)), 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def get_last_run_timestamp(self) -> Optional[datetime]:
        """Get timestamp of last run."""
        state = self.get_last_run_state()
        ts = state.get("last_run")
        if ts:
            try:
                return datetime.fromisoformat(ts)
            except (ValueError, TypeError):
                pass
        return None

    def collect_new_ifc_files(self, since: Optional[datetime],
                               transmitted: Set[str]) -> List[Dict]:
        """Collect IFC files new or updated since last run.

        Returns list of {path, doc_id, revision, action} dicts.
        action is 'New' or 'Update'.
        """
        results = []
        if not self.ifc_dir.exists():
            return results

        for f in self.ifc_dir.iterdir():
            if not f.is_file() or f.suffix.lower() != '.pdf':
                continue
            if f.name.startswith('~$'):
                continue

            doc_id = self._extract_doc_id(f.name)
            if not doc_id:
                continue

            rev = self._extract_revision_number(f.name)
            mtime = datetime.fromtimestamp(f.stat().st_mtime)

            # Determine if this file should be included
            include = False
            if since is None:
                include = True  # First run: include all
            elif mtime > since:
                include = True  # Modified since last run
            elif doc_id not in transmitted:
                include = True  # New doc-ID not previously transmitted

            if not include:
                continue

            action = "Update" if doc_id in transmitted else "New"
            results.append({
                "path": f,
                "doc_id": doc_id,
                "revision": rev,
                "action": action,
                "filename": f.name,
            })

        # Sort by doc_id for consistent ordering
        results.sort(key=lambda x: x["doc_id"])
        return results

    def determine_next_tsmt_number(self) -> int:
        """Find next transmittal number (shared sequence with IFR).

        Scans 8. Deliverables/Transmittal/ for existing TSMT files.
        """
        max_num = 0
        if self.transmittal_dir.exists():
            for f in self.transmittal_dir.iterdir():
                if not f.is_file():
                    continue
                m = re.search(r'TSMT[_-]?(\d+)', f.name, re.IGNORECASE)
                if m:
                    num = int(m.group(1))
                    if num > max_num:
                        max_num = num

        # Also check state file for last known number
        state = self.get_last_run_state()
        last_num = state.get("last_tsmt_number", 0)
        max_num = max(max_num, last_num)

        return max_num + 1

    def _get_project_number(self) -> str:
        """Extract project number (e.g. '50023') from header config or project path."""
        project_str = self.header_config.get("project", "")
        m = re.match(r'(\d{5})', project_str)
        if m:
            return m.group(1)
        # Fallback: try extracting from project path
        for part in self.project_path.parts:
            m2 = re.match(r'(\d{5})', part)
            if m2:
                return m2.group(1)
        return "00000"

    def generate_transmittal(self, files: List[Dict], tsmt_num: int,
                              run_date: datetime) -> Optional[Path]:
        """Generate transmittal Excel workbook.

        Args:
            files: List of {path, doc_id, revision, action, filename}
            tsmt_num: Transmittal sequence number
            run_date: Date for the transmittal

        Returns:
            Path to generated transmittal file, or None on error.
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl not available, cannot generate transmittal")
            return None

        project_num = self._get_project_number()
        tsmt_filename = f"{project_num}-TSMT-{tsmt_num}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transmittal"

        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 6
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30

        # ── Header section (rows 1-11) ──
        ws.cell(row=1, column=2, value=self.header_config.get("company", "")).font = self.HEADER_FONT
        ws.merge_cells('B1:G1')

        # Company details
        details = self.header_config.get("company_details", "")
        ws.cell(row=2, column=2, value=details).font = self.NORMAL_FONT
        ws.merge_cells('B2:G2')
        ws.row_dimensions[2].height = 50

        # Transmittal number and date
        ws.cell(row=4, column=2, value="TRANSMITTAL").font = self.SUBHEADER_FONT
        ws.cell(row=4, column=5, value=f"No: {project_num}-TSMT-{tsmt_num}").font = self.SUBHEADER_FONT
        ws.cell(row=5, column=5, value=f"Date: {run_date.strftime('%d/%m/%Y')}").font = self.NORMAL_FONT

        # To / From
        ws.cell(row=6, column=2, value="To:").font = self.SUBHEADER_FONT
        ws.cell(row=6, column=3, value=self.header_config.get("to", "")).font = self.NORMAL_FONT
        ws.cell(row=7, column=3, value=self.header_config.get("to_address", "")).font = self.NORMAL_FONT
        ws.row_dimensions[7].height = 30

        ws.cell(row=8, column=2, value="Attention:").font = self.SUBHEADER_FONT
        ws.cell(row=8, column=3, value=self.header_config.get("attention", "")).font = self.NORMAL_FONT

        ws.cell(row=9, column=2, value="Project:").font = self.SUBHEADER_FONT
        ws.cell(row=9, column=3, value=self.header_config.get("project", "")).font = self.NORMAL_FONT

        ws.cell(row=10, column=2, value="Subject:").font = self.SUBHEADER_FONT
        ws.cell(row=10, column=3, value=self.header_config.get("subject", "")).font = self.NORMAL_FONT

        # ── Column headers (row 12) ──
        headers = [
            (2, "Item NO."),
            (3, "Doc ID"),
            (4, "Qty"),
            (5, "File Name"),
            (6, "Comment"),
            (7, "Description"),
        ]
        for col, text in headers:
            cell = ws.cell(row=12, column=col, value=text)
            cell.font = self.COL_HEADER_FONT
            cell.fill = self.COL_HEADER_FILL
            cell.border = self.THIN_BORDER

        # ── Data rows (14, 16, 18... alternating) ──
        data_row = 14
        for idx, file_info in enumerate(files, 1):
            ws.cell(row=data_row, column=2, value=idx).border = self.THIN_BORDER
            ws.cell(row=data_row, column=3, value=file_info["doc_id"]).border = self.THIN_BORDER
            ws.cell(row=data_row, column=4, value=1).border = self.THIN_BORDER
            ws.cell(row=data_row, column=5, value=file_info["filename"]).border = self.THIN_BORDER
            ws.cell(row=data_row, column=6, value=file_info["action"]).border = self.THIN_BORDER

            # Extract description from filename
            desc = self._extract_description(file_info["filename"], file_info["doc_id"])
            ws.cell(row=data_row, column=7, value=desc).border = self.THIN_BORDER

            data_row += 2  # Alternating rows

        # Save transmittal
        self.transmittal_dir.mkdir(parents=True, exist_ok=True)
        tsmt_path = self.transmittal_dir / tsmt_filename

        if self.dry_run:
            wb.close()
            return tsmt_path

        try:
            wb.save(str(to_long_path(tsmt_path)))
            self.logger.info(f"Transmittal saved: {tsmt_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to save transmittal: {e}")
            wb.close()
            return None
        wb.close()

        # Sync to targets
        self._sync_transmittal(tsmt_path)

        return tsmt_path

    def _sync_transmittal(self, tsmt_path: Path):
        """Copy transmittal to sync target directories."""
        for rel_target in self.sync_targets:
            target_dir = self.project_path / rel_target
            if not target_dir.exists():
                try:
                    target_dir.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    self.logger.warning(f"Cannot create sync target: {target_dir} - {e}")
                    continue
            dest = target_dir / tsmt_path.name
            try:
                shutil.copy2(str(to_long_path(tsmt_path)), str(to_long_path(dest)))
                self.logger.info(f"Synced transmittal to: {rel_target}")
            except Exception as e:
                self.logger.error(f"Transmittal sync failed to {rel_target}: {e}")

    def save_run_state(self, run_date: datetime, tsmt_num: int,
                        transmitted_files: List[str]):
        """Save run state to ifc_state.json."""
        if self.dry_run:
            return
        state = self.get_last_run_state()
        # Merge transmitted files with existing
        existing = set(state.get("transmitted_files", []))
        existing.update(transmitted_files)

        new_state = {
            "last_run": run_date.isoformat(),
            "last_tsmt_number": tsmt_num,
            "transmitted_files": sorted(existing),
        }
        try:
            self.transmittal_dir.mkdir(parents=True, exist_ok=True)
            with open(str(to_long_path(self.state_file)), 'w', encoding='utf-8') as f:
                json.dump(new_state, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"Failed to save state: {e}")

    # ── 2d. Orchestrator ──

    def run(self, run_date: datetime = None) -> IFCResult:
        """Run the full IFC management workflow.

        1. Scan & deduplicate IFC files
        2. Collect new/updated files since last run
        3. Update deliverable Excel
        4. Generate transmittal
        5. Save state
        """
        if run_date is None:
            run_date = datetime.now()

        result = IFCResult()

        # 1. Scan and group IFC files
        grouped = self.scan_ifc_files()
        result.total_ifc_files = sum(len(files) for files in grouped.values())

        # 2. Identify and archive duplicates
        duplicates = self.identify_duplicates(grouped)
        result.duplicates_archived = self.move_old_to_ss(duplicates)

        # Track archived files
        for src, dest in duplicates:
            result.file_list.append({
                "name": src.name,
                "action": "Archived",
                "doc_id": self._extract_doc_id(src.name) or "",
                "revision": self._extract_revision_number(src.name),
            })

        # 3. Collect new/updated files since last run
        state = self.get_last_run_state()
        since = self.get_last_run_timestamp()
        transmitted = set(state.get("transmitted_files", []))
        new_files = self.collect_new_ifc_files(since, transmitted)
        result.new_files_since_last = len(new_files)

        if not new_files:
            # No new files to process
            return result

        # 4. Update deliverable Excel with IFC revisions
        ifc_file_map = {}
        for fi in new_files:
            ifc_file_map[fi["doc_id"]] = {
                "revision": fi["revision"],
                "filename": fi["filename"],
            }
        result.deliverable_updates = self.update_deliverable_ifc_status(ifc_file_map)

        # 5. Generate transmittal
        tsmt_num = self.determine_next_tsmt_number()
        tsmt_path = self.generate_transmittal(new_files, tsmt_num, run_date)
        if tsmt_path:
            result.transmittal_path = str(tsmt_path)
            result.transmittal_number = tsmt_num

        # Track new/updated files in result
        for fi in new_files:
            result.file_list.append({
                "name": fi["filename"],
                "action": fi["action"],
                "doc_id": fi["doc_id"],
                "revision": fi["revision"],
            })

        # 6. Save state
        transmitted_filenames = [fi["filename"] for fi in new_files]
        self.save_run_state(run_date, tsmt_num, transmitted_filenames)

        return result

    # ── Helper methods ──

    def _extract_doc_id(self, filename: str) -> Optional[str]:
        """Extract doc ID from filename using LMS pattern."""
        stem = Path(filename).stem
        m = self.RE_DOC_ID.match(stem)
        if m:
            return m.group(1)
        return None

    def _extract_revision_number(self, filename: str) -> int:
        """Extract numeric revision from filename. Returns 0 if not found."""
        stem = Path(filename).stem
        m = self.RE_IFC_REV.search(stem)
        if m:
            return int(m.group(1))
        return 0

    def _extract_description(self, filename: str, doc_id: str) -> str:
        """Extract description from filename (part between doc_id and revision)."""
        stem = Path(filename).stem
        if doc_id and stem.startswith(doc_id):
            remainder = stem[len(doc_id):].lstrip('_- ')
        else:
            remainder = stem
        # Remove revision suffix
        remainder = re.sub(r'[_\s-](?:[Rr]ev|[Rr])\.?\s*\d+.*$', '', remainder, flags=re.IGNORECASE)
        return remainder.strip('_- ') or stem


# =============================================================================
# IFC Manager - AutoCAD COM Automation (NEW in v7)
# =============================================================================

class IFCManager:
    """Automate IFR→IFC conversion via AutoCAD COM API.

    Workflow per DWG:
      1. Open DWG in AutoCAD
      2. Find title block, read latest IFR revision row
      3. Update title block: set IFC revision, clear old rows, write IFC row
      4. SaveAs new IFC DWG
      5. Export PDF (all layouts) to IFC(Client) folder
      6. Close without saving original IFR DWG
    """

    NATIVE_ROOT = "Design/Engineering/1. Drawings/1. Native"
    IFC_OUTPUT = "Design/Engineering/1. Drawings/4. IFC(Client)"

    # Title block attribute TAGs
    DEFAULT_TITLE_BLOCK = "ACE-Wanertown_Siyuan"
    REV_ROWS = 6  # max revision history rows in title block

    # Personnel TAG suffixes per row (prefixed by row number, e.g. 1DRAWN, 2DRAWN)
    PERSONNEL_TAGS = ['DRAWN', 'CHECK', 'ENGINEER', 'QA', 'PROJECT']

    def __init__(self, project_path, dry_run=False, title_block_name=None):
        self.project_path = Path(project_path)
        self.dry_run = dry_run
        self.title_block_name = title_block_name or self.DEFAULT_TITLE_BLOCK
        self._acad = None
        self._dm = DeliverableManager(self.project_path, dry_run=dry_run)

    @property
    def native_root(self) -> Path:
        return self.project_path / self.NATIVE_ROOT

    @property
    def ifc_output(self) -> Path:
        return self.project_path / self.IFC_OUTPUT

    def _get_acad(self):
        """Get or launch AutoCAD COM instance."""
        if not WIN32COM_AVAILABLE:
            raise RuntimeError("win32com 未安装，无法使用 AutoCAD COM API。请安装 pywin32。")
        if self._acad is not None:
            try:
                _ = self._acad.Visible
                return self._acad
            except Exception:
                self._acad = None
        # Try to connect to running instance first
        try:
            self._acad = win32com.client.GetActiveObject("AutoCAD.Application")
        except Exception:
            try:
                print("  正在启动 AutoCAD（可能需要30秒）...")
                self._acad = win32com.client.Dispatch("AutoCAD.Application")
                self._acad.Visible = True
            except Exception as e:
                raise RuntimeError(f"无法连接或启动 AutoCAD: {e}")
        return self._acad

    def scan_native_folders(self) -> List[Dict]:
        """Scan 1. Native/ subfolders and identify IFR DWGs ready for IFC conversion.

        Returns list of dicts:
          {doc_id, folder, latest_ifr_dwg, latest_ifr_rev,
           existing_ifc_rev, needs_ifc, description}
        """
        results = []
        native = self.native_root
        if not native.exists():
            return results

        # Also scan IFC output folder to know existing IFC revisions
        existing_ifc = self._scan_existing_ifc()

        for category_dir in sorted(native.iterdir()):
            if not category_dir.is_dir():
                continue
            if category_dir.name.lower() in ('ss', 'superseded', 'superceded'):
                continue
            for doc_folder in sorted(category_dir.iterdir()):
                if not doc_folder.is_dir():
                    continue
                if doc_folder.name.lower() in ('ss', 'superseded', 'superceded'):
                    continue
                doc_id, description = _parse_folder_name(doc_folder.name)
                if not doc_id:
                    continue

                # Find latest IFR DWG (highest letter revision)
                latest_ifr = None
                latest_rev = ''
                for f in doc_folder.iterdir():
                    if not f.is_file() or f.suffix.lower() != '.dwg':
                        continue
                    if f.name.startswith('~$'):
                        continue
                    rev_type, revision = _classify_dwg(f.name)
                    if rev_type == 'IFR' and revision > latest_rev:
                        latest_rev = revision
                        latest_ifr = f

                if not latest_ifr:
                    continue

                ifc_rev = existing_ifc.get(doc_id)
                results.append({
                    'doc_id': doc_id,
                    'folder': doc_folder,
                    'latest_ifr_dwg': latest_ifr,
                    'latest_ifr_rev': latest_rev,
                    'existing_ifc_rev': ifc_rev,
                    'needs_ifc': True,  # always allow re-conversion
                    'description': description,
                })

        return results

    def _scan_existing_ifc(self) -> Dict[str, int]:
        """Scan IFC output folder for existing IFC files, return {doc_id: max_rev_number}."""
        ifc_revs = {}
        ifc_dir = self.ifc_output
        if not ifc_dir.exists():
            return ifc_revs
        for f in ifc_dir.rglob("*"):
            if not f.is_file():
                continue
            if f.suffix.lower() not in ('.dwg', '.pdf'):
                continue
            doc_id = self._dm.extract_doc_id(f.name)
            if not doc_id:
                continue
            rev_type, revision = _classify_dwg(f.name)
            if rev_type == 'IFC':
                try:
                    rev_num = int(revision)
                    if doc_id not in ifc_revs or rev_num > ifc_revs[doc_id]:
                        ifc_revs[doc_id] = rev_num
                except ValueError:
                    pass
        return ifc_revs

    def _find_title_block(self, doc):
        """Find title block in AutoCAD document.

        Returns (block_ref, attrs_dict) where attrs_dict maps TAG -> attribute object.
        Tries exact name match first, then fuzzy fallback.
        """
        ms = doc.ModelSpace
        # Also check PaperSpace (layouts) since title blocks are often in PaperSpace
        spaces = [ms]
        try:
            for layout in doc.Layouts:
                if layout.Name.lower() != 'model':
                    block = doc.PaperSpace
                    spaces.append(block)
                    break
        except Exception:
            pass

        for space in spaces:
            for i in range(space.Count):
                entity = space.Item(i)
                try:
                    if entity.EntityName != 'AcDbBlockReference':
                        continue
                except Exception:
                    continue
                try:
                    block_name = entity.Name
                except Exception:
                    continue

                # Exact match
                if block_name == self.title_block_name:
                    return entity, self._get_attrs_dict(entity)

                # Fuzzy fallback: block with DRAWINGNUMBER + REVISION tags
                attrs = self._get_attrs_dict(entity)
                if 'DRAWINGNUMBER' in attrs and 'REVISION' in attrs:
                    return entity, attrs

        return None, {}

    def _get_attrs_dict(self, block_ref) -> Dict:
        """Extract TAG -> attribute object dict from a block reference."""
        attrs = {}
        try:
            for attr in block_ref.GetAttributes():
                attrs[attr.TagString.upper()] = attr
        except Exception:
            pass
        return attrs

    def _read_latest_ifr_row(self, attrs: Dict) -> Dict:
        """Find the highest non-empty revision row and return personnel info.

        Scans rows 6→1 looking for non-empty {N}REV tag.
        Returns {row, rev, drawn, check, engineer, qa, project, date, description}.
        """
        for row_num in range(self.REV_ROWS, 0, -1):
            rev_tag = f"{row_num}REV"
            if rev_tag in attrs:
                val = attrs[rev_tag].TextString.strip()
                if val:
                    info = {'row': row_num, 'rev': val}
                    for tag in self.PERSONNEL_TAGS:
                        full_tag = f"{row_num}{tag}"
                        if full_tag in attrs:
                            info[tag.lower()] = attrs[full_tag].TextString.strip()
                        else:
                            info[tag.lower()] = ''
                    date_tag = f"{row_num}DATE"
                    if date_tag in attrs:
                        info['date'] = attrs[date_tag].TextString.strip()
                    desc_tag = f"{row_num}DESCRIPTION"
                    if desc_tag in attrs:
                        info['description'] = attrs[desc_tag].TextString.strip()
                    return info
        return {}

    def _update_title_block(self, attrs: Dict, ifc_rev: int, personnel: Dict, date_str: str):
        """Update title block attributes for IFC conversion.

        Sets REVISION, clears all 6 rows, writes IFC info to row 1.
        """
        # Set main REVISION attribute
        if 'REVISION' in attrs:
            attrs['REVISION'].TextString = str(ifc_rev)

        # Clear all revision rows
        for row_num in range(1, self.REV_ROWS + 1):
            for suffix in ['REV', 'DATE', 'DESCRIPTION'] + self.PERSONNEL_TAGS:
                tag = f"{row_num}{suffix}"
                if tag in attrs:
                    attrs[tag].TextString = ''

        # Write row 1 with IFC data
        if '1REV' in attrs:
            attrs['1REV'].TextString = str(ifc_rev)
        if '1DESCRIPTION' in attrs:
            attrs['1DESCRIPTION'].TextString = 'FOR CONSTRUCTION'
        if '1DATE' in attrs:
            attrs['1DATE'].TextString = date_str

        # Copy personnel from latest IFR row
        for tag in self.PERSONNEL_TAGS:
            full_tag = f"1{tag}"
            if full_tag in attrs:
                attrs[full_tag].TextString = personnel.get(tag.lower(), '')

    def _export_pdf(self, doc, output_path: Path) -> bool:
        """Export PDF using -EXPORTPDF command (all layouts).

        Uses SendCommand which is async — polls CMDACTIVE until done.
        """
        output_str = str(output_path).replace('\\', '/')
        # -EXPORTPDF command sequence:
        # _-EXPORTPDF → All layouts → output path
        cmd = f'-EXPORTPDF\n_All\n{output_str}\n'
        try:
            doc.SendCommand(cmd)
            # Poll for completion (CMDACTIVE = 0 means no command running)
            max_wait = 120  # seconds
            start = time.time()
            while time.time() - start < max_wait:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        break
                except Exception:
                    break
                time.sleep(1)
            return output_path.exists()
        except Exception as e:
            logging.warning(f"PDF export failed: {e}")
            return False

    def _build_ifc_filename(self, doc_id: str, description: str, ifc_rev: int) -> str:
        """Build IFC filename: {doc_id}_{description}_Rev{N}_IFC"""
        # Clean description
        desc = re.sub(r'[<>:"/\\|?*]', '', description).strip()
        desc = re.sub(r'\s+', ' ', desc).replace(' ', '_')
        return f"{doc_id}_{desc}_Rev{ifc_rev}_IFC"

    def _wait_for_command(self, doc, timeout: int = 60):
        """Wait for AutoCAD command to finish."""
        start = time.time()
        time.sleep(0.5)
        while time.time() - start < timeout:
            try:
                if doc.GetVariable("CMDACTIVE") == 0:
                    return True
            except Exception:
                return False
            time.sleep(0.5)
        return False

    def convert_to_ifc(self, dwg_info: Dict) -> Dict:
        """Convert a single IFR DWG to IFC.

        Steps:
          1. Open DWG in AutoCAD
          2. Find title block, read latest IFR row
          3. Calculate IFC rev (0 if first, else existing + 1)
          4. Update title block
          5. SaveAs new IFC DWG
          6. Export PDF to IFC(Client)/
          7. Close without saving original

        Returns {success, dwg_path, pdf_path, ifc_rev, errors}.
        """
        result = {
            'success': False,
            'doc_id': dwg_info['doc_id'],
            'dwg_path': None,
            'pdf_path': None,
            'ifc_rev': None,
            'errors': [],
        }

        dwg_path = dwg_info['latest_ifr_dwg']
        doc_id = dwg_info['doc_id']
        description = dwg_info['description']

        # Check for lock files
        lock1 = dwg_path.with_suffix('.dwl')
        lock2 = dwg_path.with_suffix('.dwl2')
        if lock1.exists() or lock2.exists():
            result['errors'].append(f"文件被锁定: {dwg_path.name}")
            return result

        # Calculate IFC revision
        if dwg_info['existing_ifc_rev'] is not None:
            ifc_rev = dwg_info['existing_ifc_rev'] + 1
        else:
            ifc_rev = 0
        result['ifc_rev'] = ifc_rev

        if self.dry_run:
            ifc_name = self._build_ifc_filename(doc_id, description, ifc_rev)
            result['success'] = True
            result['dwg_path'] = str(dwg_info['folder'] / f"{ifc_name}.dwg")
            result['pdf_path'] = str(self.ifc_output / f"{ifc_name}.pdf")
            return result

        # --- Actual AutoCAD operations ---
        doc = None
        try:
            acad = self._get_acad()

            # Open the DWG
            dwg_str = str(dwg_path)
            try:
                doc = acad.Documents.Open(dwg_str)
            except Exception as e:
                result['errors'].append(f"无法打开 DWG: {e}")
                return result

            # Find title block
            block_ref, attrs = self._find_title_block(doc)
            if not attrs:
                result['errors'].append(f"未找到 title block（尝试: {self.title_block_name}）")
                doc.Close(False)
                return result

            # Read latest IFR row
            personnel = self._read_latest_ifr_row(attrs)
            if not personnel:
                result['errors'].append("未找到有效的 IFR revision 行")
                doc.Close(False)
                return result

            # Update title block
            date_str = datetime.now().strftime('%d/%m/%y')
            self._update_title_block(attrs, ifc_rev, personnel, date_str)

            # Build output paths
            ifc_name = self._build_ifc_filename(doc_id, description, ifc_rev)
            ifc_dwg_path = dwg_info['folder'] / f"{ifc_name}.dwg"
            ifc_pdf_path = self.ifc_output / f"{ifc_name}.pdf"

            # Ensure IFC output directory exists
            to_long_path(self.ifc_output).mkdir(parents=True, exist_ok=True)

            # SaveAs new IFC DWG (keeps original untouched)
            try:
                doc.SaveAs(str(ifc_dwg_path))
                result['dwg_path'] = str(ifc_dwg_path)
            except Exception as e:
                result['errors'].append(f"SaveAs 失败: {e}")
                doc.Close(False)
                return result

            # Export PDF
            pdf_ok = self._export_pdf(doc, ifc_pdf_path)
            if pdf_ok:
                result['pdf_path'] = str(ifc_pdf_path)
            else:
                # Retry once
                time.sleep(2)
                pdf_ok = self._export_pdf(doc, ifc_pdf_path)
                if pdf_ok:
                    result['pdf_path'] = str(ifc_pdf_path)
                else:
                    result['errors'].append("PDF 导出失败（已重试）")

            # Close document without saving (we already SaveAs'd to new file)
            doc.Close(False)
            doc = None

            # Re-open original to restore it (AutoCAD SaveAs changes the active file)
            # Actually, SaveAs already saved to new path and original is untouched
            # since we opened it read-only effectively. Just need to ensure
            # the original file is not modified.

            result['success'] = True

        except Exception as e:
            result['errors'].append(f"转换异常: {e}")
            if doc is not None:
                try:
                    doc.Close(False)
                except Exception:
                    pass

        return result

    def batch_convert(self, doc_ids: Optional[List[str]] = None,
                      update_deliverables: bool = False) -> List[Dict]:
        """Convert multiple IFR DWGs to IFC.

        Args:
            doc_ids: If provided, only convert these doc-IDs. Otherwise convert all.
            update_deliverables: If True, auto-update deliverables Excel after conversion.

        Returns list of result dicts from convert_to_ifc().
        """
        scan = self.scan_native_folders()
        if doc_ids:
            scan = [s for s in scan if s['doc_id'] in doc_ids]

        if not scan:
            print("  没有找到需要转换的文件")
            return []

        results = []
        total = len(scan)
        for idx, dwg_info in enumerate(scan, 1):
            doc_id = dwg_info['doc_id']
            print(f"  [{idx}/{total}] {doc_id} (IFR Rev{dwg_info['latest_ifr_rev']} → "
                  f"IFC Rev{dwg_info['existing_ifc_rev'] + 1 if dwg_info['existing_ifc_rev'] is not None else 0})...")

            r = self.convert_to_ifc(dwg_info)
            results.append(r)

            if r['success']:
                status = "OK"
                extra = f"DWG={Path(r['dwg_path']).name}" if r['dwg_path'] else ''
                if r['pdf_path']:
                    extra += f", PDF={Path(r['pdf_path']).name}"
                if r['errors']:
                    extra += f" (警告: {'; '.join(r['errors'])})"
                print(f"    [{status}] {extra}")
            else:
                print(f"    [FAIL] {'; '.join(r['errors'])}")

        # Summary
        ok = sum(1 for r in results if r['success'])
        fail = total - ok
        print(f"\n  转换完成: 成功={ok}, 失败={fail}")

        # Auto-update deliverables if requested
        if update_deliverables and ok > 0:
            print("\n正在更新交付物状态...")
            self.update_ifc_deliverables(results)

        return results

    def update_ifc_deliverables(self, conversion_results: List[Dict]) -> Dict:
        """Update deliverables Excel with IFC revision info from conversion results.

        Uses exact revision numbers from the conversion rather than re-scanning,
        ensuring accuracy. Updates revision, status ('Approved IFC'), and date.

        Returns summary dict: {updated: int, skipped: int, errors: list}.
        """
        summary = {'updated': 0, 'skipped': 0, 'errors': []}

        # Filter to successful conversions
        successful = [r for r in conversion_results if r['success'] and r['ifc_rev'] is not None]
        if not successful:
            summary['errors'].append("没有成功的转换结果")
            return summary

        # Find deliverable Excel
        excel_path = self._dm.find_deliverable_excel()
        if not excel_path:
            summary['errors'].append("未找到交付物 Excel 文件")
            return summary

        if not OPENPYXL_AVAILABLE:
            summary['errors'].append("openpyxl 未安装")
            return summary

        wb = self._dm._load_workbook_with_retry(excel_path)
        if wb is None:
            summary['errors'].append(f"无法打开 Excel: {excel_path}")
            return summary

        ws = wb.active
        try:
            layout = self._dm.detect_layout(ws)
        except Exception as e:
            summary['errors'].append(f"Excel 格式检测失败: {e}")
            wb.close()
            return summary

        # Build doc_id → row mapping from Excel
        excel_items, _ = self._dm.read_excel_items(ws, layout)

        # Clear previous highlights
        self._dm.clear_previous_highlights(ws, layout)

        date_str = datetime.now().strftime('%Y-%m-%d')
        updated_rows = []

        for result in successful:
            doc_id = result['doc_id']
            ifc_rev = result['ifc_rev']

            if doc_id not in excel_items:
                summary['skipped'] += 1
                print(f"    {doc_id}: 未在 Excel 中找到，跳过")
                continue

            row = excel_items[doc_id]['row']

            # Update revision (numeric IFC rev)
            ws.cell(row=row, column=layout.rev_col).value = str(ifc_rev)
            # Update status
            ws.cell(row=row, column=layout.status_col).value = 'Approved IFC'
            # Update date
            ws.cell(row=row, column=layout.date_col).value = date_str

            # Green fill on K-M
            ws.cell(row=row, column=layout.rev_col).fill = self._dm.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.date_col).fill = self._dm.STATUS_APPROVED_FILL
            ws.cell(row=row, column=layout.status_col).fill = self._dm.STATUS_APPROVED_FILL
            # Marker highlights on J and N
            ws.cell(row=row, column=10).fill = self._dm.CHANGE_MARKER_FILL
            ws.cell(row=row, column=14).fill = self._dm.CHANGE_MARKER_FILL

            updated_rows.append(doc_id)
            summary['updated'] += 1
            print(f"    {doc_id}: Rev{ifc_rev} → Approved IFC ✓")

        if not updated_rows:
            wb.close()
            return summary

        # Update file revision and last updated date
        old_file_rev = ws[layout.file_rev_cell].value
        new_file_rev = self._dm._increment_file_revision(old_file_rev)
        old_str = str(old_file_rev).strip() if old_file_rev else ''
        prefix_match = re.match(r'^((?:Revision|Rev)\s*)', old_str, re.IGNORECASE)
        if prefix_match:
            ws[layout.file_rev_cell].value = f"{prefix_match.group(1)}{new_file_rev}"
        else:
            ws[layout.file_rev_cell].value = new_file_rev
        ws[layout.last_updated_cell].value = date_str

        if not self.dry_run:
            new_path = self._dm._get_new_filename(excel_path, new_file_rev)
            self._dm._supersede_file(excel_path)
            wb.save(str(to_long_path(new_path)))
            wb.close()
            self._dm._sync_deliverable(new_path)

            # Cleanup old versions
            ss_folder = excel_path.parent / "SS"
            if ss_folder.exists():
                self._dm.cleanup_ss_folder(ss_folder, max_versions=3)

            print(f"\n  交付物 Excel 已保存: {new_path.name}")
            print(f"  更新 {summary['updated']} 行, 跳过 {summary['skipped']} 行")
        else:
            wb.close()
            print(f"\n  [DRY-RUN] 将更新 {summary['updated']} 行")

        return summary


# =============================================================================
# Panel IFC Manager (multi-page panel designs)
# =============================================================================

# Regex: split filename into doc_no and page_number
# e.g. "TSF-EN-ELE-DRG-10-01" → ("TSF-EN-ELE-DRG-10", "01")
_RE_PANEL_PAGE = re.compile(r'^(.+)-(\d{2}|0[A-Za-z])$')

# Regex: match IFC PDF names like "TSF-EN-ELE-DRG-10_Rev0_IFC.pdf"
_RE_PANEL_IFC_PDF = re.compile(r'^(.+?)_Rev(\d+)_IFC\.pdf$', re.IGNORECASE)


class PanelIFCManager:
    """IFC conversion for multi-page panel designs (one DWG per page).

    Unlike IFCManager which handles single-doc-per-DWG, this handles
    a flat folder of DWGs where multiple files belong to one document set
    (e.g. TSF-EN-ELE-DRG-10-00 through -10). Output is one multi-page
    PDF per document group via AutoCAD PUBLISH.

    Supports two title block update methods:
      - 'com': Direct COM API attribute modification (default)
      - 'utb': Lee-Mac Update Title Block AutoLISP plugin via CSV
    """

    # Max revision history rows to auto-detect (scan up to this)
    MAX_REV_ROWS = 6

    def __init__(self, source_folder, ifc_output=None, dry_run=False,
                 title_block_method='com', title_block_name=None,
                 utb_lsp_path=None):
        """
        Args:
            source_folder: Path to flat folder containing DWG files
            ifc_output: Path for PDF output (default: project's 4. IFC(Client))
            dry_run: Preview only, no changes
            title_block_method: 'com' or 'utb'
            title_block_name: Exact block name for COM mode (optional, uses fuzzy if None)
            utb_lsp_path: Path to Lee-Mac UTB .lsp file (required for 'utb' mode)
        """
        self.source_folder = Path(source_folder)
        self.ifc_output = Path(ifc_output) if ifc_output else self._detect_ifc_output()
        self.dry_run = dry_run
        self.title_block_method = title_block_method
        self.title_block_name = title_block_name
        self.utb_lsp_path = Path(utb_lsp_path) if utb_lsp_path else None
        self._acad = None

    def _detect_ifc_output(self) -> Path:
        """Walk up from source_folder to find project root, then return IFC(Client) path."""
        # Look for "1. Drawings" in parents
        for parent in self.source_folder.parents:
            ifc_dir = parent / "4. IFC(Client)"
            if ifc_dir.exists():
                return ifc_dir
            # Also check one level up from "1. Drawings"
            if parent.name == "1. Native" or parent.name.startswith("1."):
                drawings_dir = parent.parent
                ifc_dir = drawings_dir / "4. IFC(Client)"
                if ifc_dir.exists():
                    return ifc_dir
        # Fallback: sibling of source folder
        return self.source_folder.parent / "IFC_Output"

    def _get_acad(self):
        """Get or launch AutoCAD COM instance."""
        if not WIN32COM_AVAILABLE:
            raise RuntimeError("win32com 未安装，无法使用 AutoCAD COM API。请安装 pywin32。")
        if self._acad is not None:
            try:
                _ = self._acad.Visible
                return self._acad
            except Exception:
                self._acad = None
        try:
            self._acad = win32com.client.GetActiveObject("AutoCAD.Application")
        except Exception:
            try:
                print("  正在启动 AutoCAD（可能需要30秒）...")
                self._acad = win32com.client.Dispatch("AutoCAD.Application")
                self._acad.Visible = True
                # Wait for AutoCAD to be ready
                for _wait in range(60):
                    try:
                        _ = self._acad.Documents
                        break
                    except Exception:
                        time.sleep(1)
            except Exception as e:
                raise RuntimeError(f"无法连接或启动 AutoCAD: {e}")
        return self._acad

    # ── Scanning & Grouping ──────────────────────────────────────────────

    def scan_and_group(self) -> Dict[str, List[Dict]]:
        """Scan source folder DWGs and group by doc_no.

        Returns {doc_no: [{'page': '01', 'path': Path, 'filename': str}, ...]}
        sorted by page number within each group.
        """
        groups: Dict[str, List[Dict]] = {}
        for f in sorted(self.source_folder.iterdir()):
            if not f.is_file() or f.suffix.lower() != '.dwg':
                continue
            if f.name.startswith('~$'):
                continue
            m = _RE_PANEL_PAGE.match(f.stem)
            if not m:
                continue
            doc_no = m.group(1)
            page = m.group(2)
            if doc_no not in groups:
                groups[doc_no] = []
            groups[doc_no].append({
                'page': page,
                'path': f,
                'filename': f.name,
            })
        # Sort pages within each group
        for doc_no in groups:
            groups[doc_no].sort(key=lambda x: (
                # 00 first, then numeric, then 0A/0B etc at end
                0 if x['page'] == '00' else (2 if x['page'].startswith('0') and not x['page'].isdigit() else 1),
                x['page']
            ))
        return groups

    def _get_existing_ifc_rev(self, doc_no: str) -> int:
        """Check IFC(Client) folder for existing IFC PDFs of this doc_no.

        Returns next revision number (0 if no existing IFC found).
        """
        if not self.ifc_output.exists():
            return 0
        max_rev = -1
        for f in self.ifc_output.iterdir():
            if not f.is_file() or f.suffix.lower() != '.pdf':
                continue
            m = _RE_PANEL_IFC_PDF.match(f.name)
            if m and m.group(1) == doc_no:
                rev = int(m.group(2))
                if rev > max_rev:
                    max_rev = rev
        return max_rev + 1 if max_rev >= 0 else 0

    # ── Block Scanning (diagnostics) ────────────────────────────────────

    def scan_blocks(self, dwg_path: str) -> List[Dict]:
        """Scan a DWG file and return all block references with their attributes.

        Returns list of {layout, block_name, attributes: {tag: value}}.
        Used for diagnostics / remote debugging via bot.
        """
        import time as _time
        acad = self._get_acad()
        doc = acad.Documents.Open(str(dwg_path), True)  # ReadOnly
        _time.sleep(2)
        results = []
        try:
            for layout in doc.Layouts:
                layout_name = layout.Name
                block = layout.Block
                for i in range(block.Count):
                    entity = block.Item(i)
                    try:
                        if entity.EntityName != 'AcDbBlockReference':
                            continue
                    except Exception:
                        continue
                    try:
                        bname = entity.Name
                    except Exception:
                        continue
                    attrs = {}
                    try:
                        for attr in entity.GetAttributes():
                            attrs[attr.TagString] = attr.TextString
                    except Exception:
                        pass
                    results.append({
                        'layout': layout_name,
                        'block_name': bname,
                        'attributes': attrs,
                    })
        finally:
            doc.Close(False)
        return results

    # ── COM Retry Helper ───────────────────────────────────────────────

    @staticmethod
    def _com_retry(func, max_retries=5, delay=2):
        """Retry a COM call that may fail with 'Call was rejected by callee'."""
        for attempt in range(max_retries):
            try:
                return func()
            except Exception as e:
                err_str = str(e)
                # -2147418111 = RPC_E_CALL_REJECTED (callee busy)
                # -2147417848 = RPC_E_SERVERFAULT
                if '-2147418111' in err_str or '-2147417848' in err_str:
                    if attempt < max_retries - 1:
                        time.sleep(delay * (attempt + 1))
                        continue
                raise
        return None

    # ── Title Block Operations (shared helpers) ──────────────────────────

    def _find_title_block(self, doc):
        """Find title block in AutoCAD document.

        Returns (block_ref, attrs_dict, space). Search strategy:
          1. Exact name match (if title_block_name is set)
          2. Fuzzy: any block with DRAWINGNUMBER + REVISION tags
          3. Fuzzy: any block with >=5 attributes containing REV-like tags
        Searches PaperSpace first (title blocks are usually there).
        Logs all block names on failure to aid remote debugging.
        """
        # Search PaperSpace first — title blocks are almost always there
        spaces = []
        try:
            for layout in self._com_retry(lambda: doc.Layouts):
                if layout.Name.lower() != 'model':
                    spaces.append(self._com_retry(lambda: doc.PaperSpace))
                    break
        except Exception:
            pass
        try:
            ms = self._com_retry(lambda: doc.ModelSpace)
            if ms is not None:
                spaces.append(ms)
        except Exception:
            pass

        all_blocks_found = []
        best_candidate = None
        best_attr_count = 0

        for space in spaces:
            if space is None:
                continue
            count = self._com_retry(lambda: space.Count)
            if count is None:
                continue
            for i in range(count):
                try:
                    entity = self._com_retry(lambda idx=i: space.Item(idx))
                except Exception:
                    continue
                if entity is None:
                    continue
                try:
                    if entity.EntityName != 'AcDbBlockReference':
                        continue
                except Exception:
                    continue
                try:
                    block_name = entity.Name
                except Exception:
                    continue

                attrs = self._get_attrs_dict(entity)
                all_blocks_found.append((block_name, len(attrs), list(attrs.keys())[:10]))

                # Strategy 1: Exact match
                if self.title_block_name and block_name == self.title_block_name:
                    return entity, attrs, space

                # Strategy 2: Classic DRAWINGNUMBER + REVISION
                if 'DRAWINGNUMBER' in attrs and 'REVISION' in attrs:
                    return entity, attrs, space

                # Strategy 3: Any block with REV-like tags and many attributes
                rev_tags = [k for k in attrs if 'REV' in k.upper() or 'REVISION' in k.upper()]
                if rev_tags and len(attrs) >= 5 and len(attrs) > best_attr_count:
                    best_candidate = (entity, attrs, space)
                    best_attr_count = len(attrs)

        if best_candidate:
            print(f"    Title block 模糊匹配: {best_attr_count} 个属性")
            return best_candidate

        # Log all blocks found for remote debugging
        if all_blocks_found:
            print(f"    [DEBUG] 所有 block references ({len(all_blocks_found)}):")
            for bname, attr_count, sample_tags in all_blocks_found:
                print(f"      '{bname}' ({attr_count} attrs): {sample_tags}")
        else:
            print(f"    [DEBUG] 未找到任何 block reference")

        return None, {}, None

    def _get_attrs_dict(self, block_ref) -> Dict:
        """Extract TAG -> attribute object dict from a block reference."""
        attrs = {}
        try:
            raw_attrs = self._com_retry(lambda: block_ref.GetAttributes())
            if raw_attrs:
                for attr in raw_attrs:
                    tag = self._com_retry(lambda a=attr: a.TagString.upper())
                    if tag:
                        attrs[tag] = attr
        except Exception:
            pass
        return attrs

    def _detect_rev_rows(self, attrs: Dict) -> int:
        """Auto-detect how many revision history rows the title block has."""
        for n in range(self.MAX_REV_ROWS, 0, -1):
            if f"{n}REV" in attrs:
                return n
        return 0

    def _detect_personnel_tags(self, attrs: Dict) -> List[str]:
        """Auto-detect personnel TAG suffixes from row 1 attributes.

        Returns list like ['DRAWN', 'APPROVED', 'DESIGNED', 'PROJECT']
        or ['DRAWN', 'CHECK', 'ENGINEER', 'QA', 'PROJECT'].
        """
        known_suffixes = ['DRAWN', 'CHECK', 'ENGINEER', 'QA', 'PROJECT',
                          'APPROVED', 'DESIGNED', 'SUBJECT', 'DRAWINGNUMBER']
        found = []
        for suffix in known_suffixes:
            if f"1{suffix}" in attrs:
                found.append(suffix)
        return found

    def _read_latest_row(self, attrs: Dict) -> Dict:
        """Find the highest non-empty revision row and return all fields.

        Auto-detects row count and personnel tags.
        Returns {row, rev, drawn, approved, ...} with whatever tags exist.
        """
        rev_rows = self._detect_rev_rows(attrs)
        for row_num in range(rev_rows, 0, -1):
            rev_tag = f"{row_num}REV"
            if rev_tag in attrs:
                val = self._get_attr_text(attrs[rev_tag])
                if val:
                    info = {'row': row_num, 'rev': val}
                    # Collect all tags for this row
                    personnel_tags = self._detect_personnel_tags(attrs)
                    for tag in personnel_tags:
                        full_tag = f"{row_num}{tag}"
                        if full_tag in attrs:
                            info[tag.lower()] = self._get_attr_text(attrs[full_tag])
                        else:
                            info[tag.lower()] = ''
                    # Also get DATE and DESCRIPTION
                    for extra in ['DATE', 'DESCRIPTION']:
                        full_tag = f"{row_num}{extra}"
                        if full_tag in attrs:
                            info[extra.lower()] = self._get_attr_text(attrs[full_tag])
                    return info
        return {}

    def _set_attr_text(self, attr, value: str):
        """Set attribute TextString with COM retry protection."""
        self._com_retry(lambda: setattr(attr, 'TextString', value))

    def _get_attr_text(self, attr) -> str:
        """Get attribute TextString with COM retry protection."""
        result = self._com_retry(lambda: attr.TextString)
        return result.strip() if result else ''

    def _write_ifc_row(self, attrs: Dict, ifc_rev: int, personnel: Dict, date_str: str):
        """Update title block attributes for IFC conversion.

        Sets REVISION, clears all history rows, writes IFC info to row 1.
        Auto-detects row count and personnel tags.
        """
        # Set main REVISION attribute
        if 'REVISION' in attrs:
            self._set_attr_text(attrs['REVISION'], str(ifc_rev))

        # Auto-detect structure
        rev_rows = self._detect_rev_rows(attrs)
        personnel_tags = self._detect_personnel_tags(attrs)

        # Clear all revision rows
        all_suffixes = ['REV', 'DATE', 'DESCRIPTION'] + personnel_tags
        for row_num in range(1, rev_rows + 1):
            for suffix in all_suffixes:
                tag = f"{row_num}{suffix}"
                if tag in attrs:
                    self._set_attr_text(attrs[tag], '')

        # Write row 1 with IFC data
        if '1REV' in attrs:
            self._set_attr_text(attrs['1REV'], str(ifc_rev))
        if '1DESCRIPTION' in attrs:
            self._set_attr_text(attrs['1DESCRIPTION'], 'ISSUED FOR CONSTRUCTION')
        if '1DATE' in attrs:
            self._set_attr_text(attrs['1DATE'], date_str)

        # Copy personnel from latest IFR row
        for tag in personnel_tags:
            full_tag = f"1{tag}"
            if full_tag in attrs and tag.lower() in personnel:
                self._set_attr_text(attrs[full_tag], personnel[tag.lower()])

    # ── IFC Stamp ──────────────────────────────────────────────────────────

    # ── IFC Stamp constants (from Tatua_Standard_Frame.dwg 'IFR' block) ──
    STAMP_LAYER = 'IFC_STAMP'
    STAMP_X = 141.0          # MText insertion X (center of stamp)
    STAMP_Y = 589.176        # MText insertion Y (top of text, TopCenter)
    STAMP_RECT_W = 110.511   # rectangle width  (mm)
    STAMP_RECT_H = 17.745    # rectangle height (mm)
    STAMP_TEXT_Y_OFFSET = 13.182  # MText Y from rect bottom
    STAMP_TEXT_H = 7.0        # text height (mm)
    STAMP_TEXT_W = 116.419    # MText width property

    def _remove_status_stamp(self, doc):
        """Remove existing FOR CONSTRUCTION / FOR REVIEW stamps (MText + border)."""
        for get_space in [lambda: doc.PaperSpace, lambda: doc.ModelSpace]:
            try:
                space = self._com_retry(get_space)
                if space is None:
                    continue
                count = self._com_retry(lambda: space.Count)
                if count is None:
                    continue
                for i in range(count - 1, -1, -1):
                    try:
                        entity = self._com_retry(lambda idx=i: space.Item(idx))
                        if entity is None:
                            continue
                        # Remove stamp layer entities (both rect and text)
                        try:
                            layer = self._com_retry(lambda: entity.Layer)
                            if layer == self.STAMP_LAYER:
                                self._com_retry(lambda: entity.Delete())
                                continue
                        except Exception:
                            pass
                        # Also remove legacy MText stamps (no layer tag)
                        ename = self._com_retry(lambda: entity.EntityName)
                        if ename == 'AcDbMText':
                            text = self._com_retry(lambda: entity.TextString)
                            if text and ('FOR CONSTRUCTION' in text.upper()
                                         or 'FOR REVIEW' in text.upper()):
                                self._com_retry(lambda: entity.Delete())
                    except Exception:
                        continue
            except Exception:
                continue

    def _ensure_stamp_layer(self, doc):
        """Create IFC_STAMP layer if it doesn't exist."""
        try:
            layers = self._com_retry(lambda: doc.Layers)
            try:
                layers.Item(self.STAMP_LAYER)
            except Exception:
                layer = self._com_retry(lambda: layers.Add(self.STAMP_LAYER))
                if layer:
                    self._com_retry(lambda: setattr(layer, 'color', 1))  # red
        except Exception:
            pass

    def _add_ifc_stamp(self, doc, block_ref, space):
        """Add 'FOR CONSTRUCTION' stamp with border rectangle.

        Replicates the 'IFR' block style from Tatua_Standard_Frame.dwg:
          - Closed polyline rectangle (110.5mm x 17.75mm)
          - MText 'FOR CONSTRUCTION' centered, Height 7.0, TopCenter
          - Fixed position: (141, 589.176) in PaperSpace
          - All entities on 'IFC_STAMP' layer for easy removal
        """
        import pythoncom

        self._remove_status_stamp(doc)
        self._ensure_stamp_layer(doc)

        # Calculate rectangle corners from stamp constants
        rect_half_w = self.STAMP_RECT_W / 2.0
        rect_left   = self.STAMP_X - rect_half_w      # ~85.74
        rect_right  = self.STAMP_X + rect_half_w      # ~196.26
        rect_bottom = self.STAMP_Y - self.STAMP_TEXT_Y_OFFSET  # ~575.99
        rect_top    = rect_bottom + self.STAMP_RECT_H          # ~593.74

        stamp_text = "{\\fArial Narrow|b1;FOR CONSTRUCTION}"

        # Ensure correct PaperSpace layout is active
        try:
            for layout in self._com_retry(lambda: doc.Layouts):
                if layout.Name.lower() != 'model':
                    self._com_retry(lambda: setattr(doc, 'ActiveLayout', layout))
                    print(f"    印章: 已切换到 layout '{layout.Name}'")
                    break
            time.sleep(0.5)
            space = self._com_retry(lambda: doc.PaperSpace)
        except Exception as e:
            print(f"    印章: layout 切换警告: {e}")

        if space is None:
            print(f"    印章: 无法获取 PaperSpace，跳过")
            return

        # --- Add rectangle border (closed polyline) ---
        rect_ok = False
        try:
            rect_pts = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8,
                [rect_left, rect_bottom,
                 rect_right, rect_bottom,
                 rect_right, rect_top,
                 rect_left, rect_top])
            pline = self._com_retry(lambda: space.AddLightWeightPolyline(rect_pts))
            if pline is not None:
                self._com_retry(lambda: setattr(pline, 'Closed', True))
                self._com_retry(lambda: setattr(pline, 'Layer', self.STAMP_LAYER))
                rect_ok = True
                print(f"    印章: 矩形边框 ({rect_left:.1f},{rect_bottom:.1f})-({rect_right:.1f},{rect_top:.1f})")
        except Exception as e:
            print(f"    印章: 矩形边框失败: {e}")

        # --- Add MText ---
        try:
            pt = win32com.client.VARIANT(
                pythoncom.VT_ARRAY | pythoncom.VT_R8,
                [self.STAMP_X, self.STAMP_Y, 0.0])
            mtext = self._com_retry(lambda: space.AddMText(pt, self.STAMP_TEXT_W, stamp_text))
            if mtext is not None:
                self._com_retry(lambda: setattr(mtext, 'Height', self.STAMP_TEXT_H))
                self._com_retry(lambda: setattr(mtext, 'AttachmentPoint', 2))  # TopCenter
                self._com_retry(lambda: setattr(mtext, 'Layer', self.STAMP_LAYER))
                try:
                    doc.Regen(1)
                except Exception:
                    pass
                print(f"    印章: FOR CONSTRUCTION 已添加 at ({self.STAMP_X},{self.STAMP_Y})"
                      f" {'带边框' if rect_ok else '无边框'}")
                return
        except Exception as e:
            print(f"    印章 COM 失败: {e}, 尝试 SendCommand...")

        # --- Fallback: SendCommand ---
        try:
            # Rectangle via RECTANG command
            cmd_rect = (
                f'RECTANG\n'
                f'{rect_left},{rect_bottom}\n'
                f'{rect_right},{rect_top}\n'
            )
            doc.SendCommand(cmd_rect)
            time.sleep(1)

            # MText via -MTEXT command
            opp_x = self.STAMP_X - self.STAMP_TEXT_W / 2
            opp_y = self.STAMP_Y - self.STAMP_TEXT_H
            cmd_text = (
                f'-MTEXT\n'
                f'{self.STAMP_X},{self.STAMP_Y}\n'
                f'J\nTC\n'
                f'H\n{self.STAMP_TEXT_H}\n'
                f'{opp_x},{opp_y}\n'
                f'{stamp_text}\n'
                f'\n'
            )
            doc.SendCommand(cmd_text)

            for _ in range(15):
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        break
                except Exception:
                    break
                time.sleep(0.5)

            try:
                doc.Regen(1)
            except Exception:
                pass
            print(f"    印章: FOR CONSTRUCTION 已添加 (SendCommand)")
        except Exception as e:
            print(f"    印章添加失败: {e}")

    # ── COM Mode ─────────────────────────────────────────────────────────

    def _update_via_com(self, doc, ifc_rev: int, date_str: str) -> Dict:
        """Update title block via direct COM API attribute modification.

        Returns {'success': bool, 'error': str, 'personnel': dict}.
        """
        block_ref, attrs, space = self._find_title_block(doc)
        if not attrs:
            return {'success': False, 'error': '未找到 title block', 'personnel': {}}

        personnel = self._read_latest_row(attrs)
        if not personnel:
            # Cover/index sheets may have no revision rows — proceed with empty personnel
            print(f"    注意: 未找到 IFR revision 行，使用空 personnel 继续")
            personnel = {}

        self._write_ifc_row(attrs, ifc_rev, personnel, date_str)
        self._add_ifc_stamp(doc, block_ref, space)
        return {'success': True, 'error': '', 'personnel': personnel}

    # ── UTB Mode ─────────────────────────────────────────────────────────

    def _generate_ifc_csv(self, pages: List[Dict], ifc_rev: int, date_str: str,
                          existing_csv: Optional[Path] = None) -> Path:
        """Generate IFC CSV for Lee-Mac UTB plugin.

        If existing_csv is provided, reads it and transforms IFR → IFC values.
        Otherwise, generates from scratch using COM to read current attributes.
        """
        import csv

        # Try to find existing CSV in source folder
        if existing_csv is None:
            for f in self.source_folder.iterdir():
                if f.suffix.lower() == '.csv' and 'RGSTR' in f.name.upper():
                    existing_csv = f
                    break

        output_csv = self.source_folder / '_IFC_TEMP.csv'

        if existing_csv and existing_csv.exists():
            # Read existing CSV and transform values
            with open(existing_csv, 'r', encoding='utf-8-sig') as fh:
                reader = csv.DictReader(fh)
                headers = reader.fieldnames
                rows = list(reader)

            page_filenames = {p['path'].stem for p in pages}

            new_rows = []
            for row in rows:
                filename = row.get('FILENAME', '').strip()
                if not filename or filename not in page_filenames:
                    continue
                # Transform IFR → IFC values
                row['1REV'] = str(ifc_rev)
                row['1DESCRIPTION'] = 'ISSUED FOR CONSTRUCTION'
                row['1DATE'] = date_str
                row['REVISION'] = str(ifc_rev)
                # Clear rows 2-4
                for n in range(2, 5):
                    for suffix in headers:
                        if suffix.startswith(f'{n}'):
                            row[suffix] = ''
                new_rows.append(row)

            with open(output_csv, 'w', encoding='utf-8', newline='') as fh:
                writer = csv.DictWriter(fh, fieldnames=headers)
                writer.writeheader()
                writer.writerows(new_rows)
        else:
            raise FileNotFoundError(
                f"未找到 drawing register CSV 文件。UTB 模式需要 CSV。\n"
                f"请确保 {self.source_folder} 中有 *RGSTR*.csv 文件。"
            )

        return output_csv

    def _update_via_utb(self, doc, csv_path: Path) -> Dict:
        """Update title block via Lee-Mac UTB AutoLISP plugin.

        Loads UTB.lsp, runs utb command which reads the CSV and updates attributes.
        Returns {'success': bool, 'error': str}.
        """
        if not self.utb_lsp_path or not self.utb_lsp_path.exists():
            return {'success': False, 'error': f'UTB .lsp 文件不存在: {self.utb_lsp_path}'}

        try:
            lsp_str = str(self.utb_lsp_path).replace('\\', '/')
            # Load the UTB lisp file
            doc.SendCommand(f'(load "{lsp_str}")\n')
            time.sleep(1)

            # Run the UTB command
            doc.SendCommand('utb\n')

            # Wait for completion
            start = time.time()
            while time.time() - start < 30:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        break
                except Exception:
                    break
                time.sleep(0.5)

            return {'success': True, 'error': ''}
        except Exception as e:
            return {'success': False, 'error': f'UTB 执行失败: {e}'}

    # ── Fallback Logic ───────────────────────────────────────────────────

    def _update_with_fallback(self, doc, ifc_rev: int, date_str: str,
                              csv_path: Optional[Path] = None) -> Dict:
        """Try primary method, auto-fallback on failure (no interactive prompt)."""
        if self.title_block_method == 'com':
            result = self._update_via_com(doc, ifc_rev, date_str)
            if not result['success'] and csv_path:
                print(f"    COM API 失败: {result['error']}，自动切换 UTB 模式")
                self.title_block_method = 'utb'
                result = self._update_via_utb(doc, csv_path)
        else:
            if csv_path is None:
                return {'success': False, 'error': 'UTB 模式需要 CSV 文件'}
            result = self._update_via_utb(doc, csv_path)
            if not result['success']:
                print(f"    UTB 失败: {result['error']}，自动切换 COM API 模式")
                self.title_block_method = 'com'
                result = self._update_via_com(doc, ifc_rev, date_str)
        return result

    # ── PDF Export (PUBLISH multi-page) ──────────────────────────────────

    def _build_dsd(self, pages: List[Dict], pdf_path: Path, dwg_folder: Path) -> str:
        """Build DSD (Drawing Set Description) file content for PUBLISH.

        Uses per-page layout names stored in page['layout'] during batch processing.
        """
        lines = [
            '[DWF6Version]',
            'Ver=1',
            '[DWF6MinorVersion]',
            'MinorVer=1',
        ]

        for page in pages:
            ifc_dwg = dwg_folder / page['path'].name
            dwg_path = ifc_dwg if ifc_dwg.exists() else page['path']
            if not dwg_path.exists():
                continue
            sheet_name = page['path'].stem
            layout = page.get('layout', 'Layout1')
            dwg_str = str(dwg_path)
            lines.extend([
                f'[DWF6Sheet:{sheet_name}-{layout}]',
                f'DWG={dwg_str}',
                f'Layout={layout}',
                'Setup=',
                f'OriginalSheetPath={dwg_str}',
                'Has Plot Port=0',
                'Has3DDWF=0',
            ])

        pdf_str = str(pdf_path)
        out_str = str(pdf_path.parent)
        lines.extend([
            '[Target]',
            'Type=6',
            f'DWF={pdf_str}',
            f'OUT={out_str}',
            'PWD=',
            'PromptForDwfName=FALSE',
        ])

        return '\n'.join(lines)

    def _publish_group_pdf(self, doc_no: str, ifc_rev: int,
                           dwg_folder: Path, pages: List[Dict]) -> Dict:
        """Create multi-page PDF using PUBLISH command with DSD file."""
        pdf_name = f"{doc_no}_Rev{ifc_rev}_IFC.pdf"
        pdf_path = self.ifc_output / pdf_name
        result = {'success': False, 'pdf_path': str(pdf_path), 'error': ''}

        if self.dry_run:
            result['success'] = True
            return result

        to_long_path(self.ifc_output).mkdir(parents=True, exist_ok=True)

        # Build DSD file
        dsd_content = self._build_dsd(pages, pdf_path, dwg_folder)
        dsd_path = dwg_folder / f"{doc_no}_IFC.dsd"
        dsd_path.write_text(dsd_content, encoding='utf-8')

        try:
            acad = self._get_acad()
            # Close any open documents first
            try:
                while acad.Documents.Count > 0:
                    acad.Documents.Item(0).Close(False)
                    time.sleep(1)
            except Exception:
                pass
            time.sleep(2)

            # Open first DWG to run PUBLISH
            first_dwg = dwg_folder / pages[0]['path'].name
            if not first_dwg.exists():
                first_dwg = pages[0]['path']
            doc = None
            for _retry in range(3):
                try:
                    doc = self._com_retry(
                        lambda p=str(first_dwg): acad.Documents.Open(p))
                    if doc is not None:
                        _ = doc.Name
                        break
                except Exception:
                    time.sleep(3)
            if doc is None:
                result['error'] = f"无法打开 {first_dwg.name} 用于 PUBLISH"
                return result

            # Wait for document to load
            for _wait in range(30):
                try:
                    _ = doc.ModelSpace.Count
                    break
                except Exception:
                    time.sleep(1)
            time.sleep(2)

            # Suppress dialogs
            try:
                doc.SetVariable("FILEDIA", 0)
                doc.SetVariable("BACKGROUNDPLOT", 0)
            except Exception:
                pass

            dsd_str = str(dsd_path)
            doc.SendCommand(f'-PUBLISH\n{dsd_str}\n')

            # Wait for PUBLISH to complete
            start = time.time()
            max_wait = 300
            last_log = start
            timed_out = True
            while time.time() - start < max_wait:
                try:
                    if doc.GetVariable("CMDACTIVE") == 0:
                        timed_out = False
                        break
                except Exception:
                    timed_out = False
                    break
                if time.time() - last_log > 30:
                    print(f"    PUBLISH 进行中... ({int(time.time() - start)}s)")
                    last_log = time.time()
                time.sleep(2)

            if timed_out:
                print(f"    PUBLISH 超时 ({max_wait}s)")

            # Restore and close
            try:
                doc.SetVariable("FILEDIA", 1)
                doc.SetVariable("BACKGROUNDPLOT", 2)
            except Exception:
                pass
            try:
                doc.Close(False)
            except Exception:
                pass

            time.sleep(3)
            if pdf_path.exists():
                result['success'] = True
            else:
                reason = "PUBLISH 超时" if timed_out else "PUBLISH 完成但 PDF 未生成"
                result['error'] = f"{reason}: {pdf_name}"

        except Exception as e:
            result['error'] = f"PUBLISH 失败: {e}"

        return result

    # ── Main Batch Conversion ────────────────────────────────────────────

    def batch_convert_panel(self, doc_no_filter: Optional[str] = None) -> List[Dict]:
        """Convert panel design DWGs to IFC.

        Args:
            doc_no_filter: If set, only convert this specific doc_no.

        Returns list of per-group result dicts.
        """
        groups = self.scan_and_group()
        if not groups:
            print("  未找到可分组的 DWG 文件")
            return []

        if doc_no_filter:
            groups = {k: v for k, v in groups.items() if k == doc_no_filter}
            if not groups:
                print(f"  未找到 doc_no={doc_no_filter} 的 DWG 文件")
                return []

        date_str = datetime.now().strftime('%d/%m/%y')
        results = []

        # Prepare IFC subfolder
        ifc_folder = self.source_folder / "IFC"
        if not self.dry_run:
            to_long_path(ifc_folder).mkdir(parents=True, exist_ok=True)

        # Prepare UTB CSV if needed
        csv_path = None
        if self.title_block_method == 'utb':
            all_pages = [p for pages in groups.values() for p in pages]
            # Determine IFC rev (use first group's rev for CSV generation)
            first_doc_no = next(iter(groups))
            first_rev = self._get_existing_ifc_rev(first_doc_no)
            try:
                csv_path = self._generate_ifc_csv(all_pages, first_rev, date_str)
                print(f"  已生成 IFC CSV: {csv_path.name}")
            except FileNotFoundError as e:
                print(f"  错误: {e}")
                return []

        for doc_no, pages in groups.items():
            ifc_rev = self._get_existing_ifc_rev(doc_no)

            group_result = {
                'doc_no': doc_no,
                'ifc_rev': ifc_rev,
                'pages': len(pages),
                'dwgs_updated': 0,
                'dwgs_failed': 0,
                'pdf_result': None,
                'errors': [],
            }

            print(f"\n  [{doc_no}] {len(pages)} 页 -> IFC Rev{ifc_rev}")

            if self.dry_run:
                # Preview mode
                for page in pages:
                    print(f"    {page['filename']} (page {page['page']})")
                pdf_name = f"{doc_no}_Rev{ifc_rev}_IFC.pdf"
                print(f"    -> PDF: {pdf_name}")
                group_result['dwgs_updated'] = len(pages)
                group_result['pdf_result'] = {'success': True, 'pdf_path': str(self.ifc_output / pdf_name)}
                results.append(group_result)
                continue

            # --- Actual conversion ---
            acad = self._get_acad()

            # Ensure IFC output directory exists for SaveAs
            to_long_path(ifc_folder).mkdir(parents=True, exist_ok=True)

            for page in pages:
                dwg_path = page['path']
                print(f"    打开: {dwg_path.name}...", end=' ')

                # Check lock files — try to clean stale locks first
                dwl1 = dwg_path.with_suffix('.dwl')
                dwl2 = dwg_path.with_suffix('.dwl2')
                if dwl1.exists() or dwl2.exists():
                    # Check if file is actually open in AutoCAD
                    stale = True
                    try:
                        open_docs = [acad.Documents.Item(i).FullName
                                     for i in range(acad.Documents.Count)]
                        if str(dwg_path) in open_docs:
                            stale = False
                    except Exception:
                        pass
                    if stale:
                        # Remove stale lock files
                        for lf in (dwl1, dwl2):
                            try:
                                lf.unlink(missing_ok=True)
                            except Exception:
                                pass
                        print("(清理残留锁文件) ", end='')
                    else:
                        print("已锁定，跳过")
                        group_result['dwgs_failed'] += 1
                        group_result['errors'].append(f"{dwg_path.name}: 文件被锁定")
                        continue

                try:
                    # Retry Open up to 5 times with increasing delay
                    doc = None
                    for _retry in range(5):
                        try:
                            doc = self._com_retry(
                                lambda p=str(dwg_path): acad.Documents.Open(p))
                            # Validate: ensure doc is a real document, not a stale dispatch
                            if doc is None:
                                raise RuntimeError("Open 返回 None")
                            _ = doc.Name  # throws if doc is a bad dispatch
                            break
                        except Exception as open_err:
                            if _retry < 4:
                                time.sleep(3 + _retry * 2)
                            else:
                                raise open_err
                    # Wait for document to fully load
                    for _wait_load in range(30):
                        try:
                            _ = doc.ModelSpace.Count
                            break
                        except Exception:
                            time.sleep(1)
                    else:
                        raise RuntimeError("文档加载超时")
                    # Extra settle time before COM operations
                    time.sleep(2)
                    # Detect layout name for this DWG (used later in PUBLISH DSD)
                    try:
                        for layout in doc.Layouts:
                            if layout.Name.lower() != 'model':
                                page['layout'] = layout.Name
                                break
                    except Exception:
                        pass
                    if 'layout' not in page:
                        page['layout'] = 'Layout1'
                except Exception as e:
                    print(f"打开失败: {e}")
                    group_result['dwgs_failed'] += 1
                    group_result['errors'].append(f"{dwg_path.name}: {e}")
                    continue

                # Update title block
                tb_result = self._update_with_fallback(doc, ifc_rev, date_str, csv_path)

                if tb_result['success']:
                    # SaveAs to IFC subfolder
                    ifc_dwg = ifc_folder / dwg_path.name
                    try:
                        doc.SaveAs(str(ifc_dwg))
                        doc.Close(False)
                        time.sleep(5)  # Let AutoCAD fully settle before next Open
                        group_result['dwgs_updated'] += 1
                        print("OK")
                    except Exception as e:
                        try:
                            doc.Close(False)
                        except Exception:
                            pass
                        time.sleep(2)
                        group_result['dwgs_failed'] += 1
                        group_result['errors'].append(f"{dwg_path.name}: SaveAs 失败 - {e}")
                        print(f"SaveAs 失败: {e}")
                else:
                    try:
                        doc.Close(False)
                    except Exception:
                        pass
                    time.sleep(2)
                    group_result['dwgs_failed'] += 1
                    group_result['errors'].append(f"{dwg_path.name}: {tb_result['error']}")
                    print(f"TB 更新失败: {tb_result['error']}")

            # PUBLISH multi-page PDF — only include successfully processed pages
            if group_result['dwgs_updated'] > 0:
                ok_pages = [p for p in pages if (ifc_folder / p['path'].name).exists()]
                if ok_pages:
                    print(f"    正在生成多页 PDF ({len(ok_pages)}/{len(pages)} 页)...")
                    pdf_result = self._publish_group_pdf(doc_no, ifc_rev, ifc_folder, ok_pages)
                    group_result['pdf_result'] = pdf_result
                    if pdf_result['success']:
                        print(f"    PDF: {Path(pdf_result['pdf_path']).name}")
                    else:
                        print(f"    PDF 失败: {pdf_result['error']}")
                        group_result['errors'].append(pdf_result['error'])
                else:
                    print(f"    IFC 文件夹中无可用 DWG，跳过 PDF 导出")
            else:
                print(f"    所有 DWG 更新失败，跳过 PDF 导出")

            results.append(group_result)

        # Summary
        total_ok = sum(r['dwgs_updated'] for r in results)
        total_fail = sum(r['dwgs_failed'] for r in results)
        pdf_ok = sum(1 for r in results if r.get('pdf_result', {}).get('success'))
        print(f"\n  完成: DWG 更新 {total_ok} 成功 / {total_fail} 失败, PDF {pdf_ok}/{len(results)} 组")

        return results


# =============================================================================
# Pipeline Orchestrator (NEW in v7)
# =============================================================================

class PipelineOrchestrator:
    """Sequences the 3 stages per project: IFR Sync → Version Mgmt → Deliverable."""

    def __init__(self, config: ConfigManager, dry_run: bool = False,
                 stages: Optional[List[str]] = None):
        self.config = config
        self.dry_run = dry_run
        self.stages = stages or ['ifr_sync', 'version_mgmt', 'deliverable']

    def run_pipeline(self, project_path: Path, project_validation: ProjectValidation) -> Dict:
        """Run the full pipeline for a single project."""
        results = {
            'project_name': project_validation.project_name,
            'ifr_sync': None,
            'version_mgmt': None,
            'deliverable': None,
            'success': True,
        }

        UIHelper.print_separator("=", 72)
        print(f"\n  管线处理: {project_validation.project_name}")
        UIHelper.print_separator("=", 72)

        # Stage 1: IFR Sync
        if 'ifr_sync' in self.stages:
            print(f"\n  [Stage 1/3] IFR 同步...")
            try:
                automation = IFRAutomation(
                    root_path=str(project_path.parent),
                    config=self.config,
                    dry_run=self.dry_run,
                    interactive=True
                )
                sync_result = automation.process_project(project_validation)
                results['ifr_sync'] = {
                    'success': sync_result.success,
                    'folders': sync_result.folders_created,
                    'drawings': sync_result.drawings_copied,
                    'reports': sync_result.reports_copied,
                }
                if sync_result.success:
                    UIHelper.print_success(
                        f"IFR同步完成: 文件夹={sync_result.folders_created}, "
                        f"图纸={sync_result.drawings_copied}, 报告={sync_result.reports_copied}")
                else:
                    UIHelper.print_error(f"IFR同步失败: {sync_result.errors}")
                    results['success'] = False
            except Exception as e:
                UIHelper.print_error(f"IFR同步异常: {e}")
                results['ifr_sync'] = {'success': False, 'error': str(e)}
                results['success'] = False

        # Stage 2: Version Management
        if 'version_mgmt' in self.stages:
            print(f"\n  [Stage 2/3] 版本管理...")
            try:
                vm = VersionManager(str(project_path.parent), dry_run=self.dry_run)
                vm_stats = vm.process_project(project_path, show_details=True)
                results['version_mgmt'] = vm_stats
                UIHelper.print_success(
                    f"版本管理完成: 扫描={vm_stats['scanned']}, 移动={vm_stats['moved']}")
            except Exception as e:
                UIHelper.print_error(f"版本管理异常: {e}")
                results['version_mgmt'] = {'error': str(e)}
                results['success'] = False

        # Stage 3: Deliverable Cross-Check
        if 'deliverable' in self.stages:
            print(f"\n  [Stage 3/3] 交付物检查...")
            try:
                dm = DeliverableManager(project_path, dry_run=self.dry_run)
                excel_path = dm.find_deliverable_excel()
                if excel_path:
                    check_result = dm.cross_check(excel_path)
                    self._print_cross_check_summary(check_result)
                    if not self.dry_run and (check_result.items_in_folders_not_excel or
                                              check_result.revision_mismatches or
                                              check_result.doc_id_corrections or
                                              check_result.status_updates):
                        check_result = dm.apply_updates(excel_path, check_result)
                        UIHelper.print_success(
                            f"交付物更新完成: 新增={check_result.rows_inserted}, "
                            f"更新={check_result.rows_updated}")
                    results['deliverable'] = {
                        'new_items': len(check_result.items_in_folders_not_excel),
                        'rev_mismatches': len(check_result.revision_mismatches),
                        'doc_id_corrections': len(check_result.doc_id_corrections),
                        'status_updates': len(check_result.status_updates),
                        'naming_warnings': len(check_result.naming_warnings),
                        'inserted': check_result.rows_inserted,
                        'updated': check_result.rows_updated,
                    }
                else:
                    UIHelper.print_warning("未找到交付物 Excel 文件")
                    results['deliverable'] = {'skipped': True, 'reason': '未找到Excel'}
            except Exception as e:
                UIHelper.print_error(f"交付物检查异常: {e}")
                results['deliverable'] = {'error': str(e)}
                results['success'] = False

        return results

    def _print_cross_check_summary(self, result: DeliverableCrossCheckResult):
        """Print cross-check results summary."""
        if result.errors:
            for err in result.errors:
                UIHelper.print_error(err)
            return

        if result.items_in_folders_not_excel:
            print(f"\n    新增文件 (在文件夹中但不在Excel中): {len(result.items_in_folders_not_excel)}")
            for item in result.items_in_folders_not_excel[:10]:
                print(f"      + {item['doc_id']} ({item['filename']})")
            if len(result.items_in_folders_not_excel) > 10:
                print(f"      ... 还有 {len(result.items_in_folders_not_excel) - 10} 个")

        if result.items_in_excel_not_folders:
            print(f"\n    仅在Excel中 (信息): {len(result.items_in_excel_not_folders)}")
            for doc_id in result.items_in_excel_not_folders[:5]:
                print(f"      ? {doc_id}")

        if result.doc_id_corrections:
            print(f"\n    FILE NO 修正 (按描述匹配): {len(result.doc_id_corrections)}")
            for corr in result.doc_id_corrections:
                print(f"      Row {corr['row']}: {corr['old_doc_id']} -> {corr['doc_id']} ({corr['description']})")

        if result.revision_mismatches:
            print(f"\n    版本不一致: {len(result.revision_mismatches)}")
            for mm in result.revision_mismatches[:10]:
                print(f"      {mm['doc_id']}: Excel={mm['excel_rev']} -> 文件夹={mm['folder_rev']}")

        if result.status_updates:
            print(f"\n    IFC 状态更新: {len(result.status_updates)}")
            for su in result.status_updates[:10]:
                print(f"      {su['doc_id']}: {su['old_status'] or '(空)'} -> {su['new_status']}")

        if result.naming_warnings:
            print(f"\n    命名规范警告: {len(result.naming_warnings)}")
            for nw in result.naming_warnings[:10]:
                print(f"      {nw['doc_id']}: {', '.join(nw['changes'])}")
                print(f"        建议: {nw['suggested']}")

        if (not result.items_in_folders_not_excel and not result.revision_mismatches
                and not result.doc_id_corrections and not result.status_updates):
            UIHelper.print_success("交付物清单与文件夹一致，无需更新")


# =============================================================================
# Interactive Mode (Unified)
# =============================================================================

class UnifiedInteractiveMode:
    """Unified interactive mode combining IFR Sync + Version Mgmt + Deliverable."""

    def __init__(self, config: ConfigManager):
        self.config = config
        self.ui = UIHelper()
        self.validator = ProjectValidator(config)
        self.scanner = ProjectScanner(config, self.validator)
        self.safety_checker = SafetyChecker(config)
        self.report_generator = ValidationReportGenerator()
        self.root_path: Optional[Path] = None

    def run(self):
        """Main interactive loop."""
        self.ui.clear_screen()
        self.show_welcome()

        while True:
            choice = self.show_main_menu()

            if choice == '0':
                self.ui.print_info("感谢使用，再见！")
                break
            elif choice == '1':
                self.full_pipeline_mode()
            elif choice == '2':
                self.scan_and_process_projects()
            elif choice == '3':
                self.version_mgmt_mode()
            elif choice == '4':
                self.deliverable_mode()
            elif choice == '5':
                self.native_dwg_menu()
            elif choice == '6':
                self.folder_relocate_menu()
            elif choice == '7':
                self.ifc_copy_mode()
            elif choice == '8':
                self.validate_only_mode()
            elif choice == '9':
                self.modify_config()
            elif choice == '10':
                self.ifc_convert_mode()
            elif choice == '11':
                self.panel_ifc_convert_mode()
            else:
                self.ui.print_warning("无效选项，请重新选择")

    def show_welcome(self):
        """Display welcome screen."""
        self.ui.print_header(
            "工程文档自动化管线 v8.0",
            "IFR Sync + Version Mgmt + Deliverable + Panel IFC"
        )

        print("当前配置信息：")
        root = self.config.get("default_root_path", "")
        if root:
            display = root[:50] + "..." if len(root) > 50 else root
            print(f"  [DIR] 根目录: {display}")
        else:
            print(f"  [DIR] 根目录: 未设置")

        print(f"  日志级别: {self.config.get('log_level', 'INFO')}")
        print(f"  [v] 自动备份: {'已启用' if self.config.get('auto_backup', True) else '已禁用'}")

        # Show filters
        filters = self.config.get("project_filters", {})
        whitelist = filters.get("whitelist", [])
        blacklist = filters.get("blacklist", [])
        if whitelist:
            print(f"  白名单: {', '.join(whitelist[:3])}{'...' if len(whitelist) > 3 else ''}")
        if blacklist:
            print(f"  黑名单: {', '.join(blacklist[:3])}{'...' if len(blacklist) > 3 else ''}")

        print()

    def show_main_menu(self) -> str:
        """Display main menu."""
        self.ui.print_separator()

        options = [
            ("1", "完整流程 (IFR Sync + Version Mgmt + Deliverable)"),
            ("2", "仅 IFR 同步"),
            ("3", "仅版本管理 (PDF)"),
            ("4", "仅交付物检查 & 更新"),
            ("5", "Native/Reports/Schedule 版本管理"),
            ("6", "文件夹归位"),
            ("7", "复制 IFC(Client) 到目标"),
            ("8", "仅验证项目结构"),
            ("9", "配置 / 日志"),
            ("10", "IFR → IFC 转换 (AutoCAD)"),
            ("11", "Panel IFC 批量转换 (多页 DWG)"),
            ("0", "退出"),
        ]

        return self.ui.print_menu(options, "请输入选项 [0-11]")

    def get_root_path(self) -> Optional[Path]:
        """Get root path from config or user input."""
        if self.root_path:
            return self.root_path

        config_root = self.config.get("default_root_path")
        if config_root and Path(config_root).exists():
            print(f"\n当前根目录: {config_root}")
            if self.ui.confirm("使用此目录？"):
                self.root_path = Path(config_root)
                return self.root_path

        print("\n请输入项目根目录路径：")
        path_str = input("> ").strip().strip('"')

        if path_str and Path(path_str).exists():
            self.root_path = Path(path_str)
            self.config.set("default_root_path", str(self.root_path))
            return self.root_path
        else:
            self.ui.print_error("路径不存在！")
            return None

    # =========================================================================
    # V7: New Pipeline / Deliverable / Version Mgmt modes
    # =========================================================================

    def full_pipeline_mode(self):
        """Run the full pipeline (IFR Sync + Version Mgmt + Deliverable) per project."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)
        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        all_projects = self.display_projects_with_status(projects_by_region)
        selected = self.get_project_selection(projects_by_region)
        if not selected:
            return

        # Ask for dry-run
        dry_run = not self.ui.confirm("\n执行实际操作？(选N为预览模式)", default=False)

        pipeline = PipelineOrchestrator(self.config, dry_run=dry_run)

        all_results = []
        for idx, project in enumerate(selected, 1):
            print(f"\n{'='*72}")
            print(f"管线处理项目 {idx}/{len(selected)}: {project.project_name}")
            print(f"{'='*72}")

            result = pipeline.run_pipeline(Path(project.project_path), project)
            all_results.append(result)

            if result['success']:
                self.ui.print_success(f"项目 {project.project_name} 管线完成")
            else:
                self.ui.print_warning(f"项目 {project.project_name} 部分失败")

        # Summary
        self.ui.print_header("管线执行完成", f"{'预览模式' if dry_run else '已执行'}")
        success_count = sum(1 for r in all_results if r['success'])
        print(f"  成功: {success_count}/{len(all_results)}")
        input("\n按 Enter 返回主菜单...")

    def version_mgmt_mode(self):
        """PDF version management mode (from version_manager_v4)."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        vm = VersionManager(str(root))
        projects = vm.scan_for_projects()

        if not projects:
            self.ui.print_warning("未找到任何项目！")
            return

        print(f"\n找到 {len(projects)} 个项目:")
        for i, project in enumerate(projects, 1):
            print(f"  [{i}] {project.name}")
        print(f"\n  [0] 返回主菜单")
        self.ui.print_separator()

        choice = input(f"\n请选择项目 [0-{len(projects)}，输入 'all' 处理全部]: ").strip()

        if choice == '0':
            return
        elif choice.lower() == 'all':
            selected_projects = projects
        else:
            try:
                idx = int(choice)
                if 1 <= idx <= len(projects):
                    selected_projects = [projects[idx - 1]]
                else:
                    return
            except ValueError:
                return

        # Preview or execute
        print("\n选择操作:")
        print("  [1] 执行移动")
        print("  [2] 仅预览 (dry-run)")
        print("  [0] 取消返回")
        op = input("\n请选择: ").strip()

        if op == '0':
            return

        dry_run = op != '1'
        if not dry_run:
            confirm = input("\n确认执行? [Y/n]: ").strip().lower()
            if confirm not in ('', 'y', 'yes'):
                print("已取消")
                return

        vm.dry_run = dry_run
        total_stats = {"scanned": 0, "groups": 0, "moved": 0}

        for project in selected_projects:
            print(f"\n处理: {project.name}")
            self.ui.print_separator()
            stats = vm.process_project(project)
            total_stats["scanned"] += stats["scanned"]
            total_stats["groups"] += stats["groups"]
            total_stats["moved"] += stats["moved"]

        print(f"\n完成: 扫描={total_stats['scanned']}, 组={total_stats['groups']}, 移动={total_stats['moved']}")
        input("\n按 Enter 返回主菜单...")

    def deliverable_mode(self):
        """Deliverable cross-check and update mode."""
        if not OPENPYXL_AVAILABLE:
            self.ui.print_error("需要安装 openpyxl: pip install openpyxl")
            input("\n按 Enter 返回...")
            return

        root = self.get_root_path()
        if not root:
            return

        # Scan for projects with deliverable Excel files
        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)
        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        all_projects = [p for projects in projects_by_region.values() for p in projects]

        # Find which projects have deliverable Excel files
        dlv_projects = []
        for p in all_projects:
            dm = DeliverableManager(Path(p.project_path))
            excel = dm.find_deliverable_excel()
            if excel:
                dlv_projects.append((p, excel))

        if not dlv_projects:
            self.ui.print_warning("没有找到包含交付物 Excel 的项目！")
            input("\n按 Enter 返回...")
            return

        print(f"\n找到 {len(dlv_projects)} 个包含交付物 Excel 的项目:")
        for i, (p, excel) in enumerate(dlv_projects, 1):
            print(f"  [{i}] {p.project_name}")
            print(f"      Excel: {excel.name}")

        print(f"\n  [0] 返回主菜单")
        self.ui.print_separator()

        choice = input(f"\n请选择项目 [0-{len(dlv_projects)}]: ").strip()
        if choice == '0':
            return

        try:
            idx = int(choice)
            if not (1 <= idx <= len(dlv_projects)):
                return
        except ValueError:
            return

        project, excel_path = dlv_projects[idx - 1]
        project_path = Path(project.project_path)

        # Cross-check
        print(f"\n正在检查: {project.project_name}")
        dm = DeliverableManager(project_path)

        check_result = dm.cross_check(excel_path)

        if check_result.errors:
            for err in check_result.errors:
                self.ui.print_error(err)
            input("\n按 Enter 返回...")
            return

        # Display results
        print(f"\n{'='*60}")
        print(f"  交付物检查结果: {project.project_name}")
        print(f"{'='*60}")

        if check_result.items_in_folders_not_excel:
            print(f"\n  新增文件 (在文件夹中但不在Excel): {len(check_result.items_in_folders_not_excel)}")
            for item in check_result.items_in_folders_not_excel:
                print(f"    + {item['doc_id']} | {item['description'][:40]} | Rev {item['revision'] or '?'}")

        if check_result.items_in_excel_not_folders:
            print(f"\n  仅在Excel中 (信息): {len(check_result.items_in_excel_not_folders)}")
            for doc_id in check_result.items_in_excel_not_folders[:10]:
                print(f"    ? {doc_id}")

        if check_result.revision_mismatches:
            print(f"\n  版本不一致: {len(check_result.revision_mismatches)}")
            for mm in check_result.revision_mismatches:
                print(f"    {mm['doc_id']}: Excel=Rev{mm['excel_rev']} -> 文件夹=Rev{mm['folder_rev']}")

        if check_result.status_updates:
            print(f"\n  IFC 状态更新: {len(check_result.status_updates)}")
            for su in check_result.status_updates:
                print(f"    {su['doc_id']}: {su['old_status'] or '(空)'} -> {su['new_status']}")

        if check_result.naming_warnings:
            print(f"\n  命名规范警告: {len(check_result.naming_warnings)}")
            for nw in check_result.naming_warnings[:10]:
                print(f"    {nw['doc_id']}: {', '.join(nw['changes'])}")
                print(f"      当前: {nw['filename']}")
                print(f"      建议: {nw['suggested']}")

        if (not check_result.items_in_folders_not_excel and not check_result.revision_mismatches
                and not check_result.status_updates and not check_result.doc_id_corrections):
            self.ui.print_success("交付物清单与文件夹一致，无需更新")
            input("\n按 Enter 返回...")
            return

        # Ask to apply
        self.ui.print_separator()
        print("\n选择操作:")
        print("  [1] 应用更新到 Excel")
        print("  [2] 预览模式 (dry-run)")
        print("  [0] 返回 (不执行)")
        op = input("\n请选择: ").strip()

        if op == '0':
            return

        dry_run = op != '1'
        if not dry_run:
            if not self.ui.confirm("确认更新Excel文件？"):
                return

        dm.dry_run = dry_run
        updated = dm.apply_updates(excel_path, check_result)

        if dry_run:
            print(f"\n[预览] 将新增 {updated.rows_inserted} 行, 更新 {updated.rows_updated} 行")
            print(f"[预览] 文件版本将更新为: {updated.new_file_rev}")
        else:
            self.ui.print_success(
                f"完成: 新增={updated.rows_inserted}, 更新={updated.rows_updated}, "
                f"新版本={updated.new_file_rev}")
            if updated.excel_path:
                print(f"  新文件: {updated.excel_path}")

        input("\n按 Enter 返回主菜单...")

    def native_dwg_menu(self):
        """Native/Reports/Schedule version management sub-menu."""
        root = self.get_root_path()
        if not root:
            return

        vm = VersionManager(str(root))
        projects = vm.scan_for_projects()

        if not projects:
            self.ui.print_warning("未找到任何项目")
            input("\n按 Enter 返回...")
            return

        print("\n选择项目:")
        for i, project in enumerate(projects, 1):
            print(f"  [{i}] {project.name}")
        print(f"\n  [0] 返回主菜单")

        choice = input(f"\n请选择项目 [0-{len(projects)}]: ").strip()
        try:
            idx = int(choice)
            if idx == 0:
                return
            if 1 <= idx <= len(projects):
                project = projects[idx - 1]
            else:
                return
        except ValueError:
            return

        print(f"\n  项目: {project.name}")
        native_mgr = NativeVersionManager(str(project), dry_run=True)

        # Scope selection
        print("\n  扫描范围:")
        print("    [1] 全部 (Native + Reports + Schedule)")
        print("    [2] 仅 Native")
        print("    [3] 仅 Reports")
        print("    [4] 仅 Schedule")
        scope_choice = input("\n  请选择 [1-4, 默认=1]: ").strip() or '1'
        scope_map = {'1': 'all', '2': 'native', '3': 'reports', '4': 'schedule'}
        scope = scope_map.get(scope_choice, 'all')

        folder_filter = input("\n  筛选文件夹 (留空=全部): ").strip() or None

        results = native_mgr.process_all(folder_filter, scope)
        if not results:
            print("\n  没有找到文档文件夹")
            input("\n按 Enter 返回...")
            return

        total_actions = sum(len(r.actions) for _, r in results)
        folders_with_files = sum(1 for _, r in results if r.dwg_files)

        print(f"\n  [预览] 扫描 {len(results)} 个文件夹, "
              f"{folders_with_files} 个有文件, {total_actions} 个动作")
        self.ui.print_separator()

        for group_name, group_results in groupby(results, key=lambda x: x[0]):
            group_list = list(group_results)
            group_folders_with_files = [r for _, r in group_list if r.dwg_files]
            print(f"\n  === {group_name} === ({len(group_folders_with_files)} folders)")

            for _, result in group_list:
                if not result.dwg_files:
                    continue
                print(f"    {result.folder_name}")
                print(f"      doc_id: {result.doc_id}  |  {len(result.dwg_files)} files")
                if result.kept_ifr:
                    print(f"      [保留 IFR] {result.kept_ifr.filename}")
                if result.kept_ifc:
                    print(f"      [保留 IFC] {result.kept_ifc.filename}")
                for action in result.actions:
                    if action.action == 'rename':
                        print(f"      [重命名] {action.source.name} -> {action.dest.name}")
                    else:
                        print(f"      [->{action.dest.parent.name}/] {action.source.name}")

        if total_actions == 0:
            print("\n  [v] 所有文件已是最新标准格式")
            input("\n按 Enter 返回...")
            return

        renames = sum(1 for _, r in results for a in r.actions if a.action == 'rename')
        moves = sum(1 for _, r in results for a in r.actions if a.action == 'move_to_ss')
        print(f"\n  汇总: {renames} 个重命名, {moves} 个移动到 SS/")

        print("\n  选择操作:")
        print("    [1] 执行")
        print("    [0] 返回 (不执行)")
        execute_choice = input("\n  请选择: ").strip()
        if execute_choice == '1':
            confirm = input("\n  确认执行? 输入 'YES': ").strip()
            if confirm == 'YES':
                native_mgr.dry_run = False
                stats = native_mgr.execute_actions(results)
                print(f"\n  完成: {stats['renamed']} 重命名, {stats['moved']} 移动, "
                      f"{stats['ss_created']} SS/ 创建, {stats['errors']} 错误")
            else:
                print("\n  已取消")
        input("\n按 Enter 返回主菜单...")

    def folder_relocate_menu(self):
        """Folder relocation sub-menu."""
        root = self.get_root_path()
        if not root:
            return

        vm = VersionManager(str(root))
        projects = vm.scan_for_projects()

        if not projects:
            self.ui.print_warning("未找到任何项目")
            input("\n按 Enter 返回...")
            return

        print("\n选择项目:")
        for i, project in enumerate(projects, 1):
            print(f"  [{i}] {project.name}")
        print(f"\n  [0] 返回主菜单")

        choice = input(f"\n请选择项目 [0-{len(projects)}]: ").strip()
        try:
            idx = int(choice)
            if idx == 0:
                return
            if 1 <= idx <= len(projects):
                project = projects[idx - 1]
            else:
                return
        except ValueError:
            return

        print(f"\n  项目: {project.name}")
        relocator = FolderRelocator(project, dry_run=True)

        if not relocator.drawings_root.exists():
            print(f"\n  [!] 1. Drawings/ 不存在")
            input("\n按 Enter 返回...")
            return

        relocations = relocator.scan()
        if not relocations:
            self.ui.print_success("未检测到错位的文件夹，结构正确")
            input("\n按 Enter 返回...")
            return

        actionable = [r for r in relocations if r.action != 'warn']
        warnings = [r for r in relocations if r.action == 'warn']

        if actionable:
            print(f"\n  检测到 {len(actionable)} 个错位文件夹:")
            for i, rel in enumerate(actionable, 1):
                action_str = '移动' if rel.action == 'move' else '合并'
                print(f"    [{i}] [{action_str}] {rel.source.name}")
                print(f"         原因: {rel.reason}")

        if warnings:
            print(f"\n  警告 ({len(warnings)}):")
            for rel in warnings:
                print(f"    [!] {rel.reason}")

        if not actionable:
            input("\n按 Enter 返回...")
            return

        print("\n  选择操作:")
        print("    [1] 执行归位")
        print("    [0] 返回")
        if input("\n  请选择: ").strip() == '1':
            if input("\n  确认执行? 输入 'YES': ").strip() == 'YES':
                relocator.dry_run = False
                stats = relocator.execute(actionable)
                print(f"\n  完成: {stats['moved']} 移动, {stats['merged']} 合并, {stats['errors']} 错误")
            else:
                print("  已取消")
        input("\n按 Enter 返回主菜单...")

    def scan_and_process_projects(self):
        """Scan projects and process with confirmation (IFR sync only)."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)

        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        # Display projects with status
        self.display_projects_with_status(projects_by_region)

        # Get user selection
        selected = self.get_project_selection(projects_by_region)
        if not selected:
            return

        # Select operation
        operation = self.select_operation()
        if operation is None:
            return

        # IFC copy has its own flow
        if operation == 'ifc_copy':
            self.process_ifc_copy(selected)
            return

        # Process projects with confirmation
        self.process_projects_with_confirmation(selected, operation)

    def display_projects_with_status(self, projects_by_region: Dict[str, List[ProjectValidation]]):
        """Display projects grouped by region with status indicators."""
        total = sum(len(projects) for projects in projects_by_region.values())

        self.ui.print_header(f"检测到 {total} 个项目目录", "")

        all_projects = []
        project_idx = 1

        for region, projects in sorted(projects_by_region.items()):
            print(f"\n[DIR] 区域: {region}")
            self.ui.print_separator("-", 72)

            for project in projects:
                all_projects.append(project)

                # Status
                status_icon, status_text = self.ui.status_icon(
                    "ready" if project.recommended_action == "process" else
                    "confirm" if project.recommended_action == "confirm" else "legacy"
                )

                # Action
                action_text = self.ui.action_icon(project.recommended_action)

                # Source dirs count
                active_sources = sum(1 for s in project.source_dirs_found if s.exists and s.file_count > 0)

                # V4: IFC status
                if project.has_existing_ifc_client:
                    ifc_status = f"[v] {project.ifc_file_count} 个文件"
                else:
                    ifc_status = "[x] 不存在"

                print(f"\n  [{project_idx}] {project.project_name}")
                print(f"      结构状态: {status_icon} {status_text}")
                print(f"      建议操作: {action_text}")
                print(f"      源目录: {active_sources} 个有效")
                print(f"      IFC(Client): {ifc_status}")
                print(f"      最后修改: {project.last_modified}")

                if project.warning_message:
                    if COLORAMA_AVAILABLE:
                        print(f"      {Fore.YELLOW}警告: {project.warning_message}{Style.RESET_ALL}")
                    else:
                        print(f"      警告: {project.warning_message}")

                project_idx += 1

        # Summary
        print()
        self.ui.print_separator("=", 72)

        ready_projects = [p for projects in projects_by_region.values()
                        for p in projects if p.recommended_action == "process"]
        confirm_projects = [p for projects in projects_by_region.values()
                          for p in projects if p.recommended_action == "confirm"]
        skip_projects = [p for projects in projects_by_region.values()
                        for p in projects if p.recommended_action == "skip"]

        ready_nums = [str(i+1) for i, p in enumerate(all_projects) if p.recommended_action == "process"]
        confirm_nums = [str(i+1) for i, p in enumerate(all_projects) if p.recommended_action == "confirm"]
        skip_nums = [str(i+1) for i, p in enumerate(all_projects) if p.recommended_action == "skip"]

        print(f"\n自动推荐处理: 项目 #{', #'.join(ready_nums) if ready_nums else '无'} ({len(ready_projects)}个)")
        print(f"需要确认的项目: 项目 #{', #'.join(confirm_nums) if confirm_nums else '无'} ({len(confirm_projects)}个)")
        print(f"建议跳过的项目: 项目 #{', #'.join(skip_nums) if skip_nums else '无'} ({len(skip_projects)}个)")

        # V4: IFC summary
        ifc_projects = [p for projects in projects_by_region.values()
                       for p in projects if p.has_existing_ifc_client]
        if ifc_projects:
            ifc_nums = [str(i+1) for i, p in enumerate(all_projects) if p.has_existing_ifc_client]
            print(f"含 IFC(Client): 项目 #{', #'.join(ifc_nums)} ({len(ifc_projects)}个)")

        self.ui.print_separator("=", 72)

        return all_projects

    def get_project_selection(self, projects_by_region: Dict[str, List[ProjectValidation]]) -> List[ProjectValidation]:
        """Get user's project selection."""
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        ready_indices = [i for i, p in enumerate(all_projects) if p.recommended_action == "process"]

        options = [
            ("1", f"只处理推荐的项目 ({len(ready_indices)}个)"),
            ("2", "手动选择项目 (输入编号)"),
            ("3", "查看项目详细信息"),
            ("4", "处理所有项目（包括需要确认的）"),
            ("5", "生成验证报告"),
            ("0", "返回主菜单"),
        ]

        choice = self.ui.print_menu(options, "请选择")

        if choice == '0':
            return []
        elif choice == '1':
            return [all_projects[i] for i in ready_indices]
        elif choice == '2':
            print("\n输入项目编号（多选用逗号分隔，如: 1,3,4 或 1-5 或 'all'）：")
            selection = input("> ").strip()
            indices = self._parse_selection(selection, len(all_projects))
            return [all_projects[i-1] for i in indices]
        elif choice == '3':
            self._show_project_details(all_projects)
            return self.get_project_selection(projects_by_region)
        elif choice == '4':
            return [p for p in all_projects if p.recommended_action != "skip"]
        elif choice == '5':
            self._generate_and_save_report(projects_by_region)
            return self.get_project_selection(projects_by_region)

        return []

    def _parse_selection(self, selection: str, max_num: int) -> List[int]:
        """Parse user selection string."""
        indices = []
        selection = selection.strip().lower()

        if selection == 'all':
            return list(range(1, max_num + 1))

        parts = selection.replace(' ', '').split(',')
        for part in parts:
            if '-' in part:
                try:
                    start, end = part.split('-')
                    indices.extend(range(int(start), min(int(end) + 1, max_num + 1)))
                except:
                    pass
            else:
                try:
                    num = int(part)
                    if 1 <= num <= max_num:
                        indices.append(num)
                except:
                    pass

        return sorted(set(indices))

    def _show_project_details(self, projects: List[ProjectValidation]):
        """Show detailed information for a project."""
        print("\n输入项目编号查看详情（0 返回）：")
        choice = input("> ").strip()

        try:
            idx = int(choice)
            if idx == 0:
                return
            if 1 <= idx <= len(projects):
                p = projects[idx - 1]
                print()
                self.ui.print_separator("=", 72)
                print(f"[i] 项目: {p.project_name}")
                print(f"[DIR] 路径: {p.project_path}")
                print(f"[i] 区域: {p.region}")
                print(f"[i] 最后修改: {p.last_modified} ({p.days_since_modified} 天前)")
                print(f"\n[DIR] 源目录状态:")

                for source in p.source_dirs_found:
                    status = "[v]" if source.exists and source.file_count > 0 else "[x]"
                    if COLORAMA_AVAILABLE:
                        color = Fore.GREEN if source.exists and source.file_count > 0 else Fore.RED
                        print(f"   {color}{status}{Style.RESET_ALL} {source.name}", end="")
                    else:
                        print(f"   {status} {source.name}", end="")

                    if source.exists:
                        print(f" ({source.file_count} 个文件)")
                    else:
                        print(" (不存在)")

                # V4: Show IFC(Client) status
                print(f"\n[DIR] IFC(Client) 状态:")
                if p.has_existing_ifc_client:
                    ifc_path = Path(p.project_path) / self.validator.ifc_client_rel_path
                    if COLORAMA_AVAILABLE:
                        print(f"   {Fore.GREEN}[v]{Style.RESET_ALL} 存在 ({p.ifc_file_count} 个文件)")
                    else:
                        print(f"   [v] 存在 ({p.ifc_file_count} 个文件)")
                    print(f"   路径: {ifc_path}")
                else:
                    if COLORAMA_AVAILABLE:
                        print(f"   {Fore.RED}[x]{Style.RESET_ALL} 不存在")
                    else:
                        print(f"   [x] 不存在")

                print(f"\n[i] 预计操作:")
                print(f"   - 创建 {p.folders_to_create} 个文件夹")
                print(f"   - 复制 {p.drawings_count} 个图纸")
                print(f"   - 复制 {p.reports_count} 个报告")

                if p.warning_message:
                    print(f"\n[!] 警告: {p.warning_message}")

                self.ui.print_separator("=", 72)
                input("\n按 Enter 返回...")
        except:
            pass

    def _generate_and_save_report(self, projects_by_region: Dict[str, List[ProjectValidation]]):
        """Generate and save validation report."""
        report = self.report_generator.generate_report(
            projects_by_region,
            str(self.root_path)
        )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = Path(__file__).parent / "logs" / f"validation_report_{timestamp}.txt"
        report_path.parent.mkdir(exist_ok=True)

        self.report_generator.save_report(report, report_path)
        self.ui.print_success(f"报告已保存: {report_path}")

        if self.ui.confirm("是否打开报告？"):
            if os.name == 'nt':
                os.startfile(report_path)
            else:
                os.system(f'open "{report_path}"')

    def select_operation(self) -> Optional[str]:
        """Let user select operation type."""
        print("\n选择要执行的操作：")
        options = [
            ("1", "仅创建文件夹"),
            ("2", "仅镜像文件"),
            ("3", "全部执行 (推荐)"),
            ("4", "预览模式 (dry-run)"),
            ("5", "复制 IFC(Client) 到目标文件夹"),
            ("0", "返回"),
        ]

        choice = self.ui.print_menu(options, "请选择操作")

        operation_map = {
            '1': 'create_folders',
            '2': 'mirror_files',
            '3': 'full',
            '4': 'dry_run',
            '5': 'ifc_copy',
            '0': None
        }

        return operation_map.get(choice)

    # =========================================================================
    # V4: IFC(Client) Copy Mode
    # =========================================================================

    def ifc_copy_mode(self):
        """Standalone IFC(Client) copy mode from main menu."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)

        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        # Filter to only show projects that have IFC(Client)
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        ifc_projects = [p for p in all_projects if p.has_existing_ifc_client]

        if not ifc_projects:
            self.ui.print_warning("没有找到包含 4. IFC(Client) 文件夹的项目！")
            input("\n按 Enter 返回...")
            return

        # Display IFC projects
        self.ui.print_header(
            f"IFC(Client) 复制 - 找到 {len(ifc_projects)} 个项目",
            "选择要复制 IFC(Client) 的项目"
        )

        for idx, p in enumerate(ifc_projects, 1):
            print(f"  [{idx}] {p.project_name}")
            print(f"      IFC(Client): {p.ifc_file_count} 个文件")
            print(f"      路径: {p.project_path}")
            print()

        self.ui.print_separator()

        # Select projects
        print("\n输入项目编号（多选用逗号分隔，如: 1,3,4 或 1-5 或 'all'，0 返回）：")
        selection = input("> ").strip()

        if selection == '0':
            return

        indices = self._parse_selection(selection, len(ifc_projects))
        if not indices:
            self.ui.print_warning("未选择任何项目")
            return

        selected = [ifc_projects[i-1] for i in indices]
        self.process_ifc_copy(selected)

    def process_ifc_copy(self, projects: List[ProjectValidation]):
        """Process IFC(Client) copy for selected projects.

        For each project, ask user for target folder, then copy.
        """
        # Filter projects that actually have IFC(Client)
        valid_projects = []
        for p in projects:
            if p.has_existing_ifc_client:
                valid_projects.append(p)
            else:
                self.ui.print_warning(f"跳过 {p.project_name}: 没有 4. IFC(Client) 文件夹")

        if not valid_projects:
            self.ui.print_warning("没有可复制的项目（所有项目均缺少 IFC(Client) 文件夹）")
            input("\n按 Enter 返回...")
            return

        results = []

        for idx, project in enumerate(valid_projects, 1):
            print(f"\n{'='*72}")
            print(f"IFC(Client) 复制 - 项目 {idx}/{len(valid_projects)}")
            print(f"{'='*72}\n")

            print(f"[i] 项目: {project.project_name}")
            ifc_source = Path(project.project_path) / self.validator.ifc_client_rel_path
            print(f"[DIR] IFC 源目录: {ifc_source}")
            print(f"[i] 文件数量: {project.ifc_file_count} 个")

            # Ask for target path
            print(f"\n请输入目标文件夹路径（可拖拽文件夹，输入 0 跳过此项目）：")
            target_str = input("> ").strip().strip('"')

            if not target_str or target_str == '0':
                results.append(ProcessResult(
                    project_name=project.project_name,
                    project_path=project.project_path,
                    success=False,
                    skipped=True,
                    skip_reason="用户跳过"
                ))
                print("[->] 跳过此项目")
                continue

            target_path = Path(target_str)

            # Validate target path (parent must exist, or we create it)
            if not target_path.parent.exists():
                self.ui.print_error(f"目标路径的父目录不存在: {target_path.parent}")
                results.append(ProcessResult(
                    project_name=project.project_name,
                    project_path=project.project_path,
                    success=False,
                    errors=[f"目标路径父目录不存在: {target_path.parent}"]
                ))
                continue

            # Show confirmation
            print(f"\n[i] 操作确认:")
            print(f"   源: {ifc_source}")
            print(f"   目标: {target_path}")
            print(f"   文件数: {project.ifc_file_count}")
            print(f"   方式: 复制 (copy & paste)")

            if target_path.exists():
                self.ui.print_warning("目标文件夹已存在，同名文件将根据大小/日期决定是否覆盖")

            if not self.ui.confirm("\n确认开始复制？"):
                results.append(ProcessResult(
                    project_name=project.project_name,
                    project_path=project.project_path,
                    success=False,
                    skipped=True,
                    skip_reason="用户跳过"
                ))
                print("[->] 跳过此项目")
                continue

            # Execute copy
            automation = IFRAutomation(
                root_path=str(Path(project.project_path).parent),
                config=self.config,
                interactive=True
            )

            print("\n正在复制...")
            ifc_result = automation.copy_ifc_to_target(
                Path(project.project_path),
                target_path
            )

            result = ProcessResult(
                project_name=project.project_name,
                project_path=project.project_path,
                success=len(ifc_result["errors"]) == 0,
                ifc_files_copied=ifc_result["copied"],
                ifc_files_skipped=ifc_result["skipped"],
                ifc_target_path=str(target_path),
                errors=[str(e) for e in ifc_result["errors"]]
            )
            results.append(result)

            if result.success:
                self.ui.print_success(
                    f"完成: {project.project_name} -> "
                    f"复制 {ifc_result['copied']} 个文件, "
                    f"跳过 {ifc_result['skipped']} 个 (已存在)"
                )
            else:
                self.ui.print_warning(
                    f"部分完成: {project.project_name} -> "
                    f"复制 {ifc_result['copied']} 个, "
                    f"跳过 {ifc_result['skipped']} 个, "
                    f"错误 {len(ifc_result['errors'])} 个"
                )
                for err in ifc_result["errors"][:5]:
                    print(f"   错误: {err}")

        # Show IFC copy summary
        self.show_ifc_copy_summary(results)

    def show_ifc_copy_summary(self, results: List[ProcessResult]):
        """Show IFC copy completion summary."""
        print()
        self.ui.print_separator("=", 72)
        self.ui.print_header("IFC(Client) 复制完成", "")

        success = sum(1 for r in results if r.success)
        skipped = sum(1 for r in results if r.skipped)
        failed = len(results) - success - skipped

        total_copied = sum(r.ifc_files_copied for r in results)
        total_skipped_files = sum(r.ifc_files_skipped for r in results)

        print(f"  项目统计:")
        print(f"    成功: {success}")
        print(f"    跳过: {skipped}")
        print(f"    失败: {failed}")

        print(f"\n  文件统计:")
        print(f"    复制: {total_copied} 个文件")
        print(f"    跳过 (已存在): {total_skipped_files} 个文件")

        if success > 0:
            print(f"\n  复制详情:")
            for r in results:
                if r.success:
                    print(f"    {r.project_name}")
                    print(f"      -> {r.ifc_target_path}")
                    print(f"      复制: {r.ifc_files_copied}, 跳过: {r.ifc_files_skipped}")

        self.ui.print_separator()
        input("\n按 Enter 返回...")

    def ifc_convert_mode(self):
        """IFR → IFC 转换模式 (AutoCAD COM)."""
        if not WIN32COM_AVAILABLE:
            self.ui.print_error("需要安装 pywin32 才能使用此功能 (pip install pywin32)")
            input("\n按 Enter 返回...")
            return

        root = self.get_root_path()
        if not root:
            return

        # Scan projects
        print("\n正在扫描项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)
        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        all_projects = [p for projects in projects_by_region.values() for p in projects]

        # Filter to projects that have Native drawings folder
        native_projects = []
        for p in all_projects:
            native_dir = Path(p.project_path) / IFCManager.NATIVE_ROOT
            if native_dir.exists():
                native_projects.append(p)

        if not native_projects:
            self.ui.print_warning("没有找到包含 1. Native 文件夹的项目！")
            input("\n按 Enter 返回...")
            return

        # Display and select project
        self.ui.print_header(
            f"IFR → IFC 转换 - 找到 {len(native_projects)} 个项目",
            "选择要转换的项目"
        )

        for idx, p in enumerate(native_projects, 1):
            print(f"  [{idx}] {p.project_name}")
            print(f"      路径: {p.project_path}")

        print(f"\n  [0] 返回主菜单")
        self.ui.print_separator()

        choice = input(f"\n请选择项目 [0-{len(native_projects)}]: ").strip()
        if choice == '0':
            return

        try:
            idx = int(choice)
            if not (1 <= idx <= len(native_projects)):
                return
        except ValueError:
            return

        project = native_projects[idx - 1]
        project_path = Path(project.project_path)

        # Scan native folders for IFR DWGs
        print(f"\n正在扫描 Native 文件夹: {project.project_name}")
        mgr = IFCManager(project_path, dry_run=True)
        scan_results = mgr.scan_native_folders()

        if not scan_results:
            self.ui.print_warning("未找到可转换的 IFR DWG 文件")
            input("\n按 Enter 返回...")
            return

        # Display scan results
        print(f"\n{'='*60}")
        print(f"  找到 {len(scan_results)} 个 doc-ID 的 IFR DWG 文件")
        print(f"{'='*60}")

        for i, info in enumerate(scan_results, 1):
            ifc_status = f"已有 IFC Rev{info['existing_ifc_rev']}" if info['existing_ifc_rev'] is not None else "无 IFC"
            print(f"  [{i:3d}] {info['doc_id']}")
            print(f"        IFR: {info['latest_ifr_dwg'].name} (Rev{info['latest_ifr_rev']})")
            print(f"        IFC: {ifc_status}")

        self.ui.print_separator()

        # Preview (dry-run)
        print(f"\n  预览转换结果（dry-run）:")
        preview_results = []
        for info in scan_results:
            r = mgr.convert_to_ifc(info)
            preview_results.append(r)
            ifc_rev = r['ifc_rev']
            dwg_name = Path(r['dwg_path']).name if r['dwg_path'] else '?'
            pdf_name = Path(r['pdf_path']).name if r['pdf_path'] else '?'
            print(f"    {info['doc_id']}: → IFC Rev{ifc_rev}")
            print(f"      DWG: {dwg_name}")
            print(f"      PDF: {pdf_name}")

        self.ui.print_separator()

        # Confirm
        print("\n选择操作:")
        print("  [1] 执行转换（批量全部转）")
        print("  [2] 预览模式（仅显示，不执行）")
        print("  [0] 返回（不执行）")

        action = input("\n请选择 [0-2]: ").strip()

        if action == '0':
            return
        elif action == '2':
            self.ui.print_success("预览完成，未执行任何操作")
            input("\n按 Enter 返回...")
            return

        if action != '1':
            return

        # Execute conversion
        print(f"\n正在启动 AutoCAD 转换...")
        real_mgr = IFCManager(project_path, dry_run=False)
        results = real_mgr.batch_convert()

        # Summary
        ok_count = sum(1 for r in results if r['success'])
        fail_count = len(results) - ok_count

        print(f"\n{'='*60}")
        print(f"  IFC 转换完成")
        print(f"  成功: {ok_count}, 失败: {fail_count}")
        print(f"{'='*60}")

        if fail_count > 0:
            print("\n  失败详情:")
            for r in results:
                if not r['success']:
                    print(f"    {r['doc_id']}: {'; '.join(r['errors'])}")

        # Trigger deliverable sync using IFCManager's dedicated method
        if ok_count > 0:
            print("\n正在更新交付物状态...")
            real_mgr.update_ifc_deliverables(results)

        input("\n按 Enter 返回...")

    def panel_ifc_convert_mode(self):
        """Panel IFC 批量转换模式 (多页 DWG → 合并 PDF)."""
        if not WIN32COM_AVAILABLE:
            self.ui.print_error("需要安装 pywin32 才能使用此功能 (pip install pywin32)")
            input("\n按 Enter 返回...")
            return

        self.ui.print_header(
            "Panel IFC 批量转换",
            "多页 DWG Title Block 更新 + 合并 PDF 导出"
        )

        # 1. Get source folder
        print("请输入 Panel Design 文件夹路径 (包含多个 DWG 的平面目录):")
        print("  示例: .../1. Native/STS1&2 Panel Design/CAD/Project-TSF")
        folder_str = input("\n> ").strip().strip('"')
        if not folder_str or not Path(folder_str).exists():
            self.ui.print_warning("路径不存在")
            input("\n按 Enter 返回...")
            return

        source_folder = Path(folder_str)

        # 2. Scan and group
        print(f"\n正在扫描: {source_folder.name}")
        preview_mgr = PanelIFCManager(source_folder, dry_run=True)
        groups = preview_mgr.scan_and_group()

        if not groups:
            self.ui.print_warning("未找到可分组的 DWG 文件 (需要 {DocNO}-{PageNum}.dwg 格式)")
            input("\n按 Enter 返回...")
            return

        # Display groups
        print(f"\n{'='*60}")
        print(f"  找到 {len(groups)} 个文档组:")
        print(f"{'='*60}")

        for doc_no, pages in groups.items():
            ifc_rev = preview_mgr._get_existing_ifc_rev(doc_no)
            ifc_label = f"Rev{ifc_rev}" if ifc_rev == 0 else f"Rev{ifc_rev} (已有 Rev{ifc_rev-1})"
            print(f"\n  [{doc_no}] {len(pages)} 页 -> IFC {ifc_label}")
            for p in pages:
                page_type = ""
                if p['page'] == '00':
                    page_type = " (Cover)"
                elif p['page'].startswith('0') and not p['page'].isdigit():
                    page_type = " (TOC/Appendix)"
                print(f"    {p['filename']}{page_type}")
            pdf_name = f"{doc_no}_Rev{ifc_rev}_IFC.pdf"
            print(f"    -> PDF: {pdf_name}")

        self.ui.print_separator()

        # 3. Select title block update method
        print("\n选择 Title Block 更新方式:")
        print("  [1] COM API 直接修改 (推荐)")
        print("  [2] Lee-Mac UTB 插件 (需要 .lsp 文件)")
        print("  [0] 返回")
        method_choice = input("\n请选择 [0-2]: ").strip()

        if method_choice == '0':
            return
        elif method_choice == '2':
            tb_method = 'utb'
            print("\n请输入 UTB .lsp 文件路径:")
            lsp_str = input("> ").strip().strip('"')
            if not lsp_str or not Path(lsp_str).exists():
                self.ui.print_warning(".lsp 文件不存在")
                input("\n按 Enter 返回...")
                return
            utb_lsp_path = lsp_str
        else:
            tb_method = 'com'
            utb_lsp_path = None

        # 4. Dry-run preview
        print(f"\n{'='*60}")
        print(f"  预览 (dry-run) — 使用 {tb_method.upper()} 模式")
        print(f"{'='*60}")

        dry_mgr = PanelIFCManager(
            source_folder, dry_run=True,
            title_block_method=tb_method,
            utb_lsp_path=utb_lsp_path,
        )
        dry_mgr.batch_convert_panel()

        # 5. Confirm
        print("\n选择操作:")
        print("  [1] 执行转换")
        print("  [0] 返回（不执行）")
        action = input("\n请选择 [0-1]: ").strip()

        if action != '1':
            return

        # 6. Execute
        print(f"\n正在执行 IFC 转换...")
        real_mgr = PanelIFCManager(
            source_folder, dry_run=False,
            title_block_method=tb_method,
            utb_lsp_path=utb_lsp_path,
        )
        results = real_mgr.batch_convert_panel()

        # 7. Summary
        total_groups = len(results)
        ok_groups = sum(1 for r in results if r.get('pdf_result', {}).get('success'))
        total_dwgs = sum(r['dwgs_updated'] for r in results)

        print(f"\n{'='*60}")
        print(f"  Panel IFC 转换完成")
        print(f"  文档组: {ok_groups}/{total_groups} 成功")
        print(f"  DWG 更新: {total_dwgs} 个")
        print(f"{'='*60}")

        if any(r['errors'] for r in results):
            print("\n  错误详情:")
            for r in results:
                for err in r['errors']:
                    print(f"    [{r['doc_no']}] {err}")

        input("\n按 Enter 返回...")

    def process_projects_with_confirmation(self, projects: List[ProjectValidation], operation: str):
        """Process projects with per-project confirmation."""
        dry_run = operation == 'dry_run'
        create_folders_only = operation == 'create_folders'
        mirror_files_only = operation == 'mirror_files'

        results = []

        for idx, project in enumerate(projects, 1):
            print(f"\n{'='*72}")
            print(f"正在处理项目 {idx}/{len(projects)}")
            print(f"{'='*72}\n")

            # Display project details
            print(f"[i] 项目: {project.project_name}")
            print(f"[DIR] 路径: {project.project_path}")
            print(f"[i] 最后修改: {project.last_modified}")

            # Safety check
            safety = self.safety_checker.check_before_process(project)

            if not safety["safe"]:
                print(f"\n[!] 安全检查警告:")
                for warning in safety["warnings"]:
                    if COLORAMA_AVAILABLE:
                        print(f"   {Fore.YELLOW}- {warning}{Style.RESET_ALL}")
                    else:
                        print(f"   - {warning}")

            # Show operation preview
            print(f"\n[i] 操作预览:")
            if not mirror_files_only:
                print(f"   - 创建文件夹: {project.folders_to_create} 个")
            if not create_folders_only:
                print(f"   - 复制图纸: {project.drawings_count} 个文件")
                print(f"   - 复制报告: {project.reports_count} 个文件")

            if dry_run:
                print(f"\n   [i] 预览模式 - 不会执行实际操作")

            # Confirm if needed
            if project.recommended_action != "process" or not safety["safe"]:
                if not self.ui.confirm("\n是否继续处理此项目？", default=False):
                    results.append(ProcessResult(
                        project_name=project.project_name,
                        project_path=project.project_path,
                        success=False,
                        skipped=True,
                        skip_reason="用户跳过"
                    ))
                    print("[->] 跳过此项目")
                    continue
            else:
                if not self.ui.confirm("\n开始处理？"):
                    results.append(ProcessResult(
                        project_name=project.project_name,
                        project_path=project.project_path,
                        success=False,
                        skipped=True,
                        skip_reason="用户跳过"
                    ))
                    print("[->] 跳过此项目")
                    continue

            # Execute
            automation = IFRAutomation(
                root_path=str(Path(project.project_path).parent),
                config=self.config,
                dry_run=dry_run,
                create_folders_only=create_folders_only,
                mirror_files_only=mirror_files_only,
                interactive=True
            )

            result = automation.process_project(project)
            results.append(result)

            if result.success:
                self.ui.print_success(f"完成: {project.project_name}")
                print(f"   文件夹: {result.folders_created}, 图纸: {result.drawings_copied}, 报告: {result.reports_copied}")
                self.config.add_recent_project(project.project_path)
            else:
                self.ui.print_error(f"失败: {project.project_name}")
                for error in result.errors:
                    print(f"   错误: {error}")

        # Show summary
        self.show_completion_summary(results, dry_run)

    def show_completion_summary(self, results: List[ProcessResult], dry_run: bool):
        """Show processing summary."""
        print()
        self.ui.print_separator("=", 72)

        mode = "预览模式" if dry_run else "执行完成"
        self.ui.print_header(f"操作{mode}", "")

        success = sum(1 for r in results if r.success)
        skipped = sum(1 for r in results if r.skipped)
        failed = len(results) - success - skipped

        total_folders = sum(r.folders_created for r in results)
        total_drawings = sum(r.drawings_copied for r in results)
        total_reports = sum(r.reports_copied for r in results)

        print(f"  项目统计:")
        print(f"    成功: {success}")
        print(f"    跳过: {skipped}")
        print(f"    失败: {failed}")

        print(f"\n  操作统计:")
        print(f"    文件夹创建: {total_folders}")
        print(f"    图纸复制: {total_drawings}")
        print(f"    报告复制: {total_reports}")

        self.ui.print_separator()

        options = [
            ("1", "打开日志文件"),
            ("2", "返回主菜单"),
            ("0", "退出"),
        ]

        choice = self.ui.print_menu(options, "请选择")

        if choice == '1':
            self.view_logs()
        elif choice == '0':
            sys.exit(0)

    def validate_only_mode(self):
        """Validate projects without processing."""
        root = self.get_root_path()
        if not root:
            return

        print("\n正在扫描和验证项目...")
        projects_by_region = self.scanner.scan_hierarchical(root)

        if not projects_by_region:
            self.ui.print_warning("未找到任何项目！")
            return

        # Display results
        self.display_projects_with_status(projects_by_region)

        # Generate report option
        if self.ui.confirm("\n是否生成验证报告？"):
            self._generate_and_save_report(projects_by_region)

    def input_single_project(self):
        """Handle single project input."""
        print("\n请输入项目路径：")
        print("（可以直接拖拽文件夹到此窗口）")

        path_str = input("> ").strip().strip('"')

        if not path_str:
            return

        project_path = Path(path_str)

        if not project_path.exists():
            self.ui.print_error("路径不存在！")
            return

        # Validate
        validation = self.validator.validate_project(project_path)

        print()
        self.ui.print_separator()
        print(f"[i] 项目: {validation.project_name}")
        print(f"[i] 状态: {self.ui.status_icon(validation.recommended_action)[1]}")
        print(f"[i] 建议: {self.ui.action_icon(validation.recommended_action)}")

        if validation.has_existing_ifc_client:
            print(f"[i] IFC(Client): {validation.ifc_file_count} 个文件")

        if validation.warning_message:
            self.ui.print_warning(validation.warning_message)

        self.ui.print_separator()

        if not self.ui.confirm("\n是否处理此项目？"):
            return

        operation = self.select_operation()
        if operation is None:
            return

        if operation == 'ifc_copy':
            self.process_ifc_copy([validation])
        else:
            self.process_projects_with_confirmation([validation], operation)

    def modify_config(self):
        """Modify configuration and view logs."""
        while True:
            print("\n当前配置：")
            print(f"  1. 默认根目录: {self.config.get('default_root_path', '未设置')}")
            print(f"  2. 日志级别: {self.config.get('log_level', 'INFO')}")
            print(f"  3. 白名单: {self.config.get('project_filters.whitelist', [])}")
            print(f"  4. 黑名单: {self.config.get('project_filters.blacklist', [])}")
            print(f"  5. 最少源目录数: {self.config.get('validation_rules.min_source_dirs', 1)}")
            print(f"  6. 最大未修改天数: {self.config.get('validation_rules.max_days_since_modified', 180)}")
            print(f"  7. IFC(Client) 路径: {self.config.get('ifc_client_path', ProjectValidator.IFC_CLIENT_REL_PATH)}")
            print(f"  8. 查看日志")
            print(f"  0. 返回主菜单")

            choice = input("\n选择配置项 [0-8]: ").strip()

            if choice == '0':
                break
            elif choice == '1':
                new_path = input("输入新的根目录路径: ").strip().strip('"')
                if new_path and Path(new_path).exists():
                    self.config.set("default_root_path", new_path)
                    self.root_path = Path(new_path)
                    self.ui.print_success("已更新根目录")
                else:
                    self.ui.print_error("路径无效")
            elif choice == '2':
                levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR']
                print("可选级别: " + ", ".join(levels))
                level = input("输入日志级别: ").strip().upper()
                if level in levels:
                    self.config.set("log_level", level)
                    self.ui.print_success("已更新日志级别")
            elif choice == '3':
                print("输入白名单项目（逗号分隔，支持通配符如 NSW*）：")
                whitelist = input("> ").strip()
                if whitelist:
                    items = [x.strip() for x in whitelist.split(',')]
                    self.config.set("project_filters.whitelist", items)
                else:
                    self.config.set("project_filters.whitelist", [])
                self.ui.print_success("已更新白名单")
            elif choice == '4':
                print("输入黑名单项目（逗号分隔，支持通配符如 Old_*）：")
                blacklist = input("> ").strip()
                if blacklist:
                    items = [x.strip() for x in blacklist.split(',')]
                    self.config.set("project_filters.blacklist", items)
                else:
                    self.config.set("project_filters.blacklist", [])
                self.ui.print_success("已更新黑名单")
            elif choice == '5':
                try:
                    num = int(input("输入最少源目录数 (1-4): ").strip())
                    if 1 <= num <= 4:
                        self.config.set("validation_rules.min_source_dirs", num)
                        self.ui.print_success("已更新")
                except:
                    self.ui.print_error("无效输入")
            elif choice == '6':
                try:
                    num = int(input("输入最大未修改天数: ").strip())
                    if num > 0:
                        self.config.set("validation_rules.max_days_since_modified", num)
                        self.ui.print_success("已更新")
                except:
                    self.ui.print_error("无效输入")
            elif choice == '7':
                print(f"当前 IFC(Client) 路径: {self.config.get('ifc_client_path', ProjectValidator.IFC_CLIENT_REL_PATH)}")
                print("输入新的相对路径（相对于项目根目录，留空保持默认）：")
                new_path = input("> ").strip()
                if new_path:
                    self.config.set("ifc_client_path", new_path)
                    self.validator = ProjectValidator(self.config)
                    self.scanner = ProjectScanner(self.config, self.validator)
                    self.ui.print_success("已更新 IFC(Client) 路径")
            elif choice == '8':
                self.view_logs()

    def view_logs(self):
        """View log files."""
        log_dir = Path(__file__).parent / "logs"

        if not log_dir.exists():
            self.ui.print_warning("日志目录不存在")
            return

        log_files = sorted(log_dir.glob("ifr_automation_*.log"), reverse=True)

        if not log_files:
            self.ui.print_warning("没有找到日志文件")
            return

        print("\n最近的日志文件：")
        for i, log_file in enumerate(log_files[:5], 1):
            mtime = datetime.fromtimestamp(log_file.stat().st_mtime)
            print(f"  {i}. {log_file.name} ({mtime.strftime('%Y-%m-%d %H:%M')})")

        print("\n选择要查看的日志（输入序号，或 0 返回）：")
        choice = input("> ").strip()

        try:
            idx = int(choice)
            if 1 <= idx <= len(log_files[:5]):
                log_file = log_files[idx - 1]
                if os.name == 'nt':
                    os.startfile(log_file)
                else:
                    os.system(f'open "{log_file}"')
        except:
            pass


# =============================================================================
# Command Line Interface
# =============================================================================

def main():
    """Main entry point."""

    # Interactive mode if no arguments
    if len(sys.argv) == 1:
        config = ConfigManager()
        interactive = UnifiedInteractiveMode(config)
        interactive.run()
        return

    # Command line mode
    parser = argparse.ArgumentParser(
        description="Engineering Pipeline v7.0 - IFR Sync + Version Mgmt + Deliverable",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode
  python ifr_automation_v7.py

  # Full pipeline (IFR sync + version mgmt + deliverable)
  python ifr_automation_v7.py --pipeline --root "..."

  # Specific stages only
  python ifr_automation_v7.py --stages ifr_sync version_mgmt --root "..."

  # Deliverable cross-check only (dry-run)
  python ifr_automation_v7.py --deliverable-only --dry-run --root "..."

  # Deliverable check only (report, no update)
  python ifr_automation_v7.py --deliverable-check-only --root "..."

  # IFR sync only
  python ifr_automation_v7.py --root "..." --auto-safe-only

  # Version management (PDF)
  python ifr_automation_v7.py --version-mgmt --root "..."

  # Native/Reports/Schedule version management
  python ifr_automation_v7.py --native --root "..." --project "GG-31*"
        """
    )

    parser.add_argument('--root', help='Root directory containing project folders')
    parser.add_argument('--dry-run', action='store_true', help='Preview mode (no changes)')
    parser.add_argument('--create-folders', action='store_true', help='Only create folder structure')
    parser.add_argument('--mirror-files', action='store_true', help='Only mirror files')
    parser.add_argument('--project', help='Process specific project (partial match)')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'], default='INFO')
    parser.add_argument('--output-json', action='store_true', help='Generate JSON report')

    # V3 arguments
    parser.add_argument('--validate-only', action='store_true', help='Only validate, no processing')
    parser.add_argument('--whitelist', nargs='+', help='Only process matching projects')
    parser.add_argument('--blacklist', nargs='+', help='Exclude matching projects')
    parser.add_argument('--auto-safe-only', action='store_true', help='Auto-process only safe projects')
    parser.add_argument('--yes', '-y', action='store_true', dest='yes_to_all', help='Yes to all prompts')
    parser.add_argument('--export-report', type=str, metavar='PATH', help='Export validation report')

    # V4 arguments
    parser.add_argument('--ifc-copy-target', type=str, metavar='PATH',
                       help='Copy IFC(Client) to target folder')

    # V7 arguments
    parser.add_argument('--pipeline', action='store_true',
                       help='Run full pipeline (IFR sync + version mgmt + deliverable)')
    parser.add_argument('--stages', nargs='+',
                       choices=['ifr_sync', 'version_mgmt', 'deliverable'],
                       help='Run specific pipeline stages')
    parser.add_argument('--deliverable-only', action='store_true',
                       help='Run deliverable cross-check and update')
    parser.add_argument('--deliverable-check-only', action='store_true',
                       help='Run deliverable cross-check (report only, no update)')
    parser.add_argument('--version-mgmt', action='store_true',
                       help='Run PDF version management')
    parser.add_argument('--native', action='store_true',
                       help='Run Native/Reports/Schedule version management')
    parser.add_argument('--scope', choices=['native', 'reports', 'schedule', 'all'],
                       default='all', help='Scope for --native mode')
    parser.add_argument('--folder', type=str, help='Folder filter for --native mode')
    parser.add_argument('--execute', action='store_true',
                       help='Execute operations (for --native/--version-mgmt)')

    args = parser.parse_args()

    # Load config
    config = ConfigManager()

    if args.whitelist:
        config.set("project_filters.whitelist", args.whitelist)
    if args.blacklist:
        config.set("project_filters.blacklist", args.blacklist)

    root_path = args.root or config.get("default_root_path")
    if not root_path:
        print("Error: No root path specified. Use --root or set in config.")
        sys.exit(1)
    if not Path(root_path).exists():
        print(f"Error: Root path does not exist: {root_path}")
        sys.exit(1)

    # V7: Native mode (from version_manager_v4)
    if args.native:
        project_path = args.project
        if not project_path:
            # Try to find project from recent
            cfg_file = Path(__file__).parent / 'config.json'
            if cfg_file.exists():
                try:
                    with open(cfg_file, 'r', encoding='utf-8') as f:
                        cfg = json.load(f)
                    recent = cfg.get('recent_projects', [])
                    if recent:
                        project_path = recent[0]
                except Exception:
                    pass
        if not project_path:
            print("Error: --native requires --project")
            sys.exit(1)
        project_path = Path(project_path)
        if not project_path.exists():
            print(f"Error: Path not found: {project_path}")
            sys.exit(1)

        dry_run = not args.execute
        mgr = NativeVersionManager(str(project_path), dry_run=dry_run)
        results = mgr.process_all(folder_filter=args.folder, scope=args.scope)
        if not results:
            print("No document folders found")
            sys.exit(0)
        total_actions = sum(len(r.actions) for _, r in results)
        for group_name, group_results in groupby(results, key=lambda x: x[0]):
            group_list = list(group_results)
            print(f"\n  === {group_name} ===")
            for _, result in group_list:
                if not result.dwg_files:
                    continue
                print(f"    {result.folder_name} ({len(result.dwg_files)} files)")
                for action in result.actions:
                    if action.action == 'rename':
                        print(f"      [重命名] {action.source.name} -> {action.dest.name}")
                    else:
                        print(f"      [->SS] {action.source.name}")
        if not dry_run and total_actions > 0:
            stats = mgr.execute_actions(results)
            print(f"\n  Done: {stats['renamed']} renamed, {stats['moved']} moved")
        elif total_actions > 0:
            print(f"\n  Use --execute to apply {total_actions} actions")
        sys.exit(0)

    # V7: Version management mode
    if args.version_mgmt:
        vm = VersionManager(root_path, dry_run=not args.execute)
        projects = vm.scan_for_projects()
        if args.project:
            projects = [p for p in projects if args.project.lower() in p.name.lower()]
        for project in projects:
            print(f"\nProcessing: {project.name}")
            vm.process_project(project)
        sys.exit(0)

    # V7: Deliverable modes
    if args.deliverable_only or args.deliverable_check_only:
        automation = IFRAutomation(root_path=root_path, config=config, interactive=True)
        projects_by_region = automation.scan_projects()
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        if args.project:
            all_projects = [p for p in all_projects if args.project.lower() in p.project_name.lower()]
        tg_lines = []
        for project in all_projects:
            dm = DeliverableManager(Path(project.project_path), dry_run=args.dry_run)
            excel = dm.find_deliverable_excel()
            if not excel:
                print(f"[SKIP] {project.project_name}: No deliverable Excel found")
                continue
            check = dm.cross_check(excel)
            print(f"\n{project.project_name}: "
                  f"new={len(check.items_in_folders_not_excel)}, "
                  f"rev_mismatch={len(check.revision_mismatches)}")
            if check.errors:
                for e in check.errors:
                    print(f"  ERROR: {e}")
            if args.deliverable_only and not args.dry_run:
                if check.items_in_folders_not_excel or check.revision_mismatches:
                    updated = dm.apply_updates(excel, check)
                    print(f"  Updated: inserted={updated.rows_inserted}, "
                          f"updated={updated.rows_updated}, new_rev={updated.new_file_rev}")
                    tg_lines.append(
                        f"📋 <b>{project.project_name}</b>\n"
                        f"  新增={len(check.items_in_folders_not_excel)}, "
                        f"更新={len(check.revision_mismatches)}, "
                        f"inserted={updated.rows_inserted}, updated={updated.rows_updated}")
        if tg_lines:
            send_telegram_notification("📋 <b>Deliverable Update</b>\n\n" + "\n\n".join(tg_lines))
        sys.exit(0)

    # V7: Pipeline mode
    if args.pipeline or args.stages:
        stages = args.stages or ['ifr_sync', 'version_mgmt', 'deliverable']
        automation = IFRAutomation(root_path=root_path, config=config, interactive=True)
        projects_by_region = automation.scan_projects()
        all_projects = [p for projects in projects_by_region.values() for p in projects]
        if args.auto_safe_only:
            all_projects = [p for p in all_projects if p.recommended_action == "process"]
        if args.project:
            all_projects = [p for p in all_projects if args.project.lower() in p.project_name.lower()]
        pipeline = PipelineOrchestrator(config, dry_run=args.dry_run, stages=stages)
        all_results = []
        for project in all_projects:
            result = pipeline.run_pipeline(Path(project.project_path), project)
            all_results.append(result)
            status = "OK" if result['success'] else "FAIL"
            print(f"[{status}] {project.project_name}")
        # Send Telegram notification
        if all_results and not args.dry_run:
            tg_lines = [format_pipeline_result(r) for r in all_results]
            send_telegram_notification("\n\n".join(tg_lines))
        sys.exit(0)

    # Default: IFR sync mode (v6 behavior)
    automation = IFRAutomation(
        root_path=root_path, config=config,
        dry_run=args.dry_run, create_folders_only=args.create_folders,
        mirror_files_only=args.mirror_files, validate_only=args.validate_only,
        yes_to_all=args.yes_to_all, auto_safe_only=args.auto_safe_only,
        log_level=args.log_level, output_json=args.output_json
    )

    projects_by_region = automation.scan_projects()
    if not projects_by_region:
        print("No projects found!")
        sys.exit(0)

    if args.export_report or args.validate_only:
        report = automation.report_generator.generate_report(projects_by_region, root_path)
        print(report)
        if args.export_report:
            automation.report_generator.save_report(report, Path(args.export_report))
        if args.validate_only:
            sys.exit(0)

    all_projects = [p for projects in projects_by_region.values() for p in projects]
    if args.auto_safe_only:
        all_projects = [p for p in all_projects if p.recommended_action == "process"]
    if args.project:
        all_projects = [p for p in all_projects if args.project.lower() in p.project_name.lower()]

    if args.ifc_copy_target:
        target = Path(args.ifc_copy_target)
        for project in all_projects:
            if not project.has_existing_ifc_client:
                continue
            ifc_result = automation.copy_ifc_to_target(Path(project.project_path), target)
            status = "OK" if len(ifc_result["errors"]) == 0 else "WARN"
            print(f"[{status}] {project.project_name}: copied={ifc_result['copied']}")
        sys.exit(0)

    for project in all_projects:
        if not args.yes_to_all and project.recommended_action != "process":
            response = input(f"Process {project.project_name}? [y/N]: ").strip().lower()
            if response not in ('y', 'yes'):
                continue
        result = automation.process_project(project)
        status = "OK" if result.success else "FAIL"
        print(f"[{status}] {project.project_name}: folders={result.folders_created}, "
              f"drawings={result.drawings_copied}, reports={result.reports_copied}")


def keep_window_open():
    """Keep console window open."""
    print()
    input("按 Enter 键退出...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n用户中断操作")
    except Exception as e:
        print(f"\n发生错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if len(sys.argv) == 1:
            keep_window_open()
